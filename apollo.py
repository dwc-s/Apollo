"""Apollo — archery session tracker.

Flask app for logging arrow shots against a target image. A "session" is a
practice round; a "quiver" is a fixed batch of arrows within a session
(e.g. 6 arrows shot, walk to the target, repeat). Shot coordinates are
stored as physical millimeters from target center (+X right, +Y up); the
sentinel value 100000 in *both* x_coord and y_coord marks "missed target."

Storage runs through SQLAlchemy Core so the same code works against
SQLite (local default — zero setup, `python apollo.py` Just Works) or
MySQL (web/cloud — set DATABASE_URL=mysql+pymysql://user:pass@host/db).
Schema is defined once via Core Table objects and compiled to the right
dialect by metadata.create_all() at import time.
"""
import base64
import csv
import hashlib
import json
import io
import math
import secrets
import time
import zipfile
from contextlib import closing
from functools import wraps
import os
import re
import uuid
from urllib.parse import urlparse

# Local pure-Python modules (no Flask/DB) for the AGB handicap math and the
# classification resolvers. Kept standalone so they stay unit-testable.
#
# Ensure this file's own directory is importable for these sibling modules.
# When apollo.py is run as a script (``python apollo.py``) Python adds its
# directory to sys.path automatically, but when it's imported as a module
# under a WSGI server (``from apollo import app``) that doesn't happen — the
# server's sys.path may locate apollo.py without its directory being importable
# for siblings. Inserting it here is idempotent and fixes the deploy import.
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import handicap
import classifications
import form_checkpoints

# Ordered choice lists for the archer-profile form (classification category).
ARCHER_GENDERS = ['male', 'female', 'open']
ARCHER_AGE_GROUPS = ['adult', '50+', 'under 21', 'under 18', 'under 16',
                     'under 15', 'under 14', 'under 12']
ARCHER_BOWSTYLES = ['recurve', 'compound', 'barebow', 'longbow',
                    'traditional', 'flatbow']

# Display labels for bowstyles. Stored values stay the short lowercase enums
# above (DB values, BOWSTYLE_SETTINGS keys, classification categories); only the
# text shown to the archer differs. Anything not listed falls back to a plain
# capitalize in the ``bowstyle_label`` Jinja filter.
BOWSTYLE_LABELS = {
    'recurve':     'Olympic recurve',
    'barebow':     'Barebow recurve',
    'compound':    'Compound',
    'longbow':     'Longbow',
    'traditional': 'Traditional',
    'flatbow':     'Flatbow',
}

# Per-bowstyle gear settings. Single source of truth for the bowtype-specific
# options shown on the session and tournament forms — the templates render from
# it, the POST handlers validate against it, and the analyze head-to-head report
# buckets shots by the ``select`` fields here.
#
# Each field is a dict:
#   key     — form field name + JSON key (namespaced ``style_<key>`` in forms)
#   label   — human label
#   kind    — 'select' | 'number' | 'text'
#   options — list of allowed values (select only); first is the implicit blank-
#             beating default in the UI but never auto-applied server-side
#   scope   — 'dynamic' → tuned per outing; entered in session/tournament, never
#                          a property of the bow record. (All gear is dynamic
#                          now; the key is retained for the cleaner/JSON paths.)
#   help    — optional tooltip text
#
# longbow / traditional / flatbow share one field set (aiming method + brace
# height; arrow material is inherited from the chosen arrow).
_TRAD_FIELDS = [
    {'key': 'aim_method', 'label': 'Aiming method', 'kind': 'select',
     'scope': 'dynamic', 'options': ['instinctive', 'gap']},
    {'key': 'brace_height', 'label': 'Brace height (mm)', 'kind': 'number',
     'scope': 'dynamic'},
]
BOWSTYLE_SETTINGS = {
    'compound': [
        {'key': 'release_aid', 'label': 'Release aid', 'kind': 'select',
         'scope': 'dynamic',
         'options': ['index', 'thumb-trigger', 'hinge', 'back-tension',
                     'resistance']},
        {'key': 'let_off_pct', 'label': 'Let-off (%)', 'kind': 'number',
         'scope': 'dynamic'},
        {'key': 'peep_size', 'label': 'Peep aperture (mm)', 'kind': 'text',
         'scope': 'dynamic'},
        {'key': 'scope_magnification', 'label': 'Scope magnification (×)',
         'kind': 'number', 'scope': 'dynamic'},
        {'key': 'sight_type', 'label': 'Sight type', 'kind': 'select',
         'scope': 'dynamic',
         'options': ['single-pin', 'movable', 'multi-pin']},
    ],
    'recurve': [
        {'key': 'clicker', 'label': 'Clicker', 'kind': 'select',
         'scope': 'dynamic', 'options': ['yes', 'no']},
        {'key': 'finger_protection', 'label': 'Finger protection',
         'kind': 'select', 'scope': 'dynamic', 'options': ['tab', 'glove']},
        {'key': 'sight_type', 'label': 'Sight type', 'kind': 'select',
         'scope': 'dynamic',
         'options': ['aperture', 'aperture+clarifier', 'fiber-optic']},
        {'key': 'stabilizer', 'label': 'Stabilizer', 'kind': 'select',
         'scope': 'dynamic',
         'options': ['long-rod', 'long+side', 'full-V-bar']},
        {'key': 'long_rod_length', 'label': 'Long-rod length (in)',
         'kind': 'text', 'scope': 'dynamic'},
        {'key': 'sight_aperture', 'label': 'Sight aperture (mm)',
         'kind': 'number', 'scope': 'dynamic'},
        {'key': 'plunger_tension', 'label': 'Plunger tension', 'kind': 'text',
         'scope': 'dynamic'},
        {'key': 'brace_height', 'label': 'Brace height (mm)', 'kind': 'number',
         'scope': 'dynamic'},
    ],
    'barebow': [
        {'key': 'aim_method', 'label': 'Aiming method', 'kind': 'select',
         'scope': 'dynamic',
         'options': ['string-walking', 'face-walking', 'gap', 'instinctive']},
        {'key': 'finger_protection', 'label': 'Finger protection',
         'kind': 'select', 'scope': 'dynamic', 'options': ['tab', 'glove']},
        {'key': 'riser_weights', 'label': 'Riser weights', 'kind': 'text',
         'scope': 'dynamic'},
        {'key': 'shoot_off', 'label': 'Shot off', 'kind': 'select',
         'scope': 'dynamic', 'options': ['arrow rest', 'shelf', 'hand']},
        {'key': 'string_crawl', 'label': 'String crawl', 'kind': 'text',
         'scope': 'dynamic',
         'help': 'Under-nock crawl distance / tab marks for this distance.'},
        {'key': 'plunger_tension', 'label': 'Plunger tension', 'kind': 'text',
         'scope': 'dynamic'},
    ],
    'longbow': _TRAD_FIELDS,
    'traditional': _TRAD_FIELDS,
    'flatbow': _TRAD_FIELDS,
}


def _bowstyle_settings_for(bowstyle):
    """Return the list of setting fields for a bowstyle (empty if unknown)."""
    return BOWSTYLE_SETTINGS.get((bowstyle or '').strip().lower(), [])


def _collect_style_settings(form):
    """Pull all ``style_<key>`` fields off a form into a ``{key: value}`` dict.

    Style fields are namespaced in the markup so they don't collide with other
    inputs. Validation against the bow's actual type happens later in
    _clean_style_settings, so this just harvests the raw values.
    """
    return {k[len('style_'):]: v for k, v in form.items()
            if k.startswith('style_')}


def _clean_style_settings(bowstyle, raw, scope=None):
    """Validate a raw settings dict against the schema for ``bowstyle``.

    Drops keys that aren't in the schema, rejects out-of-range ``select``
    values, and skips blanks so the stored JSON only carries set fields. When
    ``scope`` is given ('static' or 'dynamic'), only fields of that scope are
    kept — lets add/edit-bow persist just the static tier. Returns a plain
    dict (possibly empty); callers JSON-encode it for storage.
    """
    out = {}
    for field in _bowstyle_settings_for(bowstyle):
        if scope is not None and field.get('scope') != scope:
            continue
        val = raw.get(field['key'])
        if val is None:
            continue
        val = str(val).strip()
        if not val:
            continue
        if field['kind'] == 'select' and val not in field.get('options', []):
            continue
        out[field['key']] = val
    return out

# Resend is the transactional-email backend. Optional at import time so
# `python apollo.py` still launches when the package isn't installed yet
# (e.g. an older venv that predates the forgot-password feature) — the
# email helper will fall back to its dev "print to stdout" branch.
try:
    import resend  # type: ignore
except ImportError:
    resend = None
from flask import (
    Flask, render_template, request, redirect, url_for, session, Response, abort, flash,
    jsonify, g, send_from_directory,
)
from markupsafe import Markup
from flask_wtf.csrf import CSRFProtect
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError, available_timezones

# Sorted IANA timezone list, computed once at import. ~600 entries; cheap
# to ship into every authenticated page render via the context processor
# so the /account dropdown doesn't have to recompute it per request.
_TIMEZONE_CHOICES = sorted(available_timezones())
from PIL import Image
from sqlalchemy import (
    create_engine, MetaData, Table, Column, inspect as sa_inspect,
    Integer, String, Text, DateTime, Float, text,
    UniqueConstraint,
)
from sqlalchemy.exc import SQLAlchemyError, IntegrityError as DBIntegrityError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix

# Sentinel stored in x_coord *and* y_coord to mark a missed target.
# Chosen to be far outside any plausible mm-from-center value.
MISS_SENTINEL = '100000'

# Bundled default target — seeded for every new user as their default so
# /sesh works out of the box without the upload wizard. It renders as a
# vector WA-style 40cm 10-ring face: the same geometry as the WA 122cm
# face scaled to a 40cm edge, drawn client-side from the canonical
# ``wa_40`` ring spec. The image_filename is only a sizing placeholder —
# the colored rings paint over it, so the user never sees a raster face.
# The legacy NASP raster default is archived in favor of this on
# register/login (see _seed_user_default_target). Calibration matches the
# placeholder image exactly (40cm physical edge, 1197 px cropped square).
DEFAULT_TARGET_NAME     = 'WA 40cm 10-ring'
DEFAULT_TARGET_FACE_KEY = 'wa_40'                  # key into TOURNAMENT_FACES
LEGACY_NASP_TARGET_NAME = '40cm NASP target'       # archived on migration
DEFAULT_TARGET_IMAGE    = 'targets/nasp_40cm.jpg'  # placeholder; covered by rings
DEFAULT_TARGET_SIZE_MM  = 400.0
DEFAULT_TARGET_SIZE_PX  = 1197

# Line-cutter scoring assumes a finite shaft thickness: if any part of the
# shaft crosses (or touches) a ring boundary the shot scores the higher
# (inner) ring. Used as the fallback when an arrow record has no
# shaft_diameter set — 6mm is a reasonable mid-weight target shaft.
DEFAULT_SHAFT_DIAMETER_MM = 6.0

# Where uploaded target images live. Served by Flask's static handler.
TARGETS_SUBDIR        = 'targets'
ALLOWED_IMAGE_EXTS    = {'.jpg', '.jpeg', '.png', '.webp'}
MAX_UPLOAD_BYTES      = 5 * 1024 * 1024   # 5 MB cap on uploaded target images
# Pixel-count ceiling on decoded target images. 5 MB of JPEG can decompress
# to multiple GB of pixel buffer ("decompression bomb"), which is plenty to
# OOM a small VPS. 25 MP is comfortably bigger than any phone-camera shot a
# user would realistically upload of a target face.
MAX_IMAGE_PIXELS      = 25_000_000
# Per-target image_filename must match this on import. The layout is
# "targets/<stem>.<ext>" — both user uploads (8-hex prefix) and the
# bundled NASP seed live under this prefix. Anything else — including
# '..' segments or absolute paths — is rejected so a crafted SQL/CSV
# import can't talk a later delete into removing files outside
# TARGETS_DIR.
_SAFE_TARGET_FILENAME_RE = re.compile(
    r'^targets/[A-Za-z0-9._-]+\.(?:jpg|jpeg|png|webp)$', re.IGNORECASE)


def _is_safe_target_image_filename(fn):
    if not fn or not isinstance(fn, str):
        return False
    return bool(_SAFE_TARGET_FILENAME_RE.match(fn))

CIRCLE_RADIUS   = 32   # in mm — not yet wired up (boundary logic TBD)
BULLSEYE_RADIUS = 10   # in mm — not yet wired up (boundary logic TBD)

# ── Tournament mode ─────────────────────────────────────────────────────
# TOURNAMENT_FACES describes every official scoring face Apollo seeds for
# tournament rounds. Each face becomes a targets-table row per user (name
# prefixed with TOURNAMENT_TARGET_NAME_PREFIX so we can find/refresh
# them deterministically), with target_zones rows whose radii match the
# published face spec. The image asset is a placeholder until proper face
# images are sourced — see documentation/tournament/targets.md. The
# template draws colored SVG ring overlays from the same zones list, so
# the user *sees* the official ring layout even when the underlying image
# is the bundled NASP placeholder.
#
# zones are listed innermost-out as (point_value, outer_radius_mm, fill_color).
# Colors follow WA / NFAA conventions for outdoor / indoor / Vegas / NFAA
# faces respectively.
# Plain-ASCII prefix on tournament face rows so MySQL backends running
# the default 3-byte `utf8` charset accept the INSERT — a Unicode prefix
# (e.g. 🏆) is a 4-byte supplementary-plane codepoint and would raise
# "Incorrect string value" on utf8 columns. Keeps the seeder portable
# across SQLite and any MySQL configuration.
TOURNAMENT_TARGET_NAME_PREFIX = '[Tournament] '


def _display_target_name(name):
    """Strip the internal `[Tournament] ` prefix for user-facing display.

    The prefix stays in the DB so the seeder can find and refresh these
    rows, but users see the bare face name (e.g. "WA 122cm 10-zone").
    """
    if name and isinstance(name, str) and name.startswith(TOURNAMENT_TARGET_NAME_PREFIX):
        return name[len(TOURNAMENT_TARGET_NAME_PREFIX):]
    return name


TOURNAMENT_PLACEHOLDER_IMAGE  = 'targets/nasp_40cm.jpg'
TOURNAMENT_PLACEHOLDER_PX     = 1197  # px edge of TOURNAMENT_PLACEHOLDER_IMAGE
# Legacy prefix used by the very first tournament-mode release. The
# seeder migrates rows from this prefix to TOURNAMENT_TARGET_NAME_PREFIX
# on first run so users who tried the old build don't end up with two
# parallel sets of tournament faces.
_LEGACY_TOURNAMENT_PREFIX     = '\U0001F3C6 '   # '🏆 '

# WA outdoor 10-zone ring colors, listed from innermost (X) outward.
# Index 0 = X ring; indices 1..10 = the 10, 9, 8, ..., 1 rings.
_WA_PALETTE = (
    '#fff44f', '#fff44f', '#fff44f',  # X / 10 / 9 → gold
    '#ff4f4f', '#ff4f4f',             # 8 / 7 → red
    '#5fb0d6', '#5fb0d6',             # 6 / 5 → blue
    '#1a1a1a', '#1a1a1a',             # 4 / 3 → black
    '#ffffff', '#ffffff',             # 2 / 1 → white
)

def _wa_zones_10(face_radius_mm, x_ring_radius_mm):
    """Build innermost-out zones list for a WA 10-zone face.

    face_radius_mm is the face's outer edge (610 for 122cm face, 400 for
    80cm, 300 for 60cm, 200 for 40cm). The 10 ring's outer radius is
    face_radius_mm / 10; rings widen linearly outward from there.
    """
    ring_w = face_radius_mm / 10.0
    out = [(10, x_ring_radius_mm, _WA_PALETTE[0])]  # X ring
    for i in range(10):
        pv = 10 - i
        r = (i + 1) * ring_w
        out.append((pv, r, _WA_PALETTE[i + 1]))
    return out


def _wa_zones_6(face_radius_mm, x_ring_radius_mm):
    """WA 80cm 6-ring face (compound 50m) — rings 10..5 visible plus the
    inner-10 X ring. Anything outside the 5 ring scores M (miss).

    Ring widths match the underlying 10-zone face (face_radius_mm / 10).
    """
    ring_w = face_radius_mm / 10.0
    out = [(10, x_ring_radius_mm, '#fff44f')]
    for i in range(6):
        pv = 10 - i
        r = (i + 1) * ring_w
        # WA palette for printed rings 10..5
        color = ('#fff44f', '#fff44f', '#ff4f4f', '#ff4f4f',
                 '#5fb0d6', '#5fb0d6')[i]
        out.append((pv, r, color))
    return out


def _wa_zones_vegas(spot_radius_mm, x_ring_radius_mm):
    """Vegas 3-spot per-spot face — only rings 10..6 are printed, so the
    spot edge is the 6 ring. Anything outside scores M.

    Ring widths come from the parent 40cm 10-zone face, so each ring is
    spot_radius_mm / 5 wide (5 visible rings filling the spot radius).
    """
    ring_w = spot_radius_mm / 5.0
    out = [(10, x_ring_radius_mm, '#fff44f')]
    for i in range(5):
        pv = 10 - i
        r = (i + 1) * ring_w
        color = ('#fff44f', '#fff44f', '#ff4f4f', '#ff4f4f',
                 '#5fb0d6')[i]
        out.append((pv, r, color))
    return out


def _wa_field_zones(face_diameter_mm):
    """WA Field face — 6 equal-width scoring rings (6..1, center→out)
    with an inner-X ring (compound tiebreak). Used on 80/60/40/20 cm
    faces at varied distances on the WA Field round.

    Ring width = face_radius_mm / 6. Outer edge of the 6 ring is the
    yellow/black boundary; the X-ring radius is half the 6-ring radius.
    """
    face_radius_mm = face_diameter_mm / 2.0
    ring_w = face_radius_mm / 6.0
    out = [(6, ring_w / 2.0, '#fff44f')]  # inner-X
    palette = ('#fff44f', '#fff44f',  # 6 / 5 yellow
               '#1a1a1a', '#1a1a1a',  # 4 / 3 black
               '#1a1a1a', '#1a1a1a')  # 2 / 1 black
    for i in range(6):
        pv = 6 - i
        r = (i + 1) * ring_w
        out.append((pv, r, palette[i]))
    return out


def _wa_field_zones_compound(face_diameter_mm):
    """WA Field face — compound variant for the 20 cm face. Outer 6
    scores 5; only the inner-6 (X) ring scores 6. All other rings
    unchanged.
    """
    face_radius_mm = face_diameter_mm / 2.0
    ring_w = face_radius_mm / 6.0
    out = [(6, ring_w / 2.0, '#fff44f')]  # inner-X = 6
    # Outer 6 demoted to 5 for compound.
    palette = ('#fff44f', '#fff44f',
               '#1a1a1a', '#1a1a1a',
               '#1a1a1a', '#1a1a1a')
    values = (5, 5, 4, 3, 2, 1)
    for i in range(6):
        pv = values[i]
        r = (i + 1) * ring_w
        out.append((pv, r, palette[i]))
    return out


# NFAA Indoor Blue face: white rings on blue background, 5/4/3/2/1
# with X inside the inner 5. Confirmed against the NFAA 2026/27
# Constitution & By-Laws (Article VI § Indoor Target Round).
_NFAA_BLUE_BG = '#1a3a5c'
_NFAA_RING_FG = '#ffffff'


def _nfaa_indoor_blue_zones():
    """40 cm NFAA Indoor Single-Spot face (§ I.D.2.a).

    Spec: dull blue face with a white 8 cm center spot containing a
    4 cm X-ring. Four equally-wide blue rings outside the spot score
    4, 3, 2, 1 in descending order — (200 - 40) / 4 = 40 mm per ring,
    so radii run 20, 40, 80, 120, 160, 200 mm innermost-out.

    All blue scoring rings paint the same navy as ``face_bg``; the JS
    overlay renders contrasting per-ring strokes so the boundaries
    visible on the real face card show up here too.
    """
    return [
        (5, 20.0,  _NFAA_RING_FG),     # X — white center disc (4 cm dia)
        (5, 40.0,  _NFAA_RING_FG),     # 5 — white spot (8 cm dia)
        (4, 80.0,  _NFAA_BLUE_BG),     # 4 — blue ring (4 cm wide)
        (3, 120.0, _NFAA_BLUE_BG),     # 3 — blue ring
        (2, 160.0, _NFAA_BLUE_BG),     # 2 — blue ring
        (1, 200.0, _NFAA_BLUE_BG),     # 1 — outer blue ring
    ]


def _nfaa_5spot_zones():
    """NFAA 5-spot — Apollo treats each 16 cm spot as a logical face.

    Per § I.D.2.b: scoring is identical to the Single-Spot face except
    the outer 3 rings are absent — only the central white spot (5/X)
    and a single 4-ring around it remain. Per spot:
      X = 20 mm radius, 5 = 40 mm radius, 4 = 80 mm radius.
    The 4-ring fills out to the 16 cm spot perimeter, same navy as
    the face background.
    """
    return [
        (5, 20.0, _NFAA_RING_FG),      # X — white center (4 cm dia)
        (5, 40.0, _NFAA_RING_FG),      # 5 — white spot (8 cm dia)
        (4, 80.0, _NFAA_BLUE_BG),      # 4 — single blue ring per spot
    ]


def _nfaa_field_zones(face_diameter_mm):
    """NFAA Field face — three scoring zones (5/4/3) plus inner X.

    Per § I.A.1.b: published outer face diameters are 65, 50, 35 and
    20 cm. Standard ring diameters (per NFAA Constitution table):
        X = 10% × face dia
        5 = 20% × face dia
        4 = 33.4% × face dia
        3 = 100% × face dia (= the face edge)
    Radii are therefore 5/10/16.7/50% of the face diameter. For 65 cm
    that yields X=32.5 mm, 5=65 mm, 4≈108.5 mm, 3=325 mm.

    Visually black-and-white only: outer 3-zone is black, the 4-zone
    annulus is white, and the inner 5/X is a black spot with a white
    X marker on top.
    """
    return [
        (5, face_diameter_mm * 0.05,   '#1a1a1a'),       # X — black inner spot
        (5, face_diameter_mm * 0.10,   '#1a1a1a'),       # 5 — black around X
        (4, face_diameter_mm * 0.167,  _NFAA_RING_FG),   # 4 — white middle
        (3, face_diameter_mm * 0.50,   '#1a1a1a'),       # 3 — outer black
    ]


TOURNAMENT_FACES = {
    # WA outdoor — 122cm 10-zone, used at 70m (recurve)
    'wa_122': {
        'name':             'WA 122cm 10-zone',
        'physical_size_mm': 1220.0,
        'x_ring_mm':        30.5,
        'face_bg':          '#ffffff',
        'zones':            _wa_zones_10(610.0, 30.5),
        'center_mark':      'cross',
    },
    # WA outdoor — 80cm 6-ring (compound 50m)
    'wa_80_6ring': {
        'name':             'WA 80cm 6-ring',
        'physical_size_mm': 800.0,
        'x_ring_mm':        20.0,
        'face_bg':          '#ffffff',
        'zones':            _wa_zones_6(400.0, 20.0),
        'center_mark':      'cross',
    },
    # WA short-distance — 80cm 10-zone (50m / 30m on the 1440)
    'wa_80': {
        'name':             'WA 80cm 10-zone',
        'physical_size_mm': 800.0,
        'x_ring_mm':        20.0,
        'face_bg':          '#ffffff',
        'zones':            _wa_zones_10(400.0, 20.0),
        'center_mark':      'cross',
    },
    # WA Indoor 25m — 60cm 10-zone
    'wa_60': {
        'name':             'WA 60cm 10-zone',
        'physical_size_mm': 600.0,
        'x_ring_mm':        15.0,
        'face_bg':          '#ffffff',
        'zones':            _wa_zones_10(300.0, 15.0),
        'center_mark':      'cross',
    },
    # WA Indoor 18m — 40cm 10-zone
    'wa_40': {
        'name':             'WA 40cm 10-zone',
        'physical_size_mm': 400.0,
        'x_ring_mm':        10.0,
        'face_bg':          '#ffffff',
        'zones':            _wa_zones_10(200.0, 10.0),
        'center_mark':      'cross',
    },
    # WA Indoor 18m compound — same face but 10 ring scores 9 (only X = 10).
    # Encoded by demoting the "10" zone to 9; X still 10.
    'wa_40_compound': {
        'name':             'WA 40cm (compound)',
        'physical_size_mm': 400.0,
        'x_ring_mm':        10.0,
        'face_bg':          '#ffffff',
        'zones':            [
            (10, 10.0,  '#fff44f'),
            (9,  20.0,  '#fff44f'),  # demoted: only X = 10 for compound
            (9,  40.0,  '#fff44f'),
            (8,  60.0,  '#ff4f4f'),
            (7,  80.0,  '#ff4f4f'),
            (6,  100.0, '#5fb0d6'),
            (5,  120.0, '#5fb0d6'),
            (4,  140.0, '#1a1a1a'),
            (3,  160.0, '#1a1a1a'),
            (2,  180.0, '#ffffff'),
            (1,  200.0, '#ffffff'),
        ],
        'center_mark':      'cross',
    },
    # NFAA Indoor Blue (single-spot) — 40cm
    'nfaa_indoor_blue': {
        'name':             'NFAA Indoor Blue',
        'physical_size_mm': 400.0,
        'x_ring_mm':        20.0,
        # The face card is white; the round blue scoring face is
        # inscribed in it (so the corners of the canvas stay white,
        # matching the printed face card).
        'face_bg':          '#ffffff',
        'zones':            _nfaa_indoor_blue_zones(),
        'center_mark':      'x',
    },
    # NFAA 5-spot — full 40 cm face card with five 16 cm spots in
    # quincunx layout per § I.D.2.b. Each spot's center is given in mm
    # relative to the face center; clicks score against the nearest
    # spot. Per-spot zones (X / 5 / 4) live on each spot.
    'nfaa_5spot': {
        'name':             'NFAA 5-spot',
        'physical_size_mm': 400.0,           # full face card (40 cm)
        'x_ring_mm':        20.0,            # 4 cm X-ring dia per spot
        'face_bg':          '#ffffff',       # face card is white
        # The "zones" key is still required by the classifier; it
        # carries the per-spot ring radii. The multi_spot.centers list
        # below tells the classifier to score against the nearest spot
        # rather than treating the canvas as a single concentric face.
        'zones':            _nfaa_5spot_zones(),
        'multi_spot':       {
            # Quincunx — center spot at face center, four corner spots
            # offset by ±120 mm on each axis. With a spot radius of
            # 80 mm and a face radius of 200 mm, that puts each corner
            # spot tangent to two face-card edges (120 + 80 = 200) and
            # leaves a ~10 mm clearance from the center spot's outer
            # ring (sqrt(2)·120 - 160 ≈ 9.7 mm), matching the printed
            # NFAA 5-spot face card.
            'centers_mm':   [
                (0.0, 0.0),
                (-120.0,  120.0), (120.0,  120.0),
                (-120.0, -120.0), (120.0, -120.0),
            ],
        },
        'center_mark':      'x',
    },
    # Vegas 40cm 3-spot — per spot is 20 cm with rings 10..6 visible
    # plus an inner-X. Each spot derives from the WA 40 cm face; ring
    # widths are 20 mm so X=10, 10=20, 9=40, 8=60, 7=80, 6=100 mm.
    # Anything outside the 6 ring on the spot is a miss.
    'vegas_3spot': {
        'name':             'Vegas 40cm (per spot)',
        'physical_size_mm': 200.0,
        'x_ring_mm':        10.0,
        'face_bg':          '#ffffff',
        'zones':            _wa_zones_vegas(100.0, 10.0),
        'center_mark':      'x',
    },
    # NFAA Field / Hunter — 4 face sizes used at different distances
    # on the 28-target course. Same scoring rings on Field (white
    # center) and Hunter (black face) variants; Apollo encodes the
    # scoring geometry once and lets the round pick which face label
    # to apply.
    # X-ring diameter is 10% of face diameter → radius is 5% of face dia.
    'nfaa_field_65': {
        'name':             'NFAA Field 65cm',
        'physical_size_mm': 650.0,
        'x_ring_mm':        650.0 * 0.05,
        'face_bg':          '#ffffff',
        'zones':            _nfaa_field_zones(650.0),
        'center_mark':      'x',
    },
    'nfaa_field_50': {
        'name':             'NFAA Field 50cm',
        'physical_size_mm': 500.0,
        'x_ring_mm':        500.0 * 0.05,
        'face_bg':          '#ffffff',
        'zones':            _nfaa_field_zones(500.0),
        'center_mark':      'x',
    },
    'nfaa_field_35': {
        'name':             'NFAA Field 35cm',
        'physical_size_mm': 350.0,
        'x_ring_mm':        350.0 * 0.05,
        'face_bg':          '#ffffff',
        'zones':            _nfaa_field_zones(350.0),
        'center_mark':      'x',
    },
    'nfaa_field_20': {
        'name':             'NFAA Field 20cm',
        'physical_size_mm': 200.0,
        'x_ring_mm':        200.0 * 0.05,
        'face_bg':          '#ffffff',
        'zones':            _nfaa_field_zones(200.0),
        'center_mark':      'x',
    },
    # WA Field 6-zone faces (rings 6..1 with inner-X for compound).
    # Used on the 24-target WA Field course at varying distances.
    'wa_field_80': {
        'name':             'WA Field 80cm',
        'physical_size_mm': 800.0,
        'x_ring_mm':        800.0 / 12.0,
        'face_bg':          '#fff44f',
        'zones':            _wa_field_zones(800.0),
        'center_mark':      'cross',
    },
    'wa_field_60': {
        'name':             'WA Field 60cm',
        'physical_size_mm': 600.0,
        'x_ring_mm':        600.0 / 12.0,
        'face_bg':          '#fff44f',
        'zones':            _wa_field_zones(600.0),
        'center_mark':      'cross',
    },
    'wa_field_40': {
        'name':             'WA Field 40cm',
        'physical_size_mm': 400.0,
        'x_ring_mm':        400.0 / 12.0,
        'face_bg':          '#fff44f',
        'zones':            _wa_field_zones(400.0),
        'center_mark':      'cross',
    },
    'wa_field_20': {
        'name':             'WA Field 20cm',
        'physical_size_mm': 200.0,
        'x_ring_mm':        200.0 / 12.0,
        'face_bg':          '#fff44f',
        'zones':            _wa_field_zones(200.0),
        'center_mark':      'cross',
    },
    # WA Field 20cm compound variant — only inner-6 (X) ring scores 6;
    # outer 6 ring demoted to 5. Per WA Book 4 § Field for compound
    # bowmen on the smallest face.
    'wa_field_20_compound': {
        'name':             'WA Field 20cm (compound)',
        'physical_size_mm': 200.0,
        'x_ring_mm':        200.0 / 12.0,
        'face_bg':          '#fff44f',
        'zones':            _wa_field_zones_compound(200.0),
        'center_mark':      'cross',
    },
    # NASP 80cm (10-ring) — face retained for backward compatibility
    # with seeded rows; NASP rounds are no longer offered in the
    # selector (out of scope per project decision).
    'nasp_80': {
        'name':             'NASP 80cm',
        'physical_size_mm': 800.0,
        'x_ring_mm':        20.0,
        'face_bg':          '#ffffff',
        'zones':            _wa_zones_10(400.0, 20.0),
    },
}

# Course-round schedules — list of (face_key, distance_value, distance_unit)
# tuples that get expanded into per-target course entries.
#
# WA Field marked-round schedule (Book 4): 24 targets, 3 arrows each.
# Six targets at each of 4 face sizes; distances picked from the
# published marked-distance bands per face.
_WA_FIELD_24_SCHEDULE = [
    # Face,       distance (m)
    ('wa_field_80', 60), ('wa_field_80', 55), ('wa_field_80', 50),
    ('wa_field_80', 45), ('wa_field_80', 40), ('wa_field_80', 35),
    ('wa_field_60', 45), ('wa_field_60', 40), ('wa_field_60', 35),
    ('wa_field_60', 30), ('wa_field_60', 25), ('wa_field_60', 20),
    ('wa_field_40', 30), ('wa_field_40', 25), ('wa_field_40', 20),
    ('wa_field_40', 18), ('wa_field_40', 15), ('wa_field_40', 12),
    ('wa_field_20', 15), ('wa_field_20', 12), ('wa_field_20', 10),
    ('wa_field_20',  8), ('wa_field_20',  6), ('wa_field_20',  5),
]

# NFAA Field-round adult schedule (28 targets, 4 arrows each). Per
# NFAA C&BL Article VI. Distances in yards, converted to metres
# at expansion time so the recorded shot rows match the existing
# numeric-metre convention.
_NFAA_FIELD_28_SCHEDULE_YD = [
    ('nfaa_field_65', 80), ('nfaa_field_65', 70), ('nfaa_field_65', 60),
    ('nfaa_field_50', 55), ('nfaa_field_50', 50), ('nfaa_field_50', 45),
    ('nfaa_field_50', 40), ('nfaa_field_35', 40), ('nfaa_field_35', 35),
    ('nfaa_field_35', 30), ('nfaa_field_35', 25), ('nfaa_field_20', 20),
    ('nfaa_field_20', 15), ('nfaa_field_20', 11),  # unit 1 = 14 targets
    ('nfaa_field_65', 80), ('nfaa_field_65', 70), ('nfaa_field_65', 60),
    ('nfaa_field_50', 55), ('nfaa_field_50', 50), ('nfaa_field_50', 45),
    ('nfaa_field_50', 40), ('nfaa_field_35', 40), ('nfaa_field_35', 35),
    ('nfaa_field_35', 30), ('nfaa_field_35', 25), ('nfaa_field_20', 20),
    ('nfaa_field_20', 15), ('nfaa_field_20', 11),  # unit 2 = 14 targets
]

# NFAA Hunter-round adult schedule (28 targets, 4 arrows each).
# Hunter distances run closer than Field per NFAA C&BL Article VI;
# Apollo reuses the Field face geometry (scoring rings are identical).
_NFAA_HUNTER_28_SCHEDULE_YD = [
    ('nfaa_field_65', 70), ('nfaa_field_65', 65), ('nfaa_field_65', 61),
    ('nfaa_field_50', 58), ('nfaa_field_50', 53), ('nfaa_field_50', 48),
    ('nfaa_field_50', 44), ('nfaa_field_35', 36), ('nfaa_field_35', 32),
    ('nfaa_field_35', 28), ('nfaa_field_35', 23), ('nfaa_field_20', 19),
    ('nfaa_field_20', 14), ('nfaa_field_20', 11),
    ('nfaa_field_65', 70), ('nfaa_field_65', 65), ('nfaa_field_65', 61),
    ('nfaa_field_50', 58), ('nfaa_field_50', 53), ('nfaa_field_50', 48),
    ('nfaa_field_50', 44), ('nfaa_field_35', 36), ('nfaa_field_35', 32),
    ('nfaa_field_35', 28), ('nfaa_field_35', 23), ('nfaa_field_20', 19),
    ('nfaa_field_20', 14), ('nfaa_field_20', 11),
]


def _course_from_schedule_m(schedule, arrows_per_target):
    """Schedule already in metres → list of course-entry dicts."""
    return [
        {'face_key': face, 'distance_m': float(dist), 'arrows': arrows_per_target}
        for (face, dist) in schedule
    ]


def _course_from_schedule_yd(schedule, arrows_per_target):
    """Schedule in yards → list of course-entry dicts in metres
    (rounded to 0.01 m for display sanity)."""
    return [
        {'face_key': face, 'distance_m': round(float(dist) * 0.9144, 2),
         'arrows': arrows_per_target}
        for (face, dist) in schedule
    ]


# TOURNAMENT_ROUNDS describes the structured rounds Apollo supports. Each
# round references a face_key from TOURNAMENT_FACES. Multi-distance rounds
# (1440, NFAA 900) use the `segments` list to define a sequence of
# (distance_m, ends, face_key) chunks; single-segment rounds set segments
# to None and use the top-level distance_m / ends / face_key.
# Course rounds (WA Field 24, NFAA Field/Hunter 28) use the `course`
# field — a list of {face_key, distance_m, arrows} entries, one per
# target, expanded into single-end segments by _round_segments().
TOURNAMENT_ROUNDS = {
    'wa_720_recurve': {
        'org':              'World Archery',
        'name':             'WA 720 (Recurve)',
        'face_key':         'wa_122',
        'distance_m':       70,
        'arrows_per_end':   6,
        'ends':             12,
        'total_arrows':     72,
        'max_score':        720,
        'end_time_s':       240,
        'equipment_class':  'recurve',
        'description':      '72 arrows at 70m on the 122cm face. Recurve outdoor qualification round.',
        'segments':         None,
    },
    'wa_720_compound': {
        'org':              'World Archery',
        'name':             'WA 720 (Compound)',
        'face_key':         'wa_80_6ring',
        'distance_m':       50,
        'arrows_per_end':   6,
        'ends':             12,
        'total_arrows':     72,
        'max_score':        720,
        'end_time_s':       240,
        'equipment_class':  'compound',
        'description':      '72 arrows at 50m on the 80cm 6-ring face. Compound outdoor qualification round.',
        'segments':         None,
    },
    'wa_indoor_18_recurve': {
        'org':              'World Archery',
        'name':             'WA Indoor 18m (Recurve)',
        'face_key':         'wa_40',
        'distance_m':       18,
        'arrows_per_end':   3,
        'ends':             20,
        'total_arrows':     60,
        'max_score':        600,
        'end_time_s':       120,
        'equipment_class':  'recurve',
        'description':      '60 arrows at 18m on the 40cm face. Both outer-10 and inner-10 (X) score 10.',
        'segments':         None,
    },
    'wa_indoor_18_compound': {
        'org':              'World Archery',
        'name':             'WA Indoor 18m (Compound)',
        'face_key':         'wa_40_compound',
        'distance_m':       18,
        'arrows_per_end':   3,
        'ends':             20,
        'total_arrows':     60,
        'max_score':        600,
        'end_time_s':       120,
        'equipment_class':  'compound',
        'description':      '60 arrows at 18m on the 40cm face. Only the inner-10 (X-ring) scores 10; outer-10 ring scores 9.',
        'segments':         None,
    },
    'wa_indoor_25': {
        'org':              'World Archery',
        'name':             'WA Indoor 25m',
        'face_key':         'wa_60',
        'distance_m':       25,
        'arrows_per_end':   3,
        'ends':             20,
        'total_arrows':     60,
        'max_score':        600,
        'end_time_s':       120,
        'equipment_class':  'any',
        'description':      '60 arrows at 25m on the 60cm face.',
        'segments':         None,
    },
    'wa_match_recurve_cum': {
        'org':              'World Archery',
        'name':             'WA Match Play — Recurve (cumulative)',
        'face_key':         'wa_122',
        'distance_m':       70,
        'arrows_per_end':   3,
        'ends':             5,
        'total_arrows':     15,
        'max_score':        150,
        'end_time_s':       120,
        'equipment_class':  'recurve',
        'description':      '5 ends of 3 arrows at 70m, cumulative scoring. Highest total wins (used in compound match play; also handy as a recurve drill).',
        'segments':         None,
        'match_scoring':    'cumulative',
    },
    'wa_match_recurve_set': {
        'org':              'World Archery',
        'name':             'WA Match Play — Recurve (set system)',
        'face_key':         'wa_122',
        'distance_m':       70,
        'arrows_per_end':   3,
        'ends':             5,
        'total_arrows':     15,
        'max_score':        150,
        'end_time_s':       120,
        'equipment_class':  'recurve',
        'description':      'Olympic recurve match: 5 sets of 3 arrows at 70m. Per set the higher total earns 2 set points (1 each on a tie). First to 6 set points wins; tie 5-5 → single-arrow shoot-off. Apollo logs your set totals so you can compare to an opponent’s scorecard.',
        'segments':         None,
        'match_scoring':    'set',
    },
    'wa_match_compound': {
        'org':              'World Archery',
        'name':             'WA Match Play — Compound',
        'face_key':         'wa_80_6ring',
        'distance_m':       50,
        'arrows_per_end':   3,
        'ends':             5,
        'total_arrows':     15,
        'max_score':        150,
        'end_time_s':       120,
        'equipment_class':  'compound',
        'description':      '5 ends of 3 arrows at 50m. Highest cumulative wins.',
        'segments':         None,
        'match_scoring':    'cumulative',
    },
    'wa_1440_recurve_m': {
        'org':              'World Archery',
        'name':             'WA 1440 (Recurve, Men)',
        'face_key':         'wa_122',
        'distance_m':       90,
        'arrows_per_end':   6,
        'ends':             6,
        'total_arrows':     144,
        'max_score':        1440,
        'end_time_s':       240,
        'equipment_class':  'recurve',
        'description':      '36 arrows at each of 90/70m on the 122cm face, then 50/30m on the 80cm face. 144 arrows total.',
        'segments': [
            {'distance_m': 90, 'ends': 6,  'face_key': 'wa_122', 'arrows_per_end': 6},
            {'distance_m': 70, 'ends': 6,  'face_key': 'wa_122', 'arrows_per_end': 6},
            {'distance_m': 50, 'ends': 12, 'face_key': 'wa_80',  'arrows_per_end': 3},
            {'distance_m': 30, 'ends': 12, 'face_key': 'wa_80',  'arrows_per_end': 3},
        ],
    },
    'wa_1440_recurve_w': {
        'org':              'World Archery',
        'name':             'WA 1440 (Recurve, Women)',
        'face_key':         'wa_122',
        'distance_m':       70,
        'arrows_per_end':   6,
        'ends':             6,
        'total_arrows':     144,
        'max_score':        1440,
        'end_time_s':       240,
        'equipment_class':  'recurve',
        'description':      '36 arrows at each of 70/60m on the 122cm face, then 50/30m on the 80cm face. 144 arrows total.',
        'segments': [
            {'distance_m': 70, 'ends': 6,  'face_key': 'wa_122', 'arrows_per_end': 6},
            {'distance_m': 60, 'ends': 6,  'face_key': 'wa_122', 'arrows_per_end': 6},
            {'distance_m': 50, 'ends': 12, 'face_key': 'wa_80',  'arrows_per_end': 3},
            {'distance_m': 30, 'ends': 12, 'face_key': 'wa_80',  'arrows_per_end': 3},
        ],
    },
    'wa_1440_compound_m': {
        'org':              'World Archery',
        'name':             'WA 1440 (Compound, Men)',
        'face_key':         'wa_122',
        'distance_m':       70,
        'arrows_per_end':   6,
        'ends':             6,
        'total_arrows':     144,
        'max_score':        1440,
        'end_time_s':       240,
        'equipment_class':  'compound',
        'description':      'Compound 1440: 70/60m on 122cm, then 50/30m on the 80cm 6-ring face. 144 arrows total.',
        'segments': [
            {'distance_m': 70, 'ends': 6,  'face_key': 'wa_122',      'arrows_per_end': 6},
            {'distance_m': 60, 'ends': 6,  'face_key': 'wa_122',      'arrows_per_end': 6},
            {'distance_m': 50, 'ends': 12, 'face_key': 'wa_80_6ring', 'arrows_per_end': 3},
            {'distance_m': 30, 'ends': 12, 'face_key': 'wa_80_6ring', 'arrows_per_end': 3},
        ],
    },
    'wa_1440_compound_w': {
        'org':              'World Archery',
        'name':             'WA 1440 (Compound, Women)',
        'face_key':         'wa_122',
        'distance_m':       60,
        'arrows_per_end':   6,
        'ends':             6,
        'total_arrows':     144,
        'max_score':        1440,
        'end_time_s':       240,
        'equipment_class':  'compound',
        'description':      'Compound 1440 women: 60/50m on 122cm, then 40/30m on the 80cm 6-ring face. 144 arrows total.',
        'segments': [
            {'distance_m': 60, 'ends': 6,  'face_key': 'wa_122',      'arrows_per_end': 6},
            {'distance_m': 50, 'ends': 6,  'face_key': 'wa_122',      'arrows_per_end': 6},
            {'distance_m': 40, 'ends': 12, 'face_key': 'wa_80_6ring', 'arrows_per_end': 3},
            {'distance_m': 30, 'ends': 12, 'face_key': 'wa_80_6ring', 'arrows_per_end': 3},
        ],
    },
    'wa_field_24': {
        'org':              'World Archery',
        'name':             'WA Field (24 targets, marked)',
        'face_key':         'wa_field_80',
        'distance_m':       60,
        'arrows_per_end':   3,
        'ends':             24,
        'total_arrows':     72,
        'max_score':        432,
        'end_time_s':       0,    # course pace; not enforced
        'equipment_class':  'any',
        'description':      '24-target field course, 3 arrows per target on 80/60/40/20 cm WA Field faces. Max 432.',
        'segments':         None,
        'course':           _course_from_schedule_m(_WA_FIELD_24_SCHEDULE, 3),
    },
    'nfaa_indoor_blue': {
        'org':              'NFAA',
        'name':             'NFAA Indoor (Blue face)',
        'face_key':         'nfaa_indoor_blue',
        'distance_m':       18.29,  # 20 yards
        'arrows_per_end':   5,
        'ends':             12,
        'total_arrows':     60,
        'max_score':        300,
        'end_time_s':       240,
        'equipment_class':  'any',
        'description':      '60 arrows at 20 yards on the NFAA blue face. Max 5 per arrow.',
        'segments':         None,
    },
    'nfaa_5spot': {
        'org':              'NFAA',
        'name':             'NFAA 5-spot',
        'face_key':         'nfaa_5spot',
        'distance_m':       18.29,
        'arrows_per_end':   5,
        'ends':             12,
        'total_arrows':     60,
        'max_score':        300,
        'end_time_s':       240,
        'equipment_class':  'any',
        'description':      '60 arrows at 20 yards on the 5-spot face. One arrow per spot — Apollo does not enforce; the user must distribute manually.',
        'segments':         None,
    },
    'nfaa_vegas': {
        'org':              'NFAA',
        'name':             'Vegas Round (40cm 3-spot)',
        'face_key':         'vegas_3spot',
        'distance_m':       18.29,
        'arrows_per_end':   3,
        'ends':             10,
        'total_arrows':     30,
        'max_score':        300,
        'end_time_s':       120,
        'equipment_class':  'any',
        'description':      '30 arrows at 20 yards on the 40cm Vegas 3-spot. One arrow per spot; rings 6-10 only.',
        'segments':         None,
    },
    'nfaa_900': {
        'org':              'NFAA',
        'name':             'NFAA 900 Round',
        'face_key':         'wa_122',
        'distance_m':       54.86,  # 60 yards — first segment
        'arrows_per_end':   6,
        'ends':             5,      # per segment
        'total_arrows':     90,
        'max_score':        900,
        'end_time_s':       240,
        'equipment_class':  'any',
        'description':      '30 arrows at each of 60, 50, and 40 yards on the 122cm face.',
        'segments': [
            {'distance_m': 54.86, 'ends': 5, 'face_key': 'wa_122'},  # 60 yd
            {'distance_m': 45.72, 'ends': 5, 'face_key': 'wa_122'},  # 50 yd
            {'distance_m': 36.58, 'ends': 5, 'face_key': 'wa_122'},  # 40 yd
        ],
    },
    'nfaa_field_28': {
        'org':              'NFAA',
        'name':             'NFAA Field Round (28 targets)',
        'face_key':         'nfaa_field_65',
        'distance_m':       round(80 * 0.9144, 2),
        'arrows_per_end':   4,
        'ends':             28,
        'total_arrows':     112,
        'max_score':        560,
        'end_time_s':       0,
        'equipment_class':  'any',
        'description':      '28-target field course, 4 arrows per target on 65/50/35/20cm NFAA field faces. Max 560.',
        'segments':         None,
        'course':           _course_from_schedule_yd(_NFAA_FIELD_28_SCHEDULE_YD, 4),
    },
    'nfaa_hunter_28': {
        'org':              'NFAA',
        'name':             'NFAA Hunter Round (28 targets)',
        'face_key':         'nfaa_field_65',
        'distance_m':       round(70 * 0.9144, 2),
        'arrows_per_end':   4,
        'ends':             28,
        'total_arrows':     112,
        'max_score':        560,
        'end_time_s':       0,
        'equipment_class':  'any',
        'description':      '28-target hunter course, 4 arrows per target on 65/50/35/20cm faces. Closer distances than Field. Max 560.',
        'segments':         None,
        'course':           _course_from_schedule_yd(_NFAA_HUNTER_28_SCHEDULE_YD, 4),
    },
    # ─── USA Archery (USAA) rounds ─────────────────────────────────────
    # USAA adopts WA rules for its sanctioned events; each round below
    # mirrors the corresponding WA round structurally but carries the
    # USAA-branded name so it lands in the USA Archery selector group.
    'usaa_indoor_nationals': {
        'org':              'USA Archery',
        'name':             'USAA Indoor Nationals',
        'face_key':         'wa_40',
        'distance_m':       18,
        'arrows_per_end':   3,
        'ends':             20,
        'total_arrows':     60,
        'max_score':        600,
        'end_time_s':       120,
        'equipment_class':  'recurve',
        'description':      'USAA Indoor Nationals — 60 arrows at 18m on the 40cm face. Same format as WA Indoor 18m (recurve).',
        'segments':         None,
    },
    'usaa_outdoor_nationals_recurve': {
        'org':              'USA Archery',
        'name':             'USAA Outdoor Nationals (Recurve)',
        'face_key':         'wa_122',
        'distance_m':       70,
        'arrows_per_end':   6,
        'ends':             12,
        'total_arrows':     72,
        'max_score':        720,
        'end_time_s':       240,
        'equipment_class':  'recurve',
        'description':      'USAA Outdoor Nationals — WA 720 round, recurve. 72 arrows at 70m on the 122cm face.',
        'segments':         None,
    },
    'usaa_outdoor_nationals_compound': {
        'org':              'USA Archery',
        'name':             'USAA Outdoor Nationals (Compound)',
        'face_key':         'wa_80_6ring',
        'distance_m':       50,
        'arrows_per_end':   6,
        'ends':             12,
        'total_arrows':     72,
        'max_score':        720,
        'end_time_s':       240,
        'equipment_class':  'compound',
        'description':      'USAA Outdoor Nationals — WA 720 round, compound. 72 arrows at 50m on the 80cm 6-ring face.',
        'segments':         None,
    },
    'usaa_collegiate_indoor': {
        'org':              'USA Archery',
        'name':             'USAA Collegiate Indoor',
        'face_key':         'wa_40',
        'distance_m':       18,
        'arrows_per_end':   3,
        'ends':             20,
        'total_arrows':     60,
        'max_score':        600,
        'end_time_s':       120,
        'equipment_class':  'recurve',
        'description':      'USAA Collegiate Indoor — same format as WA Indoor 18m. 60 arrows at 18m on the 40cm face.',
        'segments':         None,
    },
    'usaa_collegiate_outdoor_recurve': {
        'org':              'USA Archery',
        'name':             'USAA Collegiate Outdoor (Recurve)',
        'face_key':         'wa_122',
        'distance_m':       70,
        'arrows_per_end':   6,
        'ends':             12,
        'total_arrows':     72,
        'max_score':        720,
        'end_time_s':       240,
        'equipment_class':  'recurve',
        'description':      'USAA Collegiate Outdoor (Recurve) — WA 720 at 70m on 122cm.',
        'segments':         None,
    },
    'usaa_collegiate_outdoor_compound': {
        'org':              'USA Archery',
        'name':             'USAA Collegiate Outdoor (Compound)',
        'face_key':         'wa_80_6ring',
        'distance_m':       50,
        'arrows_per_end':   6,
        'ends':             12,
        'total_arrows':     72,
        'max_score':        720,
        'end_time_s':       240,
        'equipment_class':  'compound',
        'description':      'USAA Collegiate Outdoor (Compound) — WA 720 at 50m on 80cm 6-ring.',
        'segments':         None,
    },
    'usaa_joad_indoor': {
        'org':              'USA Archery',
        'name':             'USAA JOAD Indoor',
        'face_key':         'wa_40',
        'distance_m':       18,
        'arrows_per_end':   3,
        'ends':             20,
        'total_arrows':     60,
        'max_score':        600,
        'end_time_s':       120,
        'equipment_class':  'any',
        'description':      'JOAD Indoor — same format as WA Indoor 18m. Age-class pin/star awards are scored off the total; Apollo does not enforce age class.',
        'segments':         None,
    },
}


def _tournament_round_def(round_key):
    return TOURNAMENT_ROUNDS.get(round_key)


def _tournament_face_def(face_key):
    return TOURNAMENT_FACES.get(face_key)


def _tournament_tag_for_round(round_key, practice=False):
    """Session-tag string Apollo stamps on every shot in a tournament
    session so finalization and the analytics page can recover the
    round identity without a schema migration.

    When `practice` is set, also append the `practice` tag so the round
    is recognizable as a non-competition run in past-session listings and
    analytics queries."""
    base = f'tournament:{round_key}'
    if practice:
        return f'{base}, practice'
    return base


def _round_key_from_tags(tags):
    """Recover the round_key from a comma-separated session_tags string.
    Returns None if the session is not a tournament session."""
    if not tags:
        return None
    for part in tags.split(','):
        t = part.strip()
        if t.startswith('tournament:'):
            key = t.split(':', 1)[1].strip()
            if key in TOURNAMENT_ROUNDS:
                return key
    return None


def _practice_from_tags(tags):
    """True when a `practice` tag appears in a comma-separated tag string.
    Used to recover the practice flag for an in-progress tournament after
    a cookie wipe, and to badge past sessions in the listing."""
    if not tags:
        return False
    for part in tags.split(','):
        if part.strip().lower() == 'practice':
            return True
    return False


def _counts_toward_handicap(tags):
    """True only for a round that should feed the running handicap.

    An official handicap comes from a round you actually competed: match
    play, or a score sheet logged from a real competition — both carry a
    ``match:`` tag. Practice never counts: solo live Practice mode has no
    ``match:`` tag at all, and a paper practice scorecard carries the
    ``practice_scorecard`` marker (even though it reuses the match plumbing).
    """
    if not tags:
        return False
    toks = [t.strip() for t in tags.split(',')]
    has_match = any(t.startswith('match:') for t in toks)
    is_practice_scorecard = 'practice_scorecard' in toks
    return has_match and not is_practice_scorecard


def _round_segments(round_def):
    """Normalize a round definition into a list of segments.

    Every round resolves to a list of segments, each carrying:
      face_key, distance_m, arrows_per_end, ends.

    - Single-segment rounds (segments=None, course=None) → one segment
      with the round's top-level face/distance/ends/arrows_per_end.
    - Multi-segment rounds (segments=[…]) → use the list as-is, falling
      back to the round-level arrows_per_end when a segment omits it.
    - Course rounds (course=[…]) → expand each course entry into a
      one-end segment, with arrows_per_end = the target's arrow count.

    Used by the route to resolve the active segment's face and by the
    progress calculator to chunk shots into ends correctly across
    rounds whose end size varies between segments (e.g. WA 1440).
    """
    course = round_def.get('course')
    if course:
        return [
            {
                'face_key':       t['face_key'],
                'distance_m':     t['distance_m'],
                'arrows_per_end': int(t['arrows']),
                'ends':           1,
            }
            for t in course
        ]
    segs = round_def.get('segments')
    if segs:
        default_ape = int(round_def.get('arrows_per_end', 0) or 0)
        out = []
        for s in segs:
            out.append({
                'face_key':       s.get('face_key', round_def['face_key']),
                'distance_m':     s['distance_m'],
                'arrows_per_end': int(s.get('arrows_per_end', default_ape)),
                'ends':           int(s['ends']),
            })
        return out
    return [
        {
            'face_key':       round_def['face_key'],
            'distance_m':     round_def['distance_m'],
            'arrows_per_end': int(round_def['arrows_per_end']),
            'ends':           int(round_def['ends']),
        }
    ]


# ── Handicap & classification adapters ─────────────────────────────────────
# These bridge Apollo's faces/rounds to the pure handicap.py / classifications
# modules. Handicaps are computed only for AGB-recognised target rounds, using
# the scheme-standard arrow diameter (so the number matches published AGB
# tables); the raw score itself is still Apollo's real-arrow line-cutter score.

def _round_handicap_passes(round_key):
    """Build handicap.py 'passes' for a round, or None if not handicappable.

    Each pass is ``{dist_m, n_arrows, zones, arrow_d_m}``, pulling ring
    geometry from ``TOURNAMENT_FACES`` and the indoor/outdoor standard arrow
    diameter.
    """
    round_def = _tournament_round_def(round_key)
    if not round_def:
        return None
    indoor = round_key in classifications.data.INDOOR_ROUNDS
    arrow_d = handicap.ARROW_D_INDOOR if indoor else handicap.ARROW_D_OUTDOOR
    passes = []
    for seg in _round_segments(round_def):
        face = TOURNAMENT_FACES.get(seg['face_key'])
        if not face or not face.get('zones'):
            return None
        passes.append({
            'dist_m':    float(seg['distance_m']),
            'n_arrows':  int(seg['arrows_per_end']) * int(seg['ends']),
            'zones':     face['zones'],
            'arrow_d_m': arrow_d,
        })
    return passes


def _session_handicap(round_key, total_score):
    """AGB handicap (int) for a completed round score, or None.

    Only meaningful for AGB-recognised target rounds; returns None for other
    formats (NFAA, field, match) or an out-of-range score.
    """
    if round_key not in classifications.data.AGB_TARGET_ROUNDS or not total_score:
        return None
    passes = _round_handicap_passes(round_key)
    round_def = _tournament_round_def(round_key)
    if not passes or not round_def:
        return None
    hc_val = handicap.handicap_from_score(total_score, passes, round_def.get('max_score'))
    if hc_val is None:
        return None
    # Archery GB issues handicaps on a 0–150 scale (0 = best). The model can
    # produce values outside that — negative for a near-perfect score, >150 for
    # a near-zero one — but those are not handicaps as they are awarded or
    # printed, so clamp to the scale as it is used in practice.
    return max(0, min(150, hc_val))


def _archer_category(session_tags=None):
    """Build a classifications.Category from the user's profile, with an
    optional per-round bowstyle override carried in ``session_tags``
    (e.g. ``bowstyle:compound``).
    """
    user = current_user()
    bowstyle = (user.get('default_bowstyle') if user else None) or 'recurve'
    gender = (user.get('gender') if user else None) or 'male'
    age_group = (user.get('age_group') if user else None) or 'adult'
    if session_tags:
        for part in str(session_tags).split(','):
            part = part.strip()
            if part.startswith('bowstyle:'):
                bowstyle = part.split(':', 1)[1].strip() or bowstyle
    return classifications.Category(bowstyle, gender, age_group)


def _round_bowstyle(round_def):
    """The classification bowstyle a round dictates, or ``None`` if it is
    open to any bowstyle.

    WA/USAA rounds whose ``equipment_class`` names a bowstyle (e.g. "WA 720
    (Recurve)") ARE that bowstyle's round — the face and distance only make
    sense for it — so the round itself fixes the classification category,
    not the archer's profile or a per-round picker. ``'any'`` rounds (NFAA,
    WA Field, WA Indoor 25m, JOAD, …) leave the choice open.
    """
    cls = (round_def or {}).get('equipment_class')
    return cls if cls in classifications.data.AGB_BOWSTYLES else None


def _session_handicap_awards(round_key, total_score, session_tags=None):
    """Return ``{'handicap': int|None, 'awards': [...]}`` for a finished round."""
    hc_val = _session_handicap(round_key, total_score)
    category = _archer_category(session_tags)
    awards = classifications.resolve_awards(round_key, total_score, hc_val, category)
    return {'handicap': hc_val, 'awards': awards}


def _segment_for_shot(segments, shot_index):
    """Return (segment_idx, segment_def, arrow_offset_into_segment) for
    the segment containing the given zero-based shot index.

    Returns (len(segments)-1, last_segment, count_into_last) once the
    round is fully shot (so progress display still works on the final
    segment after the round completes).
    """
    cumulative = 0
    for idx, seg in enumerate(segments):
        seg_arrows = int(seg['arrows_per_end']) * int(seg['ends'])
        if shot_index < cumulative + seg_arrows:
            return idx, seg, shot_index - cumulative
        cumulative += seg_arrows
    last_idx = len(segments) - 1
    last = segments[last_idx]
    return last_idx, last, int(last['arrows_per_end']) * int(last['ends'])


def _face_ring_labels(face_key):
    """Return the ordered list of user-facing ring labels for the
    non-X zones of a face, matching the order of zones[1:] (innermost-
    out, excluding the inner X ring).

    The label is the *physical* ring number printed on the face, which
    matches the zone's point_value for most rounds. Compound 18 m and
    compound WA-Field-20 cm need explicit lists because their first
    visible ring is the demoted outer-10 (or outer-6), labeled "10"
    (or "6") on the scorecard but assigned a lower point value in the
    zone list.
    """
    overrides = {
        'wa_40_compound':       [10, 9, 8, 7, 6, 5, 4, 3, 2, 1],
        'wa_field_20_compound': [6, 5, 4, 3, 2, 1],
    }
    if face_key in overrides:
        return overrides[face_key]
    face = TOURNAMENT_FACES.get(face_key, {})
    zones = face.get('zones') or ()
    return [pv for (pv, _r, _c) in list(zones)[1:]]


def _coords_for_ring_label(face_key, label):
    """Return (x_str, y_str) — coordinates that the existing target
    classifier will resolve into the ring identified by `label`.

    label is one of:
      "X"  — innermost zone (the inner-X ring)
      "M"  — miss (uses MISS_SENTINEL on both axes)
      "1", "2", … "10" — physical ring number; falls back to MISS
                          if the label isn't valid on this face.

    The chosen coordinate is the midpoint between the outer edge of
    the next-inner zone and the outer edge of the requested zone, so
    the line-cutter slack on the existing classifier doesn't push the
    arrow into a neighboring zone.
    """
    if label == "M":
        return (MISS_SENTINEL, MISS_SENTINEL)
    face = TOURNAMENT_FACES.get(face_key)
    if face is None:
        return (MISS_SENTINEL, MISS_SENTINEL)
    zones = list(face.get('zones') or ())
    if not zones:
        return (MISS_SENTINEL, MISS_SENTINEL)
    if label == "X":
        return ("0", "0")
    labels = _face_ring_labels(face_key)
    try:
        idx = labels.index(int(label))
    except (ValueError, TypeError):
        return (MISS_SENTINEL, MISS_SENTINEL)
    zone_idx = idx + 1  # zones[0] is the X ring
    inner_r = float(zones[zone_idx - 1][1])  # outer edge of next inner zone
    this_r  = float(zones[zone_idx][1])
    midpoint = (inner_r + this_r) / 2.0
    return (str(midpoint), "0")


def _round_total_arrows(round_def):
    """Sum total arrows across all segments. Authoritative count for
    completion checks and progress display; supersedes the round def's
    stored `total_arrows` (which is kept for the selector card)."""
    return sum(
        int(s['arrows_per_end']) * int(s['ends'])
        for s in _round_segments(round_def)
    )


# Tournament face display-name aliases — old names that have been
# renamed in the spec. Each entry maps the old display name (without
# the `[Tournament] ` prefix) to the current name. The seeder uses this
# to rename any pre-existing rows so the reverse lookup
# (_tournament_face_key_for_target_name) keeps finding them.
_TOURNAMENT_FACE_NAME_ALIASES = {
    'NFAA 5-spot (per spot)': 'NFAA 5-spot',
}


def _migrate_renamed_tournament_faces(user_id):
    """Update existing tournament face rows whose display name has
    been changed by a later Apollo build.

    If the user has only the old name, rename it. If they have both
    (because the new name was seeded as a separate row after the
    rename), deactivate the old one — the new row carries the up-to-
    date scoring zones and ring spec, and we don't want the dropdown
    listing both.
    """
    if user_id is None or not _TOURNAMENT_FACE_NAME_ALIASES:
        return
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            for old_bare, new_bare in _TOURNAMENT_FACE_NAME_ALIASES.items():
                old_name = TOURNAMENT_TARGET_NAME_PREFIX + old_bare
                new_name = TOURNAMENT_TARGET_NAME_PREFIX + new_bare
                old_row = cur.execute(
                    "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                    (user_id, old_name)
                ).fetchone()
                if old_row is None:
                    continue
                new_row = cur.execute(
                    "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                    (user_id, new_name)
                ).fetchone()
                if new_row is None:
                    # Only the old row exists — rename in place. The
                    # main seed will refresh its zones from the current
                    # spec, so the renamed row picks up the new layout
                    # without dropping any historical shot data attached
                    # to it.
                    try:
                        cur.execute(
                            "UPDATE targets SET name = %s WHERE id = %s AND user_id = %s",
                            (new_name, int(old_row[0]), user_id)
                        )
                    except SQLAlchemyError:
                        continue
                else:
                    # Both exist — keep the new row (its zones already
                    # match the current spec) and hide the legacy row
                    # from the dropdown. Don't delete it: historical
                    # shots may still reference its target_id.
                    try:
                        cur.execute(
                            "UPDATE targets SET is_active = 0, is_default = 0 "
                            "WHERE id = %s AND user_id = %s",
                            (int(old_row[0]), user_id)
                        )
                    except SQLAlchemyError:
                        continue
            con.commit()
    except SQLAlchemyError:
        pass


def _migrate_legacy_tournament_prefix(user_id):
    """Best-effort rename of any 🏆-prefixed tournament face rows from
    the very first Apollo build to the current ASCII prefix.

    Runs on a dedicated short-lived connection so a failure (typical on
    MySQL where `targets.name` is `utf8mb3` and can't compare against a
    4-byte parameter — "Illegal mix of collations") doesn't poison the
    main seed transaction. The user's MySQL database never actually
    contains legacy rows in that case (the original INSERT would have
    failed too) so silently swallowing the lookup is correct.
    """
    if user_id is None:
        return
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            for face in TOURNAMENT_FACES.values():
                legacy_name = _LEGACY_TOURNAMENT_PREFIX + face['name']
                new_name = TOURNAMENT_TARGET_NAME_PREFIX + face['name']
                try:
                    legacy = cur.execute(
                        "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                        (user_id, legacy_name)
                    ).fetchone()
                except SQLAlchemyError:
                    # Charset/collation mismatch on this backend — the
                    # row literally cannot exist (the original emoji-
                    # prefix INSERT would have failed too), so skip the
                    # rest of the migration entirely.
                    return
                if legacy is None:
                    continue
                try:
                    cur.execute(
                        "UPDATE targets SET name = %s WHERE id = %s AND user_id = %s",
                        (new_name, int(legacy[0]), user_id)
                    )
                except SQLAlchemyError:
                    continue
            con.commit()
    except SQLAlchemyError:
        # Whole-connection failure: nothing to migrate, fall through.
        pass


def _seed_tournament_faces(user_id):
    """Insert any missing tournament face rows for this user.

    Each face becomes a row in `targets` (name prefixed with
    TOURNAMENT_TARGET_NAME_PREFIX so the seeder can find them on
    subsequent calls and so the user sees a distinct "tournament" group
    in the targets dropdown). Each face's scoring zones are written to
    `target_zones` — on a refresh, existing zones are deleted and
    re-inserted from the canonical spec so future rule updates can
    re-seed by bumping the constants and bouncing the route.

    Returns a dict {face_key: target_id} mapping for the user.
    """
    if user_id is None:
        return {}
    # Best-effort legacy-prefix rename on a separate connection. Runs
    # before the main seed so any rows it renamed are visible to the
    # SELECT below; runs on its own connection so a collation failure
    # (MySQL utf8mb3 column vs utf8mb4 parameter) can't abort the seed
    # transaction that follows.
    _migrate_legacy_tournament_prefix(user_id)
    out = {}
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Backfill: tournament faces seeded under the prior policy
            # have is_default=0; bring them in line with the new "all
            # tournament faces are defaults" rule before re-seeding.
            # Restricted to rows whose bare name matches a *current*
            # face so the rename migration below can deactivate stale
            # rows without this backfill re-activating them.
            cur.execute(
                "UPDATE targets SET is_default = 1, is_active = 1 "
                "WHERE user_id = %s AND name LIKE %s "
                "AND (is_default = 0 OR is_active = 0)",
                (user_id, TOURNAMENT_TARGET_NAME_PREFIX + '%')
            )
            for face_key, face in TOURNAMENT_FACES.items():
                name = TOURNAMENT_TARGET_NAME_PREFIX + face['name']
                row = cur.execute(
                    "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                    (user_id, name)
                ).fetchone()
                if row is None:
                    cur.execute(
                        "INSERT INTO targets "
                        "(user_id, name, image_filename, physical_size_mm, "
                        "image_size_px, is_active, is_default) "
                        "VALUES (%s, %s, %s, %s, %s, 1, 1)",
                        (user_id, name, TOURNAMENT_PLACEHOLDER_IMAGE,
                         float(face['physical_size_mm']),
                         TOURNAMENT_PLACEHOLDER_PX)
                    )
                    row = cur.execute(
                        "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                        (user_id, name)
                    ).fetchone()
                if row is None:
                    # INSERT silently dropped — surface a real error
                    # rather than handing the caller a None target_id and
                    # the user a useless "please try again".
                    raise SQLAlchemyError(
                        f"targets INSERT did not produce a row for {name!r}"
                    )
                target_id = int(row[0])
                out[face_key] = target_id

                # Replace zones from the spec — innermost-out → highest
                # display_order on the innermost so the editor sorts
                # them sensibly. The classifier sorts by radius_mm.
                cur.execute(
                    "DELETE FROM target_zones WHERE user_id = %s AND target_id = %s",
                    (user_id, target_id)
                )
                for idx, (pv, radius_mm, _color) in enumerate(face['zones']):
                    cur.execute(
                        "INSERT INTO target_zones "
                        "(user_id, target_id, name, point_value, "
                        "shape_type, radius_mm, display_order) "
                        "VALUES (%s, %s, %s, %s, 'circle', %s, %s)",
                        (user_id, target_id, f"{pv} pts", pv,
                         float(radius_mm), idx)
                    )
            con.commit()
    except SQLAlchemyError as e:
        # Log loudly with a traceback. The previous swallow-and-return
        # path made this error invisible to anyone troubleshooting from
        # the production logs.
        import traceback
        print(f"⚠️ Tournament face seeding failed for user {user_id}: {e}")
        traceback.print_exc()
        return {}
    # Rename / deactivate any rows that still carry a previous Apollo
    # build's display name (e.g. "NFAA 5-spot (per spot)" → "NFAA
    # 5-spot"). Runs AFTER the main seed/backfill so the deactivation
    # of duplicate stale rows isn't immediately undone by the backfill's
    # "set is_active=1 for all tournament rows" sweep above.
    _migrate_renamed_tournament_faces(user_id)
    return out


def _tournament_face_target_id(user_id, face_key):
    """Return the targets.id row id for the given tournament face for
    this user, seeding on the fly if it doesn't exist yet."""
    if user_id is None or face_key not in TOURNAMENT_FACES:
        return None
    face = TOURNAMENT_FACES[face_key]
    name = TOURNAMENT_TARGET_NAME_PREFIX + face['name']
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                (user_id, name)
            ).fetchone()
            if row is not None:
                return int(row[0])
    except SQLAlchemyError:
        return None
    seeded = _seed_tournament_faces(user_id)
    return seeded.get(face_key)


def _tournament_face_render_payload(face_key):
    """Return a dict the template's JS uses to draw the ring overlay.
    Includes the zones list with colors for visual rendering plus the
    background color shown behind the rings."""
    face = TOURNAMENT_FACES.get(face_key)
    if face is None:
        return None
    return {
        'face_key':         face_key,
        'face_bg':          face['face_bg'],
        # rings — drawn outer→inner so painter's-order produces the
        # right z-stack (innermost on top).
        'rings':            [
            {'point_value': pv, 'radius_mm': r, 'color': c}
            for (pv, r, c) in reversed(face['zones'])
        ],
        'x_ring_mm':        face['x_ring_mm'],
        # 'x' or 'cross' (or None) — the small center mark painted on
        # top of the innermost ring so users can see the exact center
        # without relying on the X-ring boundary alone.
        'center_mark':      face.get('center_mark'),
        # When present, the face renders as multiple identical spots
        # (e.g. NFAA 5-spot). centers_mm are spot centers relative to
        # the face center; rings/center_mark are drawn at each spot.
        'multi_spot':       face.get('multi_spot'),
    }


def _compute_tournament_progress(session_id, user_id, round_def,
                                 chunk_by_stored=False):
    """Compute the per-end / total / X count summary for one tournament
    session. Returns a dict the template renders into the score panel.

    Reads every shot for the session in time order, scores it via the
    target's seeded zones (same code path as analytics), groups by
    arrows_per_end into "ends" (honoring per-segment end size for
    multi-segment rounds), and reports cumulative totals plus the
    inner-10 / X count.

    `chunk_by_stored` switches end-chunking to the practice-scorecard
    model: ends are read off the stored rows (closed when a row's
    `arrows_remaining` hits 0) instead of the round's fixed segment plan,
    so a paper practice scorecard with arbitrary end sizes scores
    correctly. In this mode every shot is on the round's primary face
    (segment 0) and `arrows_planned` tracks what was actually shot, so
    the card always reads as complete.
    """
    segments = _round_segments(round_def)
    planned_total = _round_total_arrows(round_def)
    # Max-per-arrow is taken over every face used in the round, so a
    # multi-segment round whose long-distance face caps at 10 but
    # short-distance face caps at 6 reports the right "10 count" per
    # face — which we currently flatten into one counter; future work
    # could split it per-segment.
    max_per_arrow = 0
    for seg in segments:
        face = TOURNAMENT_FACES.get(seg['face_key'], {})
        zones = face.get('zones') or ()
        for pv, _r, _c in zones:
            if pv > max_per_arrow:
                max_per_arrow = pv

    out = {
        'arrows_shot':      0,
        'arrows_planned':   planned_total,
        'total_score':      0,
        'x_count':          0,
        'ten_count':        0,
        'ends':             [],
        'is_complete':      False,
        'max_per_arrow':    max_per_arrow,
    }
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            shots = cur.execute(
                "SELECT x_coord, y_coord, target_id, arrow_shaft_diameter, "
                "arrows_remaining "
                "FROM apollo WHERE session_id = %s AND user_id = %s "
                "ORDER BY timestamp, id",
                (session_id, user_id)
            ).fetchall()
    except SQLAlchemyError:
        return out

    # Cache zones per target_id (a multi-segment round may switch
    # face_key across segments; each segment's target_id is distinct
    # so the cache resolves to the right ring list per shot).
    zones_cache = {}
    def _zones(tid):
        if tid not in zones_cache:
            zones_cache[tid] = _fetch_target_zones(tid, user_id) if tid else []
        return zones_cache[tid]

    running = 0
    end_arrows = []
    out['arrows_shot'] = len(shots)
    for shot_idx, r in enumerate(shots):
        if chunk_by_stored:
            # Practice scorecard: one face (segment 0), ends read off the
            # stored rows rather than a fixed plan.
            seg_idx, seg = 0, segments[0]
        else:
            # Resolve the segment that this shot belongs to, so we use the
            # right end size for chunking and the right face's X-ring radius
            # for the X counter.
            seg_idx, seg, _into = _segment_for_shot(segments, shot_idx)
        seg_face = TOURNAMENT_FACES.get(seg['face_key'], {})
        x_ring_radius = float(seg_face.get('x_ring_mm') or 0.0)
        seg_arrows_per_end = int(seg['arrows_per_end'])
        # Multi-spot faces (NFAA 5-spot): a shot is scored against the
        # nearest spot center. Pull the per-face spec straight from
        # TOURNAMENT_FACES — this is the in-memory canonical, so the
        # data is always current even if the row's seeded copy lags.
        seg_centers = None
        seg_multi = seg_face.get('multi_spot') or None
        if seg_multi and seg_multi.get('centers_mm'):
            seg_centers = list(seg_multi['centers_mm'])

        xraw = str(r['x_coord']).strip() if r['x_coord'] is not None else ''
        yraw = str(r['y_coord']).strip() if r['y_coord'] is not None else ''
        shaft = _row_get(r, 'arrow_shaft_diameter')
        points = _score_one_shot(xraw, yraw, _zones(r['target_id']), shaft,
                                 spot_centers_mm=seg_centers)
        out['total_score'] += points
        seg_max = max((pv for pv, _r, _c in (seg_face.get('zones') or ())),
                      default=0)
        if seg_max and points == seg_max:
            out['ten_count'] += 1
        # X count: inside the X ring (using line-cutter slack). For
        # multi-spot faces, measure to the nearest spot center.
        if _shot_is_x(xraw, yraw, x_ring_radius,
                      shaft_diameter_mm=shaft, spot_centers_mm=seg_centers):
            out['x_count'] += 1
        running += points
        end_arrows.append({'points': points, 'x': xraw, 'y': yraw,
                           'miss': (xraw == MISS_SENTINEL and yraw == MISS_SENTINEL)})
        if chunk_by_stored:
            # Close the end when the stored countdown reaches 0 (the last
            # arrow of each end was written with arrows_remaining == 0).
            try:
                ar = int(_row_get(r, 'arrows_remaining'))
            except (TypeError, ValueError):
                ar = 0
            end_done = (ar <= 0)
        else:
            end_done = (len(end_arrows) == seg_arrows_per_end)
        if end_done:
            out['ends'].append({
                'arrows':       list(end_arrows),
                'end_total':    sum(a['points'] for a in end_arrows),
                'running':      running,
                'segment_idx':  seg_idx,
            })
            end_arrows = []
    # Trailing partial end — show it so the user sees what they've shot.
    if end_arrows:
        out['ends'].append({
            'arrows':    list(end_arrows),
            'end_total': sum(a['points'] for a in end_arrows),
            'running':   running,
            'partial':   True,
        })
    if chunk_by_stored:
        # Arbitrary-length practice card: "planned" == what was shot, so
        # the card always reads as complete.
        out['arrows_planned'] = out['arrows_shot']
    out['is_complete'] = out['arrows_shot'] >= out['arrows_planned']
    return out


SESSION_DT_FMT_WITH_MS = '%Y-%m-%d %H:%M:%S.%f'
SESSION_DT_FMT = '%Y-%m-%d %H:%M:%S'


def _app_now():
    """Single source of truth for "now" on every DB write.

    Returns a timezone-naive UTC datetime. Auth-side timestamps
    (locked_until, password-reset token expiry) compare against UTC, so
    every other DB-bound timestamp uses the same clock to avoid
    cross-table TZ drift: ``end_time - begin_time`` is meaningful only
    if both legs were stamped on the same clock. Display-only places
    (form defaults, backup filenames) can still use ``datetime.now()``
    to show the user wall-clock time in their server's local zone.

    Routes through ``datetime.now(timezone.utc).replace(tzinfo=None)`` so
    we stop emitting Python-3.12 DeprecationWarnings from ``utcnow()``
    while keeping the value naive (matches existing column shapes).
    """
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _parse_session_dt(raw):
    """Parse a user-supplied datetime string. Returns None on bad format."""
    if not raw:
        return None
    s = raw.strip()
    for fmt in (SESSION_DT_FMT, SESSION_DT_FMT_WITH_MS):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _user_tz():
    """Resolve the signed-in user's IANA zone, falling back to UTC.

    Called from request handlers where ``current_user()`` is cheap (memoized
    on flask.g). Anonymous requests and unknown zone names both fall back
    to UTC so callers never have to guard the return value.
    """
    name = 'UTC'
    if current_user_id() is not None:
        user = current_user()
        if user is not None:
            try:
                name = user['timezone'] or 'UTC'
            except (KeyError, IndexError, TypeError):
                name = 'UTC'
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError:
        return ZoneInfo('UTC')


def _utc_to_user(val):
    """Convert a stored UTC-naive datetime (or DT-shaped string) into the
    user's zone. Returns a tz-aware datetime, or None on unparseable input.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        dt = val
    else:
        dt = _parse_session_dt(str(val))
        if dt is None:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_user_tz())


def _format_session_dt_user(val):
    """Render a stored UTC datetime as a wall-clock string in the user's zone."""
    dt = _utc_to_user(val)
    return dt.strftime(SESSION_DT_FMT) if dt is not None else ''


def _parse_session_dt_user(raw):
    """Parse a user-typed local datetime string and return UTC-naive.

    Mirrors the storage convention: every DB datetime is naive UTC, so the
    user types in their own zone and we convert before persisting.
    """
    naive = _parse_session_dt(raw)
    if naive is None:
        return None
    return naive.replace(tzinfo=_user_tz()).astimezone(timezone.utc).replace(tzinfo=None)


def get_stats(session_id, user_id):
    """Compute summary stats for one session, scoped to one user.

    Returns a dict on success, or a ``(message, status_code)`` tuple on
    failure — callers must check ``isinstance(result, tuple)`` before
    treating it as stats. The returned dict includes session length, total
    shots, hit/miss totals, and a per-quiver hit/miss breakdown of every
    *completed* quiver (partial trailing quivers are intentionally skipped).

    user_id scoping ensures no user can compute stats for another user's
    session by guessing a session_id — the query simply returns no row.
    """
    if session_id is None:
        return "Error: get_stats() requires a session_id", 500
    if user_id is None:
        return "Error: get_stats() requires a user_id", 500
    stats = {}
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT session_begin_time, session_end_time, manual_session_length_minutes "
                "FROM session_times WHERE session_id = %s AND user_id = %s",
                (session_id, user_id)
            ).fetchone()
            if row is None:
                print(f"❌ No session_times row found for session_id={session_id}")
                return "Session times not found in get_stats()", 500
            session_begin_time, session_end_time, manual_session_length_minutes = row
            stats.update({
                "session_begin_time": session_begin_time,
                "session_end_time": session_end_time,
                "manual_session_length_minutes": manual_session_length_minutes,
            })
    except SQLAlchemyError as e:
        print(f"❌ Session times retrieval error in get_stats(): {e}")
        return "Session times retrieval error in get_stats()", 500

    fmt_with_ms = '%Y-%m-%d %H:%M:%S.%f'
    fmt_no_ms = '%Y-%m-%d %H:%M:%S'
    def to_dt(val):
        if isinstance(val, datetime):
            return val
        val = str(val).strip()
        for fmt in (fmt_with_ms, fmt_no_ms):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                continue
        return None

    dt1 = to_dt(stats["session_begin_time"])
    if dt1 is None:
        print(f"❌ Unparseable session_begin_time={stats['session_begin_time']!r} in get_stats()")
        return "Bad session begin time in get_stats()", 500

    # Manual override wins: when the user typed a session length on the
    # end-session form, ignore the clock and use their value. Used when the
    # session was paused/interrupted and the wall-clock would be misleading.
    manual_mins = manual_session_length_minutes
    try:
        manual_mins_f = float(manual_mins) if manual_mins is not None else 0.0
    except (TypeError, ValueError):
        # Corrupt/imported value — treat as "no manual override".
        manual_mins_f = 0.0
    if manual_mins_f > 0:
        total = int(manual_mins_f * 60)
        days, remainder = divmod(total, 86400)
        hours, remainder = divmod(remainder, 3600)
        minutes, seconds = divmod(remainder, 60)
        pretty_session_length = f'{hours + days * 24} hours, {minutes} minutes, {seconds} seconds'
    elif stats["session_end_time"] is None:
        # Incomplete session: zero out the numeric fields so downstream
        # templates and arithmetic don't break on None. The pretty string
        # carries the "incomplete" signal for human display.
        days = hours = minutes = seconds = 0
        pretty_session_length = "Session incomplete"
    else:
        dt2 = to_dt(stats["session_end_time"])
        if dt2 is None:
            print(f"❌ Unparseable session_end_time={stats['session_end_time']!r} in get_stats()")
            return "Bad session end time in get_stats()", 500
        elapsed = dt2 - dt1
        total = int(abs(elapsed.total_seconds()))
        days, remainder = divmod(total, 86400)
        hours, remainder = divmod(remainder, 3600)
        minutes, seconds = divmod(remainder, 60)
        pretty_session_length = f'{hours + days * 24} hours, {minutes} minutes, {seconds} seconds'
    stats.update({"days": days, "hours": hours, "minutes": minutes, "seconds": seconds})
    stats.update({"pretty_session_length": pretty_session_length})
    # Localize the date to the user's zone so a 23:30 PT session doesn't
    # show as "tomorrow" just because the stored timestamp is UTC.
    stats.update({"session_date": (_utc_to_user(dt1) or dt1).strftime('%Y-%m-%d')})
    try: # get number of arrows shot
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            res = cur.execute(
                "SELECT COUNT(*) FROM apollo WHERE session_id = %s AND user_id = %s",
                (session_id, user_id))
            row = cur.fetchone()
            arrows_shot = row[0] if row else 0
            stats.update({"arrows_shot": arrows_shot})
    except SQLAlchemyError as e:
        print(f"❌ Num arrows shot error in get_stats(): {e}")
        return "Num arrows shot error in get_stats()", 500
    try: # get percent on target
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # When the session's target has scored zones, a shot outside
            # the most peripheral zone is treated as a miss — only shots
            # whose shaft touches a ring count as hits. Without zones we
            # fall back to "anything that isn't the miss sentinel is a hit"
            # since we have no way to tell on-target from off-target.
            shot_class_rows = cur.execute(
                "SELECT x_coord, y_coord, target_id, arrow_shaft_diameter "
                "FROM apollo WHERE session_id = %s AND user_id = %s "
                "ORDER BY id ASC",
                (session_id, user_id)
            ).fetchall()
            sw_target_id = None
            for srow in shot_class_rows:
                if srow['target_id'] is not None:
                    sw_target_id = int(srow['target_id'])
                    break
            sw_zones = _fetch_target_zones(sw_target_id, user_id) \
                if sw_target_id is not None else []
            sw_scoring = _zones_define_scoring(sw_zones)
            # Multi-spot faces (NFAA 5-spot): score against the nearest
            # spot center so a corner-spot hit isn't read as a miss off
            # the face origin. Matches the tournament-progress code path.
            sw_centers = (_target_multi_spot_centers(sw_target_id, user_id)
                          if sw_target_id is not None else None)
            missed_shots = 0
            for srow in shot_class_rows:
                xraw = str(srow['x_coord']).strip() if srow['x_coord'] is not None else ''
                yraw = str(srow['y_coord']).strip() if srow['y_coord'] is not None else ''
                if sw_scoring:
                    if _classify_shot(xraw, yraw, sw_zones,
                                      _row_get(srow, 'arrow_shaft_diameter'),
                                      spot_centers_mm=sw_centers) is None:
                        missed_shots += 1
                elif xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
                    missed_shots += 1
            hit_shots = arrows_shot - missed_shots
            percent_missed = round((missed_shots / arrows_shot) * 100, 2) if arrows_shot else 0.0
            percent_hit = round((hit_shots / arrows_shot) * 100, 2) if arrows_shot else 0.0
            print(f"Percent missed: {percent_missed}; Missed shots: {missed_shots}")
            stats.update({"percent_missed": percent_missed, "missed_shots": missed_shots,
                          "percent_hit": percent_hit, "hit_shots": hit_shots})
    except SQLAlchemyError as e:
        print(f"❌ Percent missed error in get_stats(): {e}")
        return "Percent missed error in get_stats()", 500
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Quiver size is stored per-shot (the user may change it
            # *between* quivers — the lock in the session view forbids
            # mid-quiver changes), so we walk shots in insertion order
            # and group them using each row's own quiver_size. The
            # first row in each group is the "leader" — its quiver_size
            # is the size of that quiver; we close out the group when
            # we've accumulated that many rows.
            #
            # Sessions lock the target after the first shot, so one
            # zones fetch (from the first row's target_id) covers every
            # quiver below.
            shot_rows = cur.execute(
                "SELECT quiver_size, target_id, x_coord, y_coord, is_precise, "
                "       arrow_shaft_diameter "
                "FROM apollo "
                "WHERE session_id = %s AND user_id = %s "
                "ORDER BY id ASC",
                (session_id, user_id)
            ).fetchall()

            first_target_id = (int(shot_rows[0]['target_id'])
                               if shot_rows and shot_rows[0]['target_id'] is not None
                               else None)
            zones = _fetch_target_zones(first_target_id, user_id) \
                if first_target_id is not None else []
            scoring_available = _zones_define_scoring(zones)
            # Nearest-spot centers for multi-spot faces (see sw_centers above).
            spot_centers = (_target_multi_spot_centers(first_target_id, user_id)
                            if first_target_id is not None else None)
            max_zone_points = max((int(z['point_value'] or 0) for z in zones),
                                  default=0) if scoring_available else 0

            # Slice shot_rows into completed quivers. A trailing partial
            # quiver (user ended mid-batch, or current in-progress quiver)
            # is intentionally skipped — the per-quiver stats are only
            # meaningful for completed quivers. Those shots are still
            # counted by the overall hit% calc above.
            stats_by_quiver = {}
            session_total_score = 0
            session_max_score = 0
            scored_quiver_count = 0
            quiver_number = 1
            buf = []
            buf_size = 0  # the leader's quiver_size for the buffered group
            for r in shot_rows:
                try:
                    row_qs = int(r['quiver_size']) if r['quiver_size'] else 0
                except (TypeError, ValueError):
                    row_qs = 0
                if row_qs <= 0:
                    # Bad row — drop the in-progress buffer and skip. We
                    # can't slot it into a quiver of unknown size.
                    buf = []
                    buf_size = 0
                    continue
                if not buf:
                    buf_size = row_qs
                buf.append(r)
                if len(buf) >= buf_size:
                    number_hit = 0
                    number_missed = 0
                    for q in buf:
                        # When the target has scored zones, a shot that
                        # lands outside the most peripheral zone is treated
                        # as a miss (no scoring ring touched). Without zones
                        # we fall back to the sentinel-only definition.
                        if scoring_available:
                            idx = _classify_shot(
                                str(q['x_coord']).strip() if q['x_coord'] is not None else '',
                                str(q['y_coord']).strip() if q['y_coord'] is not None else '',
                                zones,
                                _row_get(q, 'arrow_shaft_diameter'),
                                spot_centers_mm=spot_centers,
                            )
                            if idx is not None:
                                number_hit += 1
                            else:
                                number_missed += 1
                        elif (str(q['x_coord']) != MISS_SENTINEL
                                and str(q['y_coord']) != MISS_SENTINEL):
                            number_hit += 1
                        else:
                            number_missed += 1
                    percent_hit = round((number_hit / buf_size) * 100, 2)
                    percent_missed = round((number_missed / buf_size) * 100, 2)
                    q_stat = {
                        "quiver_number": quiver_number,
                        "number_hit": number_hit,
                        "percent_hit": percent_hit,
                        "number_missed": number_missed,
                        "percent_missed": percent_missed,
                    }
                    # Score only when zones are configured AND every shot
                    # in the quiver was precisely measured — estimated
                    # placements aren't reliable enough to count as points.
                    all_precise = all(
                        int(q['is_precise'] or 0) == 1 for q in buf
                    )
                    if scoring_available and all_precise:
                        score = _compute_quiver_score(buf, zones,
                                                       spot_centers_mm=spot_centers)
                        max_score = buf_size * max_zone_points
                        q_stat["score_points"] = score
                        q_stat["max_score"] = max_score
                        session_total_score += score
                        session_max_score += max_score
                        scored_quiver_count += 1
                    stats_by_quiver[quiver_number] = q_stat
                    quiver_number += 1
                    buf = []
                    buf_size = 0

            stats.update({
                "stats_by_quiver": stats_by_quiver,
                "scoring_available": bool(scoring_available),
                "scored_quiver_count": scored_quiver_count,
                "session_total_score": session_total_score,
                "session_max_score": session_max_score,
            })
    except SQLAlchemyError as e:
        print(f"❌ Percent misses by quiver error in get_stats(): {e}")
        return "Percent misses by quiver error in get_stats()", 500
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT record_mode FROM apollo WHERE session_id = %s AND user_id = %s LIMIT 1",
                (session_id, user_id)
            ).fetchone()
            stats["record_mode"] = int(row[0]) if row and row[0] is not None else 0
    except SQLAlchemyError:
        stats["record_mode"] = 0
    return stats


TARGETS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', TARGETS_SUBDIR)
_STATIC_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')


def _resolve_target_image_disk_path(image_filename):
    """Resolve a target row's ``image_filename`` to an absolute disk path,
    but only if the resolved path stays inside ``TARGETS_DIR``.

    Used as the gatekeeper before any ``os.remove`` on an image — even if a
    crafted import slipped a path-traversal string into the DB, the realpath
    check refuses to act on anything outside the uploads directory.
    """
    if not image_filename or not isinstance(image_filename, str):
        return None
    if not image_filename.startswith(TARGETS_SUBDIR + '/'):
        return None
    try:
        disk_real = os.path.realpath(os.path.join(_STATIC_ROOT, image_filename))
    except (TypeError, ValueError):
        return None
    targets_real = os.path.realpath(TARGETS_DIR)
    if disk_real == targets_real:
        return None
    if not disk_real.startswith(targets_real + os.sep):
        return None
    return disk_real

# Schema is defined once via SQLAlchemy Core Table objects and compiled
# to the right dialect (SQLite AUTOINCREMENT / MySQL AUTO_INCREMENT etc.)
# by metadata.create_all(). Every table gets an explicit `id` PK — MySQL
# has no implicit rowid like SQLite does — and SELECTs throughout the app
# alias `id AS rowid` so templates/JS that read `rowid` keep working.
metadata = MetaData()

# All per-user data tables carry a user_id FK-style column. We don't declare
# it as a SQL FOREIGN KEY so that pre-existing DBs (created before multi-user
# support landed) can be migrated in place with a plain ALTER TABLE ADD
# COLUMN — see ensure_user_id_columns() in migrate_db().

users_table = Table('users', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    # username + email both unique. Username is what the user types at login;
    # email is captured for recovery flows and to discourage account farming.
    Column('username', String(64), unique=True, nullable=False),
    Column('email', String(255), unique=True, nullable=False),
    # Werkzeug's scrypt format is ~120 chars — 255 leaves room for future
    # hash schemes without another migration.
    Column('password_hash', String(255), nullable=False),
    Column('created_at', DateTime),
    Column('last_login', DateTime),
    Column('is_active', Integer, server_default=text('1')),
    # Lockout state: failed_attempts increments on each bad password; once
    # locked_until is in the future, login is refused regardless of password.
    Column('failed_attempts', Integer, server_default=text('0')),
    Column('locked_until', DateTime),
    # Root flag — admins can browse/delete users and reset their passwords or
    # emails. Bootstrapped via APOLLO_ROOT_* env vars at startup (set by
    # install.py). No UI for granting/revoking root: keep promotion out-of-band
    # to avoid an admin race in the app itself.
    Column('is_root', Integer, server_default=text('0')),
    # IANA timezone name (e.g. "America/Denver"). All DB datetimes are
    # naive-UTC; helpers _utc_to_user / _format_session_dt_user convert
    # to this zone for display, and _parse_session_dt_user converts user
    # input back to UTC. Validated against zoneinfo.available_timezones()
    # on write, so a stale row never feeds an invalid name to ZoneInfo().
    Column('timezone', String(64), server_default=text("'UTC'")),
    # Archer profile for handicap/classification. All nullable — handicaps
    # need none of these; classifications use them to pick the right Archery
    # GB threshold table. ``age_group`` is stored as a label (e.g. 'adult',
    # '50+', 'under 18') so there is no birthday math; ``gender`` is
    # 'male'/'female'/'open'; ``default_bowstyle`` seeds the per-round
    # bowstyle (recurve/compound/barebow/longbow/…).
    Column('gender', String(16)),
    Column('age_group', String(32)),
    Column('default_bowstyle', String(32)),
)

apollo_table = Table('apollo', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True),
    Column('session_id', Integer, index=True),
    Column('timestamp', DateTime),
    # ``bow`` is the bow_model name string at shot time. The other bow_*
    # columns below are denormalized snapshots of the rest of the bow's
    # configuration when the arrow was loosed — without them, later edits
    # to the bow row silently rewrite the historical attribution of every
    # prior shot. The bows table is still the source of truth for the
    # *current* config of each bow.
    Column('bow', String(255)),
    Column('arrow_type', String(255)),
    Column('quiver_size', Integer),
    Column('arrows_remaining', Integer),
    Column('distance', String(64)),
    Column('session_notes', Text),
    Column('x_coord', String(32)),
    Column('y_coord', String(32)),
    Column('is_precise', Integer, server_default=text('0')),
    Column('record_mode', Integer, server_default=text('0')),
    Column('target_id', Integer),
    Column('nock_height', String(64)),
    Column('bow_draw_weight', String(64)),
    Column('effective_draw_weight', String(64)),
    Column('bow_amo', String(64)),
    Column('bow_type', String(255)),
    # Arrow snapshot columns — same rationale as the bow_* columns above.
    # ``arrow_type`` already holds the arrow name string; these capture
    # the rest of the arrow's config at shot time so later edits or
    # renames don't rewrite historical attribution.
    Column('arrow_length', String(64)),
    Column('arrow_spine', String(64)),
    Column('arrow_shaft_weight', String(64)),
    Column('arrow_shaft_diameter', String(64)),
    Column('arrow_shaft_material', String(255)),
    Column('arrow_nock_weight', String(64)),
    Column('arrow_tip', String(255)),
    Column('arrow_tip_weight', String(64)),
    # Comma-separated session-level tags. Denormalized onto every shot row
    # for the same reason ``session_notes`` is — keeps each row
    # self-describing without a separate session-level table.
    Column('session_tags', Text),
    # JSON snapshot of the bowtype-specific settings (merged static bow defaults
    # + per-session dynamic overrides) in force when this arrow was loosed. Same
    # self-describing rationale as the bow_* snapshot columns above.
    Column('bow_style_settings', Text),
    # Client-generated idempotency key for AJAX shot submits (/api/record_shot).
    # Lets a retried POST after a lost response (the iPad keep-alive race) be
    # detected as a duplicate and folded into a success instead of inserting a
    # second row. NULL for full-page (/sesh) and offline-synced shots, which
    # never retry. UNIQUE(user_id, client_uuid) — added by
    # _ensure_apollo_client_uuid_unique_index — enforces it; NULLs are distinct
    # under both SQLite and MySQL so unlimited NULL rows coexist.
    Column('client_uuid', String(64)),
)

arrows_table = Table('arrows', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True),
    Column('arrow', String(255)),
    Column('length', String(64)),
    Column('spine', String(64)),
    Column('shaft_weight', String(64)),
    Column('shaft_diameter', String(64)),
    Column('shaft_material', String(255)),
    Column('nock_weight', String(64)),
    Column('tip', String(255)),
    Column('tip_weight', String(64)),
)

bows_table = Table('bows', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True),
    Column('bow_model', String(255)),
    Column('bow_type', String(255)),
    Column('bow_draw_weight', String(64)),
    # Effective draw weight — what the archer is *actually* drawing, after
    # any adjustments (let-off on compounds, reduced draw weight, tuned
    # limb bolts, etc.). When set, this is what data-analysis prefers over
    # ``bow_draw_weight`` for per-shot weight; left blank, fall back to
    # the rated draw weight.
    Column('effective_draw_weight', String(64)),
    Column('amo', String(64)),
    Column('nock_height', String(64)),
    # AGB bowstyle for this bow (recurve/compound/barebow/longbow/…). Used as
    # the per-round bowstyle default when shooting with this bow, feeding the
    # classification category. Nullable; falls back to the user's
    # ``default_bowstyle`` then 'recurve'.
    Column('bowstyle', String(32)),
    # JSON object of this bow's static, bowtype-specific gear defaults (compound
    # release aid / let-off, recurve clicker / stabilizer, etc.). Schema lives
    # in BOWSTYLE_SETTINGS; nullable, empty for bows added before the field set.
    Column('style_settings', Text),
)

session_times_table = Table('session_times', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True),
    Column('session_id', Integer, index=True),
    Column('session_begin_time', DateTime),
    Column('session_end_time', DateTime),
    Column('manual_session_length_minutes', Float),
)

# Targets are square. physical_size_mm = edge length of the printed face;
# image_size_px = pixel edge of the image file (auto-detected on upload).
# is_default=1 marks the target preselected for new sessions.
#
# Target names used to be globally unique; with multi-user that becomes a
# foot-gun ("Bob's Yellow 60cm" blocking Alice from naming hers the same),
# so uniqueness is enforced per-user via a CHECK in app code rather than a
# DB constraint.
# Password-reset tokens. We store only sha256(token), never the token itself,
# so a DB leak doesn't hand attackers live reset links. The plaintext token
# is mailed once to the user and never touches storage. Tokens are one-shot
# (used_at flips on first successful reset) and short-lived (see
# PASSWORD_RESET_TTL_MINUTES below).
password_resets_table = Table('password_resets', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True, nullable=False),
    Column('token_hash', String(64), unique=True, nullable=False),
    Column('created_at', DateTime),
    Column('expires_at', DateTime),
    Column('used_at', DateTime),
)

# Persistent rate-limit counters. In-memory dicts would reset on restart
# (handing the attacker a free reset window every deploy) and don't share
# state across multiple WSGI workers. Each row is one (scope, key) pair
# with the window start and the hit count; the helper bumps and trims in
# the same transaction. ``scope`` distinguishes the limiter (login vs.
# forgot-password) so the same IP can have an independent budget for each.
rate_limit_hits_table = Table('rate_limit_hits', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('scope', String(32), nullable=False, index=True),
    # ``rate_key`` (not ``key`` — that's a reserved word in MySQL and would
    # need quoting in every statement). One row per (scope, rate_key) pair;
    # the unique index below is what makes the SELECT/INSERT/UPDATE
    # sequence race-safe.
    Column('rate_key', String(64), nullable=False),
    Column('window_start', DateTime, nullable=False),
    Column('hits', Integer, nullable=False, server_default=text('0')),
    UniqueConstraint('scope', 'rate_key', name='uq_rate_limit_hits_scope_key'),
)

targets_table = Table('targets', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True),
    Column('name', String(255)),
    Column('image_filename', String(512)),
    Column('physical_size_mm', Float),
    Column('image_size_px', Integer),
    Column('is_active', Integer, server_default=text('1')),
    Column('is_default', Integer, server_default=text('0')),
)

# User-defined scoring zones for a target. Concentric circles for now —
# shape_type is reserved so we can add polygons/ellipses later without a
# schema change. radius_mm is the *outer* radius of the ring measured from
# the calibrated target center; the inner edge is implied by the next-
# smaller ring (or 0 for the innermost). Zones are independent rings, not
# necessarily nested — the score for a shot is the ring with the smallest
# radius that still contains it.
target_zones_table = Table('target_zones', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True, nullable=False),
    Column('target_id', Integer, index=True, nullable=False),
    Column('name', String(255)),
    Column('point_value', Integer, server_default=text('0')),
    Column('shape_type', String(32), server_default=text("'circle'")),
    Column('radius_mm', Float),
    Column('display_order', Integer, server_default=text('0')),
)


# Global key/value config rows shared across all users. Currently holds
# only ``server_timezone`` (set by root from /account) but kept generic so
# future single-instance settings don't need their own table. ``setting_key``
# is unique; one row per key, updated in place.
app_settings_table = Table('app_settings', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('setting_key', String(64), unique=True, nullable=False),
    Column('setting_value', Text),
    Column('updated_at', DateTime),
)


# Personal scratchpad — one row per user holding free-form notes that aren't
# tied to a session. Updated in place by /notes.
user_notes_table = Table('user_notes', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, unique=True, nullable=False),
    Column('content', Text),
    Column('updated_at', DateTime),
)


# Form-analysis captures from the /form motion-capture page. The raw video is
# NEVER stored or transmitted — pose extraction and scoring run entirely in the
# browser. Only the derived, anonymous numbers land here: the measured joint
# angles (metrics_json) and the per-checkpoint scores (scores_json), plus the
# overall score, the bowstyle the run was graded against, and which captured
# key-frame ('anchor'/'follow_through') the angles came from. One row per saved
# analysis, owned by user_id like every other per-user table.
form_captures_table = Table('form_captures', metadata,
    Column('id', Integer, primary_key=True, autoincrement=True),
    Column('user_id', Integer, index=True, nullable=False),
    Column('created_at', DateTime),
    Column('bowstyle', String(32)),
    Column('overall_score', Float),
    Column('metrics_json', Text),
    Column('scores_json', Text),
    # Optional compact, downsampled pose sequence for replay — anonymous
    # skeleton points only (no video, no image), so a saved shot can be
    # re-watched as a silhouette with the angle lines drawn over it. Nullable:
    # captures saved before replay existed (or when the payload is too big)
    # simply have no replay.
    Column('frames_json', Text),
)


# ─── Backend selector ────────────────────────────────────────────────────
# Explicit manual switch — SQLite is the default in every case, and MySQL
# only kicks in when APOLLO_BACKEND=mysql is set. A stray DATABASE_URL in
# the environment is ignored unless that flag is on, so local runs can't
# accidentally point at a remote DB.
#   APOLLO_BACKEND=mysql  → use DATABASE_URL (raises if unset)
#   anything else / unset → local apollo.db next to this file
APOLLO_BACKEND = (os.environ.get('APOLLO_BACKEND') or '').strip().lower()

_DEFAULT_SQLITE = "sqlite:///" + os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "apollo.db"
)
if APOLLO_BACKEND == 'mysql':
    DATABASE_URL = os.environ.get('DATABASE_URL')
    if not DATABASE_URL:
        raise RuntimeError(
            "APOLLO_BACKEND='mysql' but DATABASE_URL is not set. "
            "Export DATABASE_URL=mysql+pymysql://user:pass@host/db, "
            "or unset APOLLO_BACKEND to use local SQLite."
        )
elif APOLLO_BACKEND in ('', 'sqlite'):
    DATABASE_URL = _DEFAULT_SQLITE
else:
    raise RuntimeError(
        f"APOLLO_BACKEND={APOLLO_BACKEND!r} — expected 'mysql', 'sqlite', or unset."
    )
# Print which backend we're talking to at startup — invaluable when
# you're flipping APOLLO_BACKEND mid-debug. Redact the password so an
# accidental terminal screenshot doesn't leak credentials.
_redacted = re.sub(r'(://[^:/@]+:)[^@]+(@)', r'\1***\2', DATABASE_URL)
print(f"📦 Apollo DB: {_redacted}")
# future=True is the SQLAlchemy 2.x default but kept explicit so a 1.4
# environment behaves the same way (commit-as-you-go transactions).
#
# pool_pre_ping + pool_recycle are MySQL hygiene: shared-MySQL hosts
# (PythonAnywhere, etc.) silently close idle connections after a few
# minutes, but SQLAlchemy keeps them pooled. The next checkout then
# fails mid-query with "Lost connection to MySQL server during query"
# or "SSL bad record mac". pre_ping issues a cheap SELECT 1 on checkout
# and transparently swaps out a dead conn; recycle proactively retires
# any conn older than 280s (under PA's 5-minute idle cutoff).
# SQLite ignores these (single-file DB, no network), so we apply them
# unconditionally — no harm, just no benefit on the SQLite side.
engine = create_engine(
    DATABASE_URL,
    future=True,
    pool_pre_ping=True,
    pool_recycle=280,
)


# Bridges SQLAlchemy 2.x's connection.execute(text(...), {...}) shape onto
# the DB-API-ish `cur.execute(sql, args).fetchone()` shape the rest of this
# file uses. Lets the routes stay backend-agnostic without rewriting every
# call site.
#   - %s positional placeholders → :p0, :p1, ... named binds at execute time
#   - fetched rows wrapped in HybridRow so both `row['col']` and `row[0]`
#     access keep working (sqlite3.Row had both; SQLAlchemy's Row only has
#     attribute/key access).
class HybridRow(dict):
    """sqlite3.Row lookalike: supports both ``row['col']`` and ``row[0]``.

    SQLAlchemy's Row only exposes attribute/key access, so this wrapper
    backfills positional indexing and value-iteration to keep the existing
    call sites (and templates expecting tuple-style access) working.
    """
    def __init__(self, mapping):
        super().__init__(mapping)
        self._values = list(mapping.values())
    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        return super().__getitem__(key)
    def __iter__(self):
        # Match sqlite3.Row: iteration yields values, so tuple unpacking
        # like `a, b, c = row` works. Plain dict iter would yield keys.
        return iter(self._values)


def _to_named_params(sql, args):
    if not args:
        return sql, {}
    parts = sql.split('%s')
    if len(parts) - 1 != len(args):
        # Mismatch — let SQLAlchemy surface a real error rather than
        # silently mangling the query.
        return sql, args
    out_sql = parts[0]
    out_params = {}
    for i, val in enumerate(args):
        key = f'p{i}'
        out_sql += f':{key}' + parts[i + 1]
        out_params[key] = val
    return out_sql, out_params


class CompatCursor:
    def __init__(self, conn):
        self._conn = conn
        self._result = None
    def execute(self, sql, args=None):
        named_sql, params = _to_named_params(sql, args)
        self._result = self._conn.execute(text(named_sql), params)
        return self
    def fetchone(self):
        row = self._result.fetchone()
        return HybridRow(row._mapping) if row is not None else None
    def fetchall(self):
        return [HybridRow(r._mapping) for r in self._result.fetchall()]
    @property
    def lastrowid(self):
        # SQLAlchemy exposes lastrowid for backends that report it
        # (SQLite, MySQL). Used by /import_data to map old → new target ids.
        return getattr(self._result, 'lastrowid', None)
    @property
    def rowcount(self):
        # SQLAlchemy 2.x Result.rowcount: number of rows affected by the last
        # DML statement. -1 if unsupported. Used by callers that need to tell
        # "UPDATE matched a row" from "UPDATE matched nothing" — see the
        # atomic token consume in /reset_password.
        return getattr(self._result, 'rowcount', -1)
    def close(self):
        pass


class CompatConnection:
    """Holds one SQLAlchemy Connection plus a manually managed transaction.

    commit() finalizes the open tx and starts a fresh one so a sequence of
    write/commit/write/commit inside a single `with` block keeps working
    the way the existing call sites expect. close() rolls back any work
    that wasn't committed.
    """
    def __init__(self, engine):
        self._conn = engine.connect()
        self._tx = self._conn.begin()
    def cursor(self):
        return CompatCursor(self._conn)
    @property
    def connection(self):
        return self._conn
    def commit(self):
        self._tx.commit()
        self._tx = self._conn.begin()
    def rollback(self):
        # Roll back the open transaction and start a fresh one so subsequent
        # statements on this connection (e.g. a fall-back UPDATE after a race-
        # condition recount) keep working inside the same `with` block.
        try:
            if self._tx.is_active:
                self._tx.rollback()
        except Exception:
            pass
        self._tx = self._conn.begin()
    def close(self):
        try:
            if self._tx.is_active:
                self._tx.rollback()
        except Exception:
            pass
        self._conn.close()


# Tables that gained a user_id column when the app switched from
# single-user to multi-user. On a fresh DB metadata.create_all() handles
# this; on an existing DB we ALTER TABLE ADD COLUMN at startup.
_PER_USER_TABLES = ('apollo', 'arrows', 'bows', 'session_times', 'targets',
                    'target_zones', 'form_captures')


def _ensure_user_id_columns():
    """Add a `user_id` column to legacy data tables that predate multi-user.

    Idempotent: skips tables that don't exist yet (fresh DB → create_all()
    will add the column with the right type) and tables that already have
    the column. Works on both SQLite (ALTER TABLE ADD COLUMN is fine) and
    MySQL (same statement, different dialect quirks handled by the engine).
    """
    insp = sa_inspect(engine)
    existing = set(insp.get_table_names())
    # _PER_USER_TABLES is a hardcoded tuple literal, but these names are
    # interpolated as identifiers into raw ALTER/CREATE statements below
    # (SQL can't bind identifiers). Validate against an identifier regex
    # so a future refactor that draws table names from env/config can't
    # silently turn this into a SQL-injection sink.
    _IDENT = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
    for tbl in _PER_USER_TABLES:
        if not _IDENT.match(tbl):
            raise RuntimeError(f"unsafe table name in _PER_USER_TABLES: {tbl!r}")
        if tbl not in existing:
            continue
        cols = {c['name'] for c in insp.get_columns(tbl)}
        if 'user_id' in cols:
            continue
        print(f"⚙️  Migrating: adding user_id column to {tbl}")
        with engine.begin() as conn:
            conn.execute(text(f"ALTER TABLE {tbl} ADD COLUMN user_id INTEGER"))
        # Best-effort secondary index on user_id. SQLite and MySQL both
        # accept this form; failures are non-fatal (indexes are an
        # optimization, not a correctness requirement).
        try:
            with engine.begin() as conn:
                conn.execute(text(
                    f"CREATE INDEX IF NOT EXISTS idx_{tbl}_user_id "
                    f"ON {tbl}(user_id)"
                ))
        except SQLAlchemyError:
            pass


def _drop_legacy_unique_target_name():
    """Drop the legacy globally-unique constraint on targets.name.

    Pre-multi-user installs created `targets.name` as UNIQUE. With
    multiple users we want uniqueness scoped *per-user* so two users can
    each have a "Legacy 24\\" face" without colliding — in fact, the
    /register seed-target step *needs* this, otherwise the second user
    to register can't get their bundled legacy target.

    SQLite doesn't have ALTER TABLE DROP CONSTRAINT, so the canonical
    migration recipe is: create a replacement table without the
    constraint, copy rows over, drop the old, rename. MySQL is easier —
    ALTER TABLE … DROP INDEX. Both paths are idempotent.
    """
    insp = sa_inspect(engine)
    if 'targets' not in insp.get_table_names():
        return
    # Look for a unique constraint or unique index that covers exactly the
    # `name` column. Both forms show up depending on how the column was
    # declared on the original install.
    has_legacy = False
    for uq in insp.get_unique_constraints('targets'):
        if uq.get('column_names') == ['name']:
            has_legacy = True
            break
    if not has_legacy:
        for ix in insp.get_indexes('targets'):
            if ix.get('unique') and ix.get('column_names') == ['name']:
                has_legacy = True
                break
    if not has_legacy:
        return

    dialect = engine.dialect.name
    print(f"⚙️  Migrating: dropping legacy UNIQUE constraint on targets.name ({dialect})")

    if dialect == 'sqlite':
        # SQLite recipe: rebuild the table without the UNIQUE column
        # constraint. Wrapped in a single transaction so a crash mid-
        # migration doesn't leave a half-built schema behind.
        with engine.begin() as conn:
            conn.execute(text("PRAGMA foreign_keys=OFF"))
            conn.execute(text("""
                CREATE TABLE targets_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    name VARCHAR(255),
                    image_filename VARCHAR(512),
                    physical_size_mm FLOAT,
                    image_size_px INTEGER,
                    is_active INTEGER DEFAULT 1,
                    is_default INTEGER DEFAULT 0
                )
            """))
            conn.execute(text("""
                INSERT INTO targets_new
                    (id, user_id, name, image_filename, physical_size_mm,
                     image_size_px, is_active, is_default)
                SELECT id, user_id, name, image_filename, physical_size_mm,
                       image_size_px, is_active, is_default
                FROM targets
            """))
            conn.execute(text("DROP TABLE targets"))
            conn.execute(text("ALTER TABLE targets_new RENAME TO targets"))
            conn.execute(text(
                "CREATE INDEX IF NOT EXISTS idx_targets_user_id ON targets(user_id)"
            ))
            conn.execute(text("PRAGMA foreign_keys=ON"))
        return

    if dialect == 'mysql':
        # MySQL: the column-level UNIQUE creates an index named after the
        # column. The actual name may vary across MySQL versions; try the
        # obvious candidates and swallow "doesn't exist" errors.
        _IDENT = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')
        for idx_name in ('name', 'targets_name_key', 'name_2'):
            try:
                with engine.begin() as conn:
                    conn.execute(text(f"ALTER TABLE targets DROP INDEX `{idx_name}`"))
                return
            except SQLAlchemyError:
                continue
        # Fallback: introspect for the actual unique-index name. The name
        # is read from DB metadata, but MySQL identifiers can technically
        # contain backticks (doubled-up) and other special chars. Refuse
        # anything that isn't a plain identifier rather than relying on
        # backtick-quoting alone — a doubled-backtick name would break out
        # of our quoting.
        try:
            with engine.begin() as conn:
                rows = conn.execute(text(
                    "SHOW INDEX FROM targets WHERE Non_unique = 0 AND Column_name = 'name'"
                )).fetchall()
                for r in rows:
                    name = r._mapping.get('Key_name') or r[2]  # 3rd col = Key_name
                    if not name or name == 'PRIMARY':
                        continue
                    if not _IDENT.match(name):
                        print(f"⚠️  Skipping legacy index with unsafe name: {name!r}")
                        continue
                    conn.execute(text(f"ALTER TABLE targets DROP INDEX `{name}`"))
        except SQLAlchemyError as e:
            print(f"⚠️  Could not drop legacy targets.name UNIQUE on MySQL: {e}")


def _ensure_session_times_unique_index():
    """Add a UNIQUE index on session_times(user_id, session_id).

    Closes the /sesh allocation race: MAX(session_id)+1 followed by an
    INSERT is not atomic across concurrent requests (two tabs from the
    same user could pick the same id). With this index a colliding insert
    raises IntegrityError, which /sesh catches and retries.

    Best-effort: if existing data already has duplicates (shouldn't, but
    legacy installs are unpredictable) the CREATE will fail and we log
    and move on rather than crashing the import.
    """
    try:
        insp = sa_inspect(engine)
        if 'session_times' not in set(insp.get_table_names()):
            return
        existing = {ix.get('name') for ix in insp.get_indexes('session_times')}
        if 'ux_session_times_user_session' in existing:
            return
        with engine.begin() as conn:
            conn.execute(text(
                "CREATE UNIQUE INDEX ux_session_times_user_session "
                "ON session_times(user_id, session_id)"
            ))
    except SQLAlchemyError as e:
        print(f"⚠️  Could not add unique index on session_times: {e}")


def _ensure_targets_user_name_unique_index():
    """Add a UNIQUE index on targets(user_id, name).

    Closes the /add_target race: the route does a SELECT-then-INSERT to
    enforce per-user unique target names, which isn't atomic across
    concurrent requests (two tabs from the same user could both pass the
    check and create duplicates). With this index a colliding insert
    raises IntegrityError, which /add_target already catches and
    translates to the friendly "already exists" message.

    Best-effort: a legacy DB with pre-existing duplicates will fail the
    CREATE; we log and move on rather than crashing startup.
    """
    try:
        insp = sa_inspect(engine)
        if 'targets' not in set(insp.get_table_names()):
            return
        existing = {ix.get('name') for ix in insp.get_indexes('targets')}
        if 'ux_targets_user_name' in existing:
            return
        with engine.begin() as conn:
            conn.execute(text(
                "CREATE UNIQUE INDEX ux_targets_user_name "
                "ON targets(user_id, name)"
            ))
    except SQLAlchemyError as e:
        print(f"⚠️  Could not add unique index on targets(user_id, name): {e}")


def _ensure_arrow_dimension_columns():
    """Add shaft_diameter to arrows and nock_height to bows if missing.

    Idempotent: skipped on fresh DBs (create_all() handles them) and on
    DBs that already have the columns from a prior run.
    """
    insp = sa_inspect(engine)
    existing = set(insp.get_table_names())
    if 'arrows' in existing:
        arrow_cols = {c['name'] for c in insp.get_columns('arrows')}
        if 'shaft_diameter' not in arrow_cols:
            print("⚙️  Migrating: adding shaft_diameter column to arrows")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE arrows ADD COLUMN shaft_diameter VARCHAR(64)"))
        if 'shaft_material' not in arrow_cols:
            print("⚙️  Migrating: adding shaft_material column to arrows")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE arrows ADD COLUMN shaft_material VARCHAR(255)"))
    if 'bows' in existing:
        bow_cols = {c['name'] for c in insp.get_columns('bows')}
        if 'nock_height' not in bow_cols:
            print("⚙️  Migrating: adding nock_height column to bows")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE bows ADD COLUMN nock_height VARCHAR(64)"))
    if 'apollo' in existing:
        apollo_cols = {c['name'] for c in insp.get_columns('apollo')}
        if 'nock_height' not in apollo_cols:
            print("⚙️  Migrating: adding nock_height column to apollo")
            with engine.begin() as conn:
                conn.execute(text("ALTER TABLE apollo ADD COLUMN nock_height VARCHAR(64)"))


def _migrate_poundage_to_draw_weight():
    """Rename poundage columns to draw_weight on bows and apollo.

    The correct archery term is "draw weight"; "poundage" was wrong.
    Renames apply to both the source-of-truth ``bows`` row and the
    per-shot ``apollo`` snapshot. Idempotent: only ALTERs when the old
    column is present and the new one isn't. Must run before
    _ensure_shot_snapshot_columns so that function sees the post-rename
    state when deciding which columns are missing.
    """
    insp = sa_inspect(engine)
    existing = set(insp.get_table_names())
    renames = (
        ('bow_poundage',       'bow_draw_weight'),
        ('effective_poundage', 'effective_draw_weight'),
    )
    for table in ('bows', 'apollo'):
        if table not in existing:
            continue
        cols = {c['name'] for c in insp.get_columns(table)}
        for old, new in renames:
            if old in cols and new not in cols:
                print(f"⚙️  Migrating: renaming {table}.{old} → {new}")
                with engine.begin() as conn:
                    conn.execute(text(
                        f"ALTER TABLE {table} RENAME COLUMN {old} TO {new}"
                    ))


def _ensure_shot_snapshot_columns():
    """Add denormalized bow + arrow snapshot columns to apollo.

    ``bows.effective_draw_weight`` is the user's actual draw weight,
    separate from the rated ``bow_draw_weight``. Data-analysis prefers
    it when set.

    ``apollo`` gets denormalized snapshots of the bow's *and* arrow's
    configuration captured at the moment the shot is recorded. Without
    these, editing or renaming a bow/arrow silently rewrites the
    historical attribution of every prior shot. The model name strings
    (``bow``, ``arrow_type``) and ``nock_height`` were already
    snapshotted; these new columns close the gap. Idempotent.
    """
    insp = sa_inspect(engine)
    existing = set(insp.get_table_names())
    if 'bows' in existing:
        bow_cols = {c['name'] for c in insp.get_columns('bows')}
        if 'effective_draw_weight' not in bow_cols:
            print("⚙️  Migrating: adding effective_draw_weight column to bows")
            with engine.begin() as conn:
                conn.execute(text(
                    "ALTER TABLE bows ADD COLUMN effective_draw_weight VARCHAR(64)"
                ))
    if 'apollo' in existing:
        apollo_cols = {c['name'] for c in insp.get_columns('apollo')}
        # Add each missing snapshot column. New shots will populate them
        # from the bows/arrows row at insert time; pre-migration rows stay
        # NULL and analysis code treats NULL as "unknown — skip rather
        # than fall back to the current arrows/bows row" (the fallback
        # would re-introduce the edit-rewrites-history bug this fixes).
        for col_name, col_decl in (
            # Bow snapshot
            ('bow_draw_weight',       'VARCHAR(64)'),
            ('effective_draw_weight', 'VARCHAR(64)'),
            ('bow_amo',              'VARCHAR(64)'),
            ('bow_type',             'VARCHAR(255)'),
            # Arrow snapshot. ``arrow_type`` (the arrow's name) is already
            # an existing column on apollo, so it's not in this list.
            ('arrow_length',         'VARCHAR(64)'),
            ('arrow_spine',          'VARCHAR(64)'),
            ('arrow_shaft_weight',   'VARCHAR(64)'),
            ('arrow_shaft_diameter', 'VARCHAR(64)'),
            ('arrow_shaft_material', 'VARCHAR(255)'),
            ('arrow_nock_weight',    'VARCHAR(64)'),
            ('arrow_tip',            'VARCHAR(255)'),
            ('arrow_tip_weight',     'VARCHAR(64)'),
        ):
            if col_name not in apollo_cols:
                print(f"⚙️  Migrating: adding {col_name} column to apollo")
                with engine.begin() as conn:
                    conn.execute(text(
                        f"ALTER TABLE apollo ADD COLUMN {col_name} {col_decl}"
                    ))


def _ensure_session_tags_column():
    """Add the session_tags column to apollo on DBs that predate tag support.

    Idempotent: only ALTERs when the column is missing.
    """
    insp = sa_inspect(engine)
    if 'apollo' not in insp.get_table_names():
        return
    cols = {c['name'] for c in insp.get_columns('apollo')}
    if 'session_tags' in cols:
        return
    print("⚙️  Migrating: adding session_tags column to apollo")
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE apollo ADD COLUMN session_tags TEXT"))


def _ensure_client_uuid_column():
    """Add the client_uuid column to apollo on DBs that predate AJAX shots.

    Idempotent: only ALTERs when missing. Nullable so full-page (/sesh) and
    offline-synced shots — which pass NULL — keep working. The matching unique
    index is added by _ensure_apollo_client_uuid_unique_index().
    """
    insp = sa_inspect(engine)
    if 'apollo' not in insp.get_table_names():
        return
    cols = {c['name'] for c in insp.get_columns('apollo')}
    if 'client_uuid' in cols:
        return
    print("⚙️  Migrating: adding client_uuid column to apollo")
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE apollo ADD COLUMN client_uuid VARCHAR(64)"))


def _ensure_apollo_client_uuid_unique_index():
    """Add a UNIQUE index on apollo(user_id, client_uuid) for shot idempotency.

    A retried /api/record_shot POST carrying the same client_uuid hits this
    index and the duplicate INSERT raises IntegrityError, which the route
    catches and treats as success — so a dropped response can't double-record
    a shot. NULLs (full-page and synced shots) are distinct under both SQLite
    and MySQL, so unlimited NULL rows coexist.

    Best-effort like the other index migrations: a legacy DB with pre-existing
    duplicate keys would fail the CREATE; we log and move on rather than
    crashing startup.
    """
    try:
        insp = sa_inspect(engine)
        if 'apollo' not in set(insp.get_table_names()):
            return
        existing = {ix.get('name') for ix in insp.get_indexes('apollo')}
        if 'ux_apollo_user_client_uuid' in existing:
            return
        with engine.begin() as conn:
            conn.execute(text(
                "CREATE UNIQUE INDEX ux_apollo_user_client_uuid "
                "ON apollo(user_id, client_uuid)"
            ))
    except SQLAlchemyError as e:
        print(f"⚠️  Could not add unique index on apollo(user_id, client_uuid): {e}")


def _ensure_form_captures_columns():
    """Add the frames_json column to form_captures on DBs that predate replay.

    Idempotent: skips when the table doesn't exist yet (fresh DB → create_all
    makes it with the column) or already has it. frames_json holds the optional
    compact pose sequence used to replay a saved shot.
    """
    insp = sa_inspect(engine)
    if 'form_captures' not in insp.get_table_names():
        return
    cols = {c['name'] for c in insp.get_columns('form_captures')}
    if 'frames_json' in cols:
        return
    print("⚙️  Migrating: adding frames_json column to form_captures")
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE form_captures ADD COLUMN frames_json TEXT"))


def _ensure_user_timezone_column():
    """Add the timezone column to users on DBs that predate timezone support.

    Idempotent: only ALTERs when the column is missing. Default 'UTC' so
    every existing account keeps the previous behavior (server clock is
    UTC, display matched server) until the user picks a zone.
    """
    insp = sa_inspect(engine)
    if 'users' not in insp.get_table_names():
        return
    cols = {c['name'] for c in insp.get_columns('users')}
    if 'timezone' in cols:
        return
    print("⚙️  Migrating: adding timezone column to users")
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN timezone VARCHAR(64) DEFAULT 'UTC'"))


def _ensure_archer_profile_columns():
    """Add archer-profile columns used by handicap/classification.

    ``users.gender``, ``users.age_group``, ``users.default_bowstyle`` and
    ``bows.bowstyle``. Idempotent: each column is only ALTERed in when
    missing, and all are nullable so existing rows keep working — an archer
    without a profile still gets a handicap (which needs none of these), just
    no Archery GB classification until they fill the category in.
    """
    insp = sa_inspect(engine)
    tables = set(insp.get_table_names())
    wanted = {
        'users': [
            ('gender', "VARCHAR(16)"),
            ('age_group', "VARCHAR(32)"),
            ('default_bowstyle', "VARCHAR(32)"),
        ],
        'bows': [
            ('bowstyle', "VARCHAR(32)"),
        ],
    }
    for tbl, cols in wanted.items():
        if tbl not in tables:
            continue
        existing = {c['name'] for c in insp.get_columns(tbl)}
        for col_name, col_decl in cols:
            if col_name in existing:
                continue
            print(f"⚙️  Migrating: adding {col_name} column to {tbl}")
            with engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE {tbl} ADD COLUMN {col_name} {col_decl}"))


def _ensure_style_settings_columns():
    """Add the bowtype-specific settings columns.

    ``bows.style_settings`` (static gear defaults) and
    ``apollo.bow_style_settings`` (per-shot merged snapshot). Both hold JSON
    text. Idempotent and nullable so existing bows/shots keep working with no
    settings until the user fills them in.
    """
    insp = sa_inspect(engine)
    tables = set(insp.get_table_names())
    wanted = {
        'bows':   [('style_settings', 'TEXT')],
        'apollo': [('bow_style_settings', 'TEXT')],
    }
    for tbl, cols in wanted.items():
        if tbl not in tables:
            continue
        existing = {c['name'] for c in insp.get_columns(tbl)}
        for col_name, col_decl in cols:
            if col_name in existing:
                continue
            print(f"⚙️  Migrating: adding {col_name} column to {tbl}")
            with engine.begin() as conn:
                conn.execute(text(f"ALTER TABLE {tbl} ADD COLUMN {col_name} {col_decl}"))


def _ensure_is_root_column():
    """Add the is_root column to users on DBs that predate root support.

    Idempotent: only ALTERs when the column is missing. Default 0 so every
    existing account stays a regular user — root is granted explicitly by
    _ensure_root_user() based on env vars set by install.py.
    """
    insp = sa_inspect(engine)
    if 'users' not in insp.get_table_names():
        return
    cols = {c['name'] for c in insp.get_columns('users')}
    if 'is_root' in cols:
        return
    print("⚙️  Migrating: adding is_root column to users")
    with engine.begin() as conn:
        conn.execute(text("ALTER TABLE users ADD COLUMN is_root INTEGER DEFAULT 0"))


def _ensure_root_user():
    """Create or grant the root account from APOLLO_ROOT_* env vars.

    install.py writes APOLLO_ROOT_USERNAME, APOLLO_ROOT_EMAIL, and
    APOLLO_ROOT_PASSWORD into the .env file. On startup we:
      * promote an existing user with that username to is_root=1 (idempotent
        — handles re-runs after the account already exists), and
      * create the account if it's missing.

    Never overwrites an existing password. Once the root account exists the
    user can rotate APOLLO_ROOT_PASSWORD out of the .env safely; on every
    subsequent boot we'll just re-affirm is_root=1 on the existing row.

    Server-only: the local SQLite flavor is single-operator by design, so
    there's no one for root to administer. Skip silently — install.py
    won't have written APOLLO_ROOT_* in that case anyway.
    """
    if APOLLO_BACKEND != 'mysql':
        return
    username = (os.environ.get('APOLLO_ROOT_USERNAME') or '').strip()
    if not username:
        return
    email    = (os.environ.get('APOLLO_ROOT_EMAIL') or '').strip().lower()
    password = os.environ.get('APOLLO_ROOT_PASSWORD') or ''

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT id, is_root FROM users WHERE username = %s",
                (username,)
            ).fetchone()
            if row is not None:
                if not row['is_root']:
                    cur.execute(
                        "UPDATE users SET is_root = 1, is_active = 1 WHERE id = %s",
                        (int(row['id']),)
                    )
                    con.commit()
                    print(f"🔑 Granted root to existing user '{username}'")
                return

            # Account doesn't exist yet — need email + password to mint it.
            if not email or not password:
                print(
                    f"⚠️  APOLLO_ROOT_USERNAME='{username}' is set but no user exists "
                    "and APOLLO_ROOT_EMAIL/APOLLO_ROOT_PASSWORD are not both set. "
                    "Re-run install.py or set them manually to bootstrap root."
                )
                return

            pw_hash = generate_password_hash(password)
            cur.execute(
                "INSERT INTO users (username, email, password_hash, created_at, "
                "is_active, failed_attempts, is_root) "
                "VALUES (%s, %s, %s, %s, 1, 0, 1)",
                (username, email, pw_hash, _app_now())
            )
            new_row = cur.execute(
                "SELECT id FROM users WHERE username = %s", (username,)
            ).fetchone()
            con.commit()
            new_id = int(new_row[0])
        # Seed a legacy target outside the connection block so we don't nest
        # transactions on the same engine connection.
        _seed_user_default_target(new_id)
        print(f"🔑 Bootstrapped root account '{username}' (id={new_id})")
    except DBIntegrityError:
        # Email collision with a non-root account, or a parallel boot in a
        # multi-worker deploy beat us to it. Either way, the next startup
        # cycle will reconcile via the is_root grant branch above.
        print(f"⚠️  Could not create root '{username}' — email may already be in use.")
    except SQLAlchemyError as e:
        print(f"⚠️  Root bootstrap failed: {e}")


def get_app_setting(key, default=None):
    """Read a single app_settings row, or ``default`` if missing.

    Callers should be tolerant of a missing/empty value: the table is
    created by metadata.create_all() but rows only appear once an admin
    has set the value at least once.
    """
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT setting_value FROM app_settings WHERE setting_key = %s",
                (key,)
            ).fetchone()
            if row is None:
                return default
            val = row['setting_value']
            return val if val not in (None, '') else default
    except SQLAlchemyError:
        return default


def set_app_setting(key, value):
    """Upsert one app_settings row. Returns True on success.

    Race-tolerant via SELECT-then-INSERT-or-UPDATE; the unique index on
    setting_key would surface a duplicate insert anyway, but a second
    UPDATE is cheaper than an exception path.
    """
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            existing = cur.execute(
                "SELECT id FROM app_settings WHERE setting_key = %s", (key,)
            ).fetchone()
            now = _app_now()
            if existing is None:
                cur.execute(
                    "INSERT INTO app_settings (setting_key, setting_value, updated_at) "
                    "VALUES (%s, %s, %s)",
                    (key, value, now)
                )
            else:
                cur.execute(
                    "UPDATE app_settings SET setting_value = %s, updated_at = %s "
                    "WHERE setting_key = %s",
                    (value, now, key)
                )
            con.commit()
        return True
    except SQLAlchemyError as e:
        print(f"❌ set_app_setting({key!r}) failed: {e}")
        return False


def migrate_db():
    """Create tables if missing, run in-place migrations, and seed defaults.

    metadata.create_all() emits dialect-correct CREATE TABLE IF NOT EXISTS
    for whichever backend the engine points at, so this is safe to run
    every import against a fresh or existing DB.
    """
    # In-place migration for pre-multi-user installs has to happen *before*
    # create_all(), because once a column is declared in metadata SQLAlchemy
    # assumes it exists everywhere and downstream selects will blow up if
    # the live DB hasn't caught up yet.
    _ensure_user_id_columns()
    _ensure_arrow_dimension_columns()
    _migrate_poundage_to_draw_weight()
    _ensure_shot_snapshot_columns()
    _ensure_session_tags_column()
    _ensure_client_uuid_column()
    _ensure_form_captures_columns()
    _ensure_is_root_column()
    _ensure_user_timezone_column()
    _ensure_archer_profile_columns()
    _ensure_style_settings_columns()
    metadata.create_all(engine)
    _drop_legacy_unique_target_name()
    _ensure_session_times_unique_index()
    _ensure_targets_user_name_unique_index()
    _ensure_apollo_client_uuid_unique_index()
    _ensure_root_user()


def get_default_target(user_id):
    """Return the user's default target (is_default=1), or first active, or None.

    Targets are user-scoped; we never fall back to another user's targets,
    even if the caller has none. New users get a seeded NASP target row
    of their own at registration so this should always find a hit.
    """
    if user_id is None:
        return None
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Tournament faces also carry is_default=1, so prefer a
            # non-tournament default (the seeded WA 40cm face) deterministically
            # — otherwise a migrated user whose WA-40 row has a higher id than
            # their tournament rows would get a tournament face preselected.
            row = cur.execute(
                "SELECT id AS rowid, name, image_filename, physical_size_mm, image_size_px "
                "FROM targets WHERE is_default = 1 AND is_active = 1 AND user_id = %s "
                "ORDER BY CASE WHEN name LIKE %s THEN 1 ELSE 0 END, id LIMIT 1",
                (user_id, TOURNAMENT_TARGET_NAME_PREFIX + '%')
            ).fetchone()
            if row is None:
                row = cur.execute(
                    "SELECT id AS rowid, name, image_filename, physical_size_mm, image_size_px "
                    "FROM targets WHERE is_active = 1 AND user_id = %s ORDER BY id LIMIT 1",
                    (user_id,)
                ).fetchone()
            return row
    except SQLAlchemyError:
        return None


def get_target(target_id, user_id):
    """Return a single target row by rowid for the given user, or None.

    user_id is required — never look up a target without scoping by owner,
    or one user could view another's targets by guessing IDs.
    """
    if target_id is None or user_id is None:
        return None
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            return cur.execute(
                "SELECT id AS rowid, name, image_filename, physical_size_mm, image_size_px "
                "FROM targets WHERE id = %s AND user_id = %s",
                (int(target_id), user_id)
            ).fetchone()
    except (SQLAlchemyError, ValueError, TypeError):
        return None


def _seed_user_default_target(user_id):
    """Ensure the user's default target is the WA 40cm 10-ring vector face.

    Idempotent — safe to call on every register and login:
      * Archives the legacy NASP raster default if the user still has it
        (sets is_active=0, is_default=0). The row and any shot history that
        references it by id are left intact, just hidden and de-defaulted.
      * Seeds the WA 40cm 10-ring target + its scoring zones if missing, and
        marks it the active default. Geometry/zones come straight from the
        canonical ``wa_40`` face spec (WA 122cm 10-ring scaled to a 40cm
        edge) so scoring and the client-drawn ring overlay agree exactly.

    The image_filename is only a sizing placeholder; the colored rings are
    painted over it client-side (see _render_face_key_for_target_name).
    """
    if user_id is None:
        return
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Archive the legacy NASP raster default. Shots reference the
            # target by id, so they remain valid against the hidden row.
            cur.execute(
                "UPDATE targets SET is_active = 0, is_default = 0 "
                "WHERE user_id = %s AND name = %s",
                (user_id, LEGACY_NASP_TARGET_NAME)
            )
            existing = cur.execute(
                "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                (user_id, DEFAULT_TARGET_NAME)
            ).fetchone()
            if existing is not None:
                # Already migrated — just guarantee it's the active default.
                cur.execute(
                    "UPDATE targets SET is_active = 1, is_default = 1 "
                    "WHERE id = %s AND user_id = %s",
                    (int(existing[0]), user_id)
                )
                con.commit()
                return
            cur.execute(
                "INSERT INTO targets "
                "(user_id, name, image_filename, physical_size_mm, image_size_px, "
                "is_active, is_default) "
                "VALUES (%s, %s, %s, %s, %s, 1, 1)",
                (user_id, DEFAULT_TARGET_NAME, DEFAULT_TARGET_IMAGE,
                 DEFAULT_TARGET_SIZE_MM, DEFAULT_TARGET_SIZE_PX)
            )
            target_row = cur.execute(
                "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                (user_id, DEFAULT_TARGET_NAME)
            ).fetchone()
            if target_row is not None:
                target_id = int(target_row[0])
                # Zones from the canonical face spec, innermost-out — the
                # same rows _seed_tournament_faces writes for the WA faces.
                for idx, (pv, radius_mm, _color) in enumerate(
                        TOURNAMENT_FACES[DEFAULT_TARGET_FACE_KEY]['zones']):
                    cur.execute(
                        "INSERT INTO target_zones "
                        "(user_id, target_id, name, point_value, shape_type, "
                        "radius_mm, display_order) "
                        "VALUES (%s, %s, %s, %s, 'circle', %s, %s)",
                        (user_id, target_id, f"{pv} pts", pv,
                         float(radius_mm), idx)
                    )
            con.commit()
    except SQLAlchemyError as e:
        print(f"⚠️ Failed to seed/migrate default target for user {user_id}: {e}")


def _tournament_face_key_for_target_name(name):
    """Reverse-lookup: target row name → tournament face_key.

    Tournament faces are stored as rows whose name is the prefix plus the
    face's display name. Returning the face_key lets callers attach the
    canonical ring payload (colors, radii) for client-side overlay.
    Returns None when the row isn't a tournament face.
    """
    if not name or not isinstance(name, str):
        return None
    if not name.startswith(TOURNAMENT_TARGET_NAME_PREFIX):
        return None
    bare = name[len(TOURNAMENT_TARGET_NAME_PREFIX):]
    for key, face in TOURNAMENT_FACES.items():
        if face.get('name') == bare:
            return key
    return None


# Non-tournament targets that still render with a canonical vector face.
# Maps the exact target row name → TOURNAMENT_FACES key. The seeded default
# ('WA 40cm 10-ring') reuses the wa_40 ring spec so it draws as colored WA
# rings instead of showing its placeholder image.
_NONTOURNAMENT_RENDER_FACE_BY_NAME = {
    DEFAULT_TARGET_NAME: DEFAULT_TARGET_FACE_KEY,
}


def _render_face_key_for_target_name(name):
    """face_key whose ring spec the client should draw over this target.

    Covers tournament rows (matched by prefix) plus any non-tournament
    target that opts into a canonical vector face — currently the seeded
    default. Visual only: scoring always comes from the target's own
    target_zones, never from this lookup.
    """
    fk = _tournament_face_key_for_target_name(name)
    if fk:
        return fk
    if name and isinstance(name, str):
        return _NONTOURNAMENT_RENDER_FACE_BY_NAME.get(name)
    return None


def target_to_config(row):
    """Shape a target DB row into the dict templates/JS expect.

    image_filename is stored as a path relative to static/ — legacy seed
    is "target.jpg", new uploads are "targets/<unique>.jpg" — so it can
    be handed straight to url_for('static', filename=…). For tournament
    faces, ``face_render`` carries the colored ring spec so the canvas
    overlay shows the correct face regardless of the placeholder image
    underneath.
    """
    if row is None:
        return None
    face_key = _render_face_key_for_target_name(row['name'])
    face_render = (_tournament_face_render_payload(face_key)
                   if face_key else None)
    # Targets are square — width and height are both physical_size_mm /
    # image_size_px. The duplicated keys keep templates simple (and let us
    # add non-square targets later without changing the template surface).
    return {
        'target_id':        row['rowid'],
        'name':             _display_target_name(row['name']),
        'target_image':     row['image_filename'],
        'img_width':        row['image_size_px'],
        'img_height':       row['image_size_px'],
        'target_width_mm':  row['physical_size_mm'],
        'target_height_mm': row['physical_size_mm'],
        'face_render':      face_render,
    }

def get_db_connection():
    return CompatConnection(engine)


def _zone_radii_for_target(target_id, user_id):
    """Sorted (innermost→outermost) zone radii in mm for one target.

    Returns an empty list when the target has no zones or any radius is
    unparseable — callers treat that as "no scoring rings", which the JS
    interprets as "every click is a valid hit".
    """
    if target_id is None:
        return []
    radii = []
    for z in _fetch_target_zones(target_id, user_id):
        try:
            radii.append(float(z['radius_mm']))
        except (TypeError, ValueError):
            continue
    return radii


def _arrow_shaft_diameters_for_user(user_id):
    """Map arrow name → shaft diameter (mm) for the user's arrows.

    Arrows without a parseable diameter are omitted so the JS click handler
    can fall through to its own ``DEFAULT_SHAFT_DIAMETER_MM`` constant.
    Arrow names are unique per user in practice; if a name appears twice,
    the first parseable diameter wins.
    """
    out = {}
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT arrow, shaft_diameter FROM arrows WHERE user_id = %s",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError:
        return out
    for r in rows:
        name = r['arrow']
        if not name or name in out:
            continue
        try:
            d = float(str(r['shaft_diameter']).strip())
        except (TypeError, ValueError, AttributeError):
            continue
        if d > 0:
            out[name] = d
    return out

def get_past_shots(session_id, quiver_size, arrows_remaining, user_id):
    """Return shots from the current in-progress quiver for display on the target.

    Strategy: read the tail of the session's shot log and use the last
    row's *stored* quiver_size and arrows_remaining (pre-decrement) to
    work out how many shots belong to the current quiver. The position
    of that last shot within its own quiver is
    ``quiver_size - arrows_remaining + 1`` — those are the rows we
    want to redraw. If the last shot had ``arrows_remaining == 1`` it
    completed a quiver and the canvas should clear for the next one.

    Reading the count from the DB row (rather than the cookie's
    arrows_remaining) keeps the redraw correct after a recall: deleting
    a row changes which row is "last" and therefore which quiver is
    "current", regardless of how the cookie's counter has been rewound.
    The ``quiver_size`` and ``arrows_remaining`` arguments are accepted
    for callsite-compat and are only consulted when the session has no
    shot rows yet.

    Misses (sentinel coords *or* hits that fell outside the target's
    outermost zone) are filtered out — they're recorded in the DB but
    shouldn't be rendered as markers on the target image; the visual
    treatment matches the "Missed target" button.
    """
    del quiver_size, arrows_remaining  # superseded by DB-derived counts
    if session_id is None or user_id is None:
        return []
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            last = cur.execute(
                "SELECT quiver_size, arrows_remaining FROM apollo "
                "WHERE session_id = %s AND user_id = %s "
                "ORDER BY id DESC LIMIT 1",
                (session_id, user_id)
            ).fetchone()
            if not last:
                return []
            try:
                last_qs = int(last['quiver_size']
                              if 'quiver_size' in last else last[0])
                last_ar = int(last['arrows_remaining']
                              if 'arrows_remaining' in last else last[1])
            except (TypeError, ValueError):
                return []
            if last_qs <= 0 or last_ar <= 1:
                # ar == 1 on the last row means that shot just completed
                # a quiver, so the canvas should be blank for the next.
                return []
            shots_in_quiver = last_qs - last_ar + 1
            if shots_in_quiver <= 0:
                return []
            rows = cur.execute(
                """SELECT x_coord, y_coord, target_id, arrow_shaft_diameter FROM apollo
                   WHERE session_id = %s AND user_id = %s
                   ORDER BY id DESC
                   LIMIT %s""",
                (session_id, user_id, shots_in_quiver)
            ).fetchall()
        rows = list(reversed(rows))    # ORDER BY DESC + reverse = chronological
        # The session locks the target after the first shot, so a single
        # zone fetch (from the first row's target_id) covers every shot
        # in the buffer. No zones configured → fall through to the old
        # sentinel-only filter, since we can't classify off-target.
        target_id = rows[0]['target_id'] if rows else None
        zones = _fetch_target_zones(target_id, user_id) if target_id is not None else []
        spot_centers = (_target_multi_spot_centers(target_id, user_id)
                        if target_id is not None else None)
        out = []
        for row in rows:
            xraw = str(row['x_coord']).strip() if row['x_coord'] is not None else ''
            yraw = str(row['y_coord']).strip() if row['y_coord'] is not None else ''
            if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
                continue
            if zones and _classify_shot(xraw, yraw, zones,
                                        row['arrow_shaft_diameter'],
                                        spot_centers_mm=spot_centers) is None:
                continue
            out.append({"x": float(row["x_coord"]), "y": float(row["y_coord"])})
        return out
    except (SQLAlchemyError, ValueError):
        return []


migrate_db()

app = Flask(__name__, static_folder='static', template_folder='templates')

# Detect deployment mode once at startup — referenced by cookie flags and
# the ProxyFix wiring below. Anything set to anything other than empty
# behaves like dev (debugger on, Secure cookies off) so local runs of
# `python apollo.py` Just Work without extra env vars.
_IS_PRODUCTION = os.environ.get('FLASK_ENV') == 'production'

# When hosted behind a reverse proxy (PythonAnywhere, nginx, Cloudflare),
# request.remote_addr would otherwise be the proxy's IP — which makes
# the per-IP login rate limiter useless. ProxyFix trusts one hop of
# X-Forwarded-* headers from the proxy. Only enable in production: in
# dev the Werkzeug server is hit directly, and trusting forwarded
# headers from an arbitrary client would let a remote user spoof IPs.
if _IS_PRODUCTION:
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

# SECRET_KEY drives Flask session signing *and* CSRF token signing.
# Production MUST set it via env; dev gets a hardcoded fallback (with a
# loud warning) so local runs don't need a .env file. Fail loudly rather
# than silently shipping the dev key to prod.
_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    if _IS_PRODUCTION:
        raise RuntimeError("SECRET_KEY environment variable is required in production")
    # Generate a per-process random key so dev sessions can't be forged by
    # anyone with source access. The trade-off is that restarting the app
    # invalidates in-flight session cookies — fine for local dev.
    _secret_key = secrets.token_urlsafe(32)
    print("⚠️  SECRET_KEY not set — generated a random dev key. Sessions reset on restart. "
          "Set SECRET_KEY env var for production.")
app.secret_key = _secret_key

# APOLLO_BASE_URL is required in production: password-reset emails embed
# this origin in the reset link, and falling back to request.url_root makes
# the link attacker-controllable via a spoofed Host header (a remote user
# can poison a victim's reset mail by hitting /forgot_password with their
# own Host:, then capture the token when the victim clicks). Fail loudly
# at startup rather than ship a host-header oracle.
if _IS_PRODUCTION and not (os.environ.get('APOLLO_BASE_URL') or '').strip():
    raise RuntimeError(
        "APOLLO_BASE_URL environment variable is required in production "
        "(e.g. APOLLO_BASE_URL=https://apolloshoots.org). Without it, "
        "password-reset links derive from the request's Host header, which "
        "an attacker can spoof to redirect reset links to a hostile origin."
    )

# Email-config sanity check. We don't raise on missing config (the rest
# of the app works fine), but in production a silent fallback to "print
# the reset URL to the WSGI log" would mean password-reset emails just
# never arrive without anyone noticing. Make it loud at startup instead.
if os.environ.get('RESEND_API_KEY', '').strip():
    if resend is None:
        print("⚠️  RESEND_API_KEY is set but the `resend` package is not installed. "
              "Password-reset emails will fall back to stdout. "
              "Fix: pip install resend")
elif _IS_PRODUCTION:
    print("⚠️  RESEND_API_KEY not set in production — password-reset emails "
          "will be printed to the WSGI log instead of delivered. "
          "Set RESEND_API_KEY (and RESEND_FROM) to enable real email.")

# ─── Cookie / session hardening ──────────────────────────────────────────
# HttpOnly: JS can't read the cookie → XSS can't lift the session token.
# SameSite=Lax: cookie is *not* sent on cross-site POSTs (CSRF defence-in-
# depth alongside Flask-WTF's tokens) but *is* sent on top-level GETs so
# normal links from email/external sites still log the user in.
# Secure: only set in production — in dev we're typically on http://
# localhost and a Secure cookie would just never be stored.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=_IS_PRODUCTION,
    # 30-day login persistence. Long enough that casual users don't get
    # re-prompted constantly, short enough that an abandoned device
    # eventually loses access.
    PERMANENT_SESSION_LIFETIME=timedelta(days=30),
    # 5 MB request cap — matches MAX_UPLOAD_BYTES so an oversized upload
    # is rejected by Werkzeug *before* it allocates the bytes, rather
    # than after read() in add_target. Defense-in-depth, not a behavior
    # change.
    MAX_CONTENT_LENGTH=MAX_UPLOAD_BYTES,
)

# CSRFProtect requires every POST form to include a csrf_token hidden
# input rendered via {{ csrf_token() }}. Tokens are signed with secret_key
# above, so rotating SECRET_KEY invalidates in-flight tokens (expected).
#
# WTF_CSRF_TIME_LIMIT=None disables the token's *age* check (Flask-WTF
# defaults to 3600 s). Without this, a page left open longer than an hour
# — routine on a tablet mid-round, where the device sleeps between ends —
# fails its next POST with "CSRF token has expired" (surfacing as a 502
# behind the proxy). The token stays bound to the signed session cookie,
# so it remains valid only for the life of the login (30 days) and to that
# one session; dropping the age cap is not a security regression.
app.config['WTF_CSRF_TIME_LIMIT'] = None
csrf = CSRFProtect(app)


# Security headers applied to every response. The CSP is intentionally
# tight: only same-origin scripts/styles (no inline, no third-party CDNs)
# and same-origin images plus inline data: URIs (matplotlib reports embed
# PNGs as base64 data:). Frame-ancestors 'none' is the modern equivalent
# of X-Frame-Options: DENY and stops the site being iframed for click-
# jacking. HSTS is only set in production because dev runs on http://
# localhost and an HSTS header there would lock the browser into https
# for any later https-served local dev.
@app.after_request
def _set_security_headers(response):
    response.headers.setdefault('X-Frame-Options', 'DENY')
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('Referrer-Policy', 'same-origin')
    # CSP note on 'unsafe-inline': several templates use inline <script>
    # blocks and onsubmit="return confirm(...)" handlers, and splash.html
    # embeds the buymeacoffee widget from cdnjs. Until those are
    # externalized into static/*.js with nonces, 'unsafe-inline' is the
    # practical baseline. The header still provides defense-in-depth via
    # default-src 'self' (no external resources), frame-ancestors 'none'
    # (no click-jacking), and form-action 'self' (no off-site form posts).
    #
    # Allowlisted third parties:
    #   cdnjs.buymeacoffee.com — splash.html embeds the BMC widget. This is
    #     now the only third-party script origin: the webfonts, lightGallery
    #     and Chart.js are all self-hosted under /static, and the analyze
    #     heatmap is a server-rendered matplotlib SVG — so Google Fonts,
    #     jsdelivr (lightGallery / Plotly / Chart.js) are no longer referenced
    #     and everything else works fully same-origin (and offline).
    #
    # 'wasm-unsafe-eval': the /form motion-capture page instantiates the
    # vendored MediaPipe pose model (WebAssembly) in the browser. WASM
    # compilation is blocked by a plain script-src, and this is the narrow,
    # purpose-built token for it — it permits WebAssembly.instantiate only, NOT
    # arbitrary string eval() (that would be 'unsafe-eval'). The model and wasm
    # are same-origin under /static, so connect-src 'self' still covers fetch.
    # worker-src allows the pose runner's blob-URL worker.
    response.headers.setdefault(
        'Content-Security-Policy',
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'wasm-unsafe-eval' https://cdnjs.buymeacoffee.com; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: blob: https:; "
        "font-src 'self' data:; "
        "connect-src 'self'; "
        "worker-src 'self' blob:; "
        "frame-ancestors 'none'; "
        "base-uri 'self'; "
        "form-action 'self'"
    )
    if _IS_PRODUCTION:
        response.headers.setdefault(
            'Strict-Transport-Security',
            'max-age=31536000; includeSubDomains'
        )
    return response


@app.template_filter('fmt_ts')
def _fmt_ts(value):
    """Format a timestamp as 'YYYY-MM-DD HH:MM:SS' regardless of backend.
    SQLite returns timestamps as strings (with trailing '.microseconds'),
    MySQL returns real datetime objects — the previous '[:-7]' template
    slice only worked on the former and crashed on MySQL."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    s = str(value)
    # Drop trailing '.ffffff' microseconds if present.
    if '.' in s:
        s = s.split('.', 1)[0]
    return s


@app.template_filter('bowstyle_label')
def _bowstyle_label(bs):
    """Human label for a bowstyle enum (e.g. 'recurve' → 'Olympic recurve').

    Stored values stay the short lowercase enums; this only affects display.
    Unknown values fall back to a plain capitalize.
    """
    key = (bs or '').strip().lower()
    return BOWSTYLE_LABELS.get(key, (bs or '').capitalize())


# ─── Auth: password rules, lockout, helpers ──────────────────────────────

# Username/email/password constraints. Kept loose enough to be friendly,
# strict enough to keep junk and obviously-bad inputs out. Real rules of
# thumb: usernames are short and ascii so they're safe in URLs and logs;
# passwords are *only* length-checked (NIST 800-63B says don't enforce
# composition rules — length is what matters); emails get a simple shape
# check (full RFC-5322 is a tarpit, the SMTP layer will be the real test).
USERNAME_RE     = re.compile(r'^[A-Za-z0-9_.\-]{3,32}$')
EMAIL_RE        = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
MIN_PASSWORD_LEN = 8
MAX_PASSWORD_LEN = 128

# Lockout policy — 5 failed attempts in a row locks the account for
# 15 minutes. Stored on the user row so it survives process restarts
# (an in-memory counter could be defeated by killing the process).
LOCKOUT_THRESHOLD = 5
LOCKOUT_MINUTES   = 15

# Password-reset tokens are short-lived: 60 minutes is the sweet spot
# between "user got distracted by a phone call" and "attacker has time
# to brute-force out-of-band". Tokens are also one-shot.
PASSWORD_RESET_TTL_MINUTES = 60

# Per-IP rate limit on /forgot_password requests. Stops a bored attacker
# from blasting our SMTP relay (and the victim's inbox) by churning the
# endpoint.
_RESET_IP_WINDOW_SECS = 3600
_RESET_IP_LIMIT       = 5

# Per-IP login rate limiter. Stops a single attacker from churning
# through thousands of accounts even if no single account is hit hard
# enough to trip the per-account lockout. State lives in the
# rate_limit_hits DB table so it survives process restarts (an attacker
# could otherwise force a reset by triggering a deploy) and is shared
# across multiple WSGI workers.
_IP_HIT_WINDOW_SECS = 300
_IP_HIT_LIMIT       = 30

# Pre-computed hash for the invalid-username path's timing equalizer.
# Computing it once at import time means /login spends the same amount
# of work (one check_password_hash) whether the username exists or not —
# a per-request generate_password_hash would actually make the miss path
# *slower* than the hit path, re-opening the enumeration oracle.
_DUMMY_PASSWORD_HASH = generate_password_hash("dummy-for-timing-equalization")


def _bump_rate_limit(scope, key, window_secs, limit):
    """Increment the (scope, key) hit counter and return True if over limit.

    Sliding-ish window: a row whose window_start is older than
    ``window_secs`` is rolled over to "now" with hits=1. Same semantics as
    the previous in-memory counter, but persisted so it survives restarts
    and is shared across workers.

    Failure-mode: DB errors deliberately *don't* block the request. The
    rate limiter is a backstop on top of the per-user lockout and the
    SECRET_KEY-bound CSRF/session protections — degrading to "no IP cap"
    on a DB outage is preferable to handing every user a 500.
    """
    if not key:
        return False
    now = _app_now()
    cutoff = now - timedelta(seconds=window_secs)
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT id, window_start, hits FROM rate_limit_hits "
                "WHERE scope = %s AND rate_key = %s LIMIT 1",
                (scope, key)
            ).fetchone()
            if row is None:
                # Two concurrent requests can both see "no row" and both
                # INSERT — the unique (scope, rate_key) constraint will
                # raise on the loser, in which case we fall through to the
                # update path below by retrying once.
                try:
                    cur.execute(
                        "INSERT INTO rate_limit_hits (scope, rate_key, window_start, hits) "
                        "VALUES (%s, %s, %s, 1)",
                        (scope, key, now)
                    )
                    con.commit()
                    return 1 > limit
                except DBIntegrityError:
                    con.rollback()
                    row = cur.execute(
                        "SELECT id, window_start, hits FROM rate_limit_hits "
                        "WHERE scope = %s AND rate_key = %s LIMIT 1",
                        (scope, key)
                    ).fetchone()
                    if row is None:
                        # Shouldn't happen, but bail out rather than crash
                        return False
            row_id = row['id'] if 'id' in row else row[0]
            win_start = row['window_start'] if 'window_start' in row else row[1]
            hits = row['hits'] if 'hits' in row else row[2]
            # Normalize string-stored datetimes (SQLite) into datetimes.
            if isinstance(win_start, str):
                try:
                    win_start = datetime.fromisoformat(win_start)
                except ValueError:
                    win_start = cutoff  # treat unparseable as expired
            if win_start < cutoff:
                new_hits = 1
                cur.execute(
                    "UPDATE rate_limit_hits SET window_start = %s, hits = %s "
                    "WHERE id = %s",
                    (now, new_hits, row_id)
                )
            else:
                new_hits = int(hits or 0) + 1
                cur.execute(
                    "UPDATE rate_limit_hits SET hits = %s WHERE id = %s",
                    (new_hits, row_id)
                )
            con.commit()
            return new_hits > limit
    except SQLAlchemyError as e:
        print(f"⚠️  Rate-limit bump failed (scope={scope}, key={key}): {e}")
        return False


def _ip_rate_limited(ip):
    """Return True if this IP has exceeded the login attempt cap."""
    return _bump_rate_limit('login', ip, _IP_HIT_WINDOW_SECS, _IP_HIT_LIMIT)


def current_user_id():
    """Return the logged-in user's id, or None when not signed in."""
    return session.get('user_id')


def current_user():
    """Look up the current user's row, or None.

    Used by the context processor to expose user info to every template.
    Returns None (not a redirect) on missing/invalid sessions — route
    decorators handle the auth-required redirect separately.

    Result is memoized on ``flask.g`` so a single request that calls this
    from a route handler AND has it inlined by the context processor
    only hits the DB once.
    """
    uid = current_user_id()
    if uid is None:
        return None
    cached = getattr(g, '_current_user', None)
    if cached is not None and cached.get('rowid') == uid:
        return cached
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT id AS rowid, username, email, created_at, last_login, "
                "is_root, timezone, gender, age_group, default_bowstyle "
                "FROM users WHERE id = %s AND is_active = 1", (uid,)
            ).fetchone()
            # Stale session: cookie says user 42 but user 42 was deleted
            # or deactivated. Clear the cookie so the next request bumps
            # the user to /login cleanly.
            if row is None:
                session.clear()
                return None
            g._current_user = row
            return row
    except SQLAlchemyError:
        return None


def login_required(f):
    """Redirect to /login when the route is hit without a valid session.

    Captures the originally-requested path in ?next= so users land back
    where they were trying to go after authenticating. Only same-origin
    paths are echoed back into ?next= — the /login handler validates
    that again before redirecting, but defense-in-depth.
    """
    @wraps(f)
    def wrapped(*args, **kwargs):
        if current_user_id() is None:
            # Preserve the requested path (not full URL — query string is
            # dropped on purpose to avoid round-tripping CSRF tokens or
            # other sensitive query params through the login form).
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return wrapped


def root_required(f):
    """Like login_required, but also requires the current user to be root.

    Non-root signed-in users get a 403 rather than a redirect — they
    *are* authenticated, just not authorized, and bouncing them through
    /login would loop forever.

    Server-only: refuse outright on the local SQLite flavor. The local
    install is single-operator (no other accounts to administer) and
    root never gets bootstrapped there, but a flipped is_root bit in a
    hand-edited DB shouldn't open the admin surface either.
    """
    @wraps(f)
    def wrapped(*args, **kwargs):
        if APOLLO_BACKEND != 'mysql':
            abort(404)
        user = current_user()
        if user is None:
            return redirect(url_for('login', next=request.path))
        if not user['is_root']:
            abort(403)
        return f(*args, **kwargs)
    return wrapped


def _rotate_session(user_id):
    """Session-fixation defence: drop any pre-login session data and
    install a fresh session bound to the new user.

    Flask's session is itself a signed cookie, but if an attacker can
    convince a victim to authenticate while holding the attacker's
    pre-set session cookie, the attacker would inherit the logged-in
    state. Clearing everything before writing the new user_id makes the
    new session token effectively brand-new from the client's POV.
    """
    session.clear()
    session['user_id']   = user_id
    session.permanent    = True


def _account_is_locked(row):
    """Return True when user.row's locked_until is in the future.

    Fails *closed* on a malformed locked_until value (treat as locked and
    log loudly) — a parse error means we can't verify the lockout has
    expired, and silently unlocking the account would defeat the policy.
    """
    if row is None:
        return False
    locked = row['locked_until'] if 'locked_until' in row else None
    if locked is None:
        return False
    try:
        if isinstance(locked, datetime):
            return locked > _app_now()
        s = str(locked)
        # SQLite stores DateTime as ISO string when no type adapter is
        # registered. Try the common shapes — fromisoformat handles
        # "YYYY-MM-DD HH:MM:SS[.ffffff]" on 3.11+, the strptime branches
        # cover older Pythons and a "T" separator.
        try:
            parsed = datetime.fromisoformat(s)
        except ValueError:
            parsed = None
            for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S',
                        '%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S'):
                try:
                    parsed = datetime.strptime(s, fmt)
                    break
                except ValueError:
                    continue
            if parsed is None:
                raise ValueError(f"unparseable locked_until: {s!r}")
        return parsed > _app_now()
    except (ValueError, TypeError) as e:
        print(f"⚠️  Unparseable locked_until for user — failing closed: {e}")
        return True


def _record_failed_login(user_id):
    """Increment the failed-attempt counter and apply lockout when hit.

    Reset-on-expiry and increment happen in a single CASE-driven UPDATE so
    two concurrent failed logins can't both observe attempts=4 and write
    back 5 (skipping the lockout threshold). When a previous lockout has
    already expired the served time counts — the CASE branch resets to 1
    instead of leaving ``failed_attempts`` ≥ threshold (which would re-lock
    immediately on the very next failure).
    """
    if user_id is None:
        return
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            now = _app_now()
            cur.execute(
                "UPDATE users SET "
                "  failed_attempts = CASE "
                "    WHEN locked_until IS NOT NULL AND locked_until <= %s THEN 1 "
                "    ELSE COALESCE(failed_attempts, 0) + 1 "
                "  END, "
                "  locked_until = CASE "
                "    WHEN locked_until IS NOT NULL AND locked_until <= %s THEN NULL "
                "    ELSE locked_until "
                "  END "
                "WHERE id = %s",
                (now, now, user_id)
            )
            row = cur.execute(
                "SELECT failed_attempts FROM users WHERE id = %s", (user_id,)
            ).fetchone()
            attempts = int(row['failed_attempts']) if row and row['failed_attempts'] else 0
            if attempts >= LOCKOUT_THRESHOLD:
                locked_until = now + timedelta(minutes=LOCKOUT_MINUTES)
                cur.execute(
                    "UPDATE users SET locked_until = %s WHERE id = %s",
                    (locked_until, user_id)
                )
            con.commit()
    except SQLAlchemyError as e:
        print(f"⚠️  Failed to record failed login for user {user_id}: {e}")


def _record_successful_login(user_id):
    """Reset lockout counters and stamp last_login on the user row."""
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            cur.execute(
                "UPDATE users SET failed_attempts = 0, locked_until = NULL, "
                "last_login = %s WHERE id = %s",
                (_app_now(), user_id)
            )
            con.commit()
    except SQLAlchemyError as e:
        print(f"⚠️  Failed to record successful login for user {user_id}: {e}")


def _reset_ip_rate_limited(ip):
    """Return True when this IP has asked for too many password resets recently."""
    return _bump_rate_limit('forgot', ip, _RESET_IP_WINDOW_SECS, _RESET_IP_LIMIT)


def _hash_reset_token(token):
    """SHA-256 hex digest — what we store and what we look up by."""
    return hashlib.sha256(token.encode('utf-8')).hexdigest()


def _send_email(to_addr, subject, body, attachments=None):
    """Send a plain-text email via Resend, or fall back to stdout in dev.

    Reads RESEND_API_KEY at call time so a missing/rotated key doesn't
    crash the app at import. Optional RESEND_FROM controls the sender
    address (default 'onboarding@resend.dev' — Resend's shared sandbox
    sender, which only delivers to verified test recipients; set
    RESEND_FROM to an address on a domain you've verified in the Resend
    dashboard before opening this up to real users).

    ``attachments`` is an optional list of ``{filename, content}`` dicts
    where ``content`` is raw bytes; each is base64-encoded for the Resend
    API (the SDK's Attachment.content accepts a base64 string).

    When RESEND_API_KEY is unset *or* the `resend` package isn't
    installed, we print the message to stdout — the dev flow
    ("forgot password" → copy the URL out of the terminal") still works
    without any external service.
    """
    api_key = os.environ.get('RESEND_API_KEY', '').strip()
    if not api_key or resend is None:
        reason = "RESEND_API_KEY not set" if resend is not None else "resend package not installed"
        print(f"✉️  [dev] Would send email ({reason}):")
        print(f"    To:      {to_addr}")
        print(f"    Subject: {subject}")
        for line in body.splitlines():
            print(f"    | {line}")
        for a in (attachments or []):
            print(f"    + attachment: {a['filename']} ({len(a['content'])} bytes)")
        return True

    sender = os.environ.get('RESEND_FROM', '').strip() or 'onboarding@resend.dev'
    resend.api_key = api_key
    payload = {
        "from":    sender,
        "to":      to_addr,
        "subject": subject,
        "text":    body,
    }
    if attachments:
        payload["attachments"] = [
            {
                "filename": a['filename'],
                "content":  base64.b64encode(a['content']).decode('ascii'),
            }
            for a in attachments
        ]
    try:
        resend.Emails.send(payload)
        return True
    except Exception as e:
        # Resend raises its own exception types; catch broadly so a
        # transient delivery failure can't crash the request handler.
        print(f"⚠️  Failed to send email to {to_addr} via Resend: {e}")
        return False


def _password_reset_base_url():
    """Origin (scheme://host) to embed in reset emails.

    Prefer the explicit APOLLO_BASE_URL env var (set in production behind
    a reverse proxy where request.url_root may be wrong). Falls back to
    request.url_root for local dev so `python apollo.py` Just Works.
    """
    explicit = (os.environ.get('APOLLO_BASE_URL') or '').strip()
    if explicit:
        return explicit.rstrip('/')
    return request.url_root.rstrip('/')


def _validate_registration(username, email, password, confirm):
    """Return None on success, or an error string for the form."""
    if not username or not USERNAME_RE.match(username):
        return ("Username must be 3–32 characters and contain only letters, "
                "digits, underscores, dots, or hyphens.")
    if not email or not EMAIL_RE.match(email) or len(email) > 255:
        return "Please enter a valid email address."
    if not password or len(password) < MIN_PASSWORD_LEN:
        return f"Password must be at least {MIN_PASSWORD_LEN} characters."
    if len(password) > MAX_PASSWORD_LEN:
        # Cap to keep werkzeug's hash work bounded — scrypt on a multi-MB
        # password would be a DoS surface.
        return f"Password must be at most {MAX_PASSWORD_LEN} characters."
    if password != confirm:
        return "Passwords do not match."
    return None


def _claim_orphan_data(user_id):
    """Assign any pre-multi-user rows (user_id IS NULL) to the new user.

    Only runs at the moment the *first* account is created on this DB —
    that user is the de-facto original owner of any data that predates
    multi-user support. Later registrations don't claim anything: their
    data starts empty. Idempotent because each call only sees rows that
    are still NULL.
    """
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Only run if this is the only user — otherwise we'd be
            # silently appropriating data that some other user might
            # legitimately own.
            row = cur.execute("SELECT COUNT(*) FROM users").fetchone()
            user_count = int(row[0]) if row else 0
            if user_count != 1:
                return
            claimed_total = 0
            for tbl in _PER_USER_TABLES:
                res = cur.execute(
                    f"UPDATE {tbl} SET user_id = %s WHERE user_id IS NULL",
                    (user_id,)
                )
                # SQLAlchemy 2.x Result has rowcount; CompatCursor doesn't
                # surface it, so we re-count instead of relying on it.
                count_row = cur.execute(
                    f"SELECT COUNT(*) FROM {tbl} WHERE user_id = %s",
                    (user_id,)
                ).fetchone()
                claimed_total += int(count_row[0]) if count_row else 0
            con.commit()
            if claimed_total > 0:
                print(f"📦 Claimed {claimed_total} pre-existing row(s) for user {user_id}")
    except SQLAlchemyError as e:
        print(f"⚠️  Orphan-claim failed: {e}")


@app.context_processor
def inject_template_globals():
    """Expose the default target *and* the current user to every template.

    Default target only resolves when a user is signed in (it's a per-user
    row now); the splash uses ``current_user`` to decide whether to render
    the side-nav links or a "sign up / log in" CTA.

    Anonymous requests short-circuit entirely — no DB work at all on
    /login, /register, etc.
    """
    if current_user_id() is None:
        return dict(default_target=None, current_user=None,
                    apollo_backend=APOLLO_BACKEND)
    user = current_user()
    row = get_default_target(user['rowid']) if user is not None else None
    return dict(
        default_target=target_to_config(row),
        current_user=user,
        timezones=_TIMEZONE_CHOICES,
        apollo_backend=APOLLO_BACKEND,
        server_timezone=get_app_setting('server_timezone', 'UTC'),
        archer_genders=ARCHER_GENDERS,
        archer_age_groups=ARCHER_AGE_GROUPS,
        archer_bowstyles=ARCHER_BOWSTYLES,
    )

@app.route('/', methods=['GET'])
def index():
    """Render the splash/landing page.

    Public: the splash works for signed-out users too, showing a sign-up
    CTA. The template branches on ``current_user`` exposed by the context
    processor above.
    """
    return render_template('splash.html')


# ─── PWA: service worker + manifest ──────────────────────────────────────
# Both are served from the site root (not /static/) on purpose. A service
# worker only controls pages within its own URL scope, and a worker fetched
# from /static/sw.js would be scoped to /static/ — unable to intercept
# /sesh. Serving it from /sw.js gives it root scope. (The explicit
# Service-Worker-Allowed header is belt-and-suspenders for the same goal.)
# The manifest lives at the root too so its start_url / scope resolve to /.
# Both are public (no @login_required): the browser fetches them before any
# auth context and they contain nothing user-specific.

@app.route('/sw.js', methods=['GET'])
def service_worker():
    resp = app.send_static_file('sw.js')
    resp.headers['Content-Type'] = 'application/javascript'
    resp.headers['Service-Worker-Allowed'] = '/'
    # The SW file itself must never be served stale, or clients can get
    # wedged on an old cache strategy; the SW handles asset versioning.
    resp.headers['Cache-Control'] = 'no-cache'
    return resp


@app.route('/manifest.webmanifest', methods=['GET'])
def web_manifest():
    resp = app.send_static_file('manifest.webmanifest')
    resp.headers['Content-Type'] = 'application/manifest+json'
    return resp


# ─── Auth routes ─────────────────────────────────────────────────────────

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Create a new user account.

    Tries to keep error responses *generic* where possible (e.g. on login
    failure) but registration deliberately surfaces "username taken" /
    "email taken" so users aren't left guessing why their submission
    bounced — the username/email namespaces are already enumerable by
    trying to register, so withholding the reason here adds friction
    without security benefit.
    """
    if current_user_id() is not None:
        return redirect(url_for('index'))

    if request.method == 'GET':
        return render_template('register.html', error=None, form={})

    username = (request.form.get('username') or '').strip()
    email    = (request.form.get('email') or '').strip().lower()
    password = request.form.get('password') or ''
    confirm  = request.form.get('confirm_password') or ''

    err = _validate_registration(username, email, password, confirm)
    if err is not None:
        return render_template('register.html', error=err,
                               form={'username': username, 'email': email})

    pw_hash = generate_password_hash(password)

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            cur.execute(
                "INSERT INTO users (username, email, password_hash, created_at, is_active, "
                "failed_attempts) VALUES (%s, %s, %s, %s, 1, 0)",
                (username, email, pw_hash, _app_now())
            )
            # Read the just-inserted id back. LAST_INSERT_ID() is MySQL;
            # SQLite uses last_insert_rowid(). Easiest cross-dialect path:
            # SELECT by the unique username we just inserted.
            row = cur.execute(
                "SELECT id FROM users WHERE username = %s", (username,)
            ).fetchone()
            con.commit()
            new_user_id = int(row[0])
    except DBIntegrityError:
        # Either username or email collision. Re-query to tell which, so
        # the UX message is precise. (Enumeration risk is low — username
        # uniqueness is inherently observable on any auth system.)
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                if cur.execute("SELECT 1 FROM users WHERE username = %s",
                               (username,)).fetchone():
                    msg = "That username is already taken."
                else:
                    msg = "An account with that email already exists."
        except SQLAlchemyError:
            msg = "Could not create account — please try a different username or email."
        return render_template('register.html', error=msg,
                               form={'username': username, 'email': email})
    except SQLAlchemyError as e:
        print(f"❌ Register error: {e}")
        return render_template('register.html',
                               error="Could not create account — please try again.",
                               form={'username': username, 'email': email})

    # Seed the new user's data. Claim orphans *first* so that if the
    # pre-multi-user DB already has a target row, it gets adopted by
    # this user and the subsequent seed call is a no-op (rather than
    # creating a confusing duplicate).
    _claim_orphan_data(new_user_id)
    _seed_user_default_target(new_user_id)
    _seed_tournament_faces(new_user_id)

    # Log the new user in immediately — modern UX expectation, and skips
    # the awkward "now go to the login page" handoff.
    _rotate_session(new_user_id)
    _record_successful_login(new_user_id)
    return redirect(url_for('index'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Sign in an existing user.

    Generic "invalid username or password" message on failure — never
    leak which half was wrong. The per-account lockout still kicks in
    correctly under the hood since we look up by username first.
    """
    if current_user_id() is not None:
        return redirect(url_for('index'))

    # ?next= controls where to land after a successful sign-in. Only
    # accept same-origin paths to block open-redirect abuse from a
    # crafted link like /login?next=https://evil.example/.
    next_url = (request.args.get('next') or request.form.get('next') or '').strip()
    parsed = urlparse(next_url)
    if (parsed.scheme or parsed.netloc
            or not next_url.startswith('/')
            or next_url.startswith('//')
            or next_url.startswith('/\\')):
        next_url = ''

    if request.method == 'GET':
        return render_template('login.html', error=None, next_url=next_url, form={})

    # Reject obvious flooding before we touch the DB or the password hasher.
    ip = request.remote_addr or 'unknown'
    if _ip_rate_limited(ip):
        return render_template('login.html',
                               error="Too many attempts — please wait a few minutes and try again.",
                               next_url=next_url, form={}), 429

    identifier = (request.form.get('username') or '').strip()
    password   = request.form.get('password') or ''

    # Username OR email both accepted as identifier so users don't have
    # to remember which they picked.
    user_row = None
    if identifier:
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                user_row = cur.execute(
                    "SELECT id, username, password_hash, is_active, failed_attempts, "
                    "locked_until FROM users "
                    "WHERE LOWER(username) = %s OR email = %s LIMIT 1",
                    (identifier.lower(), identifier.lower())
                ).fetchone()
        except SQLAlchemyError as e:
            print(f"❌ Login lookup error: {e}")
            return render_template('login.html',
                                   error="Could not sign in — please try again.",
                                   next_url=next_url, form={'username': identifier}), 500

    generic_error = "Invalid username or password."

    if user_row is None:
        # Burn a hash compare against a pre-computed dummy to keep timing
        # roughly constant whether or not the username exists. Stops a
        # trivial username-enumeration timing oracle.
        check_password_hash(_DUMMY_PASSWORD_HASH, password)
        return render_template('login.html', error=generic_error,
                               next_url=next_url, form={'username': identifier}), 401

    if not user_row['is_active']:
        return render_template('login.html',
                               error="This account has been deactivated.",
                               next_url=next_url, form={'username': identifier}), 403

    if _account_is_locked(user_row):
        return render_template('login.html',
                               error=("This account is temporarily locked due to too many "
                                      "failed attempts. Try again in a few minutes."),
                               next_url=next_url, form={'username': identifier}), 423

    if not check_password_hash(user_row['password_hash'], password):
        _record_failed_login(int(user_row['id']))
        return render_template('login.html', error=generic_error,
                               next_url=next_url, form={'username': identifier}), 401

    _rotate_session(int(user_row['id']))
    _record_successful_login(int(user_row['id']))
    # Backfill tournament faces for pre-existing users: the seeder is
    # idempotent — no-op if the rows already exist with the right flags.
    _seed_tournament_faces(int(user_row['id']))
    # Migrate pre-existing users off the legacy NASP raster default onto the
    # WA 40cm 10-ring vector face. Idempotent — no-op once migrated.
    _seed_user_default_target(int(user_row['id']))
    return redirect(next_url or url_for('index'))


@app.route('/logout', methods=['POST'])
def logout():
    """Clear the session and bounce back to the splash.

    POST-only so a malicious image tag can't log a victim out via GET
    (annoyance, not a security issue, but easy to prevent). CSRF token
    is enforced by Flask-WTF on every POST.
    """
    session.clear()
    return redirect(url_for('index'))


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    """Start the password-reset flow.

    Accepts a username or email. The response is intentionally identical
    whether or not the identifier matches a real account — leaking that
    distinction here would re-open the enumeration hole that /login
    closes with its generic error.
    """
    if current_user_id() is not None:
        return redirect(url_for('index'))

    if request.method == 'GET':
        return render_template('forgot_password.html', error=None, sent=False, form={})

    ip = request.remote_addr or 'unknown'
    if _reset_ip_rate_limited(ip):
        return render_template('forgot_password.html',
                               error="Too many reset requests — please wait and try again later.",
                               sent=False, form={}), 429

    identifier = (request.form.get('identifier') or '').strip()
    # Always render the same "sent" page so the response shape doesn't
    # leak whether the identifier matched. We still do the lookup, send
    # the email, and burn the work — just don't tell the client.
    generic_sent = render_template('forgot_password.html',
                                   error=None, sent=True, form={})

    if not identifier:
        return generic_sent

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # LOWER() on both sides so "Dave" and "dave" both match a
            # registered "Dave". Email is already stored normalized.
            user_row = cur.execute(
                "SELECT id, username, email, is_active FROM users "
                "WHERE LOWER(username) = %s OR email = %s LIMIT 1",
                (identifier.lower(), identifier.lower())
            ).fetchone()
    except SQLAlchemyError as e:
        print(f"❌ Forgot-password lookup error: {e}")
        return generic_sent

    if user_row is None or not user_row['is_active']:
        return generic_sent

    # token_urlsafe(32) → ~43 chars of base64url, ≥256 bits of entropy.
    # Brute-forcing this within the 60-minute TTL is not realistic.
    token = secrets.token_urlsafe(32)
    token_hash = _hash_reset_token(token)
    now = _app_now()
    expires = now + timedelta(minutes=PASSWORD_RESET_TTL_MINUTES)

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Invalidate any prior outstanding tokens for this user so an
            # attacker who somehow obtained an earlier one can't keep it
            # warm by re-requesting.
            cur.execute(
                "UPDATE password_resets SET used_at = %s "
                "WHERE user_id = %s AND used_at IS NULL",
                (now, int(user_row['id']))
            )
            cur.execute(
                "INSERT INTO password_resets "
                "(user_id, token_hash, created_at, expires_at) "
                "VALUES (%s, %s, %s, %s)",
                (int(user_row['id']), token_hash, now, expires)
            )
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Forgot-password insert error: {e}")
        return generic_sent

    reset_url = f"{_password_reset_base_url()}{url_for('reset_password', token=token)}"
    body = (
        f"Hi {user_row['username']},\n\n"
        f"Someone (hopefully you) asked to reset your Apollo password.\n"
        f"Click the link below to choose a new one. It expires in "
        f"{PASSWORD_RESET_TTL_MINUTES} minutes.\n\n"
        f"{reset_url}\n\n"
        f"If you didn't request this, you can ignore this email — your "
        f"password won't change.\n"
    )
    _send_email(user_row['email'], "Reset your Apollo password", body)
    return generic_sent


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    """Land here from the email link. GET shows the form, POST applies the change.

    Token validity is checked on both verbs so a stale GET doesn't render
    a form the POST would then reject — the user sees the "expired" page
    immediately. The token is rotated out on first successful use to
    prevent replay if the email is forwarded or sits in a logged proxy.
    """
    if current_user_id() is not None:
        # Signed-in users go through /account, not the recovery flow.
        return redirect(url_for('index'))

    token_hash = _hash_reset_token(token or '')
    now = _app_now()
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT id, user_id, expires_at, used_at FROM password_resets "
                "WHERE token_hash = %s LIMIT 1",
                (token_hash,)
            ).fetchone()
    except SQLAlchemyError as e:
        print(f"❌ Reset-password lookup error: {e}")
        return render_template('reset_password.html',
                               error="Could not process reset link — please try again.",
                               token=token, valid=False, success=None), 500

    valid = row is not None and row['used_at'] is None
    if valid:
        exp = row['expires_at']
        try:
            exp_dt = exp if isinstance(exp, datetime) else datetime.fromisoformat(str(exp))
            if exp_dt <= now:
                valid = False
        except (ValueError, TypeError):
            valid = False

    if not valid:
        return render_template('reset_password.html',
                               error="This reset link is invalid or has expired. "
                                     "Please request a new one.",
                               token=token, valid=False, success=None), 400

    if request.method == 'GET':
        return render_template('reset_password.html', error=None,
                               token=token, valid=True, success=None)

    new_pw  = request.form.get('new_password') or ''
    confirm = request.form.get('confirm_new_password') or ''
    if len(new_pw) < MIN_PASSWORD_LEN or len(new_pw) > MAX_PASSWORD_LEN:
        return render_template('reset_password.html',
            error=f"Password must be {MIN_PASSWORD_LEN}–{MAX_PASSWORD_LEN} characters.",
            token=token, valid=True, success=None)
    if new_pw != confirm:
        return render_template('reset_password.html',
            error="Passwords do not match.",
            token=token, valid=True, success=None)

    new_hash = generate_password_hash(new_pw)
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Atomic consume: UPDATE only matches a row whose used_at is
            # still NULL. cur.rowcount tells us whether we actually won the
            # race — two simultaneous POSTs both see used_at=NULL on their
            # SELECT, but only one UPDATE matches.
            cur.execute(
                "UPDATE password_resets SET used_at = %s "
                "WHERE id = %s AND used_at IS NULL",
                (now, int(row['id']))
            )
            if cur.rowcount == 0:
                con.commit()
                return render_template('reset_password.html',
                    error="This reset link has already been used.",
                    token=token, valid=False, success=None), 400
            cur.execute(
                "UPDATE users SET password_hash = %s, failed_attempts = 0, "
                "locked_until = NULL WHERE id = %s",
                (new_hash, int(row['user_id']))
            )
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Reset-password update error: {e}")
        return render_template('reset_password.html',
            error="Could not reset password — please try again.",
            token=token, valid=True, success=None), 500

    return render_template('reset_password.html', error=None,
                           token=token, valid=False,
                           success="Your password has been reset. You can now sign in.")


@app.route('/account', methods=['GET', 'POST'])
@login_required
def account():
    """Account settings page — currently change-password and change-email.

    Re-prompts for the *current* password on any change. That's the
    standard "the right person is in front of the keyboard" check
    (defends against opportunistic access to an unlocked laptop).
    """
    user = current_user()
    if user is None:
        return redirect(url_for('login'))

    if request.method == 'GET':
        return render_template('account.html', error=None, success=None)

    action = request.form.get('action') or ''
    current_pw = request.form.get('current_password') or ''

    # Per-user timezone is a display-only preference — no password re-prompt.
    if action == 'change_timezone':
        new_tz = (request.form.get('timezone') or '').strip()
        if new_tz not in available_timezones():
            return render_template('account.html',
                error="Unknown timezone.", success=None), 400
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "UPDATE users SET timezone = %s WHERE id = %s",
                    (new_tz, user['rowid'])
                )
                con.commit()
        except SQLAlchemyError as e:
            print(f"❌ Change-timezone error: {e}")
            return render_template('account.html',
                error="Could not update timezone — please try again.",
                success=None), 500
        # Drop the cached current_user so the new tz is reflected on the
        # next template render in this same request.
        g.pop('_current_user', None)
        return render_template('account.html', error=None,
                               success=f"Your timezone set to {new_tz}.")

    # Archer profile (gender / age group / default bowstyle) drives the
    # classification category. Display-only preference — no password re-prompt.
    if action == 'change_archer_profile':
        gender = (request.form.get('gender') or '').strip().lower()
        age_group = (request.form.get('age_group') or '').strip().lower()
        bowstyle = (request.form.get('default_bowstyle') or '').strip().lower()
        # Blank = "leave unset"; otherwise validate against the known options.
        if gender and gender not in ('male', 'female', 'open'):
            return render_template('account.html',
                error="Unknown gender option.", success=None), 400
        if age_group and age_group not in classifications.data.AGB_AGE_STEP:
            return render_template('account.html',
                error="Unknown age group.", success=None), 400
        if bowstyle and bowstyle not in classifications.data.AGB_BOWSTYLES:
            return render_template('account.html',
                error="Unknown bowstyle.", success=None), 400
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "UPDATE users SET gender = %s, age_group = %s, "
                    "default_bowstyle = %s WHERE id = %s",
                    (gender or None, age_group or None, bowstyle or None,
                     user['rowid'])
                )
                con.commit()
        except SQLAlchemyError as e:
            print(f"❌ Change-archer-profile error: {e}")
            return render_template('account.html',
                error="Could not update archer profile — please try again.",
                success=None), 500
        g.pop('_current_user', None)
        return render_template('account.html', error=None,
                               success="Archer profile updated.")

    # Server-wide timezone is admin-only and only meaningful on the
    # multi-user MySQL flavor; reject from the local SQLite install too.
    if action == 'change_server_timezone':
        if APOLLO_BACKEND != 'mysql' or not user.get('is_root'):
            return render_template('account.html',
                error="Not authorized to change server timezone.",
                success=None), 403
        new_tz = (request.form.get('server_timezone') or '').strip()
        if new_tz not in available_timezones():
            return render_template('account.html',
                error="Unknown server timezone.", success=None), 400
        if not set_app_setting('server_timezone', new_tz):
            return render_template('account.html',
                error="Could not update server timezone — please try again.",
                success=None), 500
        return render_template('account.html', error=None,
                               success=f"Server timezone set to {new_tz}.")

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT password_hash FROM users WHERE id = %s",
                (user['rowid'],)
            ).fetchone()
    except SQLAlchemyError as e:
        print(f"❌ Account lookup error: {e}")
        return render_template('account.html',
                               error="Could not update account — please try again.",
                               success=None), 500

    if row is None or not check_password_hash(row['password_hash'], current_pw):
        return render_template('account.html',
                               error="Current password is incorrect.",
                               success=None), 401

    if action == 'change_password':
        new_pw  = request.form.get('new_password') or ''
        confirm = request.form.get('confirm_new_password') or ''
        if len(new_pw) < MIN_PASSWORD_LEN or len(new_pw) > MAX_PASSWORD_LEN:
            return render_template('account.html',
                error=f"New password must be {MIN_PASSWORD_LEN}–{MAX_PASSWORD_LEN} characters.",
                success=None)
        if new_pw != confirm:
            return render_template('account.html',
                error="New passwords do not match.", success=None)
        new_hash = generate_password_hash(new_pw)
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "UPDATE users SET password_hash = %s WHERE id = %s",
                    (new_hash, user['rowid'])
                )
                con.commit()
        except SQLAlchemyError as e:
            print(f"❌ Change-password error: {e}")
            return render_template('account.html',
                error="Could not change password — please try again.",
                success=None), 500
        # Rotate the session so any other sessions held by the same user
        # (forgotten browser, shared laptop) are invalidated on next hit.
        _rotate_session(user['rowid'])
        return render_template('account.html', error=None,
                               success="Password updated.")

    if action == 'change_email':
        new_email = (request.form.get('new_email') or '').strip().lower()
        if not EMAIL_RE.match(new_email) or len(new_email) > 255:
            return render_template('account.html',
                error="Please enter a valid email address.", success=None)
        old_email = (user.get('email') or '').strip().lower()
        if new_email == old_email:
            return render_template('account.html',
                error="That's already your email address.", success=None)
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "UPDATE users SET email = %s WHERE id = %s",
                    (new_email, user['rowid'])
                )
                con.commit()
        except DBIntegrityError:
            return render_template('account.html',
                error="An account with that email already exists.", success=None)
        except SQLAlchemyError as e:
            print(f"❌ Change-email error: {e}")
            return render_template('account.html',
                error="Could not change email — please try again.",
                success=None), 500
        # Notify the prior address so a takeover via account-page email
        # swap is at least visible to the legitimate owner. Failures are
        # swallowed — the update has already happened and email delivery
        # is best-effort.
        if old_email:
            try:
                _send_email(
                    old_email,
                    "Your Apollo account email was changed",
                    f"The email address on your Apollo account was just changed "
                    f"to {new_email}.\n\n"
                    f"If you did not make this change, reply to this message "
                    f"immediately — your account may have been accessed by "
                    f"someone else."
                )
            except Exception as e:
                print(f"⚠️  Failed to notify {old_email} of email change: {e}")
        return render_template('account.html', error=None,
                               success="Email updated. A notice was sent to your previous address.")

    return render_template('account.html',
                           error="Unknown action.", success=None), 400


@app.route('/delete_account', methods=['POST'])
@login_required
def delete_account():
    """Permanently delete the signed-in user and all their data.

    Hard delete across every per-user table so we don't leave orphan rows
    behind. Requires re-entering the password — same threat model as
    /account (someone walks up to an unlocked laptop).
    """
    user = current_user()
    if user is None:
        return redirect(url_for('login'))

    current_pw = request.form.get('current_password') or ''
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT password_hash FROM users WHERE id = %s",
                (user['rowid'],)
            ).fetchone()
    except SQLAlchemyError as e:
        print(f"❌ Delete-account lookup error: {e}")
        return render_template('account.html',
                               error="Could not delete account — please try again.",
                               success=None), 500
    if row is None or not check_password_hash(row['password_hash'], current_pw):
        return render_template('account.html',
                               error="Password incorrect — account not deleted.",
                               success=None), 401

    try:
        _purge_user(user['rowid'])
    except SQLAlchemyError as e:
        print(f"❌ Delete-account error: {e}")
        return render_template('account.html',
                               error="Could not delete account — please try again.",
                               success=None), 500

    session.clear()
    return redirect(url_for('index'))


def _purge_user(user_id):
    """Hard-delete a user and every row they own. Raises on DB error.

    Shared by self-deletion (/delete_account) and admin deletion
    (/admin/users/<id>/delete). Removes uploaded target images from disk
    after the DB commit, but only when no other user still references the
    same file — image_filename is shared by reference when users export/
    import targets between accounts.
    """
    uploaded_images = []
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        target_rows = cur.execute(
            "SELECT image_filename FROM targets WHERE user_id = %s",
            (user_id,)
        ).fetchall()
        for r in target_rows:
            fn = r['image_filename']
            # Drop anything that doesn't resolve cleanly inside the
            # uploads dir — an imported row with '..' in image_filename
            # would otherwise let this loop reach files outside static/.
            if _resolve_target_image_disk_path(fn) is None:
                continue
            other = cur.execute(
                "SELECT 1 FROM targets WHERE image_filename = %s "
                "AND user_id <> %s LIMIT 1",
                (fn, user_id)
            ).fetchone()
            if other is None:
                uploaded_images.append(fn)
        for tbl in _PER_USER_TABLES:
            cur.execute(f"DELETE FROM {tbl} WHERE user_id = %s", (user_id,))
        cur.execute("DELETE FROM password_resets WHERE user_id = %s",
                    (user_id,))
        cur.execute("DELETE FROM user_notes WHERE user_id = %s",
                    (user_id,))
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
        con.commit()

    for fn in uploaded_images:
        disk = _resolve_target_image_disk_path(fn)
        if disk is None:
            continue
        try:
            os.remove(disk)
        except OSError:
            pass


# ─── Root admin routes ───────────────────────────────────────────────────

@app.route('/admin', methods=['GET'])
@root_required
def admin_users():
    """List every account. Root-only.

    Joins the bare users table with COUNT(*) over apollo so admins can see
    activity at a glance. Sorted by created_at ascending so the oldest
    accounts (typically the original installer) appear at the top.
    """
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT u.id AS rowid, u.username, u.email, u.created_at, "
                "u.last_login, u.is_active, u.is_root, "
                "(SELECT COUNT(*) FROM apollo a WHERE a.user_id = u.id) AS shot_count "
                "FROM users u ORDER BY u.created_at"
            ).fetchall()
    except SQLAlchemyError as e:
        print(f"❌ Admin list error: {e}")
        rows = []
    return render_template('admin.html', users=rows, error=None, success=None)


def _admin_render(error=None, success=None, status=200):
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT u.id AS rowid, u.username, u.email, u.created_at, "
                "u.last_login, u.is_active, u.is_root, "
                "(SELECT COUNT(*) FROM apollo a WHERE a.user_id = u.id) AS shot_count "
                "FROM users u ORDER BY u.created_at"
            ).fetchall()
    except SQLAlchemyError:
        rows = []
    return render_template('admin.html', users=rows,
                           error=error, success=success), status


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@root_required
def admin_delete_user(user_id):
    """Hard-delete a user and all their data. Root-only.

    Forbids self-deletion via this route — root removing themselves through
    the admin UI is almost always a mistake (the next request would 403).
    A root user who really wants to delete their own account can still use
    /delete_account on the account page.
    """
    me = current_user()
    if me is not None and int(me['rowid']) == user_id:
        return _admin_render(error="Use the account page to delete your own account.",
                             status=400)
    try:
        _purge_user(user_id)
    except SQLAlchemyError as e:
        print(f"❌ Admin delete error: {e}")
        return _admin_render(error="Could not delete user — please try again.",
                             status=500)
    return _admin_render(success=f"User {user_id} deleted.")


@app.route('/admin/users/<int:user_id>/password', methods=['POST'])
@root_required
def admin_change_password(user_id):
    """Force-set a user's password. Root-only.

    Resets failed_attempts and locked_until alongside the password change so
    an admin reset also unsticks a locked-out account in one step.
    """
    new_pw  = request.form.get('new_password') or ''
    confirm = request.form.get('confirm_new_password') or ''
    if len(new_pw) < MIN_PASSWORD_LEN or len(new_pw) > MAX_PASSWORD_LEN:
        return _admin_render(
            error=f"Password must be {MIN_PASSWORD_LEN}–{MAX_PASSWORD_LEN} characters.",
            status=400)
    if new_pw != confirm:
        return _admin_render(error="Passwords do not match.", status=400)

    new_hash = generate_password_hash(new_pw)
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            cur.execute(
                "UPDATE users SET password_hash = %s, failed_attempts = 0, "
                "locked_until = NULL WHERE id = %s",
                (new_hash, user_id)
            )
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Admin change-password error: {e}")
        return _admin_render(error="Could not change password — please try again.",
                             status=500)
    return _admin_render(success=f"Password updated for user {user_id}.")


@app.route('/admin/users/<int:user_id>/email', methods=['POST'])
@root_required
def admin_change_email(user_id):
    """Set a user's email address. Root-only."""
    new_email = (request.form.get('new_email') or '').strip().lower()
    if not EMAIL_RE.match(new_email) or len(new_email) > 255:
        return _admin_render(error="Please enter a valid email address.", status=400)
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            cur.execute(
                "UPDATE users SET email = %s WHERE id = %s",
                (new_email, user_id)
            )
            con.commit()
    except DBIntegrityError:
        return _admin_render(error="An account with that email already exists.",
                             status=400)
    except SQLAlchemyError as e:
        print(f"❌ Admin change-email error: {e}")
        return _admin_render(error="Could not change email — please try again.",
                             status=500)
    return _admin_render(success=f"Email updated for user {user_id}.")

def _normalize_tags(raw):
    """Clean a comma-separated tag string: trim, drop empties, dedupe (case-
    insensitive, keeping first-seen casing). Returns a canonical
    comma-separated string suitable for storage."""
    if not raw:
        return ''
    seen = set()
    out = []
    for part in raw.split(','):
        t = part.strip()
        if not t:
            continue
        key = t.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(t)
    return ', '.join(out)


def _distinct_user_tags(user_id):
    """Return a sorted list of distinct tag strings the user has used
    previously. Powers the autocomplete on the session form."""
    if user_id is None:
        return []
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT DISTINCT session_tags FROM apollo "
                "WHERE user_id = %s AND session_tags IS NOT NULL "
                "AND session_tags <> ''",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError:
        return []
    seen = set()
    out = []
    for row in rows:
        raw = row[0] if row else ''
        if not raw:
            continue
        for part in raw.split(','):
            t = part.strip()
            if not t:
                continue
            key = t.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(t)
    out.sort(key=str.lower)
    return out


def _style_settings_by_bow(user_id):
    """Return ``{bow_model: {type, values}}`` to prefill the session and
    tournament forms' bowtype-specific settings group.

    ``values`` is the bow's saved gear (legacy static blob, if any) overlaid
    with the *most recent shot's* full snapshot, so every field prefills to its
    last-used value and tweaks persist across shots within a session. On a bow's
    first-ever shot the gear starts blank. Empty dict on error so the template
    still renders.
    """
    if user_id is None:
        return {}
    out = {}
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            for row in cur.execute(
                "SELECT bow_model, bow_type, style_settings FROM bows "
                "WHERE user_id = %s", (user_id,)
            ).fetchall() or []:
                model = row['bow_model'] if 'bow_model' in row else row[0]
                btype = row['bow_type']  if 'bow_type'  in row else row[1]
                raw   = row['style_settings'] if 'style_settings' in row else row[2]
                if not model:
                    continue
                try:
                    static = json.loads(raw) if raw else {}
                except (ValueError, TypeError):
                    static = {}
                out[str(model)] = {'type': btype, 'values': dict(static)}
            # Overlay each bow's most recent shot snapshot (merged static +
            # dynamic) so the form repopulates the last-used values.
            for row in cur.execute(
                "SELECT a.bow, a.bow_style_settings FROM apollo a "
                "INNER JOIN ("
                "    SELECT bow, MAX(id) AS max_id FROM apollo "
                "    WHERE user_id = %s "
                "      AND bow_style_settings IS NOT NULL AND bow_style_settings <> '' "
                "      AND bow IS NOT NULL AND bow <> '' "
                "    GROUP BY bow"
                ") latest ON latest.bow = a.bow AND latest.max_id = a.id "
                "WHERE a.user_id = %s", (user_id, user_id)
            ).fetchall() or []:
                bow  = row['bow'] if 'bow' in row else row[0]
                blob = row['bow_style_settings'] if 'bow_style_settings' in row else row[1]
                if not bow or str(bow) not in out:
                    continue
                try:
                    snap = json.loads(blob) if blob else {}
                except (ValueError, TypeError):
                    snap = {}
                if isinstance(snap, dict):
                    out[str(bow)]['values'].update(snap)
    except SQLAlchemyError:
        return out
    return out


def _last_session_tags(user_id, session_id):
    """Return the most recently stored session_tags string for this session,
    so a mid-session reload repopulates the tags input."""
    if user_id is None or session_id is None:
        return ''
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT session_tags FROM apollo "
                "WHERE user_id = %s AND session_id = %s "
                "AND session_tags IS NOT NULL AND session_tags <> '' "
                "ORDER BY id DESC LIMIT 1",
                (user_id, session_id)
            ).fetchone()
    except SQLAlchemyError:
        return ''
    if not row:
        return ''
    return row[0] or ''


def _last_session_bow_arrow(user_id, session_id):
    """Return ``(bow, arrow_type)`` from the latest shot in this session so a
    mid-session GET reload repopulates the bow/arrow dropdowns instead of
    snapping them back to the first option. Unlike distance/quiver-size (which
    the client hydrates from localStorage), these are session-scoped, so the
    server has to restore them. Returns ('', '') for a fresh session."""
    if user_id is None or session_id is None:
        return ('', '')
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT bow, arrow_type FROM apollo "
                "WHERE user_id = %s AND session_id = %s "
                "ORDER BY id DESC LIMIT 1",
                (user_id, session_id)
            ).fetchone()
    except SQLAlchemyError:
        return ('', '')
    if not row:
        return ('', '')
    return (row[0] or '', row[1] or '')


def _insert_shot(cur, *, user_id, session_id, timestamp, bow, arrow_type,
                 quiver_size, arrows_remaining, distance, session_notes,
                 x, y, is_precise, record_mode, target_id,
                 session_tags, style_settings_session=None,
                 client_uuid=None):
    """Snapshot the current bow/arrow config and INSERT one shot row.

    Single source of truth for the per-shot insert, shared by the live
    /sesh POST path and the offline-sync batch endpoint (/api/sync_shots).
    The caller owns the connection and the commit; ``cur`` must be an open
    cursor.

    ``timestamp`` and ``arrows_remaining`` are passed in rather than
    computed here: the live path passes ``_app_now()`` plus the running
    cookie counter; the sync path passes the real on-device shot time plus
    the counter the client tracked while offline. ``arrows_remaining`` is
    always the *pre-decrement* value, which get_past_shots() and
    recall_arrow() depend on. Effective draw weight is an intrinsic property
    of the bow now, so it's snapshotted from the bow row like draw weight.
    """
    # Snapshot the bow's current config onto the shot row. Without this, a
    # later edit to the bow (draw weight changed, AMO measured more
    # carefully, rename) would silently rewrite the historical attribution
    # of every shot ever taken with it. bow_type is included so the row
    # remains self-describing even if the bow row is later deleted.
    bow_row = cur.execute(
        "SELECT effective_draw_weight, bow_draw_weight, "
        "amo, bow_type, style_settings FROM bows "
        "WHERE bow_model = %s AND user_id = %s LIMIT 1",
        (bow, user_id)
    ).fetchone()
    if bow_row is not None:
        shot_effective_dw    = bow_row['effective_draw_weight'] if 'effective_draw_weight' in bow_row else bow_row[0]
        shot_bow_draw_weight = bow_row['bow_draw_weight']       if 'bow_draw_weight'       in bow_row else bow_row[1]
        shot_bow_amo         = bow_row['amo']                   if 'amo'                   in bow_row else bow_row[2]
        shot_bow_type        = bow_row['bow_type']              if 'bow_type'              in bow_row else bow_row[3]
        bow_style_raw        = bow_row['style_settings']        if 'style_settings'        in bow_row else bow_row[4]
    else:
        shot_effective_dw = shot_bow_draw_weight = shot_bow_amo = shot_bow_type = None
        bow_style_raw = None

    # Snapshot the bowtype-specific settings: the bow's saved gear (legacy bows
    # may still carry a static blob) overlaid with whatever the form posted this
    # session (validated against the schema for the bow's type). Stored as JSON
    # so later edits to the bow don't rewrite a past shot's gear. None when
    # nothing applies.
    try:
        bow_static = json.loads(bow_style_raw) if bow_style_raw else {}
    except (ValueError, TypeError):
        bow_static = {}
    merged_style = dict(bow_static)
    if style_settings_session:
        merged_style.update(
            _clean_style_settings(shot_bow_type, style_settings_session))
    shot_style_settings = json.dumps(merged_style) if merged_style else None

    # Same snapshot for the arrow side: capture every field on the arrows
    # row so a later rename/edit doesn't rewrite the per-shot record. spine
    # is snapshotted too even though it's locked in the UI — keeps the
    # apollo row self-describing if the arrow row is ever deleted.
    arrow_row = cur.execute(
        "SELECT length, spine, shaft_weight, shaft_diameter, "
        "shaft_material, nock_weight, tip, tip_weight FROM arrows "
        "WHERE arrow = %s AND user_id = %s LIMIT 1",
        (arrow_type, user_id)
    ).fetchone()
    if arrow_row is not None:
        shot_arrow_length    = arrow_row['length']         if 'length'         in arrow_row else arrow_row[0]
        shot_arrow_spine     = arrow_row['spine']          if 'spine'          in arrow_row else arrow_row[1]
        shot_arrow_shaft_w   = arrow_row['shaft_weight']   if 'shaft_weight'   in arrow_row else arrow_row[2]
        shot_arrow_shaft_d   = arrow_row['shaft_diameter'] if 'shaft_diameter' in arrow_row else arrow_row[3]
        shot_arrow_shaft_m   = arrow_row['shaft_material'] if 'shaft_material' in arrow_row else arrow_row[4]
        shot_arrow_nock_w    = arrow_row['nock_weight']    if 'nock_weight'    in arrow_row else arrow_row[5]
        shot_arrow_tip       = arrow_row['tip']            if 'tip'            in arrow_row else arrow_row[6]
        shot_arrow_tip_w     = arrow_row['tip_weight']     if 'tip_weight'     in arrow_row else arrow_row[7]
    else:
        shot_arrow_length = shot_arrow_spine = shot_arrow_shaft_w = None
        shot_arrow_shaft_d = shot_arrow_shaft_m = shot_arrow_nock_w = None
        shot_arrow_tip = shot_arrow_tip_w = None

    cur.execute("""
        INSERT INTO apollo (user_id, session_id, timestamp, bow,
        arrow_type, quiver_size, arrows_remaining,
        distance, session_notes, x_coord, y_coord, is_precise,
        record_mode, target_id,
        bow_draw_weight, effective_draw_weight, bow_amo, bow_type,
        arrow_length, arrow_spine, arrow_shaft_weight,
        arrow_shaft_diameter, arrow_shaft_material,
        arrow_nock_weight, arrow_tip, arrow_tip_weight,
        session_tags, bow_style_settings, client_uuid)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s)""",
        (user_id, session_id, timestamp, bow, arrow_type, quiver_size,
         arrows_remaining, distance, session_notes, x, y, is_precise,
         record_mode, target_id,
         shot_bow_draw_weight, shot_effective_dw, shot_bow_amo, shot_bow_type,
         shot_arrow_length, shot_arrow_spine, shot_arrow_shaft_w,
         shot_arrow_shaft_d, shot_arrow_shaft_m,
         shot_arrow_nock_w, shot_arrow_tip, shot_arrow_tip_w,
         session_tags, shot_style_settings, client_uuid)
    )


class _ShotRejected(Exception):
    """A per-shot submission the server refuses to record.

    Carries a user-facing ``message`` and HTTP ``code`` so the two callers
    can shape their own response — the full-page /sesh POST returns the
    message as plain-text, the AJAX /api/record_shot returns it as JSON.
    """
    def __init__(self, message, code=400):
        super().__init__(message)
        self.message = message
        self.code = code


def _session_counters_from_db(user_id, session_id):
    """Re-derive (effective_quiver_size, arrows_remaining, quivers_completed)
    for a session from its shot rows alone.

    Used to heal the Flask-session counters after an *idempotent* (retried)
    shot whose original response — and its Set-Cookie — was lost: the cookie
    may or may not reflect the already-saved shot, so we recompute the truth
    from the DB instead of trusting it.

    ``quivers_completed`` is the count of shots whose stored (pre-decrement)
    ``arrows_remaining`` is 1 — each such shot is the one that closed a
    quiver. The current ``arrows_remaining`` is the last shot's pre-decrement
    value, decremented and wrapped exactly as the live path computes it.
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        last = cur.execute(
            "SELECT quiver_size, arrows_remaining FROM apollo "
            "WHERE session_id = %s AND user_id = %s ORDER BY id DESC LIMIT 1",
            (session_id, user_id)
        ).fetchone()
        if not last:
            return 0, 0, 0
        try:
            eff = int(last['quiver_size'] if 'quiver_size' in last else last[0])
            ar  = int(last['arrows_remaining'] if 'arrows_remaining' in last else last[1])
        except (TypeError, ValueError):
            return 0, 0, 0
        qc_row = cur.execute(
            "SELECT COUNT(*) FROM apollo "
            "WHERE session_id = %s AND user_id = %s AND arrows_remaining = 1",
            (session_id, user_id)
        ).fetchone()
        quivers_completed = int(qc_row[0]) if qc_row and qc_row[0] is not None else 0
    ar -= 1
    if ar <= 0:
        ar = eff
    return eff, ar, quivers_completed


def _apply_shot(*, user_id, session_id, bow, arrow_type, distance,
                session_notes, session_tags, style_settings_session,
                x, y, is_precise, record_mode, target_id,
                quiver_size_int, mid_quiver, current_quiver_size,
                arrows_remaining, quivers_completed, client_uuid=None):
    """Validate, insert, and advance the counters for one live shot.

    Single source of truth for the per-shot mutation shared by the full-page
    /sesh POST and the AJAX /api/record_shot endpoint. Returns the tuple
    ``(effective_quiver_size, arrows_remaining, quivers_completed,
    past_shots)``; the caller owns the Flask-session writes and the response
    shape. Raises :class:`_ShotRejected` for a submission the server won't
    accept (missing coords, missing/zero quiver size, a mid-quiver size
    change) and may raise ``SQLAlchemyError`` on a DB failure.

    Idempotency: when ``client_uuid`` is supplied and a row with the same
    ``(user_id, client_uuid)`` already exists — a retried submit after a
    dropped response — no second row is inserted and the counters are
    recomputed from the DB (see :func:`_session_counters_from_db`) rather
    than advanced again, so the shot can't be double-recorded or
    double-counted.
    """
    if x == '' or y == '':
        raise _ShotRejected("Error: missing shot coordinates", 400)
    if quiver_size_int <= 0:
        # A shot with quiver_size=0 would inflate quivers_completed by one per
        # shot (arrows_remaining -= 1 → -1 ≤ 0 → reset forever). Reject.
        raise _ShotRejected("Error: quiver size must be a positive integer", 400)
    if mid_quiver and quiver_size_int != current_quiver_size:
        # Mid-quiver the size is locked; the user must finish the quiver (or
        # end the session) before changing it. The UI disables the input, so
        # reaching here implies a hand-crafted POST.
        raise _ShotRejected(
            "Error: quiver size cannot change mid-quiver — "
            "finish the current quiver first.", 400)

    # Between quivers (or at session start) the submitted value becomes the
    # new lock and the new arrows_remaining; mid-quiver keeps the locked size.
    if mid_quiver:
        effective_quiver_size = current_quiver_size
    else:
        effective_quiver_size = quiver_size_int
        arrows_remaining = effective_quiver_size

    duplicate = False
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        if client_uuid:
            existing = cur.execute(
                "SELECT id FROM apollo WHERE user_id = %s AND client_uuid = %s LIMIT 1",
                (user_id, client_uuid)
            ).fetchone()
            duplicate = existing is not None
        if not duplicate:
            try:
                _insert_shot(
                    cur,
                    user_id=user_id, session_id=session_id,
                    timestamp=_app_now(), bow=bow, arrow_type=arrow_type,
                    quiver_size=effective_quiver_size,
                    arrows_remaining=arrows_remaining,
                    distance=distance, session_notes=session_notes,
                    x=x, y=y, is_precise=is_precise,
                    record_mode=record_mode, target_id=target_id,
                    session_tags=session_tags,
                    style_settings_session=style_settings_session,
                    client_uuid=client_uuid,
                )
                con.commit()
            except DBIntegrityError:
                # A concurrent retry carrying the same client_uuid won the
                # race to the unique index — treat as a duplicate.
                con.rollback()
                duplicate = True

    if duplicate:
        # Retried submit (lost response): don't re-record or double-count —
        # recompute the authoritative counters from the DB.
        eff, ar, qc = _session_counters_from_db(user_id, session_id)
        return eff, ar, qc, get_past_shots(session_id, eff, ar, user_id)

    # Quiver bookkeeping: each saved shot decrements the remaining counter;
    # hitting zero closes a quiver and refills with the locked size.
    arrows_remaining -= 1
    if arrows_remaining <= 0:
        quivers_completed += 1
        arrows_remaining = effective_quiver_size
    past_shots = get_past_shots(
        session_id, effective_quiver_size, arrows_remaining, user_id)
    return effective_quiver_size, arrows_remaining, quivers_completed, past_shots


@app.route('/sesh', methods=['GET', 'POST'])
@login_required
def sesh():
    """Main active-session page. Handles both the initial GET and per-shot POSTs.

    Session state (current session_id, quivers_completed, arrows_remaining,
    record_mode) lives in the Flask session cookie so the user can navigate
    away (e.g. add a new bow) and come back to the in-progress round.

    POST with ``arrow_shot`` field = "shot recorded", which inserts an
    apollo row and re-renders. Coordinate values arrive as mm strings;
    the JS in session.html does the pixel→mm conversion client-side.
    """
    user_id = current_user_id()
    try:
        # First visit to /sesh in this browser session: mint a new session_id
        # by reading MAX(session_id)+1 from *this user's* apollo + session_times
        # rows. The session_id namespace is now per-user, so two different users
        # can both have a session 1 — there's no collision because every query
        # in the app filters by user_id.
        #
        # Two concurrent first-visit requests (multiple tabs) could pick the
        # same MAX+1, so the INSERT is guarded by the unique index on
        # (user_id, session_id) and we retry on IntegrityError up to a few
        # times before giving up.
        if not session.get('session_id'):
            new_session_id = None
            # 12 retries with small random backoff (≤ ~180 ms total) — gives
            # six concurrent tabs from the same user plenty of headroom even
            # under unlucky scheduling without making the user wait long if
            # all retries fail.
            for _attempt in range(12):
                with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                    res = cur.execute(
                        "SELECT MAX(session_id) FROM apollo WHERE user_id = %s",
                        (user_id,)
                    ).fetchone()
                    max_apollo = res[0] if res and res[0] is not None else 0
                    res = cur.execute(
                        "SELECT MAX(session_id) FROM session_times WHERE user_id = %s",
                        (user_id,)
                    ).fetchone()
                    max_st = res[0] if res and res[0] is not None else 0
                    candidate = max(int(max_apollo or 0), int(max_st or 0)) + 1
                    try:
                        cur.execute(
                            "INSERT INTO session_times "
                            "(user_id, session_id, session_begin_time) "
                            "VALUES (%s, %s, %s)",
                            (user_id, candidate, _app_now())
                        )
                        con.commit()
                        new_session_id = candidate
                        break
                    except DBIntegrityError:
                        # Another tab grabbed this id between our SELECT
                        # and INSERT — recompute and try again, with a small
                        # randomized backoff to desynchronize concurrent
                        # requests so they don't lockstep on the same MAX+1.
                        time.sleep(0.005 + secrets.randbelow(20) / 1000.0)
                        continue
            if new_session_id is None:
                return "Could not allocate session id — please try again", 500
            session['session_id'] = new_session_id
            session['quivers_completed'] = 0
            session['arrows_remaining'] = 0
            # ``current_quiver_size`` is the quiver size locked in at the
            # start of the in-progress quiver. None means no quiver has
            # started yet (session is fresh). It only changes when a
            # quiver completes — mid-quiver POSTs that submit a different
            # value are rejected with HTTP 400 below.
            session['current_quiver_size'] = None
            session['record_mode'] = 0
    except SQLAlchemyError as e:
        print(f"❌ Session ID error: {e}")
        return "Error with session ID entry", 500

    # Fetch arrow types from database
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        res = cur.execute(
            "SELECT DISTINCT arrow FROM arrows WHERE user_id = %s",
            (user_id,)
        ).fetchall()
        arrow_types = [row[0] for row in res] if res else []

    # Fetch bows from database
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        res = cur.execute(
            "SELECT DISTINCT bow_model FROM bows WHERE user_id = %s",
            (user_id,)
        ).fetchall()
        bow_models = [row[0] for row in res] if res else []

    # Fetch active targets for the dropdown; only active ones are pickable.
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        target_rows = cur.execute(
            "SELECT id AS rowid, name FROM targets "
            "WHERE is_active = 1 AND user_id = %s ORDER BY name",
            (user_id,)
        ).fetchall()
        targets_list = [{'rowid': r['rowid'], 'name': _display_target_name(r['name'])} for r in target_rows]

    # Has this session already saved any shots? If so, the target is locked
    # in — switching mid-session would invalidate the replay's single-image
    # render assumption.
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        shot_count_row = cur.execute(
            "SELECT COUNT(*) FROM apollo WHERE session_id = %s AND user_id = %s",
            (session['session_id'], user_id)
        ).fetchone()
        session_shot_count = shot_count_row[0] if shot_count_row else 0
        if session_shot_count > 0:
            existing = cur.execute(
                "SELECT target_id FROM apollo WHERE session_id = %s AND user_id = %s LIMIT 1",
                (session['session_id'], user_id)
            ).fetchone()
            if existing and existing[0] is not None:
                session['target_id'] = existing[0]
    target_locked = session_shot_count > 0

    if request.method == 'POST':
        def _form_int(name, default=0):
            # Coerce a form field to int, falling back to default on bad
            # input (empty string, "abc", missing key). Replaces a bare
            # int() that would 500 on any non-integer payload.
            try:
                return int(request.form.get(name, str(default)))
            except (TypeError, ValueError):
                return default

        session_id       = session['session_id']
        bow              = request.form.get('bow', '')
        arrow_type       = request.form.get('arrow_type', '')
        quiver_size      = request.form.get('quiver_size', '')
        distance         = request.form.get('distance', '')
        session_notes    = request.form.get('session_notes', '')
        session_tags     = _normalize_tags(request.form.get('session_tags', ''))
        # Effective draw weight is a property of the bow now (set on add/edit
        # bow); _insert_shot snapshots it from the bow row.
        # Bowtype-specific settings posted this session (namespaced style_<key>).
        # _insert_shot validates them against the bow's actual type and merges
        # with the bow's saved static gear before snapshotting.
        style_settings_session = _collect_style_settings(request.form)
        x                = request.form.get("x_coord", "")
        y                = request.form.get("y_coord", "")
        is_precise       = _form_int('is_precise', 0)
        record_mode      = _form_int('record_mode', 0)

        # Target is locked once a session has any shots — accept the form
        # value only on the very first shot, otherwise stick with whatever
        # we already loaded into the session above.
        if target_locked:
            target_id = session.get('target_id')
        else:
            posted = request.form.get('target_id', '').strip()
            try:
                target_id = int(posted) if posted else session.get('target_id')
            except ValueError:
                target_id = session.get('target_id')
            # Guard: only accept a target_id the current user actually owns.
            # Otherwise a crafted form could bind another user's target to
            # this user's shots (information leak via target image).
            if target_id is not None and get_target(target_id, user_id) is None:
                target_id = None
            if target_id is None:
                default_row = get_default_target(user_id)
                target_id = default_row['rowid'] if default_row is not None else None
            if target_id is not None:
                session['target_id'] = target_id

        session['record_mode'] = record_mode

        try:
            quiver_size_int = int(quiver_size) if quiver_size else 0
        except (TypeError, ValueError):
            quiver_size_int = 0
        quivers_completed = session.get('quivers_completed', 0)

        # Quiver-size lock: once a quiver starts, its size is fixed until
        # it completes. State machine:
        #   - session start          → arrows_remaining == 0
        #   - between quivers        → arrows_remaining == current_quiver_size
        #                              (reset on the previous quiver's last shot)
        #   - mid-quiver             → 0 < arrows_remaining < current_quiver_size
        # In the mid-quiver state, a submitted quiver_size that differs from
        # the locked value is rejected (see the "arrow_shot" branch below).
        # Otherwise the submitted value becomes the new lock.
        arrows_remaining = session.get('arrows_remaining', 0)
        current_quiver_size = session.get('current_quiver_size') or 0
        mid_quiver = 0 < arrows_remaining < current_quiver_size

        if "arrow_shot" in request.form:
            # ── Per-shot submit ─────────────────────────────────────────────
            # ``effective_quiver_size`` is the size we'll actually record on
            # the row and use for bookkeeping. Default it now so fallthrough
            # paths (missing coords) don't NameError; the else-branch below
            # overrides it when the user has supplied a valid new size
            # between quivers.
            effective_quiver_size = current_quiver_size if mid_quiver else quiver_size_int
            past_shots = []
            if x == '' or y == '':
                # No coordinates posted — re-render without saving (matches
                # the long-standing behavior of this no-JS fallback path).
                print("⚠️ Missing coordinates — arrow not saved")
            else:
                # Delegate the validate → insert → bookkeep state machine to
                # the shared helper (also used by /api/record_shot), then mirror
                # the resulting counters back into the session cookie.
                try:
                    (effective_quiver_size, arrows_remaining,
                     quivers_completed, past_shots) = _apply_shot(
                        user_id=user_id, session_id=session_id, bow=bow,
                        arrow_type=arrow_type, distance=distance,
                        session_notes=session_notes, session_tags=session_tags,
                        style_settings_session=style_settings_session,
                        x=x, y=y, is_precise=is_precise, record_mode=record_mode,
                        target_id=target_id, quiver_size_int=quiver_size_int,
                        mid_quiver=mid_quiver,
                        current_quiver_size=current_quiver_size,
                        arrows_remaining=arrows_remaining,
                        quivers_completed=quivers_completed)
                except _ShotRejected as e:
                    print(f"⚠️ {e.message}")
                    return e.message, e.code
                except SQLAlchemyError as e:
                    print(f"❌ Database error: {e}")
                    return "Error saving entry", 500
                session['current_quiver_size'] = effective_quiver_size
                session['arrows_remaining'] = arrows_remaining
                session['quivers_completed'] = quivers_completed
                print(f"✅ Entry saved for session {session_id}")
            # After the first saved shot the target is locked; reflect that
            # in the response so the dropdown disables without a roundtrip.
            target_locked = True
            target_config = target_to_config(get_target(session.get('target_id'), user_id))
            if target_config is not None:
                target_config['zone_radii_mm'] = _zone_radii_for_target(
                    session.get('target_id'), user_id)
                target_config['default_shaft_diameter_mm'] = DEFAULT_SHAFT_DIAMETER_MM
            # Quiver-size input is locked whenever a quiver is in progress
            # (decremented at least once but not yet completed). The template
            # uses this to render the input ``readonly`` and show a hint.
            quiver_size_locked = (
                effective_quiver_size > 0
                and 0 < arrows_remaining < effective_quiver_size
            )
            # Show the locked value in the form on the way back so the next
            # submission round-trips the same number (the field is readonly
            # when locked, but we still want the value present and visible).
            display_quiver_size = (str(effective_quiver_size)
                                   if effective_quiver_size > 0 else quiver_size)
            return render_template('session.html',
                                   session_id=session_id,
                                   arrow_type=arrow_type,
                                   quiver_size=display_quiver_size,
                                   quiver_size_locked=quiver_size_locked,
                                   quivers_completed=quivers_completed,
                                   arrows_remaining=arrows_remaining,
                                   bow=bow,
                                   bows=bow_models,
                                   bowstyle_settings=BOWSTYLE_SETTINGS,
                                   style_settings_by_bow=_style_settings_by_bow(user_id),
                                   session_notes=session_notes,
                                   session_tags=session_tags,
                                   tag_suggestions=_distinct_user_tags(user_id),
                                   distance=distance,
                                   x_coord=x,
                                   y_coord=y,
                                   arrow_types=arrow_types,
                                   arrow_shaft_diameters=_arrow_shaft_diameters_for_user(user_id),
                                   past_shots=past_shots,
                                   is_precise=is_precise,
                                   record_mode=record_mode,
                                   targets_list=targets_list,
                                   selected_target_id=session.get('target_id'),
                                   target_locked=target_locked,
                                   is_tournament=bool(session.get('tournament_round_key')
                                                      or session.get('match_id')),
                                   current_quiver_size=session.get('current_quiver_size') or 0,
                                   target_config=target_config)

    record_mode = session.get('record_mode', 0)
    # GET path: pick the session's locked target if any, else the chosen
    # one stashed in the cookie, else fall back to the default target.
    if session.get('target_id') is None:
        default_row = get_default_target(user_id)
        if default_row is not None:
            session['target_id'] = default_row['rowid']
    target_config = target_to_config(get_target(session.get('target_id'), user_id))
    if target_config is not None:
        target_config['zone_radii_mm'] = _zone_radii_for_target(
            session.get('target_id'), user_id)
        target_config['default_shaft_diameter_mm'] = DEFAULT_SHAFT_DIAMETER_MM
    # On GET (reload mid-session), repopulate quiver_size from the locked
    # value so the field doesn't go blank — and compute the same lock flag
    # the POST path emits, so the input stays readonly mid-quiver across
    # reloads.
    get_current_qs = session.get('current_quiver_size') or 0
    get_arrows_remaining = session.get('arrows_remaining', 0)
    get_quiver_size_display = str(get_current_qs) if get_current_qs > 0 else ''
    get_quiver_size_locked = (
        get_current_qs > 0
        and 0 < get_arrows_remaining < get_current_qs
    )
    # Repopulate the in-progress quiver's markers on GET so a mid-session
    # navigation (e.g. a Recall arrow that GET-redirects back to /sesh)
    # restores the past-shot dots instead of clearing the target.
    get_past_shots_list = get_past_shots(
        session['session_id'], get_current_qs, get_arrows_remaining, user_id
    ) if get_current_qs > 0 else []
    # Restore the bow/arrow picked earlier in this session so a mid-session
    # GET reload (Recall arrow redirect, PWA/browser refresh, SW re-fetch)
    # doesn't snap the dropdowns back to their first option.
    get_bow, get_arrow_type = _last_session_bow_arrow(
        user_id, session['session_id'])
    return render_template('session.html',
                           session_id=session['session_id'],
                           arrow_type=get_arrow_type,
                           quiver_size=get_quiver_size_display,
                           quiver_size_locked=get_quiver_size_locked,
                           quivers_completed=session.get('quivers_completed', 0),
                           arrows_remaining=get_arrows_remaining,
                           bow=get_bow,
                           bows=bow_models,
                           bowstyle_settings=BOWSTYLE_SETTINGS,
                           style_settings_by_bow=_style_settings_by_bow(user_id),
                           session_notes='',
                           session_tags=_last_session_tags(user_id, session['session_id']),
                           tag_suggestions=_distinct_user_tags(user_id),
                           distance='',
                           x_coord='',
                           y_coord='',
                           arrow_types=arrow_types,
                           arrow_shaft_diameters=_arrow_shaft_diameters_for_user(user_id),
                           past_shots=get_past_shots_list,
                           is_precise=0,
                           record_mode=record_mode,
                           targets_list=targets_list,
                           selected_target_id=session.get('target_id'),
                           target_locked=target_locked,
                           is_tournament=bool(session.get('tournament_round_key')
                                              or session.get('match_id')),
                           current_quiver_size=get_current_qs,
                           target_config=target_config)


@app.route('/api/record_shot', methods=['POST'])
@login_required
def api_record_shot():
    """Record one live shot via AJAX and return the new session state as JSON.

    The /sesh screen submits each shot here with ``fetch`` instead of doing a
    full-page navigation POST to /sesh. That removes the per-shot 58 KB
    re-render and — crucially — lets a dropped response be *retried* in the
    client rather than stranding the user on a 502 (the iPad keep-alive race
    that motivated this). Retries are made safe by a per-shot ``client_uuid``:
    a duplicate insert is folded into a success via _apply_shot's idempotency.

    Mirrors the /sesh POST ``arrow_shot`` branch field-for-field (target
    resolution, the quiver lock state machine, the shared _apply_shot
    mutation) so a shot saved here is byte-for-byte equivalent to one saved
    full-page. Practice sessions only — tournaments have their own route.

    CSRF: protected by CSRFProtect; the client sends the page's csrf_token via
    the X-CSRFToken header (and in the form body), exactly like /recall_arrow.
    """
    user_id = current_user_id()
    session_id = session.get('session_id')
    if not session_id:
        return jsonify(ok=False, msg='No active session.'), 400

    form = request.form

    def _form_int(name, default=0):
        try:
            return int(form.get(name, str(default)))
        except (TypeError, ValueError):
            return default

    bow                    = form.get('bow', '')
    arrow_type             = form.get('arrow_type', '')
    quiver_size            = form.get('quiver_size', '')
    distance               = form.get('distance', '')
    session_notes          = form.get('session_notes', '')
    session_tags           = _normalize_tags(form.get('session_tags', ''))
    style_settings_session = _collect_style_settings(form)
    x                      = form.get('x_coord', '')
    y                      = form.get('y_coord', '')
    is_precise             = _form_int('is_precise', 0)
    record_mode            = _form_int('record_mode', 0)
    client_uuid            = (form.get('client_uuid', '') or '').strip()[:64] or None

    # Has this session already saved any shots? If so, the target is locked.
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        cnt = cur.execute(
            "SELECT COUNT(*) FROM apollo WHERE session_id = %s AND user_id = %s",
            (session_id, user_id)
        ).fetchone()
        session_shot_count = cnt[0] if cnt else 0
    target_locked = session_shot_count > 0

    # Resolve the target exactly as /sesh does: accept the posted value only on
    # the first shot, otherwise stick with the session's locked target.
    if target_locked:
        if session.get('target_id') is None:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                existing = cur.execute(
                    "SELECT target_id FROM apollo WHERE session_id = %s AND user_id = %s LIMIT 1",
                    (session_id, user_id)
                ).fetchone()
            if existing and existing[0] is not None:
                session['target_id'] = existing[0]
        target_id = session.get('target_id')
    else:
        posted = form.get('target_id', '').strip()
        try:
            target_id = int(posted) if posted else session.get('target_id')
        except ValueError:
            target_id = session.get('target_id')
        # Only accept a target_id the current user actually owns.
        if target_id is not None and get_target(target_id, user_id) is None:
            target_id = None
        if target_id is None:
            default_row = get_default_target(user_id)
            target_id = default_row['rowid'] if default_row is not None else None
        if target_id is not None:
            session['target_id'] = target_id

    session['record_mode'] = record_mode

    try:
        quiver_size_int = int(quiver_size) if quiver_size else 0
    except (TypeError, ValueError):
        quiver_size_int = 0
    quivers_completed   = session.get('quivers_completed', 0)
    arrows_remaining    = session.get('arrows_remaining', 0)
    current_quiver_size = session.get('current_quiver_size') or 0
    mid_quiver = 0 < arrows_remaining < current_quiver_size

    try:
        (effective_quiver_size, arrows_remaining, quivers_completed,
         past_shots) = _apply_shot(
            user_id=user_id, session_id=session_id, bow=bow,
            arrow_type=arrow_type, distance=distance,
            session_notes=session_notes, session_tags=session_tags,
            style_settings_session=style_settings_session,
            x=x, y=y, is_precise=is_precise, record_mode=record_mode,
            target_id=target_id, quiver_size_int=quiver_size_int,
            mid_quiver=mid_quiver, current_quiver_size=current_quiver_size,
            arrows_remaining=arrows_remaining,
            quivers_completed=quivers_completed, client_uuid=client_uuid)
    except _ShotRejected as e:
        return jsonify(ok=False, msg=e.message), e.code
    except SQLAlchemyError as e:
        print(f"❌ Database error (record_shot): {e}")
        return jsonify(ok=False, msg='Error saving entry'), 500

    session['current_quiver_size'] = effective_quiver_size
    session['arrows_remaining']    = arrows_remaining
    session['quivers_completed']   = quivers_completed
    print(f"✅ Entry saved (ajax) for session {session_id}")

    quiver_size_locked = (
        effective_quiver_size > 0
        and 0 < arrows_remaining < effective_quiver_size
    )
    display_quiver_size = (str(effective_quiver_size)
                           if effective_quiver_size > 0 else quiver_size)

    return jsonify(
        ok=True,
        arrows_remaining=arrows_remaining,
        quivers_completed=quivers_completed,
        quiver_size_locked=quiver_size_locked,
        display_quiver_size=display_quiver_size,
        current_quiver_size=session.get('current_quiver_size') or 0,
        target_locked=True,
        target_id=session.get('target_id'),
        past_shots=past_shots,
    )


@app.route('/api/target_config/<int:target_id>', methods=['GET'])
@login_required
def api_target_config(target_id):
    """Return the target config JSON for a given target_id.

    Used by /sesh's target dropdown to hot-swap the canvas image and
    physical dimensions when the user picks a different target without
    posting the form. Scoped to the current user — a request for a
    target id that doesn't belong to the caller returns 404 rather than
    leaking another user's target metadata.
    """
    user_id = current_user_id()
    cfg = target_to_config(get_target(target_id, user_id))
    if cfg is None:
        return jsonify(ok=False), 404
    cfg['zone_radii_mm'] = _zone_radii_for_target(target_id, user_id)
    cfg['default_shaft_diameter_mm'] = DEFAULT_SHAFT_DIAMETER_MM
    cfg['target_image_url'] = url_for('static',
                                      filename=cfg['target_image'])
    return jsonify(ok=True, config=cfg)


@app.route('/recall_arrow', methods=['POST'])
@login_required
def recall_arrow():
    """Undo the most recently submitted shot in the active session.

    Deletes the latest apollo row for this user's current session and
    rewinds the in-memory counters (arrows_remaining, quivers_completed,
    current_quiver_size, and the tournament segment index when applicable)
    to the state they held before that shot was saved.

    The row's stored ``arrows_remaining`` is the *pre-decrement* value,
    so restoring the cookie to that number undoes the per-shot decrement;
    if it was 1, that shot also closed a quiver so quivers_completed
    drops by one as well. Used by the "Recall arrow" button on /sesh
    and /tournament for accidental submissions.
    """
    user_id = current_user_id()
    session_id = session.get('session_id')
    if session_id is None:
        return jsonify(ok=False, msg='No active session.'), 400

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT id, quiver_size, arrows_remaining FROM apollo "
                "WHERE session_id = %s AND user_id = %s "
                "ORDER BY id DESC LIMIT 1",
                (session_id, user_id)
            ).fetchone()
            if not row:
                return jsonify(ok=False, msg='No arrow to recall.'), 400
            row_id = row['id']            if 'id'            in row else row[0]
            row_qs = row['quiver_size']   if 'quiver_size'   in row else row[1]
            row_ar = row['arrows_remaining'] if 'arrows_remaining' in row else row[2]
            cur.execute(
                "DELETE FROM apollo WHERE id = %s AND user_id = %s",
                (row_id, user_id)
            )
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Recall arrow error: {e}")
        return jsonify(ok=False, msg='Database error.'), 500

    try:
        row_qs_int = int(row_qs) if row_qs is not None else 0
    except (TypeError, ValueError):
        row_qs_int = 0
    try:
        row_ar_int = int(row_ar) if row_ar is not None else 0
    except (TypeError, ValueError):
        row_ar_int = 0

    quivers_completed = int(session.get('quivers_completed', 0) or 0)
    if row_ar_int == 1 and quivers_completed > 0:
        quivers_completed -= 1
    session['quivers_completed'] = quivers_completed
    session['arrows_remaining']  = row_ar_int
    if row_qs_int > 0:
        session['current_quiver_size'] = row_qs_int

    # Tournament-only: a recalled shot that previously crossed a segment
    # boundary should roll the segment index back too, so the next shot's
    # distance matches what it had when the recalled shot was fired.
    # Derive the correct segment from the new shot count after the recall.
    round_key = session.get('tournament_round_key')
    round_def = TOURNAMENT_ROUNDS.get(round_key) if round_key else None
    if round_def is not None:
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                _row = cur.execute(
                    "SELECT COUNT(*) FROM apollo "
                    "WHERE session_id = %s AND user_id = %s",
                    (session.get('session_id') or 0, current_user_id())
                ).fetchone()
                shots_after_recall = int(_row[0]) if _row and _row[0] is not None else 0
        except SQLAlchemyError:
            shots_after_recall = 0
        segments = _round_segments(round_def)
        total = _round_total_arrows(round_def)
        if total > 0:
            pointer = min(shots_after_recall, total - 1)
            new_idx, _seg, _ = _segment_for_shot(segments, pointer)
            session['tournament_segment_idx'] = new_idx

    return jsonify(ok=True)


@app.route('/api/sync_shots', methods=['POST'])
@login_required
def sync_shots():
    """Batch-insert shots captured offline by the PWA shot queue.

    The client (static/apollo-offline.js) records shots into an IndexedDB
    queue while ``navigator.onLine`` is false, tracking the quiver state
    machine locally. When the connection returns it POSTs the whole queue
    here as JSON ``{"shots": [ {...}, ... ]}`` and, on a 2xx + ``ok``,
    clears the synced rows.

    CSRF: Flask-WTF's CSRFProtect honours the ``X-CSRFToken`` header, which
    the client sends with the page's csrf_token — so this endpoint stays
    protected with no exemption.

    Each shot carries the values the live /sesh POST would have computed
    client-side: coords, equipment selections, the *pre-decrement*
    ``arrows_remaining`` and locked ``quiver_size`` (so get_past_shots /
    recall stay correct), and a real on-device ``ts``. Equipment config is
    snapshotted server-side via the shared _insert_shot() helper, so a
    synced shot is byte-for-byte equivalent to one saved live — except its
    timestamp reflects when the arrow was actually shot, not sync time.

    Scope (v1): practice sessions only. The session must already exist
    (its session_times row was created by the online GET /sesh that
    started the round); an unknown session_id rejects the whole batch so
    the client keeps the queue rather than silently dropping shots.
    """
    user_id = current_user_id()
    payload = request.get_json(silent=True) or {}
    shots = payload.get('shots')
    if not isinstance(shots, list):
        return jsonify(ok=False, msg='Expected {"shots": [...]}.'), 400
    if not shots:
        return jsonify(ok=True, inserted=0)
    # Defensive cap: a legitimate range trip is hundreds of arrows, not
    # tens of thousands. Anything larger is almost certainly a bug or abuse.
    if len(shots) > 5000:
        return jsonify(ok=False, msg='Too many shots in one batch.'), 413

    def _client_ts(raw):
        # Parse the client's ISO shot time into a naive-UTC datetime that
        # matches the shape _app_now() writes. Fall back to now() on any
        # malformed value so a bad clock never strands the queue.
        if not raw:
            return _app_now()
        try:
            dt = datetime.fromisoformat(str(raw).strip().replace('Z', '+00:00'))
        except (ValueError, TypeError):
            return _app_now()
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt.replace(microsecond=0)

    # Validate every referenced session belongs to this user *before*
    # inserting anything — reject the batch as a unit so a half-synced
    # queue can't happen.
    try:
        wanted = {int(s.get('session_id')) for s in shots
                  if str(s.get('session_id', '')).strip() != ''}
    except (TypeError, ValueError):
        return jsonify(ok=False, msg='Bad session_id in batch.'), 400
    if not wanted:
        return jsonify(ok=False, msg='No session_id in batch.'), 400
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT session_id FROM session_times WHERE user_id = %s",
            (user_id,)
        ).fetchall()
        owned = {int(r[0]) for r in rows} if rows else set()
    missing = wanted - owned
    if missing:
        return jsonify(ok=False,
                       msg='Unknown session(s) for this user; not synced.',
                       missing=sorted(missing)), 409

    # Cache target-ownership lookups across the batch (a session is locked
    # to one target, so this is usually a single id).
    target_ok = {}

    def _coerce_target(raw):
        if raw is None or str(raw).strip() == '':
            return None
        try:
            tid = int(raw)
        except (TypeError, ValueError):
            return None
        if tid not in target_ok:
            target_ok[tid] = get_target(tid, user_id) is not None
        return tid if target_ok[tid] else None

    inserted = 0
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            for s in shots:
                session_id = int(s.get('session_id'))
                try:
                    quiver_size = int(s.get('quiver_size'))
                except (TypeError, ValueError):
                    quiver_size = 0
                # Mirror the live path's guard: a non-positive quiver size
                # would corrupt the counters, so skip rather than insert.
                if quiver_size <= 0:
                    continue
                try:
                    arrows_remaining = int(s.get('arrows_remaining'))
                except (TypeError, ValueError):
                    arrows_remaining = quiver_size

                _insert_shot(
                    cur,
                    user_id=user_id, session_id=session_id,
                    timestamp=_client_ts(s.get('ts')),
                    bow=s.get('bow', '') or '',
                    arrow_type=s.get('arrow_type', '') or '',
                    quiver_size=quiver_size,
                    arrows_remaining=arrows_remaining,
                    distance=s.get('distance', '') or '',
                    session_notes=s.get('session_notes', '') or '',
                    x=s.get('x_coord', ''), y=s.get('y_coord', ''),
                    is_precise=int(s.get('is_precise', 1) or 0),
                    record_mode=int(s.get('record_mode', 1) or 0),
                    target_id=_coerce_target(s.get('target_id')),
                    session_tags=_normalize_tags(s.get('session_tags', '') or ''),
                    # Offline clients may attach the bowtype-specific settings
                    # as a dict; validated/merged server-side in _insert_shot.
                    style_settings_session=(s.get('style_settings')
                        if isinstance(s.get('style_settings'), dict) else None),
                )
                inserted += 1
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ sync_shots error: {e}")
        return jsonify(ok=False, msg='Database error during sync.'), 500

    # Reconcile the live cookie counters. The /sesh POST tracks quiver
    # state (arrows_remaining / quivers_completed / current_quiver_size) in
    # the Flask session cookie, *not* the posted form fields — so after
    # offline shots land, the cookie is frozen at its pre-offline value. If
    # the user then takes a shot online without this fix, the server would
    # re-use a stale arrows_remaining and write a duplicate counter. The
    # client (which ran the same state machine offline) reports its final
    # counters in ``active``; adopt them, but only when they belong to the
    # session that's still active in this cookie — a synced batch for an
    # older, since-replaced session must not clobber the new one's state.
    reconciled = False
    active = payload.get('active') or {}
    try:
        if active and int(active.get('session_id')) == int(session.get('session_id') or 0):
            session['arrows_remaining'] = int(active.get('arrows_remaining'))
            session['quivers_completed'] = int(active.get('quivers_completed'))
            cqs = active.get('current_quiver_size')
            session['current_quiver_size'] = int(cqs) if cqs else None
            reconciled = True
    except (TypeError, ValueError):
        pass

    print(f"✅ Offline sync: inserted {inserted} shot(s) for user {user_id}")
    return jsonify(ok=True, inserted=inserted, reconciled=reconciled)


# ── Tournament mode routes ──────────────────────────────────────────────
# A tournament session is a /sesh session with an extra `session_tags`
# marker (`tournament:<round_key>`) so post-hoc code can recover which
# round it was. The /tournament route enforces the round's arrows-per-end
# and target lock; everything else flows through the existing shot-save
# and end-session machinery so analytics, replay, and exports keep
# working without changes.

@app.route('/tournament', methods=['GET', 'POST'])
@login_required
def tournament():
    """Tournament mode. GET shows either the round-selector or the
    in-progress shot UI; POST records a shot the same way /sesh does
    but with the round's arrows-per-end and target locked.

    Round identity lives in `session_tags` (tag `tournament:<key>`),
    so the route recovers state across reloads by reading the latest
    apollo row for the active session. The Flask cookie carries a
    cached round_key for the GET path to avoid a DB hit when the
    session is mid-round.
    """
    user_id = current_user_id()

    # ── GET path: figure out whether a tournament is in progress ───────
    # An "in-progress tournament" means: there's a session_id in the
    # cookie, AND its latest shot row carries a tournament:* tag. The
    # session_tags lookup also lets us survive a cookie wipe — if the
    # cookie's tournament_round_key was lost but session_id remains
    # (e.g. the user cleared local storage), we can still recover it
    # from the DB.
    # The active round comes from either:
    #   - the Flask cookie's cached `tournament_round_key` (set on
    #     /tournament/start, holds before any shot is recorded), or
    #   - the latest apollo row's session_tags (authoritative once any
    #     shot exists; survives a cookie wipe).
    # If session_id is set but neither source identifies a tournament,
    # the user is in a regular /sesh — fall through to the selector and
    # let them either resume that session via /sesh or end it first.
    active_round_key = session.get('tournament_round_key')
    if not active_round_key and session.get('session_id') is not None:
        latest_tags = _last_session_tags(user_id, session['session_id'])
        active_round_key = _round_key_from_tags(latest_tags)
    if active_round_key and active_round_key not in TOURNAMENT_ROUNDS:
        # Stale key from an older deploy — clear and bounce to selector.
        session.pop('tournament_round_key', None)
        active_round_key = None
    if active_round_key:
        session['tournament_round_key'] = active_round_key

    if request.method == 'GET' and active_round_key is None:
        # Round selector page. Group by org for the UI.
        by_org = {}
        for key, rd in TOURNAMENT_ROUNDS.items():
            is_course = bool(rd.get('course'))
            is_multi  = bool(rd.get('segments'))
            by_org.setdefault(rd['org'], []).append({
                'key':              key,
                'name':             rd['name'],
                'description':      rd['description'],
                'distance_m':       rd['distance_m'],
                'arrows_per_end':   rd['arrows_per_end'],
                'ends':             rd['ends'],
                'total_arrows':     rd['total_arrows'],
                'max_score':        rd['max_score'],
                'equipment_class':  rd['equipment_class'],
                'is_course':        is_course,
                'is_multi_segment': is_multi,
            })
        return render_template(
            'tournament.html',
            view='selector',
            rounds_by_org=by_org,
        )

    # From here on we expect a round_key. POSTs use the cookie's cached
    # key (mid-round shot submit). If that's also missing on POST,
    # something's wrong with the cookie — bounce to the selector.
    round_key = active_round_key or session.get('tournament_round_key')
    if not round_key:
        return redirect(url_for('tournament'))
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        # Stale key from an older deploy — clear and bounce.
        session.pop('tournament_round_key', None)
        return redirect(url_for('tournament'))

    # Resolve the *active* segment — for single-segment rounds this is
    # always segment 0 with the round's top-level face/distance/end
    # size; for multi-segment and course rounds the segment that
    # contains the next-arrow-to-shoot drives the active face, distance,
    # and end size.
    segments = _round_segments(round_def)
    round_total_arrows = _round_total_arrows(round_def)
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            _row = cur.execute(
                "SELECT COUNT(*) FROM apollo "
                "WHERE session_id = %s AND user_id = %s",
                (session.get('session_id') or 0, user_id)
            ).fetchone()
            shots_so_far_route = int(_row[0]) if _row and _row[0] is not None else 0
    except SQLAlchemyError:
        shots_so_far_route = 0
    # Clamp to the planned count so the "completed" state lands on the
    # last segment instead of pointing one past the end.
    shot_pointer = min(shots_so_far_route, round_total_arrows - 1) if round_total_arrows > 0 else 0
    active_seg_idx, active_seg, _arrows_into_seg = _segment_for_shot(
        segments, shot_pointer)
    session['tournament_segment_idx'] = active_seg_idx

    face_key = active_seg['face_key']
    face_def = TOURNAMENT_FACES[face_key]
    target_id = _tournament_face_target_id(user_id, face_key)
    if target_id is None:
        # Seeding failed earlier; the seeder already logged the traceback.
        # Surface the offending face_key so the user has something to
        # share when reporting.
        print(f"❌ Tournament face seed missing — user={user_id} face={face_key}")
        return (f"Could not seed tournament target face "
                f"(face_key={face_key}). Check server logs for the "
                f"underlying database error."), 500
    session['target_id'] = target_id

    arrows_per_end = int(active_seg['arrows_per_end'])

    # Lazy session_id mint (same retry-with-backoff trick as /sesh) when
    # the user is starting a fresh round. Stamps the round tag onto the
    # first shot row written below.
    if not session.get('session_id'):
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            new_session_id = _mint_session_id(con, cur, user_id)
        if new_session_id is None:
            return "Could not allocate session id — please try again", 500
        session['session_id'] = new_session_id
        session['quivers_completed'] = 0
        session['arrows_remaining'] = 0
        session['current_quiver_size'] = arrows_per_end
        session['record_mode'] = 0

    session_id = session['session_id']

    # ── Live match-play state ──────────────────────────────────────────
    # When `match_id` is in the cookie we're running 2–4 archers on this
    # device. `match_archers` holds the roster in seat order; `match_idx`
    # points at the archer whose turn it is — their session_id is mirrored
    # into session['session_id'] above, so everything below renders for
    # the active archer. We only diverge to (a) tag shots with the match,
    # (b) swap the active archer at end boundaries, and (c) build a
    # combined scoreboard for the GET render.
    match_id      = session.get('match_id')
    match_archers = session.get('match_archers') or []
    match_idx     = int(session.get('match_idx', 0) or 0)
    if match_id and match_archers:
        if match_idx >= len(match_archers):
            match_idx = 0
        active_archer = match_archers[match_idx]
        # Keep session_id locked to the active archer (defensive against a
        # cookie that drifted between the two keys).
        session_id = active_archer.get('sid', session_id)
        session['session_id'] = session_id
    else:
        match_id = None
        active_archer = None

    # All tournament rounds in Apollo are practice by definition — real
    # competitions don't use a phone-based score logger. The historical
    # `practice` tag is retained on past sessions, but new shots no
    # longer get tagged with it.
    tournament_tag = _tournament_tag_for_round(round_key)
    # Per-round bowstyle override (set at round start) — carried on each shot
    # so the classification category reflects what was actually shot, even if
    # it differs from the archer's profile default.
    _bs_override = session.get('tournament_bowstyle')
    if _bs_override:
        tournament_tag += f', bowstyle:{_bs_override}'
    if match_id and active_archer:
        tournament_tag += f', match:{match_id}, participant:{active_archer.get("name","")}'
        if active_archer.get('email'):
            tournament_tag += f', email:{active_archer["email"]}'
        tournament_tag += f', seat:{active_archer.get("seat", match_idx)}'

    # ── POST: record a shot ─────────────────────────────────────────────
    if request.method == 'POST':
        if 'arrow_shot' not in request.form:
            return redirect(url_for('tournament'))
        x = request.form.get("x_coord", "")
        y = request.form.get("y_coord", "")
        is_precise = 0
        try:
            is_precise = int(request.form.get('is_precise', '0'))
        except (TypeError, ValueError):
            is_precise = 0
        bow         = request.form.get('bow', '')
        arrow_type  = request.form.get('arrow_type', '')
        record_mode = 0
        try:
            record_mode = int(request.form.get('record_mode', '0'))
        except (TypeError, ValueError):
            record_mode = 0
        session['record_mode'] = record_mode

        # Distance comes from the active segment. The user can't
        # override it — locking it keeps the recorded shot data
        # consistent with the published round structure.
        distance_m = active_seg['distance_m']
        distance = f"{distance_m}"

        # Refuse further shots once the round is complete. The template
        # hides the form when complete, but a stale POST shouldn't add
        # a 73rd arrow to a 72-arrow round.
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            shot_row = cur.execute(
                "SELECT COUNT(*) FROM apollo "
                "WHERE session_id = %s AND user_id = %s",
                (session_id, user_id)
            ).fetchone()
            shots_so_far = int(shot_row[0]) if shot_row and shot_row[0] is not None else 0
        if shots_so_far >= round_total_arrows:
            return redirect(url_for('tournament'))
        if x == '' or y == '':
            return redirect(url_for('tournament'))
        # In a live match, refuse further shots once the match is already
        # decided — set play (12.1.4.1) ends the moment an archer reaches
        # 6 set points, which can happen before all ends are shot.
        if match_id and len(match_archers) > 1:
            decided_check = [{
                'name':     a.get('name', ''),
                'seat':     a.get('seat'),
                'progress': _compute_tournament_progress(a.get('sid'), user_id, round_def),
            } for a in match_archers]
            if _match_set_scoring(decided_check, round_def).get('decided'):
                return redirect(url_for('tournament'))

        # Recover quiver bookkeeping (same logic as /sesh, but the size
        # is fixed at arrows_per_end — no user-supplied value to lock).
        arrows_remaining = int(session.get('arrows_remaining', 0) or 0)
        quivers_completed = int(session.get('quivers_completed', 0) or 0)
        if arrows_remaining <= 0:
            arrows_remaining = arrows_per_end
        session['current_quiver_size'] = arrows_per_end

        # Equipment-row snapshots go through _insert_shot — the single
        # source of truth the live /sesh path also uses, so a future
        # column add can't drift between the two insert sites. The
        # tournament specifics: quiver_size is fixed at arrows_per_end,
        # there are no session notes, and the tag carries the round/match
        # context instead of user tags.
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                _insert_shot(
                    cur,
                    user_id=user_id, session_id=session_id,
                    timestamp=_app_now(), bow=bow, arrow_type=arrow_type,
                    quiver_size=arrows_per_end, arrows_remaining=arrows_remaining,
                    distance=distance, session_notes='', x=x, y=y,
                    is_precise=is_precise, record_mode=record_mode,
                    target_id=target_id, session_tags=tournament_tag,
                    style_settings_session=_collect_style_settings(request.form),
                )
                con.commit()
        except SQLAlchemyError as e:
            print(f"❌ Tournament shot save error: {e}")
            return "Error saving entry", 500

        arrows_remaining -= 1
        end_completed = False
        if arrows_remaining <= 0:
            end_completed = True
            quivers_completed += 1
            # Refill arrows_remaining to the *next* segment's end size,
            # which differs from the current one on rounds where the
            # end size changes mid-round (e.g. WA 1440 drops from 6 to
            # 3 at the 50m boundary, NFAA Field cycles through targets).
            next_shot_idx = shots_so_far + 1
            if next_shot_idx < round_total_arrows:
                next_seg_idx, next_seg, _ = _segment_for_shot(segments, next_shot_idx)
                arrows_remaining = int(next_seg['arrows_per_end'])
                if next_seg_idx != active_seg_idx:
                    session['tournament_segment_idx'] = next_seg_idx
            else:
                arrows_remaining = arrows_per_end
        session['arrows_remaining'] = arrows_remaining
        session['quivers_completed'] = quivers_completed

        # ── Live match: hand the canvas to the next archer ─────────────
        # An end just finished; AB-CD detail shooting means the next
        # archer (who still has arrows left in the round) steps to the
        # line. Cycle forward from the current seat; if every archer has
        # finished the whole round, leave state put so the GET render
        # shows the winner banner. A switched-in archer always begins a
        # fresh end, so arrows_remaining=0 (route refills on their next
        # shot) is correct.
        if end_completed and match_id and len(match_archers) > 1:
            try:
                with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                    counts = {}
                    for a in match_archers:
                        row = cur.execute(
                            "SELECT COUNT(*) FROM apollo "
                            "WHERE session_id = %s AND user_id = %s",
                            (a['sid'], user_id)
                        ).fetchone()
                        counts[a['sid']] = int(row[0]) if row and row[0] is not None else 0
            except SQLAlchemyError:
                counts = {}
            n = len(match_archers)
            next_idx = None
            for step in range(1, n + 1):
                cand = (match_idx + step) % n
                if counts.get(match_archers[cand]['sid'], 0) < round_total_arrows:
                    next_idx = cand
                    break
            if next_idx is not None:
                session['match_idx'] = next_idx
                session['session_id'] = match_archers[next_idx]['sid']
                session['arrows_remaining'] = 0
                session['quivers_completed'] = 0
                session['tournament_segment_idx'] = 0

        return redirect(url_for('tournament'))

    # ── GET path with active round: render the shot UI ─────────────────
    # Pull the user's bows / arrows (same as /sesh) so the right rail's
    # equipment selectors stay populated.
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        arrow_types = [r[0] for r in cur.execute(
            "SELECT DISTINCT arrow FROM arrows WHERE user_id = %s",
            (user_id,)).fetchall() or []]
        bow_models = [r[0] for r in cur.execute(
            "SELECT DISTINCT bow_model FROM bows WHERE user_id = %s",
            (user_id,)).fetchall() or []]

    progress = _compute_tournament_progress(session_id, user_id, round_def)

    # ── Live match scoreboard ──────────────────────────────────────────
    # One progress per archer (reuse the active archer's, already
    # computed), scored via _match_set_scoring for set points / winner.
    match_ctx = None
    if match_id and match_archers:
        archers_for_scoring = []
        for i, a in enumerate(match_archers):
            a_prog = (progress if a.get('sid') == session_id
                      else _compute_tournament_progress(a.get('sid'), user_id, round_def))
            archers_for_scoring.append({
                'name':      a.get('name', ''),
                'seat':      a.get('seat', i),
                'is_active': (i == match_idx),
                'progress':  a_prog,
            })
        match_ctx = {
            'score':       _match_set_scoring(archers_for_scoring, round_def),
            'idx':         match_idx,
            'count':       len(match_archers),
            'active_name': active_archer.get('name', '') if active_archer else '',
            'match_id':    match_id,
        }

    # Surface the active segment's distance / index to the template
    # (already resolved above from the shot count).
    current_distance_m = active_seg['distance_m']
    current_segment_idx = active_seg_idx

    target_config = target_to_config(get_target(target_id, user_id))
    if target_config is not None:
        target_config['zone_radii_mm'] = _zone_radii_for_target(target_id, user_id)
        target_config['default_shaft_diameter_mm'] = DEFAULT_SHAFT_DIAMETER_MM

    # At session start (and between completed ends) the cookie stores
    # arrows_remaining=0 — the POST handler treats that as "refill to
    # arrows_per_end" before saving the next shot. Mirror that here so
    # the right-rail counter shows the upcoming end's size instead of 0.
    arrows_remaining_display = (
        int(session.get('arrows_remaining') or 0) or arrows_per_end
    )
    past_shots = get_past_shots(session_id, arrows_per_end,
                                arrows_remaining_display, user_id)

    is_course_round = bool(round_def.get('course'))
    is_practice = False  # see note above tournament_tag — kept for template compat
    # Maximum arrows-per-end across the round — drives the scorecard
    # column count. For single-segment rounds this equals the round's
    # arrows_per_end; for mixed rounds (1440, Field) it's the largest
    # end-size used in any segment.
    max_arrows_per_end = max(int(s['arrows_per_end']) for s in segments)

    # In match mode the cookie's quiver counters belong to whichever
    # archer was active when they were written — derive the active
    # archer's end count from their own progress instead so the right-
    # rail counter and shot-clock end index follow the canvas handoff.
    if match_id:
        _completed_ends = sum(1 for e in (progress.get('ends') or [])
                              if not e.get('partial'))
        quivers_completed_display = _completed_ends
        current_end_index = _completed_ends + 1
    else:
        quivers_completed_display = session.get('quivers_completed', 0)
        current_end_index = int(session.get('quivers_completed', 0) or 0) + 1

    # Once the round is complete, compute the AGB handicap and any
    # classification awards for the single-archer score. Skipped for live
    # match play (short set/cumulative formats aren't handicap rounds).
    handicap_info = None
    if progress.get('is_complete') and not match_id:
        handicap_info = _session_handicap_awards(
            round_key, progress.get('total_score'),
            _last_session_tags(user_id, session_id))

    return render_template(
        'tournament.html',
        view='shoot',
        session_id=session_id,
        round_key=round_key,
        round_def=round_def,
        handicap_info=handicap_info,
        face_def=face_def,
        face_render=_tournament_face_render_payload(face_key),
        target_config=target_config,
        past_shots=past_shots,
        arrow_shaft_diameters=_arrow_shaft_diameters_for_user(user_id),
        arrow_types=arrow_types,
        bow_models=bow_models,
        bowstyle_settings=BOWSTYLE_SETTINGS,
        style_settings_by_bow=_style_settings_by_bow(user_id),
        arrows_per_end=arrows_per_end,
        max_arrows_per_end=max_arrows_per_end,
        arrows_remaining=arrows_remaining_display,
        quivers_completed=quivers_completed_display,
        progress=progress,
        current_distance_m=current_distance_m,
        current_segment_idx=current_segment_idx,
        is_course_round=is_course_round,
        segment_count=len(segments),
        is_practice=is_practice,
        match=match_ctx,
        # End index used by the shot-clock to detect when a new end has
        # started — JS resets the countdown when this number changes.
        # 1-based so the on-screen label matches the scorecard rows.
        current_end_index=current_end_index,
    )


@app.route('/tournament/start', methods=['POST'])
@login_required
def tournament_start():
    """Begin a new tournament round.

    Hard-requires that the user not have an in-progress regular or
    tournament session (it would step on the session_id cookie),
    matching the constraint /sesh implicitly imposes. The user can
    end the existing session via the modal flow if needed.
    """
    user_id = current_user_id()
    round_key = (request.form.get('round_key') or '').strip()
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown tournament round", 400

    if session.get('session_id'):
        # An in-progress session is in the cookie. Reuse it only if it
        # has zero shots (the user clicked something and bounced back);
        # otherwise refuse and route them through end-session.
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                existing = cur.execute(
                    "SELECT COUNT(*) FROM apollo "
                    "WHERE session_id = %s AND user_id = %s",
                    (session['session_id'], user_id)
                ).fetchone()
                shots = int(existing[0]) if existing and existing[0] is not None else 0
        except SQLAlchemyError:
            shots = 0
        if shots > 0:
            return ("A session is already in progress. End it first via "
                    "<a href='/end_session'>End session</a>.", 409)
        # Reuse the empty session row — clear per-round keys to start
        # fresh.
        for k in ('quivers_completed', 'arrows_remaining',
                  'current_quiver_size', 'record_mode',
                  'tournament_segment_idx'):
            session.pop(k, None)

    _seed_tournament_faces(user_id)
    session['tournament_round_key'] = round_key
    session['tournament_segment_idx'] = 0
    session['target_id'] = _tournament_face_target_id(user_id, round_def['face_key'])
    # Bowstyle for AGB classification. A round that names a bowstyle (e.g.
    # "WA 720 (Recurve)") fixes it — its face/distance only make sense for
    # that style. Only 'any' rounds honour the per-round picker, falling
    # back to the profile bowstyle when blank.
    forced_bs = _round_bowstyle(round_def)
    if forced_bs:
        session['tournament_bowstyle'] = forced_bs
    else:
        bs = (request.form.get('bowstyle') or '').strip().lower()
        if bs in classifications.data.AGB_BOWSTYLES:
            session['tournament_bowstyle'] = bs
        else:
            session.pop('tournament_bowstyle', None)
    return redirect(url_for('tournament'))


def _owner_name_email(user_id):
    """(username, email) for the device owner — used to pre-fill the
    first participant row in the score-sheet and match-setup forms."""
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT username, email FROM users WHERE id = %s",
                (user_id,)
            ).fetchone()
            if row is not None:
                name  = row['username'] if 'username' in row else row[0]
                email = row['email']    if 'email'    in row else row[1]
                return (name or ''), (email or '')
    except SQLAlchemyError:
        pass
    return '', ''


def _finalize_match_sessions(user_id):
    """Ending a live match: stamp session_end_time on every archer
    session (the end-session routes only finalize the cookie's active
    session). Reads the roster from the cookie; no-op outside match
    mode. Leaves an already-set end time untouched."""
    archers = session.get('match_archers') or []
    if not archers:
        return
    now = _app_now()
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            for a in archers:
                sid = a.get('sid')
                if sid is None:
                    continue
                cur.execute(
                    "UPDATE session_times SET session_end_time = %s "
                    "WHERE session_id = %s AND user_id = %s "
                    "AND session_end_time IS NULL",
                    (now, sid, user_id)
                )
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Match finalize error: {e}")


@app.route('/tournament/match/setup', methods=['POST'])
@login_required
def tournament_match_setup():
    """Render the live match-play setup form for a chosen round.

    Live match play runs 2–4 archers on this one device: they take
    turns at the canvas, the active archer switching after each
    completed end (AB-CD detail shooting). Each archer keeps a separate
    scorecard; set points / a winner are computed per the round's
    scoring system. This route just collects archer names + emails;
    /tournament/match/start does the work.
    """
    user_id = current_user_id()
    round_key = (request.form.get('round_key') or '').strip()
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown tournament round", 400
    _seed_tournament_faces(user_id)
    owner_name, owner_email = _owner_name_email(user_id)
    return render_template(
        'tournament.html',
        view='match_setup',
        round_key=round_key,
        round_def=round_def,
        owner_name=owner_name,
        owner_email=owner_email,
        max_archers=4,
    )


@app.route('/tournament/match/start', methods=['POST'])
@login_required
def tournament_match_start():
    """Begin a live multi-archer match.

    Mints one session_id per archer (all under the device owner), tags
    them with a shared match id, and stores the archer roster +
    active-archer pointer in the cookie. The active archer's session_id
    is mirrored into `session['session_id']` so the existing
    single-archer /tournament render path Just Works; the POST handler
    swaps it at each end boundary.
    """
    user_id = current_user_id()
    round_key = (request.form.get('round_key') or '').strip()
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown tournament round", 400

    # Refuse if a non-empty session is already in progress (same guard
    # as /tournament/start) — a live match owns the session_id cookie.
    if session.get('session_id'):
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                existing = cur.execute(
                    "SELECT COUNT(*) FROM apollo "
                    "WHERE session_id = %s AND user_id = %s",
                    (session['session_id'], user_id)
                ).fetchone()
                shots = int(existing[0]) if existing and existing[0] is not None else 0
        except SQLAlchemyError:
            shots = 0
        if shots > 0:
            return ("A session is already in progress. End it first via "
                    "<a href='/end_session'>End session</a>.", 409)

    participants = _parse_participants_from_form(request.form)
    if len(participants) < 2:
        return "A live match needs at least two archers.", 400
    if len(participants) > 4:
        participants = participants[:4]

    _seed_tournament_faces(user_id)
    match_id = _match_id_for(user_id, round_key)

    archers = []
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            for seat, p in enumerate(participants):
                sid = _mint_session_id(con, cur, user_id)
                if sid is None:
                    return "Could not allocate session id — please try again", 500
                archers.append({
                    'sid':   sid,
                    'name':  p['name'],
                    'email': p['email'],
                    'seat':  seat,
                })
    except SQLAlchemyError as e:
        print(f"❌ Match start error: {e}")
        return "Error starting match — please try again", 500

    # Clear any stale per-round state, then install the match roster.
    for k in ('quivers_completed', 'arrows_remaining', 'current_quiver_size',
              'record_mode', 'tournament_segment_idx'):
        session.pop(k, None)
    session['match_id']       = match_id
    session['match_archers']  = archers
    session['match_idx']      = 0
    session['session_id']     = archers[0]['sid']
    session['tournament_round_key'] = round_key
    session['tournament_segment_idx'] = 0
    session['arrows_remaining'] = 0
    session['quivers_completed'] = 0
    # A bowstyle-specific round (e.g. "WA 720 (Recurve)") fixes the AGB
    # classification category; clear any stale override otherwise.
    forced_bs = _round_bowstyle(round_def)
    if forced_bs:
        session['tournament_bowstyle'] = forced_bs
    else:
        session.pop('tournament_bowstyle', None)
    session['target_id'] = _tournament_face_target_id(user_id, round_def['face_key'])
    return redirect(url_for('tournament'))


def _scorecard_equipment_ctx(user_id):
    """Context for the scorecard equipment menu — the same bow/arrow/settings
    data the live shoot view feeds ``_bow_style_settings.html``. Lets a
    scorecard attribute its rows to a bow/arrow + bowtype settings, exactly
    like a plotted session, so analyze/handicap filters can see the gear.
    Scorecard rows are still kept out of the dispersion fit (see
    ``_predict_row_is_scorecard``)."""
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        arows = cur.execute(
            "SELECT DISTINCT arrow FROM arrows WHERE user_id = %s",
            (user_id,)).fetchall()
        brows = cur.execute(
            "SELECT DISTINCT bow_model FROM bows WHERE user_id = %s",
            (user_id,)).fetchall()
    return {
        'bow_models':           [r[0] for r in brows] if brows else [],
        'arrow_types':          [r[0] for r in arows] if arows else [],
        'bowstyle_settings':    BOWSTYLE_SETTINGS,
        'style_settings_by_bow': _style_settings_by_bow(user_id),
    }


@app.route('/tournament/score_sheet/setup', methods=['POST'])
@login_required
def tournament_score_sheet_setup():
    """Render the score-sheet form for a chosen round.

    Score-sheet mode is for logging a real-world competition after the
    fact. The user (or scorekeeper) adds 1+ participants, then types
    each arrow's ring value into a per-end / per-participant grid.
    Unlike live mode, no canvas placement is required.
    """
    user_id = current_user_id()
    round_key = (request.form.get('round_key') or '').strip()
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown tournament round", 400
    _seed_tournament_faces(user_id)

    # Pre-fill the first participant row with the device owner's
    # username + email so a solo logger doesn't have to retype.
    owner_name, owner_email = _owner_name_email(user_id)

    segments = _round_segments(round_def)
    # Per-end metadata for the template — labels, valid ring labels,
    # and end count per segment. The template uses this to render one
    # mini-scorecard per segment (since end size and face can change
    # mid-round, e.g. WA 1440 / Field rounds).
    sheet_segments = []
    points_by_segment = {}  # seg_idx → { label: point_value }
    end_counter = 0
    for seg_idx, seg in enumerate(segments):
        face_key = seg['face_key']
        face = TOURNAMENT_FACES.get(face_key, {})
        labels = _face_ring_labels(face_key)
        # Ring labels for the dropdown — X first, then physical ring
        # numbers in descending order (10/9/8/.../1), then M.
        choices = ['X'] + [str(l) for l in labels] + ['M']
        seg_ends = []
        for _ in range(int(seg['ends'])):
            end_counter += 1
            seg_ends.append({
                'end_number':     end_counter,
                'arrows_per_end': int(seg['arrows_per_end']),
            })
        sheet_segments.append({
            'segment_idx':    seg_idx,
            'face_key':       face_key,
            'face_name':      face.get('name', face_key),
            'distance_m':     float(seg['distance_m']),
            'arrows_per_end': int(seg['arrows_per_end']),
            'choices':        choices,
            'ends':           seg_ends,
        })
        # Build the label → point-value lookup. zones[0] is the X ring;
        # zones[1:] map to the labels list 1:1 (same order, innermost-
        # out). Use string keys so the JS lookup matches input.value.
        zones = list(face.get('zones') or ())
        seg_points = {'X': int(zones[0][0]) if zones else 0, 'M': 0}
        for i, label in enumerate(labels):
            if 1 + i < len(zones):
                seg_points[str(label)] = int(zones[1 + i][0])
        points_by_segment[seg_idx] = seg_points

    return render_template(
        'tournament.html',
        view='score_sheet',
        round_key=round_key,
        round_def=round_def,
        sheet_segments=sheet_segments,
        points_by_segment_json=json.dumps(points_by_segment),
        total_arrows=_round_total_arrows(round_def),
        owner_name=owner_name,
        owner_email=owner_email,
        today_iso=_app_now().strftime('%Y-%m-%d'),
        **_scorecard_equipment_ctx(user_id),
    )


def _mint_session_id(con, cur, user_id):
    """Allocate a fresh session_id for `user_id` and insert its
    session_times row, committing on success. Returns the new id, or
    None if all retries lose the race. Same retry-with-backoff trick
    used by /sesh, /tournament, and the score-sheet/match flows —
    factored here so they share one implementation. Caller supplies the
    open connection + cursor (the cursor stays usable for follow-up
    inserts in the same transaction scope)."""
    for _attempt in range(12):
        res = cur.execute(
            "SELECT MAX(session_id) FROM apollo WHERE user_id = %s",
            (user_id,)
        ).fetchone()
        max_apollo = res[0] if res and res[0] is not None else 0
        res = cur.execute(
            "SELECT MAX(session_id) FROM session_times WHERE user_id = %s",
            (user_id,)
        ).fetchone()
        max_st = res[0] if res and res[0] is not None else 0
        candidate = max(int(max_apollo or 0), int(max_st or 0)) + 1
        try:
            cur.execute(
                "INSERT INTO session_times "
                "(user_id, session_id, session_begin_time) "
                "VALUES (%s, %s, %s)",
                (user_id, candidate, _app_now())
            )
            con.commit()
            return candidate
        except DBIntegrityError:
            time.sleep(0.005 + secrets.randbelow(20) / 1000.0)
            continue
    return None


def _match_id_for(user_id, round_key):
    """Mint a short, opaque match id. Each score-sheet submission
    becomes one "match" — a group of participant sessions tagged with
    `match:<id>` so post-hoc result pages can find them again."""
    raw = f"{user_id}:{round_key}:{_app_now().isoformat()}:{secrets.token_hex(4)}"
    return hashlib.sha256(raw.encode()).hexdigest()[:16]


def _parse_participants_from_form(form):
    """Pull the `participant_<i>_name` and `participant_<i>_email`
    fields out of the score-sheet form, in order. Skips rows whose
    name field is empty. Returns a list of dicts."""
    out = []
    # Cap at 16 to fence off a malformed bulk POST.
    for i in range(16):
        name  = (form.get(f'participant_{i}_name')  or '').strip()
        email = (form.get(f'participant_{i}_email') or '').strip()
        if not name:
            continue
        out.append({'idx': i, 'name': name, 'email': email})
    return out


def _match_set_scoring(archers, round_def):
    """Score a live (or finished) match across `archers`.

    `archers` is a list of dicts in seat order, each carrying:
      {'name', 'seat', 'is_active'(opt), 'progress'} where `progress`
    is a _compute_tournament_progress() result.

    Returns a dict the templates render:
      {'mode': 'set'|'cumulative'|'leaderboard',
       'rows': [{'name','seat','total','x_count','set_points'(or None),
                 'is_active','is_complete'}, ...],   # seat order
       'decided': bool, 'winner_name': str|None, 'is_tie': bool,
       'summary': str}

    Set scoring (WA 12.1.4.1) is only defined head-to-head, so it
    applies when the round declares match_scoring=='set' AND there are
    exactly two archers; otherwise we fall back to a cumulative-total
    leaderboard. Compound / cumulative match rounds rank by total.
    """
    scoring = (round_def or {}).get('match_scoring')

    def _end_totals(p):
        return [e['end_total'] for e in (p.get('ends') or [])
                if not e.get('partial')]

    rows = []
    for a in archers:
        p = a.get('progress') or {}
        rows.append({
            'name':        a.get('name', ''),
            'seat':        a.get('seat'),
            'total':       int(p.get('total_score', 0) or 0),
            'x_count':     int(p.get('x_count', 0) or 0),
            'set_points':  None,
            'is_active':   bool(a.get('is_active')),
            'is_complete': bool(p.get('is_complete')),
            '_end_totals': _end_totals(p),
        })

    all_complete = bool(rows) and all(r['is_complete'] for r in rows)

    # ── Set system: two archers, recurve/barebow set play ──────────────
    if scoring == 'set' and len(rows) == 2:
        a, b = rows[0], rows[1]
        a['set_points'] = 0
        b['set_points'] = 0
        # Award per end both archers have completed (the trailing archer
        # may be one end behind mid-cycle — only score settled ends).
        settled_ends = min(len(a['_end_totals']), len(b['_end_totals']))
        for i in range(settled_ends):
            ta, tb = a['_end_totals'][i], b['_end_totals'][i]
            if ta > tb:
                a['set_points'] += 2
            elif tb > ta:
                b['set_points'] += 2
            else:
                a['set_points'] += 1
                b['set_points'] += 1
        decided = False
        winner_name = None
        is_tie = False
        # First to 6 set points wins (WA 12.1.4.1).
        if a['set_points'] >= 6 or b['set_points'] >= 6:
            decided = True
            winner_name = (a['name'] if a['set_points'] >= b['set_points']
                           else b['name'])
        elif all_complete:
            decided = True
            if a['set_points'] > b['set_points']:
                winner_name = a['name']
            elif b['set_points'] > a['set_points']:
                winner_name = b['name']
            else:
                is_tie = True  # 5–5 → single-arrow shoot-off
        if is_tie:
            summary = (f"Tied {a['set_points']}–{b['set_points']} on set "
                       f"points — shoot-off")
        elif decided:
            hi = max(a['set_points'], b['set_points'])
            lo = min(a['set_points'], b['set_points'])
            summary = f"{winner_name} wins {hi}–{lo} on set points"
        else:
            lead = (a if a['set_points'] >= b['set_points'] else b)
            trail = b if lead is a else a
            if lead['set_points'] == trail['set_points']:
                summary = f"Level {a['set_points']}–{b['set_points']} on set points"
            else:
                summary = (f"{lead['name']} leads "
                           f"{lead['set_points']}–{trail['set_points']} on set points")
        for r in rows:
            r.pop('_end_totals', None)
        return {'mode': 'set', 'rows': rows, 'decided': decided,
                'winner_name': winner_name, 'is_tie': is_tie,
                'summary': summary}

    # ── Cumulative / leaderboard: rank by total ────────────────────────
    mode = 'cumulative' if scoring == 'cumulative' else 'leaderboard'
    ranked = sorted(rows, key=lambda r: r['total'], reverse=True)
    decided = False
    winner_name = None
    is_tie = False
    summary = ''
    if all_complete and ranked:
        top = ranked[0]
        tied_top = [r for r in ranked if r['total'] == top['total']]
        if len(tied_top) > 1:
            is_tie = True
            decided = True
            summary = (f"Tied at {top['total']} — shoot-off "
                       f"({', '.join(r['name'] for r in tied_top)})")
        else:
            decided = True
            winner_name = top['name']
            margin = top['total'] - ranked[1]['total'] if len(ranked) > 1 else top['total']
            summary = f"{winner_name} wins by {margin} ({top['total']})"
    elif ranked:
        top = ranked[0]
        leaders = [r for r in ranked if r['total'] == top['total']]
        if len(leaders) > 1 or top['total'] == 0:
            summary = "All level" if top['total'] == 0 else (
                "Tied at " + str(top['total']))
        else:
            margin = top['total'] - ranked[1]['total'] if len(ranked) > 1 else top['total']
            summary = f"{top['name']} leads by {margin} ({top['total']})"
    for r in rows:
        r.pop('_end_totals', None)
    return {'mode': mode, 'rows': rows, 'decided': decided,
            'winner_name': winner_name, 'is_tie': is_tie,
            'summary': summary}


@app.route('/tournament/score_sheet', methods=['POST'])
@login_required
def tournament_score_sheet_submit():
    """Process a completed score-sheet form.

    For each participant, allocate a fresh session_id under the
    device-owner's user_id, write synthetic apollo shot rows whose
    coordinates resolve to the ring the user entered, and tag every
    row with `tournament:<round_key>, match:<id>, participant:<name>`.
    Redirects to the match results page.
    """
    user_id = current_user_id()
    round_key = (request.form.get('round_key') or '').strip()
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown tournament round", 400

    participants = _parse_participants_from_form(request.form)
    if not participants:
        return "Add at least one participant before submitting.", 400

    # Competition-level metadata — entered once, applied to every
    # participant's session and printed on every scorecard.
    comp_name     = (request.form.get('competition_name')     or '').strip()
    comp_date     = (request.form.get('competition_date')     or '').strip()
    comp_location = (request.form.get('competition_location') or '').strip()

    segments = _round_segments(round_def)
    match_id = _match_id_for(user_id, round_key)

    # Resolve target_ids once per face_key — the same target rows are
    # shared across participants since they're all on the device
    # owner's user_id.
    target_ids = {}
    for seg in segments:
        fk = seg['face_key']
        if fk not in target_ids:
            target_ids[fk] = _tournament_face_target_id(user_id, fk)

    inserted_session_ids = []
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            for p in participants:
                # Mint a fresh session_id for this participant. Retry
                # on integrity errors the same way /sesh + /tournament
                # already do.
                session_id = _mint_session_id(con, cur, user_id)
                if session_id is None:
                    return "Could not allocate session id — please try again", 500

                tag = (f'tournament:{round_key}, '
                       f'match:{match_id}, '
                       f'participant:{p["name"]}')
                if p['email']:
                    tag += f', email:{p["email"]}'
                # A bowstyle-specific round fixes the AGB classification
                # category for every participant's scorecard.
                forced_bs = _round_bowstyle(round_def)
                if forced_bs:
                    tag += f', bowstyle:{forced_bs}'
                # Competition-level meta — stored on every shot row so
                # the results page and report generators can recover it
                # without a new table. URL-encode the values so commas in
                # event names don't collide with the tag separator.
                from urllib.parse import quote
                if comp_name:
                    tag += f', competition_name:{quote(comp_name, safe="")}'
                if comp_date:
                    tag += f', competition_date:{quote(comp_date, safe="")}'
                if comp_location:
                    tag += f', competition_location:{quote(comp_location, safe="")}'

                # Per-participant notes go into session_notes (per-row).
                # Same URL-encoding for symmetry.
                per_p_notes = (request.form.get(
                    f'participant_{p["idx"]}_notes') or '').strip()

                # Optional per-participant equipment (bow/arrow/effective draw
                # weight + bowtype settings). _insert_shot snapshots the gear
                # the same way a plotted session does; blanks → empty snapshot.
                p_pref = f'participant_{p["idx"]}_'
                p_bow = (request.form.get(p_pref + 'bow') or '').strip()
                p_arrow = (request.form.get(p_pref + 'arrow_type') or '').strip()
                p_style_pref = p_pref + 'style_'
                p_style = {k[len(p_style_pref):]: v
                           for k, v in request.form.items()
                           if k.startswith(p_style_pref)}

                # Walk the segment plan; for each end's worth of arrows
                # read the user-entered ring label from the form and
                # write a synthetic apollo row whose coords score it.
                end_number = 0
                ts = _app_now()
                for seg_idx, seg in enumerate(segments):
                    face_key = seg['face_key']
                    target_id = target_ids.get(face_key)
                    for _ in range(int(seg['ends'])):
                        end_number += 1
                        for arrow_n in range(1, int(seg['arrows_per_end']) + 1):
                            field_name = (f"participant_{p['idx']}_end_"
                                          f"{end_number}_arrow_{arrow_n}")
                            raw = (request.form.get(field_name) or '').strip().upper()
                            if raw == '':
                                # Empty cell → record as miss (forgiving:
                                # the user can leave gaps for arrows they
                                # didn't shoot or weren't reported).
                                raw = 'M'
                            x, y = _coords_for_ring_label(face_key, raw)
                            _insert_shot(
                                cur, user_id=user_id, session_id=session_id,
                                timestamp=ts, bow=p_bow, arrow_type=p_arrow,
                                quiver_size=int(seg['arrows_per_end']),
                                arrows_remaining=(
                                    int(seg['arrows_per_end']) - arrow_n),
                                distance=str(seg['distance_m']),
                                session_notes=per_p_notes, x=x, y=y,
                                is_precise=0, record_mode=0,
                                target_id=target_id,
                                session_tags=tag,
                                style_settings_session=p_style)
                con.commit()
                inserted_session_ids.append(session_id)
    except SQLAlchemyError as e:
        print(f"❌ Score-sheet submit error: {e}")
        return "Error saving scorecard — please try again", 500

    return redirect(url_for('tournament_match_results', match_id=match_id))


@app.route('/tournament/practice_scorecard/setup', methods=['POST'])
@login_required
def tournament_practice_scorecard_setup():
    """Render the flexible practice-scorecard form for a chosen round.

    Like the competition score-sheet (`/tournament/score_sheet/setup`),
    but single-archer and with arbitrary ends: the grid defaults to the
    round's ends × arrows-per-end, yet the user can change any end's
    arrow count and add or remove ends. The target face is fixed to the
    round's primary face — it is *not* arbitrary."""
    user_id = current_user_id()
    round_key = (request.form.get('round_key') or '').strip()
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown tournament round", 400
    _seed_tournament_faces(user_id)

    # The practice card uses one face — the round's primary segment.
    primary = _round_segments(round_def)[0]
    face_key = primary['face_key']
    face = TOURNAMENT_FACES.get(face_key, {})
    labels = _face_ring_labels(face_key)
    # Ring labels for the inputs — X first, then physical ring numbers in
    # descending order, then M.
    choices = ['X'] + [str(l) for l in labels] + ['M']
    # label → point-value lookup (zones[0] is the X ring; zones[1:] map
    # to the labels list 1:1). String keys so the JS matches input.value.
    zones = list(face.get('zones') or ())
    points_map = {'X': int(zones[0][0]) if zones else 0, 'M': 0}
    for i, label in enumerate(labels):
        if 1 + i < len(zones):
            points_map[str(label)] = int(zones[1 + i][0])

    owner_name, owner_email = _owner_name_email(user_id)

    return render_template(
        'tournament.html',
        view='practice_scorecard',
        round_key=round_key,
        round_def=round_def,
        face_key=face_key,
        face_name=face.get('name', face_key),
        distance_m=float(primary['distance_m']),
        choices=choices,
        points_map_json=json.dumps(points_map),
        default_arrows_per_end=int(primary['arrows_per_end']),
        default_ends=int(primary['ends']),
        owner_name=owner_name,
        today_iso=_app_now().strftime('%Y-%m-%d'),
        **_scorecard_equipment_ctx(user_id),
    )


@app.route('/tournament/practice_scorecard', methods=['POST'])
@login_required
def tournament_practice_scorecard_submit():
    """Process a submitted practice scorecard.

    Single archer (the device owner), arbitrary ends. Writes one
    synthetic apollo row per arrow — coords resolved by
    `_coords_for_ring_label` so the existing classifier scores them —
    recording each end's arrow count in `quiver_size`/`arrows_remaining`
    so the results pipeline can chunk the ends back out without a fixed
    round plan. Reuses the match machinery (one participant) so the
    existing results page + downloads work unchanged."""
    user_id = current_user_id()
    round_key = (request.form.get('round_key') or '').strip()
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown tournament round", 400

    primary = _round_segments(round_def)[0]
    face_key = primary['face_key']
    distance_m = primary['distance_m']
    target_id = _tournament_face_target_id(user_id, face_key)

    owner_name, owner_email = _owner_name_email(user_id)
    archer_name = (request.form.get('archer_name') or owner_name or 'Me').strip()

    # Read the dynamic end plan: end_count, then per end end_<e>_arrows
    # and end_<e>_arrow_<a> ring labels.
    try:
        end_count = int(request.form.get('end_count') or 0)
    except (TypeError, ValueError):
        end_count = 0
    if end_count < 1:
        return "Add at least one end before submitting.", 400
    end_count = min(end_count, 200)  # fence off a malformed bulk POST

    # Competition-level meta (optional) — stored on every row like the
    # score-sheet flow so the results page / reports can recover it.
    comp_name     = (request.form.get('competition_name')     or '').strip()
    comp_date     = (request.form.get('competition_date')     or '').strip()
    comp_location = (request.form.get('competition_location') or '').strip()
    notes         = (request.form.get('notes')                or '').strip()

    match_id = _match_id_for(user_id, round_key)

    from urllib.parse import quote
    tag = (f'tournament:{round_key}, practice, practice_scorecard, '
           f'match:{match_id}, participant:{archer_name}')
    if owner_email:
        tag += f', email:{owner_email}'
    forced_bs = _round_bowstyle(round_def)
    if forced_bs:
        tag += f', bowstyle:{forced_bs}'
    if comp_name:
        tag += f', competition_name:{quote(comp_name, safe="")}'
    if comp_date:
        tag += f', competition_date:{quote(comp_date, safe="")}'
    if comp_location:
        tag += f', competition_location:{quote(comp_location, safe="")}'

    # Optional equipment for this single-archer card (bow/arrow + bowtype
    # settings) — snapshotted via _insert_shot like a plotted session. Effective
    # draw weight comes from the bow row. Single archer, so the fields aren't
    # participant-namespaced.
    sc_bow = (request.form.get('bow') or '').strip()
    sc_arrow = (request.form.get('arrow_type') or '').strip()
    sc_style = _collect_style_settings(request.form)

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            session_id = _mint_session_id(con, cur, user_id)
            if session_id is None:
                return "Could not allocate session id — please try again", 500
            ts = _app_now()
            wrote_any = False
            for end_n in range(1, end_count + 1):
                try:
                    arrows = int(request.form.get(f'end_{end_n}_arrows') or 0)
                except (TypeError, ValueError):
                    arrows = 0
                arrows = max(0, min(arrows, 24))  # sane per-end cap
                for arrow_n in range(1, arrows + 1):
                    raw = (request.form.get(
                        f'end_{end_n}_arrow_{arrow_n}') or '').strip().upper()
                    if raw == '':
                        raw = 'M'  # empty cell → miss (forgiving)
                    x, y = _coords_for_ring_label(face_key, raw)
                    _insert_shot(
                        cur, user_id=user_id, session_id=session_id,
                        timestamp=ts, bow=sc_bow, arrow_type=sc_arrow,
                        quiver_size=arrows, arrows_remaining=arrows - arrow_n,
                        distance=str(distance_m), session_notes=notes,
                        x=x, y=y, is_precise=0, record_mode=0,
                        target_id=target_id,
                        session_tags=tag, style_settings_session=sc_style)
                    wrote_any = True
            if not wrote_any:
                return "Add at least one arrow before submitting.", 400
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Practice scorecard submit error: {e}")
        return "Error saving scorecard — please try again", 500

    return redirect(url_for('tournament_match_results', match_id=match_id))


def _match_participants(user_id, match_id):
    """Return participants of a tournament match, with their session_id,
    name, email, round_key, and competition meta (name/date/location).
    Reads from the apollo table — every participant's shots carry
    `match:<id>` in `session_tags`. Per-row `session_notes` carries the
    participant's notes (same value on every row of that participant's
    session)."""
    from urllib.parse import unquote
    out = []
    seen = set()
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT DISTINCT session_id, session_tags, session_notes "
                "FROM apollo WHERE user_id = %s AND session_tags LIKE %s",
                (user_id, f'%match:{match_id}%')
            ).fetchall() or []
    except SQLAlchemyError:
        return out
    for r in rows:
        sid = int(r['session_id']) if 'session_id' in r else int(r[0])
        if sid in seen:
            continue
        seen.add(sid)
        tags = r['session_tags'] if 'session_tags' in r else r[1]
        notes = r['session_notes'] if 'session_notes' in r else (r[2] if len(r) > 2 else '')
        round_key = _round_key_from_tags(tags) or ''
        name = ''
        email = ''
        comp_name = ''
        comp_date = ''
        comp_location = ''
        seat = None
        is_practice_scorecard = False
        for part in (tags or '').split(','):
            p = part.strip()
            if p == 'practice_scorecard':
                is_practice_scorecard = True
            elif p.startswith('participant:'):
                name = p.split(':', 1)[1].strip()
            elif p.startswith('email:'):
                email = p.split(':', 1)[1].strip()
            elif p.startswith('seat:'):
                try:
                    seat = int(p.split(':', 1)[1].strip())
                except (TypeError, ValueError):
                    seat = None
            elif p.startswith('competition_name:'):
                comp_name = unquote(p.split(':', 1)[1].strip())
            elif p.startswith('competition_date:'):
                comp_date = unquote(p.split(':', 1)[1].strip())
            elif p.startswith('competition_location:'):
                comp_location = unquote(p.split(':', 1)[1].strip())
        out.append({
            'session_id':            sid,
            'name':                  name,
            'email':                 email,
            'seat':                  seat,
            'round_key':             round_key,
            'notes':                 notes or '',
            'competition_name':      comp_name,
            'competition_date':      comp_date,
            'competition_location':  comp_location,
            'practice_scorecard':    is_practice_scorecard,
        })
    return out


def _participant_scorecard(user_id, session_id, round_def, round_key=None,
                           practice=False):
    """Return the canonical scorecard for one participant:
    {'ends': [{'arrows':[label,...], 'end_total':n, 'running':n, 'segment_idx':i}],
     'total_score': n, 'x_count': n, 'arrows_shot': n, 'handicap': n, 'awards': [...]}.

    Uses _compute_tournament_progress under the hood (already segment-
    aware) and decorates each arrow with the ring label the user
    originally entered, reverse-mapped from the stored coordinates. When
    ``round_key`` is supplied and the round is complete, the AGB handicap and
    classification awards are included too.

    ``practice`` flags a paper practice scorecard: ends are chunked from
    the stored rows (arbitrary end sizes), and no AGB handicap or
    classification is attached — a practice/range log is never scored as
    an official round, regardless of its end layout."""
    progress = _compute_tournament_progress(session_id, user_id, round_def,
                                            chunk_by_stored=practice)
    segments = _round_segments(round_def)
    # Reverse-map points → ring label. This is approximate for compound
    # faces where two zones share a point value; we display the score
    # plus an 'X' badge for inner-X shots.
    ends_out = []
    for end in progress.get('ends', []):
        out_arrows = []
        for a in end['arrows']:
            if a.get('miss'):
                out_arrows.append('M')
            else:
                # Detect X by checking coords distance against the X-ring
                # of this end's segment face. Multi-spot faces measure to
                # the nearest spot center (shared with the x_count total).
                seg_idx = end.get('segment_idx', 0)
                seg_face = segments[seg_idx]['face_key'] if 0 <= seg_idx < len(segments) else None
                face = TOURNAMENT_FACES.get(seg_face, {}) if seg_face else {}
                x_ring = float(face.get('x_ring_mm') or 0.0)
                seg_multi = face.get('multi_spot') or None
                seg_centers = (list(seg_multi['centers_mm'])
                               if seg_multi and seg_multi.get('centers_mm') else None)
                is_x = _shot_is_x(a['x'], a['y'], x_ring,
                                  spot_centers_mm=seg_centers)
                out_arrows.append('X' if is_x else str(a['points']))
        ends_out.append({
            'arrows':      out_arrows,
            'end_total':   end['end_total'],
            'running':     end['running'],
            'segment_idx': end.get('segment_idx', 0),
            'partial':     end.get('partial', False),
        })
    handicap = None
    awards = []
    # A practice scorecard is a casual / range log, not an official round —
    # so it never carries a handicap or classification, even when its end
    # layout happens to match the round's standard plan.
    if round_key and progress.get('is_complete') and not practice:
        info = _session_handicap_awards(round_key, progress['total_score'])
        handicap = info['handicap']
        awards = info['awards']
    return {
        'total_score': progress['total_score'],
        'x_count':     progress['x_count'],
        'arrows_shot': progress['arrows_shot'],
        'planned':     progress['arrows_planned'],
        'ends':        ends_out,
        'handicap':    handicap,
        'awards':      awards,
    }


def _scorecard_max_ape(round_def, scorecard):
    """Widest arrow column the scorecard tables need. Takes the max over
    the round's segment end sizes *and* the actual ends shot — a practice
    scorecard may have ends larger than the round's standard end size."""
    seg_max = max((int(s['arrows_per_end']) for s in _round_segments(round_def)),
                  default=0)
    end_max = max((len(e['arrows']) for e in scorecard.get('ends', [])),
                  default=0)
    return max(seg_max, end_max, 1)


def _scorecard_csv(round_def, participant, scorecard):
    """Return CSV bytes for one participant's scorecard."""
    import csv as _csv
    import io
    buf = io.StringIO()
    w = _csv.writer(buf)
    if participant.get('competition_name'):
        w.writerow([f"Competition: {participant['competition_name']}"])
    w.writerow([f"Round: {round_def['name']}"])
    if participant.get('competition_date'):
        w.writerow([f"Date: {participant['competition_date']}"])
    if participant.get('competition_location'):
        w.writerow([f"Location: {participant['competition_location']}"])
    w.writerow([f"Competitor: {participant['name']}"])
    if participant.get('email'):
        w.writerow([f"Email: {participant['email']}"])
    w.writerow([f"Logged: {_app_now().strftime('%Y-%m-%d %H:%M')}"])
    w.writerow([])
    max_ape = _scorecard_max_ape(round_def, scorecard)
    w.writerow(['End'] + [f'A{i+1}' for i in range(max_ape)] + ['End total', 'Running'])
    for idx, end in enumerate(scorecard['ends'], start=1):
        row = [idx] + list(end['arrows'])
        # Pad to max_ape
        while len(row) - 1 < max_ape:
            row.append('')
        row += [end['end_total'], end['running']]
        w.writerow(row)
    w.writerow([])
    w.writerow(['Total score', scorecard['total_score'], f"/ {round_def['max_score']}"])
    w.writerow(['X count', scorecard['x_count']])
    w.writerow(['Arrows shot', f"{scorecard['arrows_shot']} / {scorecard['planned']}"])
    if scorecard.get('handicap') is not None:
        w.writerow(['AGB handicap', scorecard['handicap']])
    for a in scorecard.get('awards') or []:
        w.writerow([f"{a['scheme']} classification", a['name']])
    if participant.get('notes'):
        w.writerow([])
        w.writerow(['Notes', participant['notes']])
    return buf.getvalue().encode('utf-8')


def _scorecard_xlsx(round_def, participant, scorecard):
    """Return XLSX bytes for one participant's scorecard."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = 'Scorecard'
    # Title block — competition info, round, competitor.
    row_n = 1
    if participant.get('competition_name'):
        ws.cell(row=row_n, column=1, value=participant['competition_name']).font = Font(bold=True, size=14)
        row_n += 1
    ws.cell(row=row_n, column=1, value=f"Round: {round_def['name']}").font = Font(bold=True, size=12)
    row_n += 1
    if participant.get('competition_date'):
        ws.cell(row=row_n, column=1, value=f"Date: {participant['competition_date']}")
        row_n += 1
    if participant.get('competition_location'):
        ws.cell(row=row_n, column=1, value=f"Location: {participant['competition_location']}")
        row_n += 1
    ws.cell(row=row_n, column=1, value=f"Competitor: {participant['name']}").font = Font(bold=True)
    row_n += 1
    if participant.get('email'):
        ws.cell(row=row_n, column=1, value=f"Email: {participant['email']}")
        row_n += 1
    ws.cell(row=row_n, column=1, value=f"Logged: {_app_now().strftime('%Y-%m-%d %H:%M')}")
    row_n += 1

    max_ape = _scorecard_max_ape(round_def, scorecard)
    header_row = row_n + 1
    ws.cell(row=header_row, column=1, value='End').font = Font(bold=True)
    for i in range(max_ape):
        c = ws.cell(row=header_row, column=2 + i, value=f'A{i+1}')
        c.font = Font(bold=True); c.alignment = Alignment(horizontal='center')
    ws.cell(row=header_row, column=2 + max_ape, value='End total').font = Font(bold=True)
    ws.cell(row=header_row, column=3 + max_ape, value='Running').font = Font(bold=True)
    for idx, end in enumerate(scorecard['ends'], start=1):
        r = header_row + idx
        ws.cell(row=r, column=1, value=idx)
        for ai, val in enumerate(end['arrows']):
            ws.cell(row=r, column=2 + ai, value=val).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2 + max_ape, value=end['end_total'])
        ws.cell(row=r, column=3 + max_ape, value=end['running'])
    summary_r = header_row + len(scorecard['ends']) + 2
    ws.cell(row=summary_r, column=1, value='Total score').font = Font(bold=True)
    ws.cell(row=summary_r, column=2, value=scorecard['total_score'])
    ws.cell(row=summary_r, column=3, value=f"/ {round_def['max_score']}")
    ws.cell(row=summary_r + 1, column=1, value='X count').font = Font(bold=True)
    ws.cell(row=summary_r + 1, column=2, value=scorecard['x_count'])
    ws.cell(row=summary_r + 2, column=1, value='Arrows shot').font = Font(bold=True)
    ws.cell(row=summary_r + 2, column=2, value=f"{scorecard['arrows_shot']} / {scorecard['planned']}")
    extra_r = summary_r + 3
    if scorecard.get('handicap') is not None:
        ws.cell(row=extra_r, column=1, value='AGB handicap').font = Font(bold=True)
        ws.cell(row=extra_r, column=2, value=scorecard['handicap'])
        extra_r += 1
    for a in scorecard.get('awards') or []:
        ws.cell(row=extra_r, column=1, value=f"{a['scheme']} class").font = Font(bold=True)
        ws.cell(row=extra_r, column=2, value=a['name'])
        extra_r += 1
    if participant.get('notes'):
        notes_r = extra_r + 1
        ws.cell(row=notes_r, column=1, value='Notes').font = Font(bold=True)
        ws.cell(row=notes_r, column=2, value=participant['notes']).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=notes_r, start_column=2,
                       end_row=notes_r, end_column=2 + max_ape)
    ws.column_dimensions['A'].width = 16
    for col in range(2, 4 + max_ape):
        ws.column_dimensions[chr(ord('A') + col - 1)].width = 9
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _scorecard_pdf(round_def, participant, scorecard):
    """Return PDF bytes for one participant's scorecard. Uses
    matplotlib (already a dep) — renders the scorecard as a table
    figure and saves to PDF in memory."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    import io

    max_ape = _scorecard_max_ape(round_def, scorecard)
    header = ['End'] + [f'A{i+1}' for i in range(max_ape)] + ['End', 'Run']
    rows = []
    for idx, end in enumerate(scorecard['ends'], start=1):
        row = [str(idx)] + list(end['arrows'])
        while len(row) - 1 < max_ape:
            row.append('')
        row += [str(end['end_total']), str(end['running'])]
        rows.append(row)

    # Size the figure so wide rounds (WA 1440 with 24 ends + 6 arrows)
    # remain readable. Cap height; matplotlib will auto-shrink fonts.
    fig_h = max(7.0, min(14.0, 0.25 * (len(rows) + 8)))
    fig, ax = plt.subplots(figsize=(8.5, fig_h))
    ax.set_axis_off()
    # Try Quantico to match the rest of the site; fall back to sans-serif
    # if the font isn't installed on the host. matplotlib silently picks
    # a substitute when the named family isn't found.
    font_family = ['Quantico', 'sans-serif']

    # Header block — competition info, round, competitor.
    header_lines = []
    if participant.get('competition_name'):
        header_lines.append(participant['competition_name'])
    header_lines.append(f"Round: {round_def['name']}")
    meta_bits = []
    if participant.get('competition_date'):
        meta_bits.append(f"Date: {participant['competition_date']}")
    if participant.get('competition_location'):
        meta_bits.append(f"Location: {participant['competition_location']}")
    if meta_bits:
        header_lines.append(' · '.join(meta_bits))
    header_lines.append(f"Competitor: {participant['name']}")
    if participant.get('email'):
        header_lines.append(f"Email: {participant['email']}")
    header_lines.append(f"Logged: {_app_now().strftime('%Y-%m-%d %H:%M')}")
    fig.suptitle('\n'.join(header_lines), fontsize=11, ha='left', x=0.05,
                 y=0.98, fontfamily=font_family)

    table = ax.table(cellText=rows, colLabels=header, loc='center',
                     cellLoc='center')
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1.05, 1.3)  # a bit more horizontal + vertical breathing room
    for cell in table.get_celld().values():
        cell.set_text_props(fontfamily=font_family)
        # Wider horizontal padding inside each cell
        cell.PAD = 0.06

    # Summary line
    summary = (f"Total: {scorecard['total_score']} / {round_def['max_score']}    "
               f"X: {scorecard['x_count']}    "
               f"Arrows: {scorecard['arrows_shot']} / {scorecard['planned']}")
    if scorecard.get('handicap') is not None:
        summary += f"    Handicap: {scorecard['handicap']}"
    ax.text(0.5, -0.02, summary, transform=ax.transAxes, ha='center',
            fontsize=11, fontweight='bold', fontfamily=font_family)
    if scorecard.get('awards'):
        award_txt = '   '.join(f"{a['scheme']}: {a['name']}"
                               for a in scorecard['awards'])
        ax.text(0.5, -0.05, award_txt, transform=ax.transAxes, ha='center',
                fontsize=9, fontfamily=font_family)
    # Notes below the summary
    if participant.get('notes'):
        ax.text(0.05, -0.08, f"Notes: {participant['notes']}",
                transform=ax.transAxes, ha='left', va='top', fontsize=9,
                wrap=True, fontfamily=font_family)

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        pdf.savefig(fig, bbox_inches='tight')
    plt.close(fig)
    return buf.getvalue()


@app.route('/tournament/match/<match_id>', methods=['GET'])
@login_required
def tournament_match_results(match_id):
    """Render the results page for a finished score-sheet match.

    Shows each participant's scorecard summary, with per-participant
    Download buttons (PDF / CSV / Excel) and a one-click "Email
    everyone their report" action."""
    user_id = current_user_id()
    participants = _match_participants(user_id, match_id)
    if not participants:
        return "Match not found.", 404
    # Round identity is on the first participant's tag.
    round_key = participants[0]['round_key']
    round_def = _tournament_round_def(round_key)
    if round_def is None:
        return "Unknown round.", 400

    # A solo practice scorecard is single-archer with arbitrary ends —
    # no head-to-head match scoring, and nobody to email.
    is_practice = bool(participants[0].get('practice_scorecard'))

    # Score the match (set points / cumulative / leaderboard) from each
    # participant's raw progress, in seat order so set play pairs the
    # archers consistently. (Skipped for a practice scorecard.)
    match_score = None
    if not is_practice:
        by_seat = sorted(participants,
                         key=lambda q: (q.get('seat') if q.get('seat') is not None else 999))
        archers_for_scoring = [{
            'name':      p['name'],
            'seat':      p.get('seat'),
            'is_active': False,
            'progress':  _compute_tournament_progress(p['session_id'], user_id, round_def),
        } for p in by_seat]
        match_score = _match_set_scoring(archers_for_scoring, round_def)

    results = []
    for p in participants:
        scorecard = _participant_scorecard(user_id, p['session_id'], round_def,
                                           round_key=p['round_key'],
                                           practice=p.get('practice_scorecard', False))
        results.append({**p, 'scorecard': scorecard})

    # Sort highest score first (ranking)
    results.sort(key=lambda r: (r['scorecard']['total_score'],
                                r['scorecard']['x_count']),
                 reverse=True)

    email_enabled = (bool(os.environ.get('RESEND_API_KEY', '').strip())
                     and (resend is not None) and not is_practice)
    return render_template(
        'tournament.html',
        view='results',
        match_id=match_id,
        round_def=round_def,
        results=results,
        match_score=match_score,
        email_enabled=email_enabled,
        is_practice=is_practice,
    )


@app.route('/tournament/match/<match_id>/download/<int:session_id>/<fmt>',
           methods=['GET'])
@login_required
def tournament_match_download(match_id, session_id, fmt):
    """Stream one participant's scorecard as PDF / CSV / XLSX."""
    user_id = current_user_id()
    participants = _match_participants(user_id, match_id)
    p = next((q for q in participants if q['session_id'] == session_id), None)
    if p is None:
        return "Participant not found.", 404
    round_def = _tournament_round_def(p['round_key'])
    if round_def is None:
        return "Round not found.", 400
    scorecard = _participant_scorecard(user_id, session_id, round_def,
                                       round_key=p['round_key'],
                                       practice=p.get('practice_scorecard', False))
    safe_name = ''.join(c for c in p['name']
                        if c.isalnum() or c in ('-', '_')) or f's{session_id}'
    base = f"apollo-{p['round_key']}-{safe_name}"
    if fmt == 'csv':
        data = _scorecard_csv(round_def, p, scorecard)
        return (data, 200, {
            'Content-Type':        'text/csv; charset=utf-8',
            'Content-Disposition': f'attachment; filename="{base}.csv"',
        })
    if fmt == 'xlsx':
        data = _scorecard_xlsx(round_def, p, scorecard)
        return (data, 200, {
            'Content-Type':
              'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="{base}.xlsx"',
        })
    if fmt == 'pdf':
        data = _scorecard_pdf(round_def, p, scorecard)
        return (data, 200, {
            'Content-Type':        'application/pdf',
            'Content-Disposition': f'attachment; filename="{base}.pdf"',
        })
    return "Unknown format.", 400


@app.route('/tournament/match/<match_id>/email', methods=['POST'])
@login_required
def tournament_match_email(match_id):
    """Email each participant their own PDF + CSV + Excel attachments.

    Skips participants without an email address; reports the per-
    recipient result back to the device owner."""
    user_id = current_user_id()
    participants = _match_participants(user_id, match_id)
    if not participants:
        return jsonify(ok=False, error='match not found'), 404
    round_def = _tournament_round_def(participants[0]['round_key'])
    if round_def is None:
        return jsonify(ok=False, error='round not found'), 400

    sent = []
    skipped = []
    for p in participants:
        if not p.get('email'):
            skipped.append({'name': p['name'], 'reason': 'no email on file'})
            continue
        scorecard = _participant_scorecard(user_id, p['session_id'], round_def,
                                           round_key=p['round_key'],
                                           practice=p.get('practice_scorecard', False))
        attachments = [
            {'filename': 'scorecard.pdf',  'content': _scorecard_pdf(round_def, p, scorecard)},
            {'filename': 'scorecard.csv',  'content': _scorecard_csv(round_def, p, scorecard)},
            {'filename': 'scorecard.xlsx', 'content': _scorecard_xlsx(round_def, p, scorecard)},
        ]
        body_lines = [f"Your Apollo tournament results — {round_def['name']}", '']
        if p.get('competition_name'):
            body_lines.append(f"Competition: {p['competition_name']}")
        if p.get('competition_date'):
            body_lines.append(f"Date: {p['competition_date']}")
        if p.get('competition_location'):
            body_lines.append(f"Location: {p['competition_location']}")
        body_lines += [
            f"Competitor: {p['name']}",
            '',
            f"Total score: {scorecard['total_score']} / {round_def['max_score']}",
            f"X count: {scorecard['x_count']}",
            f"Arrows shot: {scorecard['arrows_shot']} / {scorecard['planned']}",
        ]
        if scorecard.get('handicap') is not None:
            body_lines.append(f"AGB handicap: {scorecard['handicap']}")
        for a in scorecard.get('awards') or []:
            body_lines.append(f"{a['scheme']} classification: {a['name']}")
        if p.get('notes'):
            body_lines += ['', f"Notes: {p['notes']}"]
        body_lines += ['', 'Full per-end breakdown is attached as PDF / CSV / Excel.']
        body = '\n'.join(body_lines)
        ok = _send_email(
            to_addr=p['email'],
            subject=f"Apollo results — {round_def['name']}",
            body=body,
            attachments=attachments,
        )
        if ok:
            sent.append(p['email'])
        else:
            skipped.append({'name': p['name'], 'reason': 'send failed'})
    return jsonify(ok=True, sent=sent, skipped=skipped)


@app.route("/previous_sessions", methods=['GET'])
@login_required
def previous_sessions():
    """List the user's past sessions with rows and computed stats.

    Each session that yields a get_stats() error is silently skipped so a
    single corrupt session doesn't break the whole page.

    Optional GET query params filter the returned sessions:
      q_notes, q_target, q_bow, q_arrow  — case-insensitive substring matches
      date_from, date_to                  — inclusive YYYY-MM-DD bounds on row timestamps
    A session is kept if at least one of its rows satisfies every supplied
    filter (target is session-scoped so it's checked once).
    """
    user_id = current_user_id()

    filters = {
        'q_notes':   (request.args.get('q_notes')  or '').strip(),
        'q_tags':    (request.args.get('q_tags')   or '').strip(),
        'q_target':  (request.args.get('q_target') or '').strip(),
        'q_bow':     (request.args.get('q_bow')    or '').strip(),
        'q_arrow':   (request.args.get('q_arrow')  or '').strip(),
        'date_from': (request.args.get('date_from') or '').strip(),
        'date_to':   (request.args.get('date_to')   or '').strip(),
    }

    # Batched fetch: pull every shot for this user in a single query and
    # group by session_id. Avoids the previous N+1 (one query per session)
    # which got painful around the 100-session mark.
    session_rows = {}
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            all_rows = cur.execute(
                "SELECT session_id, timestamp, bow, arrow_type, quiver_size, "
                "arrows_remaining, session_notes, session_tags, x_coord, y_coord, "
                "is_precise, target_id, arrow_shaft_diameter "
                "FROM apollo WHERE user_id = %s ORDER BY session_id, timestamp",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError as e:
        print(f"Previous sessions error: {e}")
        return "Previous sessions error", 500

    for row in all_rows:
        sid = int(row['session_id'])
        session_rows.setdefault(sid, []).append(row)
    session_ids = sorted(session_rows.keys())

    # Memoize targets across sessions — many users reuse a small handful.
    target_cache = {}
    def _cached_target(tid):
        if tid not in target_cache:
            target_cache[tid] = target_to_config(get_target(tid, user_id))
        return target_cache[tid]

    # Memoize zone lookups too — same justification, plus _classify_shot
    # gets called once per shot in the replay payload.
    zones_cache = {}
    def _cached_zones(tid):
        if tid not in zones_cache:
            zones_cache[tid] = _fetch_target_zones(tid, user_id) \
                if tid is not None else []
        return zones_cache[tid]

    session_data = {}
    for session_id in session_ids:
        res = session_rows[session_id]
        stats = get_stats(session_id, user_id)
        # get_stats() returns a (msg, code) tuple on failure;
        # treat that as "skip this session" rather than 500.
        if isinstance(stats, tuple):
            continue
        # Sessions lock target after first shot, so every row
        # in a session points at the same target — read once.
        target_cfg = _cached_target(res[0]['target_id'])
        if not _session_matches_filters(res, target_cfg, filters):
            continue
        # Pre-classify each shot for the replay canvas so the JS doesn't
        # need to know about zones or line-cutter rules — it just reads
        # the boolean. Out-of-zone hits then render with the miss marker.
        session_zones = _cached_zones(res[0]['target_id'])
        replay_shots = []
        for r in res:
            xraw = str(r['x_coord']).strip() if r['x_coord'] is not None else ''
            yraw = str(r['y_coord']).strip() if r['y_coord'] is not None else ''
            is_miss = (xraw == MISS_SENTINEL and yraw == MISS_SENTINEL)
            if not is_miss and session_zones:
                is_miss = _classify_shot(
                    xraw, yraw, session_zones,
                    _row_get(r, 'arrow_shaft_diameter')) is None
            replay_shots.append({'x': xraw, 'y': yraw, 'miss': is_miss})
        session_data[session_id] = {
            'rows': res,
            'stats': stats,
            'target': target_cfg,
            'replay_shots': replay_shots,
        }
        # For a completed tournament round, surface its AGB handicap and any
        # classification awards so the history list mirrors the round-complete
        # banner. Skipped for practice/regular sessions and partial rounds.
        tags0 = res[0]['session_tags'] if res else None
        rkey = _round_key_from_tags(tags0)
        if rkey:
            rdef = _tournament_round_def(rkey)
            prog = _compute_tournament_progress(session_id, user_id, rdef)
            if prog.get('is_complete'):
                info = _session_handicap_awards(rkey, prog.get('total_score'), tags0)
                session_data[session_id]['round_name'] = rdef.get('name')
                session_data[session_id]['handicap'] = info['handicap']
                session_data[session_id]['awards'] = info['awards']

    # Dropdown options for the filter bar — derived from every shot the
    # user has ever recorded, so the dropdowns surface the full history
    # regardless of which filter is currently active.
    bow_options = sorted({(r['bow'] or '').strip() for r in all_rows
                          if (r['bow'] or '').strip()}, key=str.lower)
    arrow_options = sorted({(r['arrow_type'] or '').strip() for r in all_rows
                            if (r['arrow_type'] or '').strip()}, key=str.lower)
    target_options = sorted({(cfg or {}).get('name', '').strip()
                             for cfg in target_cache.values()
                             if cfg and (cfg.get('name') or '').strip()},
                            key=str.lower)

    return render_template('previous_sessions.html',
                           session_data=session_data,
                           filters=filters,
                           bow_options=bow_options,
                           arrow_options=arrow_options,
                           target_options=target_options,
                           tag_suggestions=_distinct_user_tags(user_id),
                           tournament_rounds=TOURNAMENT_ROUNDS)


def _session_matches_filters(rows, target_cfg, filters):
    """True if this session passes every supplied search filter.

    Per-row matchers (notes/bow/arrow) succeed when ANY row matches,
    so a single quiver-note hit surfaces the whole session. Target name
    is constant per session. Date bounds compare the YYYY-MM-DD prefix of
    each row's timestamp (sqlite stores strings, mysql returns datetimes
    — _ts_date_str normalizes both)."""
    q_notes  = filters['q_notes'].lower()
    q_target = filters['q_target'].lower()
    q_bow    = filters['q_bow'].lower()
    q_arrow  = filters['q_arrow'].lower()
    date_from = filters['date_from']
    date_to   = filters['date_to']
    # Tag query: comma-separated list. Each query term must appear as a
    # whole tag (case-insensitive) somewhere in the session's tags. All
    # query terms must match (AND across the query).
    q_tag_terms = [
        t.strip().lower() for t in filters['q_tags'].split(',') if t.strip()
    ]

    if q_target:
        target_name = (target_cfg or {}).get('name', '') or ''
        if q_target not in target_name.lower():
            return False

    # Date range is a session-level filter: a session matches if any of its
    # shots land in the range. Keeping it out of the per-row content matcher
    # below means a session that straddles midnight isn't rejected just
    # because the row that happens to mention the searched bow falls on the
    # "wrong" side of date_from.
    if date_from or date_to:
        in_range = False
        for r in rows:
            d = _ts_date_str(r['timestamp'])
            if date_from and d < date_from:
                continue
            if date_to and d > date_to:
                continue
            in_range = True
            break
        if not in_range:
            return False

    def _row_matches(row):
        if q_notes and q_notes not in (row['session_notes'] or '').lower():
            return False
        if q_bow and q_bow not in (row['bow'] or '').lower():
            return False
        if q_arrow and q_arrow not in (row['arrow_type'] or '').lower():
            return False
        if q_tag_terms:
            row_tags = {
                t.strip().lower()
                for t in (row['session_tags'] or '').split(',')
                if t.strip()
            }
            if not all(term in row_tags for term in q_tag_terms):
                return False
        return True

    # No per-row content filter means every session in range passes.
    if not (q_notes or q_bow or q_arrow or q_tag_terms):
        return True
    return any(_row_matches(r) for r in rows)


def _ts_date_str(value):
    """Coerce a stored timestamp to a YYYY-MM-DD string for date filtering.

    Returns the *user's local* date — the date-from/date-to bounds in the
    search bar are typed in the user's wall-clock, so a late-evening shot
    in UTC-7 must compare against the date the user actually shot it on,
    not the UTC date.
    """
    if value is None:
        return ''
    dt = _utc_to_user(value)
    return dt.strftime('%Y-%m-%d') if dt is not None else str(value)[:10]


@app.route('/delete_session/<int:session_id>', methods=['POST'])
@login_required
def delete_session(session_id):
    """Hard-delete a single past session's shots and timing row for this user."""
    user_id = current_user_id()
    # The session currently held in the cookie shows up in the Previous
    # Sessions list too, so users naturally try to delete it from there.
    # We allow it — but must also clear the in-progress keys from the
    # cookie afterwards. Otherwise the cookie would still point at a
    # session_id with no session_times row: subsequent /sesh POSTs would
    # orphan shots into it and the unique (user_id, session_id) index would
    # block /end_session from re-allocating the same id. Clearing the keys
    # makes the next /sesh GET allocate a fresh session_id cleanly.
    deleting_current = (session.get('session_id') == session_id)
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            owner_row = cur.execute(
                "SELECT 1 FROM apollo WHERE session_id = %s AND user_id = %s LIMIT 1",
                (session_id, user_id)
            ).fetchone()
            if owner_row is None:
                abort(404)
            cur.execute(
                "DELETE FROM apollo WHERE session_id = %s AND user_id = %s",
                (session_id, user_id)
            )
            cur.execute(
                "DELETE FROM session_times WHERE session_id = %s AND user_id = %s",
                (session_id, user_id)
            )
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Delete-session error: {e}")
        flash("Could not delete session — please try again.")
        return redirect(url_for('previous_sessions'))
    # Drop the in-progress keys if we just deleted the active session, so
    # the next /sesh allocates a fresh id instead of writing into the now
    # orphaned one. Mirrors the key list end_session pops.
    if deleting_current:
        for k in ('session_id', 'quivers_completed', 'arrows_remaining',
                  'record_mode', 'target_id', 'current_quiver_size',
                  'tournament_round_key', 'tournament_segment_idx', 'tournament_bowstyle',
                  'tournament_practice',
                  'match_id', 'match_archers', 'match_idx'):
            session.pop(k, None)
    return redirect(url_for('previous_sessions'))


@app.route('/edit_session/<int:session_id>', methods=['GET', 'POST'])
@login_required
def edit_session(session_id):
    """Edit a past session's session-level attribution and per-shot rows.

    Session-level edits (bow / arrow / distance / notes / tags / target)
    apply to every row in the session and re-snapshot the bow_* / arrow_*
    denormalized columns from the user's current bows/arrows tables — same
    snapshot logic the active /sesh path uses on insert.
    """
    user_id = current_user_id()
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT id, session_id, timestamp, bow, arrow_type, quiver_size, "
                "arrows_remaining, distance, session_notes, session_tags, "
                "x_coord, y_coord, is_precise, target_id, effective_draw_weight "
                "FROM apollo WHERE session_id = %s AND user_id = %s "
                "ORDER BY timestamp, id",
                (session_id, user_id)
            ).fetchall()
            if not rows:
                abort(404)

            if request.method == 'POST':
                first = rows[0]
                new_bow      = (request.form.get('bow') or first['bow'] or '').strip()
                new_arrow    = (request.form.get('arrow_type') or first['arrow_type'] or '').strip()
                new_distance = (request.form.get('distance') or '').strip()
                new_notes    = request.form.get('session_notes', '')
                new_tags     = _normalize_tags(request.form.get('session_tags', ''))
                posted_tid   = (request.form.get('target_id') or '').strip()
                try:
                    new_target_id = int(posted_tid) if posted_tid else first['target_id']
                except ValueError:
                    new_target_id = first['target_id']
                # Reject a target_id the current user doesn't own.
                if new_target_id is not None and get_target(new_target_id, user_id) is None:
                    new_target_id = first['target_id']

                # Refresh bow snapshot from the current bows row, if any.
                # Effective draw weight is an intrinsic bow property, so it's
                # snapshotted from the bows row like draw weight / AMO.
                bow_row = cur.execute(
                    "SELECT bow_draw_weight, effective_draw_weight, "
                    "amo, bow_type FROM bows WHERE bow_model = %s AND user_id = %s LIMIT 1",
                    (new_bow, user_id)
                ).fetchone()
                if bow_row is not None:
                    bow_snap = (
                        bow_row['bow_draw_weight'], bow_row['effective_draw_weight'],
                        bow_row['amo'], bow_row['bow_type'],
                    )
                else:
                    bow_snap = (None, None, None, None)

                # Refresh arrow snapshot from the current arrows row, if any.
                arrow_row = cur.execute(
                    "SELECT length, spine, shaft_weight, shaft_diameter, "
                    "shaft_material, nock_weight, tip, tip_weight FROM arrows "
                    "WHERE arrow = %s AND user_id = %s LIMIT 1",
                    (new_arrow, user_id)
                ).fetchone()
                if arrow_row is not None:
                    arrow_snap = (
                        arrow_row['length'], arrow_row['spine'],
                        arrow_row['shaft_weight'], arrow_row['shaft_diameter'],
                        arrow_row['shaft_material'], arrow_row['nock_weight'],
                        arrow_row['tip'], arrow_row['tip_weight'],
                    )
                else:
                    arrow_snap = (None, None, None, None, None, None, None, None)

                # Per-shot deletes — checkbox name "delete_<row id>".
                shot_ids = {int(r['id']) for r in rows}
                delete_ids = []
                for key in request.form.keys():
                    if not key.startswith('delete_'):
                        continue
                    try:
                        rid = int(key.split('_', 1)[1])
                    except ValueError:
                        continue
                    if rid in shot_ids:
                        delete_ids.append(rid)
                if delete_ids:
                    placeholders = ','.join(['%s'] * len(delete_ids))
                    cur.execute(
                        f"DELETE FROM apollo WHERE user_id = %s AND session_id = %s "
                        f"AND id IN ({placeholders})",
                        tuple([user_id, session_id] + delete_ids)
                    )

                # Per-shot hit→miss conversion — checkbox name "miss_<row id>".
                # Going the other way (miss→hit) needs new coords, so it isn't
                # offered here. The form just marks a shot as a miss.
                for r in rows:
                    if int(r['id']) in delete_ids:
                        continue
                    if request.form.get(f"miss_{r['id']}") == '1':
                        rx = str(r['x_coord']).strip() if r['x_coord'] is not None else ''
                        ry = str(r['y_coord']).strip() if r['y_coord'] is not None else ''
                        if rx != MISS_SENTINEL or ry != MISS_SENTINEL:
                            cur.execute(
                                "UPDATE apollo SET x_coord = %s, y_coord = %s, "
                                "is_precise = 0 WHERE id = %s AND user_id = %s",
                                (MISS_SENTINEL, MISS_SENTINEL, int(r['id']), user_id)
                            )

                # Session-level update — bow/arrow plus their denormalized snapshots,
                # distance, notes, tags, target. quiver/arrows_remaining/x/y are
                # left untouched.
                cur.execute(
                    "UPDATE apollo SET bow = %s, arrow_type = %s, distance = %s, "
                    "session_notes = %s, session_tags = %s, target_id = %s, "
                    "bow_draw_weight = %s, effective_draw_weight = %s, "
                    "bow_amo = %s, bow_type = %s, "
                    "arrow_length = %s, arrow_spine = %s, arrow_shaft_weight = %s, "
                    "arrow_shaft_diameter = %s, arrow_shaft_material = %s, "
                    "arrow_nock_weight = %s, arrow_tip = %s, arrow_tip_weight = %s "
                    "WHERE session_id = %s AND user_id = %s",
                    (new_bow, new_arrow, new_distance, new_notes, new_tags,
                     new_target_id, *bow_snap, *arrow_snap,
                     session_id, user_id)
                )
                con.commit()
                flash("Session updated.")
                return redirect(url_for('previous_sessions'))

            # GET — gather dropdown options.
            arrow_rows = cur.execute(
                "SELECT DISTINCT arrow FROM arrows WHERE user_id = %s ORDER BY arrow",
                (user_id,)
            ).fetchall()
            arrow_types = [r[0] for r in arrow_rows] if arrow_rows else []
            bow_rows = cur.execute(
                "SELECT DISTINCT bow_model FROM bows WHERE user_id = %s ORDER BY bow_model",
                (user_id,)
            ).fetchall()
            bow_models = [r[0] for r in bow_rows] if bow_rows else []
            target_rows = cur.execute(
                "SELECT id AS rowid, name FROM targets "
                "WHERE is_active = 1 AND user_id = %s ORDER BY name",
                (user_id,)
            ).fetchall()
            targets_list = [{'rowid': r['rowid'], 'name': _display_target_name(r['name'])} for r in target_rows]
    except SQLAlchemyError as e:
        print(f"❌ Edit-session error: {e}")
        flash("Could not load session for editing — please try again.")
        return redirect(url_for('previous_sessions'))

    first = rows[0]
    current = {
        'bow':           first['bow'] or '',
        'arrow_type':    first['arrow_type'] or '',
        'distance':      first['distance'] or '',
        'session_notes': first['session_notes'] or '',
        'session_tags':  first['session_tags'] or '',
        'target_id':     first['target_id'],
    }
    # Per-row miss flag for the template — same rule as the replay payload.
    shot_view = []
    for r in rows:
        xraw = str(r['x_coord']).strip() if r['x_coord'] is not None else ''
        yraw = str(r['y_coord']).strip() if r['y_coord'] is not None else ''
        shot_view.append({
            'id':        int(r['id']),
            'timestamp': _format_session_dt_user(r['timestamp']) or r['timestamp'],
            'x':         xraw,
            'y':         yraw,
            'is_miss':   (xraw == MISS_SENTINEL and yraw == MISS_SENTINEL),
            'is_precise': r['is_precise'],
        })
    return render_template(
        'edit_session.html',
        session_id=session_id,
        current=current,
        shots=shot_view,
        bow_models=bow_models,
        arrow_types=arrow_types,
        targets_list=targets_list,
        tag_suggestions=_distinct_user_tags(user_id),
    )


def _end_session_noun():
    """Word the end-session page for the just-finished outing's context.

    A live match sets ``match_id``; a tournament round sets
    ``tournament_round_key``; anything else is a standalone session. Read
    before the per-round keys are cleared on finalize.
    """
    if session.get('match_id'):
        return 'Match'
    if session.get('tournament_round_key'):
        return 'Round'
    return 'Session'


@app.route('/end_session', methods=['GET', 'POST'])
@login_required
def end_session():
    """Two-phase: GET shows the confirmation/manual-length form, POST finalizes.

    GET also handles the "user clicked End Session without shooting
    anything" case by deleting the orphan session_times row and clearing
    the Flask session — keeps the apollo table free of zero-shot sessions.
    """
    user_id = current_user_id()
    if request.method == 'GET':
        if not session.get('session_id'):
            return render_template('splash.html')
        session_id = session.get('session_id')

        # If no arrows were shot, clean up and return to splash
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            res = cur.execute(
                "SELECT timestamp FROM apollo WHERE session_id = %s AND user_id = %s",
                (session_id, user_id)
            ).fetchone()
            if res is None:
                cur.execute(
                    "DELETE FROM session_times WHERE session_id = %s AND user_id = %s",
                    (session_id, user_id)
                )
                con.commit()
                # Drop the session_id (and other per-round keys) but keep
                # the user logged in.
                for k in ('session_id', 'quivers_completed', 'arrows_remaining',
                          'record_mode', 'target_id', 'current_quiver_size',
                  'tournament_round_key', 'tournament_segment_idx', 'tournament_practice',
                  'tournament_bowstyle',
                  'match_id', 'match_archers', 'match_idx'):
                    session.pop(k, None)
                return render_template('splash.html')

        begin_time = None
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                row = cur.execute(
                    "SELECT session_begin_time FROM session_times "
                    "WHERE session_id = %s AND user_id = %s",
                    (session_id, user_id)
                ).fetchone()
                begin_time = row[0] if row else None
                # Fall back to the earliest shot's timestamp when the
                # session_times row is missing or has a NULL begin_time,
                # so the field is never blank on the form.
                if not begin_time:
                    earliest = cur.execute(
                        "SELECT MIN(timestamp) FROM apollo "
                        "WHERE session_id = %s AND user_id = %s",
                        (session_id, user_id)
                    ).fetchone()
                    begin_time = earliest[0] if earliest and earliest[0] else _app_now()
        except SQLAlchemyError:
            begin_time = _app_now()

        end_noun = _end_session_noun()
        return render_template('end_session.html',
                               session_id=session_id,
                               begin_time=_format_session_dt_user(begin_time),
                               end_time=_format_session_dt_user(_app_now()),
                               end_noun=end_noun,
                               stats=None)

    # POST: finalize session. Capture the context noun now, before the
    # per-round Flask-session keys are cleared below, so the results view
    # can word itself "Match" / "Round" / "Session".
    end_noun = _end_session_noun()
    session_id_str = request.form.get('session_id', '').strip()
    if not session_id_str:
        session_id_str = str(session.get('session_id', ''))
    if not session_id_str:
        return redirect(url_for('index'))
    try:
        session_id = int(session_id_str)
    except ValueError:
        return redirect(url_for('index'))

    begin_time_raw = request.form.get('session_begin_time', '').strip()
    end_time_raw = request.form.get('session_end_time', '').strip()
    parsed_begin = _parse_session_dt_user(begin_time_raw)
    parsed_end = _parse_session_dt_user(end_time_raw)
    if parsed_begin is None or parsed_end is None:
        return render_template(
            'end_session.html',
            session_id=session_id,
            begin_time=begin_time_raw or _format_session_dt_user(_app_now()),
            end_time=end_time_raw or _format_session_dt_user(_app_now()),
            error="Times must be in the format YYYY-MM-DD HH:MM:SS.",
            end_noun=end_noun,
            stats=None,
        )

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            existing = cur.execute(
                "SELECT session_begin_time FROM session_times "
                "WHERE session_id = %s AND user_id = %s",
                (session_id, user_id)
            ).fetchone()
            # Recovery path: the session_times row should exist (we insert
            # it on first /sesh GET), but if it's missing — e.g. someone
            # imported apollo rows manually — back-fill begin_time from
            # the earliest shot's timestamp so stats can still compute.
            if existing is None:
                # Verify the user actually owns this session_id before
                # creating a session_times row for it. Without this check a
                # crafted form could end another user's session — though
                # the get_stats() call after would refuse to read it.
                owner_row = cur.execute(
                    "SELECT 1 FROM apollo WHERE session_id = %s AND user_id = %s LIMIT 1",
                    (session_id, user_id)
                ).fetchone()
                if owner_row is None:
                    return "Session not found", 404
                print(f"⚠️ session_times row missing for session {session_id} — reconstructing")
                cur.execute(
                    "INSERT INTO session_times "
                    "(user_id, session_id, session_begin_time, session_end_time, "
                    "manual_session_length_minutes) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (user_id, session_id, parsed_begin, parsed_end, None)
                )
            else:
                cur.execute(
                    "UPDATE session_times SET session_begin_time = %s, "
                    "session_end_time = %s, manual_session_length_minutes = %s "
                    "WHERE session_id = %s AND user_id = %s",
                    (parsed_begin, parsed_end, None, session_id, user_id)
                )
            con.commit()
    except SQLAlchemyError as e:
        print(f"❌ Error writing session end time: {e}")
        return "Error ending session", 500

    stats = get_stats(session_id, user_id)
    if isinstance(stats, tuple):
        return stats
    stats["session_id"] = session_id
    # A live match also has the other archers' sessions open — close them.
    _finalize_match_sessions(user_id)
    # Clear the in-progress keys but keep the user logged in.
    for k in ('session_id', 'quivers_completed', 'arrows_remaining',
              'record_mode', 'target_id', 'current_quiver_size',
                  'tournament_round_key', 'tournament_segment_idx', 'tournament_practice',
                  'tournament_bowstyle',
                  'match_id', 'match_archers', 'match_idx'):
        session.pop(k, None)
    return render_template('end_session.html', stats=stats, session_id=None,
                           begin_time=None, end_noun=end_noun)


@app.route('/end_session_silent', methods=['POST'])
@login_required
def end_session_silent():
    """End the current session without showing stats, then redirect to ``next``.

    Called by the session page's leave-warning modal when the user picks
    "Go to link" — we close the session server-side so subsequent shots
    don't get attributed to a session the user has walked away from.
    ``next`` must be a same-origin path; anything else falls back to '/'
    to avoid open-redirect abuse.
    """
    user_id = current_user_id()
    # Open-redirect guard: accept only same-origin paths. urlparse handles
    # the easy cases (scheme/netloc present), and we also reject leading
    # backslashes and "//" which some browsers normalize into a host.
    next_url = (request.form.get('next') or '/').strip()
    parsed = urlparse(next_url)
    if (parsed.scheme or parsed.netloc
            or not next_url.startswith('/')
            or next_url.startswith('//')
            or next_url.startswith('/\\')):
        next_url = '/'

    session_id = session.get('session_id')
    if session_id is not None:
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                shot_row = cur.execute(
                    "SELECT timestamp FROM apollo WHERE session_id = %s AND user_id = %s LIMIT 1",
                    (session_id, user_id)
                ).fetchone()
                if shot_row is None:
                    # Zero-shot session — discard the session_times row so the
                    # previous-sessions list doesn't fill with empty entries.
                    cur.execute(
                        "DELETE FROM session_times WHERE session_id = %s AND user_id = %s",
                        (session_id, user_id)
                    )
                else:
                    existing = cur.execute(
                        "SELECT session_begin_time FROM session_times "
                        "WHERE session_id = %s AND user_id = %s",
                        (session_id, user_id)
                    ).fetchone()
                    if existing is None:
                        earliest = cur.execute(
                            "SELECT MIN(timestamp) FROM apollo "
                            "WHERE session_id = %s AND user_id = %s",
                            (session_id, user_id)
                        ).fetchone()
                        begin_time = earliest[0] if earliest and earliest[0] else _app_now()
                        cur.execute(
                            "INSERT INTO session_times "
                            "(user_id, session_id, session_begin_time, session_end_time) "
                            "VALUES (%s, %s, %s, %s)",
                            (user_id, session_id, begin_time, _app_now())
                        )
                    else:
                        cur.execute(
                            "UPDATE session_times SET session_end_time = %s "
                            "WHERE session_id = %s AND user_id = %s",
                            (_app_now(), session_id, user_id)
                        )
                con.commit()
        except SQLAlchemyError as e:
            print(f"❌ Silent end-session error: {e}")
    # A live match also has the other archers' sessions open — close them.
    _finalize_match_sessions(user_id)
    # Drop per-round keys but keep the user logged in. session.clear()
    # would log them out, which isn't what the leave-warning modal means.
    for k in ('session_id', 'quivers_completed', 'arrows_remaining',
              'record_mode', 'target_id', 'current_quiver_size',
                  'tournament_round_key', 'tournament_segment_idx', 'tournament_practice',
                  'tournament_bowstyle',
                  'match_id', 'match_archers', 'match_idx'):
        session.pop(k, None)
    return redirect(next_url)


@app.route('/add_arrow', methods=['GET', 'POST'])
@login_required
def add_arrow():
    """Add a new arrow definition. GET shows the form; POST inserts a row.

    On success, redirect back into the active session if one exists,
    otherwise to the splash page.
    """
    user_id = current_user_id()
    if request.method == 'GET':
        return render_template('add_arrow.html')
    if request.method == 'POST':
        try:
            arrow_type     = request.form.get('new_arrow')
            spine          = request.form.get('new_spine')
            length         = request.form.get('new_length')
            shaft_weight   = request.form.get('new_shaft_weight')
            shaft_diameter = request.form.get('new_shaft_diameter')
            shaft_material = request.form.get('new_shaft_material')
            tip            = request.form.get('new_tip')
            tip_weight     = request.form.get('new_tip_weight')
            nock_weight    = request.form.get('new_nock_weight')

            if not arrow_type:
                return "Error: Arrow name is required", 400

            # Numeric fields are stored as text historically but downstream
            # analysis does float(...) on them — reject garbage at the door
            # so a typo doesn't corrupt later draw-weight / spine reports.
            positive_numeric = {
                'length': length,
                'spine': spine,
                'tip_weight': tip_weight,
            }
            nonneg_numeric = {
                'shaft_weight': shaft_weight,
                'shaft_diameter': shaft_diameter,
                'nock_weight': nock_weight,
            }
            for name, val in positive_numeric.items():
                if val not in (None, '') and _parse_float(val) is None:
                    return f"Error: {name} must be a positive number", 400
            for name, val in nonneg_numeric.items():
                if val not in (None, '') and _parse_nonneg_float(val) is None:
                    return f"Error: {name} must be a number", 400

            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "INSERT INTO arrows "
                    "(user_id, arrow, length, spine, shaft_weight, shaft_diameter, "
                    "shaft_material, nock_weight, tip, tip_weight) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (user_id, arrow_type, length, spine, shaft_weight, shaft_diameter,
                     shaft_material, nock_weight, tip, tip_weight)
                )
                con.commit()
            if session.get('session_id') is None:
                return redirect(url_for('index'))
            else:
                return redirect(url_for('sesh'))
        except SQLAlchemyError as e:
            print(f"Add arrow error: {e}")
            return "Error adding arrow", 500


@app.route('/add_bow', methods=['GET', 'POST'])
@login_required
def add_bow():
    """Add a new bow definition. GET shows the form; POST inserts a row.

    On success, redirect back into the active session if one exists,
    otherwise to the splash page.
    """
    user_id = current_user_id()
    if request.method == 'GET':
        return render_template('add_bow.html',
                               archer_bowstyles=ARCHER_BOWSTYLES)
    elif request.method == 'POST':
        try:
            new_bow_model         = request.form.get('new_bow_model')
            new_bow_type          = request.form.get('new_bow_type')
            new_bow_draw_weight   = request.form.get('new_bow_draw_weight')
            new_effective_dw      = request.form.get('new_effective_draw_weight')
            new_bow_amo           = request.form.get('new_bow_amo')

            if not new_bow_model:
                return "Error: New bow model field required", 400

            # Reject non-numeric draw weights etc. so analyze's float() calls
            # don't blow up later. Effective draw weight is an intrinsic per-
            # archer property of the bow (draw length rarely changes), so it
            # lives on the bow record and is snapshotted onto each shot.
            if new_bow_draw_weight not in (None, '') and _parse_float(new_bow_draw_weight) is None:
                return "Error: bow draw weight must be a positive number", 400
            if new_effective_dw not in (None, '') and _parse_float(new_effective_dw) is None:
                return "Error: effective draw weight must be a positive number", 400
            if new_bow_amo not in (None, '') and _parse_float(new_bow_amo) is None:
                return "Error: AMO must be a positive number", 400

            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "INSERT INTO bows (user_id, bow_model, bow_type, bow_draw_weight, "
                    "effective_draw_weight, amo) "
                    "VALUES (%s, %s, %s, %s, %s, %s)",
                    (user_id, new_bow_model, new_bow_type, new_bow_draw_weight,
                     new_effective_dw, new_bow_amo)
                )
                con.commit()
            if session.get('session_id') is None:
                return redirect(url_for('index'))
            else:
                return redirect(url_for('sesh'))
        except SQLAlchemyError as e:
            print(f"Add bow error: {e}")
            return "Error adding bow", 500


@app.route('/edit_bows', methods=['GET', 'POST'])
@login_required
def edit_bows():
    """List/edit/delete bows. POST with ``delete`` removes a row; POST
    without it updates the row identified by ``rowid``.

    All queries are scoped to current_user — so a crafted form supplying
    another user's bow ``rowid`` simply updates/deletes zero rows.
    """
    user_id = current_user_id()
    if request.method == 'GET':
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                res = cur.execute(
                    "SELECT DISTINCT bow_model FROM bows WHERE user_id = %s",
                    (user_id,)
                ).fetchall()
                bow_models = [row[0] for row in res] if res else []
                bow_data = []
                for bow in bow_models:
                    rows = cur.execute(
                        "SELECT id AS rowid, bow_model, bow_type, bow_draw_weight, "
                        "effective_draw_weight, amo "
                        "FROM bows WHERE bow_model = %s AND user_id = %s",
                        (bow, user_id)
                    ).fetchall()
                    if rows:
                        bow_data.append(list(rows))
                num_records = len(bow_data)
            return render_template('edit_bows.html',
                                   bow_models=bow_models,
                                   bow_data=bow_data,
                                   num_records=num_records)
        except Exception as e:
            print(f"Retrieving bows error: {e}")
            return "Error retrieving bows", 500

    if request.method == 'POST':
        try:
            rowid = int(request.form.get('rowid', ''))
        except (TypeError, ValueError):
            flash("Invalid bow id.")
            return redirect(url_for('edit_bows'))
        if "delete" in request.form:
            try:
                with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                    cur.execute(
                        "DELETE FROM bows WHERE id = %s AND user_id = %s",
                        (rowid, user_id)
                    )
                    con.commit()
                print(f"🐒 Bow removed: {rowid}")
            except Exception as e:
                print(f"Delete of bow error: {e}")
                return "Error deleting bow", 500
            # Mirror the edit_arrows pattern: redirect after delete so the
            # update branch below can't be re-entered by accident.
            return redirect(url_for('edit_bows'))
        else:
            bow_model             = request.form.get('bow_model')
            bow_draw_weight       = request.form.get('bow_draw_weight')
            effective_dw          = request.form.get('effective_draw_weight')
            bow_amo               = request.form.get('bow_amo')
            # NB: ``bow_type`` is intentionally *not* in the UPDATE — it's
            # an immutable property of a bow (longbow vs. recurve vs.
            # compound is a kind, not a tuning). The form renders it as a
            # read-only display so an attacker can't smuggle it in by
            # crafting a POST either. Bowtype-specific gear isn't on the bow
            # anymore — it's tuned per session in the gear collapsible.
            try:
                with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                    cur.execute(
                        "UPDATE bows SET bow_model = %s, bow_draw_weight = %s, "
                        "effective_draw_weight = %s, amo = %s "
                        "WHERE id = %s AND user_id = %s",
                        (bow_model, bow_draw_weight,
                         effective_dw, bow_amo, rowid, user_id)
                    )
                    con.commit()
            except Exception as e:
                print(f"Update of bow error: {e}")
                return "Error updating bow", 500

        # POST/redirect/GET so a page reload doesn't resubmit the form.
        return redirect(url_for('edit_bows'))


@app.route('/edit_arrows', methods=['GET', 'POST'])
@login_required
def edit_arrows():
    """List/edit/delete arrows. POST with ``delete`` removes a row; POST
    without it updates the row identified by ``rowid``.

    All queries are scoped to current_user — see edit_bows for the same
    pattern.
    """
    user_id = current_user_id()
    if request.method == 'GET':
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                res = cur.execute(
                    "SELECT DISTINCT arrow FROM arrows WHERE user_id = %s",
                    (user_id,)
                ).fetchall()
                arrows = [row[0] for row in res] if res else []
                arrow_data = []
                for arrow in arrows:
                    rows = cur.execute(
                        "SELECT id AS rowid, arrow, length, spine, shaft_weight, "
                        "shaft_diameter, shaft_material, nock_weight, "
                        "tip, tip_weight FROM arrows WHERE arrow = %s AND user_id = %s",
                        (arrow, user_id)
                    ).fetchall()
                    if rows:
                        arrow_data.append(list(rows))
                num_records = len(arrow_data)
            return render_template('edit_arrows.html',
                                   arrows=arrows,
                                   arrow_data=arrow_data,
                                   num_records=num_records)
        except Exception as e:
            print(f"Retrieving arrows error: {e}")
            return "Error retrieving arrows", 500

    if request.method == 'POST':
        try:
            rowid = int(request.form.get('rowid', ''))
        except (TypeError, ValueError):
            flash("Invalid arrow id.")
            return redirect(url_for('edit_arrows'))

        # FIX: same delete-falls-through-to-update bug as edit_bows
        if "delete" in request.form:
            try:
                with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                    cur.execute(
                        "DELETE FROM arrows WHERE id = %s AND user_id = %s",
                        (rowid, user_id)
                    )
                    con.commit()
                    print(f"🐒 DELETE ARROW {rowid}")
            except Exception as e:
                print(f"Delete of arrow error: {e}")
                return "Error deleting arrow", 500
            return redirect(url_for('edit_arrows'))

        arrow          = request.form.get('arrow')
        length         = request.form.get('length')
        shaft_weight   = request.form.get('shaft_weight')
        shaft_diameter = request.form.get('shaft_diameter')
        shaft_material = request.form.get('shaft_material')
        nock_weight    = request.form.get('nock_weight')
        tip            = request.form.get('tip')
        tip_weight     = request.form.get('tip_weight')
        # NB: ``spine`` is intentionally *not* in the UPDATE — spine is an
        # intrinsic, immutable property of a shaft (a 500-spine doesn't
        # become a 600 because you changed the tip). The form renders it
        # as read-only display so a crafted POST can't smuggle it in either.
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "UPDATE arrows SET arrow = %s, length = %s, shaft_weight = %s, "
                    "shaft_diameter = %s, shaft_material = %s, "
                    "nock_weight = %s, tip = %s, tip_weight = %s "
                    "WHERE id = %s AND user_id = %s",
                    (arrow, length, shaft_weight, shaft_diameter, shaft_material,
                     nock_weight, tip, tip_weight,
                     rowid, user_id)
                )
                con.commit()
        except Exception as e:
            print(f"Update of arrow error: {e}")
            return "Error updating arrow", 500

        # POST/redirect/GET so a page reload doesn't resubmit the form.
        return redirect(url_for('edit_arrows'))


def _safe_target_filename(original):
    """Return a sanitized, unique filename for an uploaded target image.

    Strips any path components from the user-supplied name, keeps only the
    extension (validated against the allowlist), and prefixes a UUID to
    sidestep collisions and path-traversal attempts.
    """
    base = os.path.basename(original or '')
    _, ext = os.path.splitext(base.lower())
    if ext not in ALLOWED_IMAGE_EXTS:
        return None
    safe_stem = re.sub(r'[^A-Za-z0-9_-]', '_', os.path.splitext(base)[0])[:40] or 'target'
    return f"{uuid.uuid4().hex[:8]}_{safe_stem}{ext}"


def _parse_float(s):
    try:
        v = float(s)
        return v if v > 0 else None
    except (TypeError, ValueError):
        return None


def _parse_nonneg_float(s):
    # Coordinate fields legitimately include 0 (clicking the top-left pixel),
    # which _parse_float would reject. Keep _parse_float strict for size-like
    # inputs that must be > 0.
    try:
        v = float(s)
        return v if v >= 0 else None
    except (TypeError, ValueError):
        return None


@app.route('/add_target', methods=['GET', 'POST'])
@login_required
def add_target():
    """Add a new target via the calibration wizard.

    GET renders the four-step wizard. POST receives the uploaded image
    plus the wizard's calibration coords (center, boundary point, and two
    scale points + real-world distance), validates and crops the image to
    a square around the calibrated center, and stores the result.
    """
    user_id = current_user_id()
    if request.method == 'GET':
        return render_template('add_target.html')

    name     = (request.form.get('name') or '').strip()
    make_def = 1 if request.form.get('make_default') else 0
    upload   = request.files.get('image')

    if not name or upload is None or not upload.filename:
        return "Error: name and image file are required", 400

    # Calibration coordinates are captured by the wizard in the original
    # image's natural pixel space — center + boundary define the square crop,
    # and two arbitrary points + a real-world distance define the scale.
    coords = {}
    for key in ('center_x', 'center_y', 'boundary_x', 'boundary_y',
                'p1_x', 'p1_y', 'p2_x', 'p2_y', 'mm_distance'):
        v = _parse_float(request.form.get(key)) if key == 'mm_distance' \
            else _parse_nonneg_float(request.form.get(key))
        if v is None:
            return f"Error: calibration field '{key}' missing or invalid", 400
        coords[key] = v

    safe_name = _safe_target_filename(upload.filename)
    if safe_name is None:
        return "Error: image must be jpg, png, or webp", 400

    blob = upload.read(MAX_UPLOAD_BYTES + 1)
    if len(blob) > MAX_UPLOAD_BYTES:
        return f"Error: image exceeds {MAX_UPLOAD_BYTES // (1024*1024)} MB limit", 400

    # Decode in-memory first so we can validate crop bounds *before* writing
    # anything to disk — avoids leaving a stray upload on a 400. The pixel-
    # count check between verify() and load() defends against decompression
    # bombs: a 5 MB JPEG can advertise a 100k×100k frame that explodes to
    # tens of GB once load() materializes the pixel buffer. PIL's default
    # MAX_IMAGE_PIXELS is ~89 MP and only emits a warning, so we enforce a
    # tighter ceiling ourselves and bail before allocating.
    try:
        with Image.open(io.BytesIO(blob)) as im:
            im.verify()
        with Image.open(io.BytesIO(blob)) as im:
            w, h = im.size
            if w <= 0 or h <= 0 or w * h > MAX_IMAGE_PIXELS:
                return (f"Error: image is too large "
                        f"({w}×{h} exceeds the "
                        f"{MAX_IMAGE_PIXELS // 1_000_000} MP limit)"), 400
            im.load()
            src = im.copy()
    except Exception as e:
        print(f"Add target — image read error: {e}")
        return "Error: uploaded file is not a valid image", 400

    cx, cy = coords['center_x'], coords['center_y']
    bx, by = coords['boundary_x'], coords['boundary_y']
    radius_px = math.hypot(bx - cx, by - cy)
    if radius_px < 5:
        return "Error: target radius is too small — re-click the boundary further from center", 400

    left, top    = cx - radius_px, cy - radius_px
    right, bot   = cx + radius_px, cy + radius_px
    if left < 0 or top < 0 or right > w or bot > h:
        return ("Error: the resulting crop extends past the image edge. "
                "Re-click with the center and boundary further inside the photo."), 400

    pixel_dist = math.hypot(coords['p2_x'] - coords['p1_x'],
                            coords['p2_y'] - coords['p1_y'])
    if pixel_dist < 5:
        return "Error: the two calibration points are too close together — pick points further apart", 400

    mm_per_pixel     = coords['mm_distance'] / pixel_dist
    image_size_px    = int(round(2 * radius_px))
    physical_size_mm = (2 * radius_px) * mm_per_pixel

    os.makedirs(TARGETS_DIR, exist_ok=True)
    disk_path = os.path.join(TARGETS_DIR, safe_name)
    try:
        cropped = src.crop((int(round(left)), int(round(top)),
                            int(round(right)), int(round(bot))))
        cropped.save(disk_path)
    except Exception as e:
        try: os.remove(disk_path)
        except OSError: pass
        print(f"Add target — crop/save error: {e}")
        return "Error: failed to crop image", 500

    rel_path = f"{TARGETS_SUBDIR}/{safe_name}"
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            # Per-user uniqueness check, since the DB-level UNIQUE on
            # targets.name is no longer in play (or, on a pre-multi-user
            # SQLite DB, it's still there but globally — which would block
            # this user from picking a name another user already used,
            # giving a confusing error). Explicit check + matching message
            # is friendlier either way.
            existing = cur.execute(
                "SELECT 1 FROM targets WHERE name = %s AND user_id = %s LIMIT 1",
                (name, user_id)
            ).fetchone()
            if existing is not None:
                try: os.remove(disk_path)
                except OSError: pass
                return f"Error: a target named '{name}' already exists", 400
            if make_def:
                cur.execute(
                    "UPDATE targets SET is_default = 0 WHERE user_id = %s",
                    (user_id,)
                )
            cur.execute(
                "INSERT INTO targets "
                "(user_id, name, image_filename, physical_size_mm, image_size_px, "
                "is_active, is_default) "
                "VALUES (%s, %s, %s, %s, %s, 1, %s)",
                (user_id, name, rel_path, physical_size_mm, image_size_px, make_def)
            )
            # lastrowid is the canonical post-INSERT id on both SQLite and
            # MySQL. Fall back to a name lookup if the driver doesn't expose
            # it (vanishingly rare here, but cheaper than aborting the wizard).
            new_target_id = cur.lastrowid
            if not new_target_id:
                row = cur.execute(
                    "SELECT id FROM targets WHERE user_id = %s AND name = %s LIMIT 1",
                    (user_id, name)
                ).fetchone()
                new_target_id = int(row[0]) if row is not None else None
            con.commit()
    except DBIntegrityError:
        # Legacy global UNIQUE on targets.name (pre-multi-user DBs) can
        # still fire here. Translate to the same per-user message — the
        # user doesn't need to know about the historical constraint.
        try: os.remove(disk_path)
        except OSError: pass
        return f"Error: a target named '{name}' already exists", 400
    except SQLAlchemyError as e:
        print(f"Add target — DB error: {e}")
        try: os.remove(disk_path)
        except OSError: pass
        return "Error saving target", 500

    # Final step of the add-target wizard is scoring-zone setup. Carry the
    # post-wizard destination through as ?next= so the zones page can return
    # the user where they were headed once they Save or Skip.
    next_dest = 'index' if session.get('session_id') is None else 'sesh'
    if new_target_id is None:
        return redirect(url_for(next_dest))
    return redirect(url_for('target_zones',
                            target_id=new_target_id, next=next_dest))


@app.route('/edit_targets', methods=['GET', 'POST'])
@login_required
def edit_targets():
    """List/edit/delete targets, scoped to the signed-in user.

    Deletes are soft when shots reference the target (is_active=0) so the
    replay image stays available; hard delete + image-file removal happen
    only when nothing references the row.
    """
    user_id = current_user_id()
    if request.method == 'POST':
        rowid = request.form.get('rowid')
        if 'delete' in request.form:
            try:
                with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                    # Soft-delete (is_active=0) when shots reference it —
                    # hard-deleting would 404 the replay image. Hard-delete
                    # only when nothing depends on it.
                    #
                    # The race we're avoiding here: concurrent /sesh tab from
                    # the same user inserts an apollo row referencing this
                    # target between our COUNT and DELETE, leaving an orphan.
                    # On MySQL we lock the target row with FOR UPDATE so
                    # concurrent writes that read the target serialize behind
                    # us; on SQLite the database is single-writer so the
                    # race window doesn't exist. As a final belt-and-braces
                    # check we recount after the deletes and roll back if
                    # any reference snuck in.
                    is_mysql = (engine.dialect.name == 'mysql')
                    if is_mysql:
                        cur.execute(
                            "SELECT 1 FROM targets WHERE id = %s AND user_id = %s FOR UPDATE",
                            (rowid, user_id)
                        )
                    refs = cur.execute(
                        "SELECT COUNT(*) FROM apollo WHERE target_id = %s AND user_id = %s",
                        (rowid, user_id)
                    ).fetchone()[0]
                    if refs > 0:
                        cur.execute(
                            "UPDATE targets SET is_active = 0, is_default = 0 "
                            "WHERE id = %s AND user_id = %s",
                            (rowid, user_id)
                        )
                    else:
                        row = cur.execute(
                            "SELECT image_filename FROM targets WHERE id = %s AND user_id = %s",
                            (rowid, user_id)
                        ).fetchone()
                        # Drop any zones first — no FK cascade is declared,
                        # so an orphaned target_zones row would otherwise
                        # linger forever pointing at a deleted target.
                        cur.execute(
                            "DELETE FROM target_zones WHERE target_id = %s AND user_id = %s",
                            (rowid, user_id)
                        )
                        cur.execute(
                            "DELETE FROM targets WHERE id = %s AND user_id = %s",
                            (rowid, user_id)
                        )
                        # Recount inside the same transaction: if a racing
                        # request slipped a shot in pointing at this target,
                        # back out and soft-delete instead so we don't strand
                        # the apollo row with a dangling target_id.
                        post_refs = cur.execute(
                            "SELECT COUNT(*) FROM apollo WHERE target_id = %s AND user_id = %s",
                            (rowid, user_id)
                        ).fetchone()[0]
                        if post_refs > 0:
                            con.rollback()
                            with closing(get_db_connection()) as con2, closing(con2.cursor()) as cur2:
                                cur2.execute(
                                    "UPDATE targets SET is_active = 0, is_default = 0 "
                                    "WHERE id = %s AND user_id = %s",
                                    (rowid, user_id)
                                )
                                con2.commit()
                            return redirect(url_for('edit_targets'))
                        # Only remove the image file for uploads — never the
                        # bundled legacy target.jpg, which lives at the
                        # static/ root and may still be referenced as fallback.
                        # _resolve_target_image_disk_path enforces both that
                        # rule and a realpath-based traversal guard, so a
                        # crafted image_filename ('targets/../foo') cannot
                        # reach files outside the uploads directory.
                        # Also, we must NOT delete an image file that another
                        # user still references (e.g. if the user uploaded a
                        # target and somehow it ended up shared — paranoid
                        # belt-and-braces check).
                        disk = _resolve_target_image_disk_path(
                            row['image_filename'] if row else None
                        )
                        if disk is not None:
                            other_ref = cur.execute(
                                "SELECT 1 FROM targets WHERE image_filename = %s LIMIT 1",
                                (row['image_filename'],)
                            ).fetchone()
                            if other_ref is None:
                                try: os.remove(disk)
                                except OSError: pass
                    con.commit()
            except SQLAlchemyError as e:
                print(f"Delete target error: {e}")
                return "Error deleting target", 500
        else:
            name     = (request.form.get('name') or '').strip()
            size_mm  = _parse_float(request.form.get('physical_size_mm'))
            make_def = 1 if request.form.get('make_default') else 0
            if not name or size_mm is None:
                return "Error: name and physical size are required", 400
            try:
                with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                    # Per-user name-uniqueness check (excluding the row being
                    # edited so saving an unchanged name doesn't blow up).
                    dup = cur.execute(
                        "SELECT 1 FROM targets WHERE name = %s AND user_id = %s "
                        "AND id <> %s LIMIT 1",
                        (name, user_id, rowid)
                    ).fetchone()
                    if dup is not None:
                        return f"Error: a target named '{name}' already exists", 400
                    if make_def:
                        cur.execute(
                            "UPDATE targets SET is_default = 0 WHERE user_id = %s",
                            (user_id,)
                        )
                    cur.execute(
                        "UPDATE targets SET name = %s, physical_size_mm = %s, is_default = %s "
                        "WHERE id = %s AND user_id = %s",
                        (name, size_mm, make_def, rowid, user_id)
                    )
                    con.commit()
            except DBIntegrityError:
                return f"Error: a target named '{name}' already exists", 400
            except SQLAlchemyError as e:
                print(f"Update target error: {e}")
                return "Error updating target", 500
        return redirect(url_for('edit_targets'))

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT id AS rowid, name, image_filename, physical_size_mm, image_size_px, "
                "is_active, is_default FROM targets "
                "WHERE user_id = %s ORDER BY is_active DESC, name",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError as e:
        print(f"Retrieving targets error: {e}")
        return "Error retrieving targets", 500
    return render_template('edit_targets.html', targets=rows)


def _fetch_target_zones(target_id, user_id):
    """Return all zones for one target, ordered innermost-out.

    Sorted by radius_mm so the renderer can walk smallest → largest and
    apply the highest-point ring first when scoring a shot.
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        return cur.execute(
            "SELECT id AS rowid, name, point_value, shape_type, radius_mm, display_order "
            "FROM target_zones WHERE target_id = %s AND user_id = %s "
            "ORDER BY radius_mm ASC, id ASC",
            (target_id, user_id)
        ).fetchall()


def _sample_zone_colors(zones, target_cfg, fallback_count=None):
    """Return a list of hex color strings sampled from the target image,
    one per zone, ordered to match ``zones`` (innermost ring first).

    For each ring we sample the pixel at the mid-radius between this
    ring and the next inward one, along the +y axis (top of the face).
    The sample is averaged over a small neighbourhood to smooth out
    JPEG noise. Returns a synthetic warm-to-cool gradient if the image
    can't be loaded so the report still renders.
    """
    n_rings = fallback_count if fallback_count is not None else len(zones)

    def _fallback_gradient(n):
        # Synthetic warm→cool ramp used only when the image can't be
        # sampled. ``t`` runs 0→1 across the rings; the channel slopes
        # (R 255→145, G 170→110, B 40→190) walk a warm gold toward a cool
        # blue so adjacent bars stay visually distinct.
        out = []
        for i in range(n):
            t = i / max(n - 1, 1)
            r = int(255 - 110 * t)
            g = int(170 - 60 * t)
            b = int(40 + 150 * t)
            out.append(f'#{r:02x}{g:02x}{b:02x}')
        return out

    if not zones:
        return _fallback_gradient(n_rings)

    img_disk_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 'static',
        target_cfg.get('target_image', '')
    )
    try:
        from PIL import Image as PILImage
        bg = PILImage.open(img_disk_path).convert('RGB')
    except (FileNotFoundError, OSError, ValueError):
        return _fallback_gradient(n_rings)

    width_px, height_px = bg.size
    half_mm = float(target_cfg['target_width_mm']) / 2.0
    if half_mm <= 0:
        return _fallback_gradient(n_rings)
    px_per_mm = (width_px / 2.0) / half_mm
    cx_px = width_px / 2.0
    cy_px = height_px / 2.0

    # zones are sorted innermost-out; iterate the same way so colors
    # line up with the bar order (innermost first).
    colors = []
    inner_radius_mm = 0.0
    for z in zones:
        try:
            outer_radius_mm = float(z['radius_mm'])
        except (TypeError, ValueError, KeyError):
            colors.append(None)
            continue
        # Sample 85% of the way from the inner ring boundary to the
        # outer one — close enough to the outer edge to avoid the X mark
        # / central cross / spider that often sits at the bullseye, but
        # still inside the ring band that owns this radius.
        mid_radius_mm = inner_radius_mm + 0.85 * (outer_radius_mm - inner_radius_mm)
        inner_radius_mm = outer_radius_mm

        # Sample along +y (above center). PIL's y-axis points down, so
        # +y in target space → smaller pixel-y. Average a 3×3 patch for
        # JPEG noise robustness.
        sample_y_px = cy_px - mid_radius_mm * px_per_mm
        sample_x_px = cx_px
        if not (0 <= sample_x_px < width_px and 0 <= sample_y_px < height_px):
            colors.append(None)
            continue
        r_sum = g_sum = b_sum = 0
        samples = 0
        for dy in (-1, 0, 1):
            for dx in (-1, 0, 1):
                px = int(round(sample_x_px + dx))
                py = int(round(sample_y_px + dy))
                if 0 <= px < width_px and 0 <= py < height_px:
                    pr, pg, pb = bg.getpixel((px, py))
                    r_sum += pr
                    g_sum += pg
                    b_sum += pb
                    samples += 1
        if samples == 0:
            colors.append(None)
            continue
        r = r_sum // samples
        g = g_sum // samples
        b = b_sum // samples
        colors.append(f'#{r:02x}{g:02x}{b:02x}')

    # Patch any None entries with the fallback gradient at the same index
    # so the bar chart never gets a missing color.
    fg = _fallback_gradient(n_rings)
    return [c if c is not None else fg[i] for i, c in enumerate(colors)]


def _zones_define_scoring(zones):
    """True when this set of zones can be used to tally points.

    Requires at least one zone and a non-null integer point_value on
    every zone. A zone with point_value=0 (e.g. a "white" outer ring)
    is allowed as long as *some* zone has a positive value — otherwise
    every shot would score zero and the tally would be meaningless.
    """
    if not zones:
        return False
    try:
        values = [int(z['point_value']) for z in zones
                  if z['point_value'] is not None]
    except (TypeError, ValueError):
        return False
    if len(values) != len(zones):
        return False
    return any(v > 0 for v in values)


def _parse_shaft_diameter_mm(raw):
    """Return a usable shaft diameter (mm) from a stored snapshot value.

    Falls back to ``DEFAULT_SHAFT_DIAMETER_MM`` when ``raw`` is missing,
    blank, non-numeric, or non-positive — so a single arrow without a
    diameter recorded doesn't silently disable line-cutter scoring.
    """
    try:
        d = float(str(raw).strip())
    except (TypeError, ValueError, AttributeError):
        return DEFAULT_SHAFT_DIAMETER_MM
    if d <= 0:
        return DEFAULT_SHAFT_DIAMETER_MM
    return d


def _classify_shot(xraw, yraw, zones, shaft_diameter_mm=None,
                   spot_centers_mm=None):
    """Return the zone index a shot lands in, or ``None`` for a miss.

    Innermost zone is index 0. A return of ``None`` covers both the
    sentinel miss and a hit that falls outside the outermost zone (with
    line-cutter slack applied). Line-cutter rule: any part of the shaft
    crossing or touching a ring counts the shot in that ring, so the
    effective distance from center is ``hypot(x, y) - shaft_radius``.

    ``zones`` must be sorted innermost-out.

    ``spot_centers_mm`` — when given, the face has multiple identical
    spots (e.g. NFAA 5-spot). The shot is scored against the nearest
    spot center; effective distance is measured from that spot, so a
    click anywhere on the face lands on the closest scoring spot.
    """
    if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
        return None
    try:
        x = float(xraw)
        y = float(yraw)
    except (TypeError, ValueError):
        return None
    shaft_radius = _parse_shaft_diameter_mm(shaft_diameter_mm) / 2.0
    if spot_centers_mm:
        # Distance from each spot center; smallest wins. The classifier
        # then runs against that spot's local zones (same `zones` for
        # every spot on these faces — they're identical by definition).
        best = None
        for (sx, sy) in spot_centers_mm:
            d = math.sqrt((x - sx) ** 2 + (y - sy) ** 2)
            if best is None or d < best:
                best = d
        dist = (best or 0.0) - shaft_radius
    else:
        dist = math.sqrt(x * x + y * y) - shaft_radius
    if dist < 0:
        dist = 0.0
    for i, z in enumerate(zones):
        # A malformed radius_mm (NULL, blank, or non-numeric) shouldn't
        # crash the entire quiver-score render — skip the bad zone and
        # keep scanning outward instead.
        try:
            r = float(z['radius_mm'])
        except (TypeError, ValueError):
            continue
        if dist <= r:
            return i
    return None


def _score_one_shot(xraw, yraw, zones, shaft_diameter_mm=None,
                    spot_centers_mm=None):
    """Points for a single shot. Misses and out-of-zone hits score 0."""
    idx = _classify_shot(xraw, yraw, zones, shaft_diameter_mm,
                         spot_centers_mm=spot_centers_mm)
    if idx is None:
        return 0
    try:
        return int(zones[idx]['point_value'] or 0)
    except (TypeError, ValueError):
        return 0


def _target_multi_spot_centers(target_id, user_id):
    """Return spot centers (mm) for a multi-spot tournament face, else None.

    Looks up the target row's name, reverse-maps it to a tournament
    face_key, and pulls the centers from TOURNAMENT_FACES. Used by
    scoring/filter paths so a click anywhere on a multi-spot face
    (e.g. NFAA 5-spot) scores against the nearest spot.
    """
    if target_id is None or user_id is None:
        return None
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT name FROM targets WHERE id = %s AND user_id = %s LIMIT 1",
                (target_id, user_id)
            ).fetchone()
    except SQLAlchemyError:
        return None
    if not row:
        return None
    name = row['name'] if 'name' in row else row[0]
    key = _tournament_face_key_for_target_name(name)
    if not key:
        return None
    face = TOURNAMENT_FACES.get(key)
    if not face:
        return None
    ms = face.get('multi_spot') or {}
    centers = ms.get('centers_mm')
    return list(centers) if centers else None


def _shot_effective_draw_weight(shot_row):
    """Return the draw weight to use for analysis on an apollo (shot) row.

    Rule: prefer ``effective_draw_weight`` (the user's actual draw weight)
    when it's set, otherwise fall back to ``bow_draw_weight`` (the rated
    weight). Returns ``None`` when neither is known — typically a pre-
    snapshot historical shot (see _ensure_shot_snapshot_columns); the
    caller is responsible for deciding whether to drop the row or fall
    back to the *current* bows.bow_draw_weight (which would re-introduce
    the "edit rewrites history" problem this snapshotting fixes).

    Accepts any mapping-like row (dict, sqlalchemy Row) and returns a
    float when parseable, ``None`` otherwise.
    """
    if shot_row is None:
        return None
    def _get(key):
        # Support: dict, sqlalchemy.Row (via ._mapping), sqlite3.Row (subscript-by-name).
        mapping = getattr(shot_row, '_mapping', None)
        if mapping is not None:
            try:
                return mapping[key]
            except (KeyError, TypeError):
                return None
        try:
            return shot_row[key]
        except (KeyError, IndexError, TypeError):
            return None
    raw = _get('effective_draw_weight')
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        raw = _get('bow_draw_weight')
    if raw is None:
        return None
    try:
        return float(str(raw).strip())
    except (TypeError, ValueError):
        return None


def _row_get(row, key, default=None):
    """Read a column from a sqlalchemy/sqlite3 Row or a plain dict."""
    mapping = getattr(row, '_mapping', None)
    if mapping is not None:
        try:
            return mapping[key]
        except (KeyError, TypeError):
            return default
    try:
        return row[key]
    except (KeyError, IndexError, TypeError):
        return default


def _shot_is_x(xraw, yraw, x_ring_radius, shaft_diameter_mm=None,
               spot_centers_mm=None):
    """Return True when a shot lands inside the inner-X ring.

    Mirrors the line-cutter slack of ``_classify_shot``: effective
    distance is ``hypot(x, y) - shaft_radius`` (or distance to the
    nearest spot center, for multi-spot faces) and the shot counts as an
    X when that is ``<= x_ring_radius``. Shared by the tournament progress
    counter and the scorecard's per-arrow label so the two can't drift.
    """
    if x_ring_radius is None or x_ring_radius <= 0:
        return False
    if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
        return False
    try:
        x = float(xraw)
        y = float(yraw)
    except (TypeError, ValueError):
        return False
    shaft_r = _parse_shaft_diameter_mm(shaft_diameter_mm) / 2.0
    if spot_centers_mm:
        d = min(math.sqrt((x - sx) ** 2 + (y - sy) ** 2)
                for (sx, sy) in spot_centers_mm)
    else:
        d = math.sqrt(x * x + y * y)
    return max(0.0, d - shaft_r) <= x_ring_radius


def _compute_quiver_score(shot_rows, zones, spot_centers_mm=None):
    """Sum points across the shots in one quiver."""
    total = 0
    for r in shot_rows:
        xraw = str(r['x_coord']).strip() if r['x_coord'] is not None else ''
        yraw = str(r['y_coord']).strip() if r['y_coord'] is not None else ''
        shaft = _row_get(r, 'arrow_shaft_diameter')
        total += _score_one_shot(xraw, yraw, zones, shaft,
                                 spot_centers_mm=spot_centers_mm)
    return total


@app.route('/target_zones/<int:target_id>', methods=['GET', 'POST'])
@login_required
def target_zones(target_id):
    """Manage scoring zones for one target.

    GET renders the zones editor. POST receives a JSON payload
    ``{"zones": [{"name": str, "point_value": int, "radius_mm": float}, ...]}``
    and replaces the full set for this target — simpler than diffing
    inserts/updates/deletes against arbitrary client-side reorderings.
    """
    user_id = current_user_id()
    target = get_target(target_id, user_id)
    if target is None:
        abort(404)

    if request.method == 'GET':
        zones = _fetch_target_zones(target_id, user_id)
        # ?next= is set by the add-target wizard so we can route the user
        # back to where they were going after Save/Skip. Whitelist the
        # accepted values — never trust a free-form redirect target.
        raw_next = (request.args.get('next') or '').strip()
        next_dest = raw_next if raw_next in ('sesh', 'index') else None
        next_url = url_for(next_dest) if next_dest else None
        return render_template(
            'target_zones.html',
            target=target_to_config(target),
            zones=zones,
            next_url=next_url,
        )

    payload = request.get_json(silent=True) or {}
    raw_zones = payload.get('zones')
    if not isinstance(raw_zones, list):
        return jsonify(ok=False, error="Expected a 'zones' array."), 400

    target_size_mm = float(target['physical_size_mm'] or 0)
    # Allow rings slightly past the calibrated edge — users sometimes draw a
    # "miss" ring that hugs the outside of the printed face — but cap at 2x
    # the target's edge so a stray click on the canvas can't insert a
    # nonsense radius (e.g. several meters).
    max_radius_mm = target_size_mm * 2 if target_size_mm > 0 else 1e6

    cleaned = []
    for idx, raw in enumerate(raw_zones):
        if not isinstance(raw, dict):
            return jsonify(ok=False, error="Each zone must be an object."), 400
        name = (raw.get('name') or '').strip()
        if not name:
            return jsonify(ok=False,
                           error=f"Zone {idx+1}: name is required."), 400
        try:
            radius_mm = float(raw.get('radius_mm'))
        except (TypeError, ValueError):
            return jsonify(ok=False,
                           error=f"Zone {idx+1}: radius must be a number."), 400
        if radius_mm <= 0 or radius_mm > max_radius_mm:
            return jsonify(ok=False,
                           error=f"Zone {idx+1}: radius is out of range."), 400
        try:
            point_value = int(raw.get('point_value') or 0)
        except (TypeError, ValueError):
            return jsonify(ok=False,
                           error=f"Zone {idx+1}: point value must be an integer."), 400
        shape_type = (raw.get('shape_type') or 'circle').strip().lower()
        if shape_type not in ('circle',):
            return jsonify(ok=False,
                           error=f"Zone {idx+1}: unsupported shape '{shape_type}'."), 400
        cleaned.append({
            'name': name[:255],
            'point_value': point_value,
            'shape_type': shape_type,
            'radius_mm': radius_mm,
            'display_order': idx,
        })

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            cur.execute(
                "DELETE FROM target_zones WHERE target_id = %s AND user_id = %s",
                (target_id, user_id)
            )
            for z in cleaned:
                cur.execute(
                    "INSERT INTO target_zones "
                    "(user_id, target_id, name, point_value, shape_type, "
                    "radius_mm, display_order) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (user_id, target_id, z['name'], z['point_value'],
                     z['shape_type'], z['radius_mm'], z['display_order'])
                )
            con.commit()
    except SQLAlchemyError as e:
        print(f"Save target zones error: {e}")
        return jsonify(ok=False, error="Database error saving zones."), 500

    return jsonify(ok=True, count=len(cleaned))


# Tables included in /export_data. Order matches FK-style dependencies so a
# straight INSERT replay against an empty DB stays valid (parents before
# children).
EXPORT_TABLES = ['targets', 'target_zones', 'bows', 'arrows', 'session_times', 'apollo']


def _sql_literal(v):
    """Render a Python value as a SQL literal for the export's INSERT lines."""
    if v is None:
        return 'NULL'
    if isinstance(v, bool):
        return '1' if v else '0'
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, datetime):
        return "'" + v.strftime('%Y-%m-%d %H:%M:%S.%f') + "'"
    return "'" + str(v).replace("'", "''") + "'"


def _collect_export_data(user_id):
    """Return ``{table: (columns, rows)}`` for every exported table.

    Scoped to one user — never returns another user's rows even if a
    crafted query string asked. Columns come from the metadata so an
    empty table still yields its header, keeping the export shape
    stable regardless of what's in the DB.
    """
    data = {}
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        for tbl in EXPORT_TABLES:
            cols = [c.name for c in metadata.tables[tbl].columns]
            rows = cur.execute(
                f"SELECT {', '.join(cols)} FROM {tbl} WHERE user_id = %s",
                (user_id,)
            ).fetchall()
            data[tbl] = (cols, rows)
    return data


@app.route('/export_data', methods=['GET'])
@login_required
def export_data():
    """Stream every row of every table back to the user.

    ``format=sql|csv|xlsx``: SQL is one file of INSERTs (parents first so
    replay against an empty DB works); CSV is a zip of one file per table;
    xlsx is one workbook with a sheet per table.
    """
    fmt = (request.args.get('format') or '').strip().lower()
    if fmt not in ('sql', 'csv', 'xlsx'):
        return "Error: format must be sql, csv, or xlsx", 400

    try:
        data = _collect_export_data(current_user_id())
    except SQLAlchemyError as e:
        print(f"❌ Export read error: {e}")
        return "Error reading data for export", 500

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    if fmt == 'sql':
        lines = [
            "-- Apollo data export",
            f"-- Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
        ]
        for tbl, (cols, rows) in data.items():
            lines.append(f"-- Table: {tbl} ({len(rows)} rows)")
            for r in rows:
                vals = ', '.join(_sql_literal(r[c]) for c in cols)
                lines.append(
                    f"INSERT INTO {tbl} ({', '.join(cols)}) VALUES ({vals});"
                )
            lines.append("")
        body = '\n'.join(lines)
        return Response(body, mimetype='application/sql', headers={
            'Content-Disposition': f'attachment; filename=apollo_export_{timestamp}.sql'
        })

    if fmt == 'csv':
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for tbl, (cols, rows) in data.items():
                sbuf = io.StringIO()
                w = csv.writer(sbuf)
                w.writerow(cols)
                for r in rows:
                    w.writerow([r[c] for c in cols])
                zf.writestr(f'{tbl}.csv', sbuf.getvalue())
        return Response(buf.getvalue(), mimetype='application/zip', headers={
            'Content-Disposition': f'attachment; filename=apollo_export_{timestamp}.zip'
        })

    # xlsx — imported lazily so a missing openpyxl only breaks this branch,
    # not the whole app at import time.
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    for tbl, (cols, rows) in data.items():
        ws = wb.create_sheet(title=tbl[:31])  # Excel caps sheet names at 31 chars
        ws.append(cols)
        for r in rows:
            ws.append([r[c] for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    return Response(buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename=apollo_export_{timestamp}.xlsx'
        })


# ─── Import ────────────────────────────────────────────────────────────────
# /import_data accepts the same three formats /export_data emits (sql / csv-
# zip / xlsx). Per-row strategy is "merge, not replace":
#   * `id` is always dropped — the DB autoincrements, so old ids never
#     collide with whatever the user already has.
#   * `user_id` is always overridden with the importer's id, so a file
#     exported from another account still lands as the current user's data.
#   * `session_id` is shifted by MAX(session_id) of the current user so the
#     imported sessions sit *after* whatever's already there. session_times
#     and apollo share the same offset so cross-table references stay
#     coherent.
#   * targets get a fresh id; apollo.target_id is rewritten through a map
#     of old_target_id → new_target_id built as the targets insert.

def _split_sql_statements(text_in):
    """Split a SQL dump into individual statements at unquoted semicolons.
    Handles '' inside string literals so a session_notes containing ';' or
    a newline doesn't truncate its INSERT."""
    out = []
    cur = []
    i = 0
    n = len(text_in)
    in_str = False
    while i < n:
        c = text_in[i]
        if in_str:
            cur.append(c)
            if c == "'":
                if i + 1 < n and text_in[i+1] == "'":
                    cur.append("'")
                    i += 2
                    continue
                in_str = False
            i += 1
            continue
        if c == "'":
            cur.append(c)
            in_str = True
            i += 1
            continue
        if c == ';':
            stmt = ''.join(cur).strip()
            if stmt:
                out.append(stmt)
            cur = []
            i += 1
            continue
        cur.append(c)
        i += 1
    last = ''.join(cur).strip()
    if last:
        out.append(last)
    return out


def _parse_sql_value(s, i):
    """Parse one SQL literal at s[i]; return (value, next_i)."""
    n = len(s)
    while i < n and s[i] in ' \t\r\n':
        i += 1
    if i >= n:
        raise ValueError("Unexpected end of SQL values")
    if s[i] == "'":
        i += 1
        out = []
        while i < n:
            if s[i] == "'":
                if i + 1 < n and s[i+1] == "'":
                    out.append("'")
                    i += 2
                else:
                    return ''.join(out), i + 1
            else:
                out.append(s[i])
                i += 1
        raise ValueError("Unterminated string literal")
    j = i
    while j < n and s[j] not in ',)':
        j += 1
    token = s[i:j].strip()
    if token.upper() == 'NULL':
        return None, j
    try:
        return int(token), j
    except ValueError:
        try:
            return float(token), j
        except ValueError:
            return token, j


def _parse_sql_insert(stmt):
    """Parse 'INSERT INTO t (cols) VALUES (vals);' → (table, cols, values)."""
    m = re.match(r'\s*INSERT\s+INTO\s+(\w+)\s*\(([^)]+)\)\s*VALUES\s*\(',
                 stmt, re.IGNORECASE)
    if not m:
        return None
    table = m.group(1)
    cols = [c.strip() for c in m.group(2).split(',')]
    i = m.end()
    n = len(stmt)
    values = []
    while True:
        v, i = _parse_sql_value(stmt, i)
        values.append(v)
        while i < n and stmt[i] in ' \t\r\n':
            i += 1
        if i >= n:
            raise ValueError("Unexpected end of INSERT")
        if stmt[i] == ',':
            i += 1
            continue
        if stmt[i] == ')':
            break
        raise ValueError(f"Unexpected char {stmt[i]!r} in INSERT")
    return table, cols, values


def _strip_leading_sql_comments(s):
    """Drop any number of leading '-- ...' comment lines (plus blank lines)
    so a statement that's preceded by an export's '-- Table: …' header still
    parses as the INSERT it actually is."""
    while True:
        s = s.lstrip()
        if not s.startswith('--'):
            return s
        nl = s.find('\n')
        if nl == -1:
            return ''
        s = s[nl+1:]


def _parse_sql_import(text_in):
    """Convert SQL dump text to {table: [row_dict, ...]}."""
    out = {}
    for stmt in _split_sql_statements(text_in):
        s = _strip_leading_sql_comments(stmt)
        if not s[:6].upper().startswith('INSERT'):
            continue
        parsed = _parse_sql_insert(s)
        if parsed is None:
            continue
        table, cols, values = parsed
        out.setdefault(table, []).append(dict(zip(cols, values)))
    return out


def _parse_csv_zip_import(blob):
    """Convert a zip-of-csvs blob to {table: [row_dict, ...]}.

    Zip entries are read as data only — no extraction to disk — but we
    still reject path-traversal segments and absolute paths so a crafted
    archive can't confuse the table-name parser into shadowing a real
    table (e.g. an entry named `x/../targets.csv` would otherwise resolve
    to `targets` and silently merge attacker rows into that table).
    """
    out = {}
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        for name in zf.namelist():
            if not name.lower().endswith('.csv'):
                continue
            # Block zip-slip: reject entries with parent-traversal segments,
            # absolute paths, or backslash separators (Windows paths in a
            # zip created by a hostile tool).
            if (name.startswith('/') or name.startswith('\\')
                    or '..' in name.replace('\\', '/').split('/')
                    or '\x00' in name):
                continue
            table = name.rsplit('/', 1)[-1][:-4]
            # Resulting table name must be a plain SQL identifier — the
            # _apply_import allowlist will catch it later, but cheaper to
            # skip the read here.
            if not _safe_ident(table):
                continue
            with zf.open(name) as f:
                reader = csv.DictReader(io.TextIOWrapper(f, encoding='utf-8'))
                out[table] = [dict(row) for row in reader]
    return out


def _parse_xlsx_import(blob):
    """Convert an xlsx workbook blob to {table: [row_dict, ...]}."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    out = {}
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        headers = None
        body = []
        for row in rows_iter:
            if headers is None:
                headers = [str(h) if h is not None else '' for h in row]
                continue
            body.append(dict(zip(headers, row)))
        out[ws.title] = body
    return out


def _coerce_value(table, col_name, raw):
    """Coerce a raw cell value into the type the column expects.
    Handles CSV-style stringified ints/floats/datetimes and treats '' / 'NULL'
    as None so blanks don't end up as the literal string 'NULL' in the DB."""
    if raw is None:
        return None
    if isinstance(raw, str):
        stripped = raw.strip()
        if stripped == '' or stripped.upper() == 'NULL':
            return None
    column = metadata.tables[table].columns.get(col_name)
    if column is None:
        return raw
    try:
        pytype = column.type.python_type
    except (AttributeError, NotImplementedError):
        return raw if not isinstance(raw, str) else raw
    try:
        if pytype is int:
            if isinstance(raw, bool):
                return 1 if raw else 0
            return int(float(raw)) if isinstance(raw, str) else int(raw)
        if pytype is float:
            return float(raw)
        if pytype is datetime:
            if isinstance(raw, datetime):
                return raw
            s = str(raw).strip()
            for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S',
                        '%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S',
                        '%Y-%m-%d'):
                try:
                    return datetime.strptime(s, fmt)
                except ValueError:
                    continue
            return None
        if pytype is bool:
            if isinstance(raw, str):
                return raw.strip().lower() in ('1', 'true', 't', 'yes', 'y')
            return bool(raw)
    except (TypeError, ValueError):
        return None
    return str(raw) if not isinstance(raw, str) else raw


IMPORT_TABLES = ['targets', 'target_zones', 'bows', 'arrows', 'session_times', 'apollo']

# Defense-in-depth: imported column names are already intersected against
# the SQLAlchemy schema (valid_cols), but the INSERT below interpolates
# them as identifiers (not as bind parameters — DBAPI can't bind names).
# Reject anything that doesn't look like a bare SQL identifier before it
# reaches the f-string, so a future refactor that loosens valid_cols can't
# accidentally turn this into an injection vector.
_SAFE_SQL_IDENT_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')


def _safe_ident(name):
    return isinstance(name, str) and bool(_SAFE_SQL_IDENT_RE.match(name))


def _apply_import(data, user_id):
    """Merge parsed rows into the current user's data.
    Returns {table: inserted_count}.

    The whole import runs in a single transaction: a failure mid-loop
    triggers a rollback so the caller never sees a partially-imported
    state where (say) targets landed but their zones didn't.
    """
    counts = {t: 0 for t in IMPORT_TABLES}
    target_id_map = {}

    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
      try:
        # session_id offset — pick max across both tables so apollo and
        # session_times stay aligned even if one is ahead of the other.
        res = cur.execute(
            "SELECT MAX(session_id) FROM apollo WHERE user_id = %s", (user_id,)
        ).fetchone()
        max_apollo = (res[0] if res and res[0] is not None else 0) or 0
        res = cur.execute(
            "SELECT MAX(session_id) FROM session_times WHERE user_id = %s",
            (user_id,)
        ).fetchone()
        max_st = (res[0] if res and res[0] is not None else 0) or 0
        session_offset = max(int(max_apollo or 0), int(max_st or 0))

        for table in IMPORT_TABLES:
            # Belt-and-braces: table comes from a hardcoded literal list,
            # but the next line interpolates it into raw SQL — re-validate
            # against an identifier regex so a future refactor can't make
            # this an injection point.
            if not _safe_ident(table):
                raise ValueError(f"refusing to import into unsafe table name: {table!r}")
            rows = data.get(table) or []
            valid_cols = {c.name for c in metadata.tables[table].columns}
            for raw in rows:
                clean = {}
                for k, v in raw.items():
                    if k == 'id' or k not in valid_cols:
                        continue
                    # Defense-in-depth: k is already in valid_cols (which
                    # comes from the SQLAlchemy schema), but the column
                    # name reaches the f-string below as an identifier.
                    # Reject anything that isn't a plain SQL identifier.
                    if not _safe_ident(k):
                        continue
                    clean[k] = _coerce_value(table, k, v)
                clean['user_id'] = user_id

                if 'session_id' in clean and clean['session_id'] is not None:
                    try:
                        clean['session_id'] = int(clean['session_id']) + session_offset
                    except (TypeError, ValueError):
                        clean['session_id'] = None

                if table in ('apollo', 'target_zones'):
                    old_t = raw.get('target_id')
                    if old_t in (None, '', 'NULL'):
                        clean['target_id'] = None
                    else:
                        try:
                            clean['target_id'] = target_id_map.get(int(float(old_t)))
                        except (TypeError, ValueError):
                            clean['target_id'] = None
                    # A zone whose target_id failed to remap would be an
                    # orphan; drop it rather than insert a stranded row.
                    if table == 'target_zones' and clean.get('target_id') is None:
                        continue

                # Refuse attacker-controlled image_filename: an import is the
                # one place a string can land in this column without going
                # through _safe_target_filename, so without this guard a
                # crafted SQL/CSV dump could store '../../apollo.py' and have
                # a later /edit_targets delete remove arbitrary files. NULL
                # out anything that doesn't match the known-safe shapes; the
                # row still imports but its replay just won't render.
                if table == 'targets':
                    fn = clean.get('image_filename')
                    if fn is not None and not _is_safe_target_image_filename(fn):
                        clean['image_filename'] = None

                col_names = list(clean.keys())
                placeholders = ', '.join(['%s'] * len(col_names))
                cur.execute(
                    f"INSERT INTO {table} ({', '.join(col_names)}) "
                    f"VALUES ({placeholders})",
                    [clean[c] for c in col_names]
                )
                counts[table] += 1

                if table == 'targets':
                    old_id = raw.get('id')
                    new_id = cur.lastrowid
                    if old_id not in (None, '') and new_id is not None:
                        try:
                            target_id_map[int(float(old_id))] = int(new_id)
                        except (TypeError, ValueError):
                            pass
        con.commit()
      except Exception:
        # Any failure mid-import (constraint violation, parse error, etc.)
        # rolls back the whole batch so the user never sees a partial state
        # — earlier-table rows committed, later-table rows missing.
        try:
            con.rollback()
        except SQLAlchemyError:
            pass
        raise
    return counts


# ─── Data analysis / visualization ────────────────────────────────────────
# Reports render a matplotlib SVG (base64 inline) and a tabular dataset the
# template can show in a <table> and offer as CSV/Excel download. Each entry
# in REPORTS is a function user_id → {title, svg_b64, columns, rows} or None
# (None = not enough data to render). matplotlib is imported lazily so a
# missing package only breaks /analyze, not the whole app.


def _render_matplotlib_svg(fig):
    """Serialize a Matplotlib figure to a base64-encoded SVG data URL.

    SVG (vector) rather than PNG so the chart stays crisp when the analyze
    lightbox blows it up to fill the screen's longest axis. Text is rendered
    as paths (svg.fonttype='path') so the output doesn't depend on the
    viewing browser having our fonts installed."""
    import base64
    import matplotlib.pyplot as plt
    buf = io.BytesIO()
    with plt.rc_context({'svg.fonttype': 'path'}):
        fig.savefig(buf, format='svg', bbox_inches='tight')
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode('ascii')


def _report_arrows_vs_time(user_id):
    """Per-session arrow count plotted against session start time."""
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT st.session_id, st.session_begin_time, COUNT(a.id) AS shots "
            "FROM session_times st "
            "LEFT JOIN apollo a "
            "  ON a.session_id = st.session_id AND a.user_id = st.user_id "
            "WHERE st.user_id = %s AND st.session_begin_time IS NOT NULL "
            "GROUP BY st.session_id, st.session_begin_time "
            "HAVING COUNT(a.id) > 0 "
            "ORDER BY st.session_begin_time ASC",
            (user_id,)
        ).fetchall()

    sessions = []
    for r in rows:
        begin_dt = _utc_to_user(r['session_begin_time'])
        if begin_dt is None:
            continue
        # Strip tzinfo so matplotlib date axes (and pure-date bucketing)
        # work on a homogeneous naive local-wall-clock value.
        begin_dt = begin_dt.replace(tzinfo=None)
        sessions.append({
            'session_id': int(r['session_id']),
            'begin_time': begin_dt,
            'arrows_shot': int(r['shots'] or 0),
        })

    if not sessions:
        return None

    # Bar chart: arrows per calendar day, with zero-fill between the
    # first and last days so practice gaps are visible. Session count is
    # intentionally ignored — the question is volume per day.
    from collections import Counter
    per_day = Counter()
    for s in sessions:
        per_day[s['begin_time'].date()] += s['arrows_shot']
    days_sorted = sorted(per_day.keys())
    first_day, last_day = days_sorted[0], days_sorted[-1]
    span = (last_day - first_day).days
    day_series = [
        (first_day + timedelta(days=i),
         per_day.get(first_day + timedelta(days=i), 0))
        for i in range(span + 1)
    ]

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.ticker import MaxNLocator

    fig, ax = plt.subplots(figsize=(9, 4.5))
    xs = [datetime.combine(d, datetime.min.time()) for d, _ in day_series]
    ys = [n for _, n in day_series]
    ax.bar(xs, ys, width=0.8, color='#4d6da6', edgecolor='#1a3a5c')
    ax.set_xlabel('Day')
    ax.set_ylabel('Arrows shot')
    ax.set_title('Arrows shot per day')
    ax.grid(True, axis='y', linestyle='--', alpha=0.4)
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))
    locator = mdates.AutoDateLocator()
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))
    fig.autofmt_xdate()
    svg_b64 = _render_matplotlib_svg(fig)

    columns = ['Session', 'Start time', 'Arrows shot']
    rows_out = [
        [s['session_id'],
         s['begin_time'].strftime(SESSION_DT_FMT),
         s['arrows_shot']]
        for s in sessions
    ]
    return {
        'key': 'arrows_vs_time',
        'title': 'Arrows shot vs time',
        'svg_b64': svg_b64,
        'columns': columns,
        'rows': rows_out,
    }


def _report_hits_by_boundaries(user_id):
    """Per-target hit distribution across user-defined scoring zones.

    Skips targets that have no zones (the chart would just be a single
    "Outside zones" bar — uninteresting).
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        target_rows = cur.execute(
            "SELECT t.id AS rowid, t.name, t.image_filename, "
            "       t.physical_size_mm, t.image_size_px "
            "FROM targets t "
            "WHERE t.user_id = %s "
            "  AND EXISTS (SELECT 1 FROM target_zones z "
            "              WHERE z.target_id = t.id AND z.user_id = %s) "
            "ORDER BY t.id ASC",
            (user_id, user_id)
        ).fetchall()

    if not target_rows:
        return None

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    panels = []
    for trow in target_rows:
        target_id = int(trow['rowid'])
        target_name = trow['name']
        target_cfg = target_to_config(trow)
        zones = _fetch_target_zones(target_id, user_id)
        if not zones:
            continue

        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            shots = cur.execute(
                "SELECT x_coord, y_coord, record_mode, arrow_shaft_diameter "
                "FROM apollo "
                "WHERE user_id = %s AND target_id = %s "
                "ORDER BY id ASC",
                (user_id, target_id)
            ).fetchall()

        if not shots:
            continue

        # Zone names are not unique; index by row position so a duplicate
        # label doesn't merge two distinct rings in the count. Out-of-zone
        # hits are folded into the miss bucket — touching no scoring ring
        # is functionally the same as not hitting the target.
        zone_counts = [0] * len(zones)
        miss = 0
        replay_shots = []
        for s in shots:
            xraw = str(s['x_coord']).strip() if s['x_coord'] is not None else ''
            yraw = str(s['y_coord']).strip() if s['y_coord'] is not None else ''
            idx = _classify_shot(xraw, yraw, zones,
                                 _row_get(s, 'arrow_shaft_diameter'))
            replay_shots.append({'x': xraw, 'y': yraw, 'miss': idx is None})
            if idx is None:
                miss += 1
            else:
                zone_counts[idx] += 1

        total = sum(zone_counts) + miss
        if total == 0:
            continue

        # Innermost (highest-value) zone first in the chart so the rings
        # read left-to-right the way they sit on the target visually.
        labels = [z['name'] or f'Zone {i + 1}' for i, z in enumerate(zones)]
        counts = list(zone_counts)
        labels.append('Miss')
        counts.append(miss)

        # Bar colors sampled directly from the target image at each ring's
        # mid-radius — works for any face the user uploads (NASP, FITA,
        # 3D animal, custom) instead of guessing with a synthetic
        # gradient. Falls back to the old warm-to-cool ramp if the image
        # can't be loaded.
        colors = _sample_zone_colors(
            zones, target_cfg, fallback_count=len(zones)
        )
        colors.append('#e53935')   # miss bar — vivid red regardless

        fig, ax = plt.subplots(figsize=(7.5, 4.5))
        bars = ax.bar(labels, counts, color=colors, edgecolor='#1a3a5c')
        ax.set_ylabel('Hits')
        ax.set_title(f'Hits by zone — {target_name}')
        ax.grid(True, axis='y', linestyle='--', alpha=0.4)
        from matplotlib.ticker import MaxNLocator
        ax.yaxis.set_major_locator(MaxNLocator(integer=True))
        for bar, n in zip(bars, counts):
            if n > 0:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_height(), str(n),
                        ha='center', va='bottom', fontsize=9)
        fig.autofmt_xdate(rotation=20)
        svg_b64 = _render_matplotlib_svg(fig)

        # Table: one row per ring, then miss footer.
        columns = ['Zone', 'Points', 'Radius (mm)', 'Hits', 'Percent']
        rows_out = []
        for i, z in enumerate(zones):
            pct = round(zone_counts[i] / total * 100, 1)
            rows_out.append([
                z['name'] or f'Zone {i + 1}',
                int(z['point_value'] or 0),
                round(float(z['radius_mm']), 1),
                zone_counts[i],
                f'{pct}%',
            ])
        rows_out.append(['Miss', 0, '—', miss,
                         f'{round(miss / total * 100, 1)}%'])

        # record_mode is per-shot in the schema but realistically constant
        # within one target's history — first shot is a fine source.
        first_record_mode = int(shots[0]['record_mode'] or 0) \
            if shots[0]['record_mode'] is not None else 0
        replay = {
            'shots': replay_shots,
            'record_mode': first_record_mode,
            'target_image_url': url_for('static',
                                         filename=target_cfg['target_image']),
            'target_width_mm': target_cfg['target_width_mm'],
            'target_name': target_name,
            # Vector ring spec (None for raster-only targets) so the replay
            # renders crisp SVG rings instead of the placeholder image.
            'face_render': target_cfg.get('face_render'),
        }

        panels.append({
            'key': f'hits_by_boundaries__{target_id}',
            'title': f'{target_name} — hits by zone',
            'svg_b64': svg_b64,
            'columns': columns,
            'rows': rows_out,
            'replay': replay,
        })

    if not panels:
        return None

    return {
        'key': 'hits_by_boundaries',
        'title': 'Hits by boundaries',
        'panels': panels,
    }


def _report_all_shots_per_target(user_id, date_from=None, date_to=None):
    """Scatter every shot the user has ever taken on each target, with
    centroid + dispersion overlay so trends (left bias, vertical stringing,
    grouping size) jump out at a glance.

    Optional `date_from` / `date_to` (YYYY-MM-DD strings) restrict the
    scatter to shots whose `timestamp` falls within the inclusive range.
    Either side may be omitted for an open-ended range; both omitted
    reproduces the original "all shots ever" behavior.
    """
    # Parse the date bounds once. Invalid input is silently ignored —
    # an out-of-format date shouldn't kill the report.
    range_from = None
    range_to = None
    if date_from:
        try:
            range_from = datetime.strptime(date_from, '%Y-%m-%d')
        except ValueError:
            range_from = None
    if date_to:
        try:
            # Inclusive upper bound — include the whole "to" day.
            range_to = datetime.strptime(date_to, '%Y-%m-%d') \
                + timedelta(days=1) - timedelta(microseconds=1)
        except ValueError:
            range_to = None

    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        target_rows = cur.execute(
            "SELECT id AS rowid, name, image_filename, "
            "       physical_size_mm, image_size_px "
            "FROM targets WHERE user_id = %s "
            "ORDER BY id ASC",
            (user_id,)
        ).fetchall()

    if not target_rows:
        return None

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.patches import Circle, Ellipse, FancyArrow

    panels = []
    for trow in target_rows:
        target_id = int(trow['rowid'])
        target_name = trow['name']
        target_cfg = target_to_config(trow)

        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            sql = ("SELECT x_coord, y_coord FROM apollo "
                   "WHERE user_id = %s AND target_id = %s")
            params = [user_id, target_id]
            if range_from is not None:
                sql += " AND timestamp >= %s"
                params.append(range_from)
            if range_to is not None:
                sql += " AND timestamp <= %s"
                params.append(range_to)
            sql += " ORDER BY id ASC"
            shots = cur.execute(sql, tuple(params)).fetchall()

        if not shots:
            continue

        xs, ys = [], []
        miss = 0
        for s in shots:
            xraw = str(s['x_coord']).strip() if s['x_coord'] is not None else ''
            yraw = str(s['y_coord']).strip() if s['y_coord'] is not None else ''
            if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
                miss += 1
                continue
            try:
                xs.append(float(xraw))
                ys.append(float(yraw))
            except ValueError:
                continue

        hits = len(xs)
        total = hits + miss
        if total == 0:
            continue

        half = float(target_cfg['target_width_mm']) / 2.0

        # Load the target image straight off disk for matplotlib — url_for
        # is a server URL, but imshow wants pixel data.
        img_disk_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'static',
            target_cfg['target_image']
        )
        try:
            from PIL import Image as PILImage
            bg = PILImage.open(img_disk_path)
        except (FileNotFoundError, OSError):
            bg = None

        # 7 inches at 110 dpi ≈ 770px — substantially bigger than the
        # 360px animated replays, which is the whole point of this report.
        fig, ax = plt.subplots(figsize=(7, 7))
        if bg is not None:
            ax.imshow(bg, extent=[-half, half, -half, half], origin='upper')
        else:
            ax.set_facecolor('#1a3a5c')
        ax.set_xlim(-half, half)
        ax.set_ylim(-half, half)
        ax.set_aspect('equal')

        stats = _archery_stats(xs, ys) if hits > 0 else None
        if stats is not None:
            ax.scatter(xs, ys, s=42,
                       facecolors='#fcba03', edgecolors='#1a3a5c',
                       linewidths=0.6, alpha=0.55, zorder=3)
            mean_x, mean_y = stats['centroid']
            # Concentric precision rings: solid Mean Radius (typical
            # group size), dashed R95 (95% containment). Both are
            # measured about the *centroid*, so they describe precision
            # independent of how far the centroid drifted from the bull.
            ax.add_patch(Circle((mean_x, mean_y), stats['mr'],
                                fill=False, edgecolor='#a1d8ed',
                                linewidth=1.6, linestyle='-', zorder=4,
                                label='MR'))
            ax.add_patch(Circle((mean_x, mean_y), stats['r95'],
                                fill=False, edgecolor='#95abcf',
                                linewidth=1.4, linestyle='--', zorder=4,
                                label='R95'))
            # Faint 1σ covariance ellipse — exposes stringing (vertical or
            # horizontal elongation) that the radial circles hide.
            if stats['sigma_x'] > 0 and stats['sigma_y'] > 0:
                # Eigen-decompose the 2×2 covariance to get the ellipse
                # axes/angle (same math the elliptical R95 branch uses).
                sxx = stats['sigma_x'] ** 2
                syy = stats['sigma_y'] ** 2
                sxy = stats['rho'] * stats['sigma_x'] * stats['sigma_y']
                tr = sxx + syy
                disc = max(0.0, tr * tr / 4.0 - (sxx * syy - sxy * sxy))
                lam1 = tr / 2.0 + math.sqrt(disc)
                lam2 = tr / 2.0 - math.sqrt(disc)
                if abs(sxy) < 1e-12 and abs(sxx - syy) < 1e-12:
                    angle_deg = 0.0
                else:
                    angle_deg = math.degrees(
                        0.5 * math.atan2(2 * sxy, sxx - syy)
                    )
                ax.add_patch(Ellipse(
                    (mean_x, mean_y),
                    2 * math.sqrt(max(0.0, lam1)),
                    2 * math.sqrt(max(0.0, lam2)),
                    angle=angle_deg,
                    fill=False, edgecolor='#ffffff', alpha=0.55,
                    linewidth=1.0, linestyle=':', zorder=4))
            # Bias arrow from bullseye to centroid — direction the sight
            # needs to move (or the form needs to compensate for).
            if stats['mpi'] > 0:
                ax.add_patch(FancyArrow(
                    0, 0, mean_x, mean_y,
                    length_includes_head=True,
                    width=max(half * 0.004, 0.4),
                    head_width=max(half * 0.018, 2.5),
                    head_length=max(half * 0.025, 3.5),
                    color='#c79b5a', alpha=0.85, zorder=5))
            ax.plot(mean_x, mean_y, marker='x', color='#c79b5a',
                    markersize=14, markeredgewidth=2.4, zorder=6)
        else:
            mean_x = mean_y = 0

        # Crosshair through target center for left/right and high/low bias.
        ax.axhline(0, color=(1, 1, 1, 0.25), linewidth=0.6, linestyle=':')
        ax.axvline(0, color=(1, 1, 1, 0.25), linewidth=0.6, linestyle=':')
        # Title suffix reflects the active date range so the chart is
        # self-describing once exported / zoomed in lightGallery.
        if range_from is not None and range_to is not None:
            range_label = f' [{date_from} → {date_to}]'
        elif range_from is not None:
            range_label = f' [from {date_from}]'
        elif range_to is not None:
            range_label = f' [through {date_to}]'
        else:
            range_label = ''
        ax.set_title(f'{target_name} — every shot{range_label} ({hits} hit'
                     f'{"" if miss == 0 else f", {miss} missed"})')
        ax.set_xlabel('X (mm from center)')
        ax.set_ylabel('Y (mm from center)')
        svg_b64 = _render_matplotlib_svg(fig)

        # Summary stats table — split into Accuracy (where the centroid
        # sits relative to the bull) and Precision (how tight the group
        # is, independent of where it sits). Misses don't contribute to
        # either — they're already in the count row above.
        hit_rate = round(hits / total * 100, 1) if total else 0.0
        columns = ['Metric', 'Value']
        rows_out = [
            ['Total shots', total],
            ['Hits', hits],
            ['Misses', miss],
            ['Hit rate', f'{hit_rate}%'],
        ]
        if stats is not None:
            mpi_mm = stats['mpi']
            bx_mm, by_mm = stats['bias_xy']
            r95_mm = stats['r95']
            mr_mm = stats['mr']
            sx_mm = stats['sigma_x']
            sy_mm = stats['sigma_y']
            es_mm = stats['extreme_spread']
            mpi_tip = (
                "MPI (Mean Point of Impact): magnitude of the group's "
                "centroid offset from the bullseye. Pure accuracy — "
                "fix with sight or anchor adjustments, not tuning."
            )
            bias_tip = (
                "Signed bias of the centroid: (Δx, Δy). Direction tells "
                "you which way to move the sight, not just by how much."
            )
            r95_tip = (
                "R95: radius about the group's own centroid that contains "
                "95% of its shots. The headline precision number — "
                "independent of where the centroid sits."
            )
            r95_empirical_tip = (
                f"R95 from empirical percentile because n={stats['n']} is "
                "below the Rayleigh-fit threshold of 10 shots. Treat as "
                "approximate until you have more data."
            )
            mr_tip = (
                "Mean Radius: average distance of each shot from the "
                "group's own centroid. Secondary precision number, "
                "canonical in shooting-sport testing."
            )
            sigma_tip = (
                "σ_x / σ_y: per-axis standard deviation about the centroid. "
                "Large σ_y with small σ_x is vertical stringing (release "
                "or anchor variation); the reverse is horizontal stringing "
                "(bow torque or grip)."
            )
            es_tip = (
                "Extreme spread: largest pairwise distance between any two "
                "shots in the group. Familiar from firearm and 3D archery "
                "scoring — sensitive to one bad shot."
            )
            if stats['is_empirical']:
                r95_label = _tip(f"R95 — empirical, n={stats['n']}",
                                 r95_empirical_tip)
            else:
                r95_label = _tip('R95', r95_tip)
            rows_out.extend([
                [Markup('<em>— Accuracy —</em>'), ''],
                [_tip('MPI', mpi_tip),
                 Markup(f'{_mm_val(mpi_mm)} {_mm_unit()}')],
                [_tip('Bias Δx, Δy (signed)', bias_tip),
                 _mm_pair(bx_mm, by_mm)],
                [Markup('<em>— Precision —</em>'), ''],
                [r95_label, Markup(f'{_mm_val(r95_mm)} {_mm_unit()}')],
                [_tip('Mean Radius', mr_tip),
                 Markup(f'{_mm_val(mr_mm)} {_mm_unit()}')],
                [_tip('σ_x, σ_y', sigma_tip),
                 _mm_pair(sx_mm, sy_mm)],
                [_tip('Extreme spread', es_tip),
                 Markup(f'{_mm_val(es_mm)} {_mm_unit()}')],
            ])

        panels.append({
            'key': f'all_shots_per_target__{target_id}',
            'title': f'{target_name} — every shot{range_label}',
            'svg_b64': svg_b64,
            'columns': columns,
            'rows': rows_out,
            # No animated replay here — the static scatter *is* the
            # visualization, and stacking an animated copy next to it would
            # just compete for attention.
        })

    if not panels:
        return None

    # Mirror the per-panel range label on the report card heading.
    if range_from is not None and range_to is not None:
        report_title = f'Shots per target ({date_from} → {date_to})'
    elif range_from is not None:
        report_title = f'Shots per target (from {date_from})'
    elif range_to is not None:
        report_title = f'Shots per target (through {date_to})'
    else:
        report_title = 'Shots per target (all time)'

    return {
        'key': 'all_shots_per_target',
        'title': report_title,
        'panels': panels,
    }


# Minimum shots a piece of equipment needs before it's eligible for a
# head-to-head comparison. Below this the stats are too noisy to be
# worth surfacing (and Welch's t-test loses what little power it had).
_HEAD_TO_HEAD_MIN_SHOTS = 5

# Bowtype-specific settings exposed as head-to-head comparison dimensions. Each
# is addressed as a 'style:<key>' column. Limited to the high-value categorical
# choices — comparing free-text tunings (plunger tension, string crawl) would
# just scatter into singleton groups. Labels reuse the schema's field labels.
_HEAD_TO_HEAD_STYLE_KEYS = ['release_aid', 'sight_type', 'aim_method',
                            'finger_protection', 'stabilizer', 'shoot_off']
_HEAD_TO_HEAD_STYLE_LABELS = {
    f['key']: f['label']
    for fields in BOWSTYLE_SETTINGS.values() for f in fields
    if f['key'] in _HEAD_TO_HEAD_STYLE_KEYS
}


def _regularized_incomplete_beta(x, a, b):
    """I_x(a, b) via the standard continued-fraction expansion.

    Used as a scipy-free fallback for survival-function p-values on the
    t and F distributions. Accurate to ~1e-7 for the parameter ranges
    /analyze produces, which is well inside the precision we display.
    """
    if x <= 0:
        return 0.0
    if x >= 1:
        return 1.0
    log_beta = (math.lgamma(a) + math.lgamma(b) - math.lgamma(a + b))
    front = math.exp(math.log(x) * a + math.log(1 - x) * b - log_beta) / a
    # Continued fraction (Lentz's method)
    fpmin = 1e-300
    qab = a + b
    qap = a + 1.0
    qam = a - 1.0
    c = 1.0
    d = 1.0 - qab * x / qap
    if abs(d) < fpmin:
        d = fpmin
    d = 1.0 / d
    h = d
    for m in range(1, 200):
        m2 = 2 * m
        aa = m * (b - m) * x / ((qam + m2) * (a + m2))
        d = 1.0 + aa * d
        if abs(d) < fpmin:
            d = fpmin
        c = 1.0 + aa / c
        if abs(c) < fpmin:
            c = fpmin
        d = 1.0 / d
        h *= d * c
        aa = -(a + m) * (qab + m) * x / ((a + m2) * (qap + m2))
        d = 1.0 + aa * d
        if abs(d) < fpmin:
            d = fpmin
        c = 1.0 + aa / c
        if abs(c) < fpmin:
            c = fpmin
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < 3e-7:
            break
    return front * h


def _std_normal_cdf(z):
    """Standard normal CDF via math.erf — no scipy dependency."""
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


def _f_sf(f, df1, df2):
    """Right-tail (survival) probability of an F(df1, df2) random variable.

    Falls back to a regularized incomplete beta when scipy isn't installed.
    Returns None when the parameters are out of range.
    """
    if f is None or f <= 0 or df1 <= 0 or df2 <= 0:
        return None
    try:
        from scipy.stats import f as _f  # type: ignore
        return float(_f.sf(f, df1, df2))
    except ImportError:
        x = df2 / (df2 + df1 * f)
        try:
            return _regularized_incomplete_beta(x, df2 / 2.0, df1 / 2.0)
        except (ValueError, ZeroDivisionError):
            return None


def _median(xs):
    """Sample median: middle value for odd n, mean of the two middle
    values for even n. Returns 0.0 for an empty sequence."""
    s = sorted(xs)
    n = len(s)
    if n == 0:
        return 0.0
    if n % 2:
        return s[n // 2]
    return 0.5 * (s[n // 2 - 1] + s[n // 2])


def _mann_whitney_u(a, b):
    """Two-sided Mann-Whitney U on samples ``a`` and ``b``.

    Returns ``(U, p)`` where U is the smaller of the two conventional U
    statistics. p is the two-sided p-value (scipy when available,
    otherwise the standard normal approximation with tie + continuity
    correction). Returns ``(None, None)`` for samples too small to test.

    Preferred over Welch's t for shot-distance data because the
    underlying distribution is right-skewed and bounded below at 0 —
    rank-based tests don't care about either.
    """
    n1, n2 = len(a), len(b)
    if n1 < 2 or n2 < 2:
        return None, None
    try:
        from scipy.stats import mannwhitneyu  # type: ignore
        res = mannwhitneyu(a, b, alternative='two-sided')
        return float(res.statistic), float(res.pvalue)
    except ImportError:
        pass

    # Combined-sample average ranks (1-indexed; ties get the midpoint).
    combined = [(v, 0) for v in a] + [(v, 1) for v in b]
    combined.sort(key=lambda x: x[0])
    ranks = [0.0] * len(combined)
    i = 0
    while i < len(combined):
        j = i
        while j + 1 < len(combined) and combined[j + 1][0] == combined[i][0]:
            j += 1
        avg_rank = (i + j) / 2.0 + 1
        for k in range(i, j + 1):
            ranks[k] = avg_rank
        i = j + 1
    r1 = sum(r for r, (_, g) in zip(ranks, combined) if g == 0)
    u1 = r1 - n1 * (n1 + 1) / 2.0
    u2 = n1 * n2 - u1
    u = min(u1, u2)

    # Normal approximation with continuity + tie correction.
    from collections import Counter
    counts = Counter(v for v, _ in combined)
    tie_term = sum(t * (t * t - 1) for t in counts.values())
    N = n1 + n2
    var = (n1 * n2 / 12.0) * ((N + 1) - tie_term / (N * (N - 1)))
    if var <= 0:
        return float(u), None
    diff = u1 - (n1 * n2 / 2.0)
    if diff > 0:
        diff -= 0.5
    elif diff < 0:
        diff += 0.5
    z = diff / math.sqrt(var)
    p = 2.0 * (1.0 - _std_normal_cdf(abs(z)))
    return float(u), max(0.0, min(1.0, p))


def _cliffs_delta(a, b):
    """Cliff's δ — nonparametric effect size paired with Mann-Whitney.

    δ = P(a > b) − P(a < b) ∈ [−1, +1]. For distance-from-center data,
    δ > 0 means ``a``'s shots tend to land farther from center than
    ``b``'s (i.e., ``b`` is more accurate); δ < 0 means the opposite.
    Naive O(n·m) implementation — fine at the dataset sizes /analyze sees.
    """
    n1, n2 = len(a), len(b)
    if n1 == 0 or n2 == 0:
        return None
    gt = lt = 0
    for x in a:
        for y in b:
            if x > y:
                gt += 1
            elif x < y:
                lt += 1
    return (gt - lt) / float(n1 * n2)


def _brown_forsythe(a, b):
    """Brown-Forsythe (median-centered Levene's) test for equality of
    spread between two samples.

    For 2 groups this reduces algebraically to a two-sample equal-variance
    t-test on the absolute deviations from each group's median; we report
    the equivalent F = t² with df=(1, n1+n2-2). Returns
    ``(F, df1, df2, p)``, or all-None when too small / no spread.
    """
    n1, n2 = len(a), len(b)
    if n1 < 2 or n2 < 2:
        return None, None, None, None
    med1 = _median(a)
    med2 = _median(b)
    z1 = [abs(x - med1) for x in a]
    z2 = [abs(x - med2) for x in b]
    m1 = sum(z1) / n1
    m2 = sum(z2) / n2
    df2 = n1 + n2 - 2
    pooled_ss = sum((x - m1) ** 2 for x in z1) + sum((x - m2) ** 2 for x in z2)
    s_pooled = pooled_ss / df2
    if s_pooled <= 0:
        return None, None, None, None
    se = math.sqrt(s_pooled * (1.0 / n1 + 1.0 / n2))
    if se == 0:
        return None, None, None, None
    t = (m1 - m2) / se
    f_stat = t * t
    return f_stat, 1, df2, _f_sf(f_stat, 1, df2)


def _hotelling_t2(a_xy, b_xy):
    """Two-sample Hotelling's T² on 2D shot vectors (cx, cy per shot).

    Multivariate analogue of the t-test for the centroid; answers
    "does one piece systematically push shots in a particular direction
    relative to the other?" Returns ``(T², F, df1, df2, p)`` with df1=2
    (the two coordinate dimensions) and df2=n1+n2−3.

    Returns all-None when either sample is too small or the pooled
    covariance is singular (e.g., all shots on a line).
    """
    n1, n2 = len(a_xy), len(b_xy)
    if n1 < 3 or n2 < 3:
        return None, None, None, None, None
    mx1 = sum(p[0] for p in a_xy) / n1
    my1 = sum(p[1] for p in a_xy) / n1
    mx2 = sum(p[0] for p in b_xy) / n2
    my2 = sum(p[1] for p in b_xy) / n2

    def _cov(pts, mx, my, n):
        sxx = sum((p[0] - mx) ** 2 for p in pts) / (n - 1)
        syy = sum((p[1] - my) ** 2 for p in pts) / (n - 1)
        sxy = sum((p[0] - mx) * (p[1] - my) for p in pts) / (n - 1)
        return sxx, syy, sxy

    s1xx, s1yy, s1xy = _cov(a_xy, mx1, my1, n1)
    s2xx, s2yy, s2xy = _cov(b_xy, mx2, my2, n2)
    df2_pool = n1 + n2 - 2
    pxx = ((n1 - 1) * s1xx + (n2 - 1) * s2xx) / df2_pool
    pyy = ((n1 - 1) * s1yy + (n2 - 1) * s2yy) / df2_pool
    pxy = ((n1 - 1) * s1xy + (n2 - 1) * s2xy) / df2_pool
    det = pxx * pyy - pxy * pxy
    if det <= 0:
        return None, None, None, None, None
    inv_xx, inv_yy, inv_xy = pyy / det, pxx / det, -pxy / det
    dx, dy = mx1 - mx2, my1 - my2
    t2 = (n1 * n2 / (n1 + n2)) * (
        dx * dx * inv_xx + 2 * dx * dy * inv_xy + dy * dy * inv_yy
    )
    df1 = 2
    df2 = n1 + n2 - df1 - 1
    if df2 <= 0:
        return None, None, None, None, None
    f_stat = t2 * df2 / (df1 * df2_pool)
    return t2, f_stat, df1, df2, _f_sf(f_stat, df1, df2)


def _holm_bonferroni(p_values):
    """Step-down Holm correction. Input order is preserved in the output.

    ``None`` entries are passed through unchanged and don't count toward
    the family size. Adjusted values are clamped to [0, 1] and made
    monotone in the rank ordering of the originals — the usual guarantee
    that a smaller raw p never produces a larger adjusted p.
    """
    indexed = [(i, p) for i, p in enumerate(p_values) if p is not None]
    if not indexed:
        return list(p_values)
    indexed.sort(key=lambda x: x[1])
    m = len(indexed)
    adj = list(p_values)
    running = 0.0
    for rank, (orig_i, p) in enumerate(indexed):
        scaled = min(1.0, p * (m - rank))
        running = max(running, scaled)
        adj[orig_i] = running
    return adj


# Chi-square critical values for a 2-DOF distribution at common percentiles.
# Used as the scipy-free fallback when computing CEP / R95 from a fitted 2D
# Gaussian. R = sqrt(chi2_ppf(q, df=2) * λ) for the isotropic case; for the
# general case we substitute the geometric mean of the covariance eigenvalues.
# (chi2_ppf(0.5, 2) = 2·ln(2) ≈ 1.3863, chi2_ppf(0.95, 2) = 2·ln(20) ≈ 5.9915.)
_CHI2_2DOF_50 = 2.0 * math.log(2.0)
_CHI2_2DOF_95 = 2.0 * math.log(20.0)


def _chi2_ppf_df2(q):
    """Inverse CDF of χ²(df=2) at quantile q. Closed-form for df=2:
    F(x) = 1 - exp(-x/2)  ⇒  x = -2·ln(1 - q). No scipy needed."""
    if q <= 0:
        return 0.0
    if q >= 1:
        return float('inf')
    return -2.0 * math.log(1.0 - q)


def _tip(label, tip):
    """Wrap a table-cell or column-header label in a CSS-tooltip span.

    The ``[data-tip]`` selector in analyze.html shows the tip on hover
    without browser delay. Returned as Markup so Jinja's autoescape
    leaves the span intact; the export sanitizer in ``analyze_export``
    strips the tag back out for CSV/XLSX downloads.
    """
    from html import escape
    return Markup(
        f'<span data-tip="{escape(tip)}">{escape(label)}</span>'
    )


def _mm_val(mm, decimals=1):
    """Render a millimetre measurement as a client-toggleable span.

    The analyze.html ``applyUnits()`` JS reads ``data-mm`` and rewrites
    the inner text on the imperial toggle (mm/25.4, two-decimal inches).
    The export sanitizer strips the span tag so CSV/XLSX cells keep the
    raw mm value — server-side downloads always speak metric.
    """
    if mm is None:
        return Markup('—')
    try:
        v = float(mm)
    except (TypeError, ValueError):
        return Markup('—')
    return Markup(
        f'<span class="metric-len" data-mm="{v:.4f}">'
        f'{v:.{decimals}f}</span>'
    )


def _mm_unit():
    """Render the "mm" unit label as a client-toggleable span.

    On imperial toggle the same JS swaps the text to "in", so labels
    like "MPI (mm)" become "MPI (in)" without a page reload.
    """
    return Markup('<span class="metric-unit">mm</span>')


def _mm_pair(mm_x, mm_y, decimals=1):
    """Render a paired (Δx, Δy) mm measurement with one toggleable unit
    label at the end. Used for centroid bias and σ_x/σ_y tuple cells."""
    return Markup(
        f'({_mm_val(mm_x, decimals)}, {_mm_val(mm_y, decimals)}) '
        f'{_mm_unit()}'
    )


def _archery_stats(xs, ys):
    """Per-group accuracy + precision stats from a 2D shot cloud.

    Inputs are parallel x/y arrays in any consistent coordinate frame
    (the caller handles mm vs normalized). Splits the two archery error
    modes cleanly:

      * Accuracy  = bias of the group centroid relative to (0, 0). MPI is
        the magnitude; (bias_x, bias_y) carries the direction so the user
        knows which way to move the sight.
      * Precision = spread *about the group's own centroid* — independent
        of bias. Reported as Mean Radius (MR), σ-about-centroid (sigma_r,
        unbiased), per-axis σ_x / σ_y for stringing, and R95 from a fitted
        bivariate normal.

    R95 / CEP use the elliptical formula when the fit shows tilt or
    elongation; otherwise the Rayleigh closed form. For samples below
    ``min_n_for_fit`` we fall back to empirical percentiles.

    Returns ``None`` if no shots supplied.
    """
    n = len(xs)
    if n == 0 or n != len(ys):
        return None

    cx = sum(xs) / n
    cy = sum(ys) / n
    mpi = math.sqrt(cx * cx + cy * cy)

    # Distances from the group's *own* centroid — the precision raw material.
    dists_from_centroid = [
        math.sqrt((x - cx) ** 2 + (y - cy) ** 2) for x, y in zip(xs, ys)
    ]
    mr = sum(dists_from_centroid) / n
    extreme_spread = 0.0
    if n >= 2:
        # Naive O(n²) pairwise — fine at the dataset sizes /analyze sees.
        for i in range(n):
            for j in range(i + 1, n):
                d = math.sqrt((xs[i] - xs[j]) ** 2 + (ys[i] - ys[j]) ** 2)
                if d > extreme_spread:
                    extreme_spread = d

    if n >= 2:
        var_x = sum((x - cx) ** 2 for x in xs) / (n - 1)
        var_y = sum((y - cy) ** 2 for y in ys) / (n - 1)
        cov_xy = sum((x - cx) * (y - cy) for x, y in zip(xs, ys)) / (n - 1)
        sigma_x = math.sqrt(max(0.0, var_x))
        sigma_y = math.sqrt(max(0.0, var_y))
        denom_rho = sigma_x * sigma_y
        rho = cov_xy / denom_rho if denom_rho > 0 else 0.0
        # 1σ radius about centroid: RMS of distance-from-centroid, n-1 normed.
        sigma_r = math.sqrt(max(0.0, var_x + var_y))
    else:
        sigma_x = sigma_y = sigma_r = 0.0
        cov_xy = 0.0
        rho = 0.0

    # Bounding box for plot scaling.
    bbox_w = (max(xs) - min(xs)) if n else 0.0
    bbox_h = (max(ys) - min(ys)) if n else 0.0

    # CEP / R95: fit-based for n ≥ 10, empirical otherwise. The Rayleigh
    # closed form (R_q = σ·√(-2·ln(1-q))) needs σ_x ≈ σ_y and ρ ≈ 0; when
    # those don't hold we use the eigenvalue product of the covariance
    # matrix, equivalent to the standard bivariate-normal ellipse area.
    is_empirical = n < 10
    if is_empirical:
        sorted_d = sorted(dists_from_centroid)
        if n == 0:
            cep = 0.0
            r95 = 0.0
        else:
            # Nearest-rank percentile; with small n this is the honest answer.
            def _pct(p):
                if n == 1:
                    return sorted_d[0]
                k = max(0, min(n - 1, int(math.ceil(p * n)) - 1))
                return sorted_d[k]
            cep = _pct(0.50)
            r95 = _pct(0.95)
    else:
        sx2, sy2 = sigma_x ** 2, sigma_y ** 2
        rayleigh_ok = (
            abs(rho) < 0.2
            and max(sigma_x, sigma_y) <= 1.2 * max(min(sigma_x, sigma_y), 1e-12)
        )
        if rayleigh_ok:
            sigma_iso_sq = (sx2 + sy2) / 2.0
            cep = math.sqrt(_CHI2_2DOF_50 * sigma_iso_sq)
            r95 = math.sqrt(_CHI2_2DOF_95 * sigma_iso_sq)
        else:
            # Eigenvalues of the 2×2 covariance: trace ± √(tr² − 4·det) / 2.
            tr = sx2 + sy2
            det = sx2 * sy2 - cov_xy * cov_xy
            disc = max(0.0, tr * tr / 4.0 - det)
            lam1 = tr / 2.0 + math.sqrt(disc)
            lam2 = tr / 2.0 - math.sqrt(disc)
            geom = math.sqrt(max(0.0, lam1) * max(0.0, lam2))
            cep = math.sqrt(_CHI2_2DOF_50 * geom)
            r95 = math.sqrt(_CHI2_2DOF_95 * geom)

    return {
        'n': n,
        'centroid': (cx, cy),
        'bias_xy': (cx, cy),
        'mpi': mpi,
        'mr': mr,
        'sigma_r': sigma_r,
        'sigma_x': sigma_x,
        'sigma_y': sigma_y,
        'rho': rho,
        'cep': cep,
        'r95': r95,
        'extreme_spread': extreme_spread,
        'bbox_w': bbox_w,
        'bbox_h': bbox_h,
        'dists_from_centroid': dists_from_centroid,
        'is_empirical': is_empirical,
    }


def _equipment_shot_samples(user_id, column):
    """Pull every shot for ``user_id`` grouped by an equipment column.

    ``column`` is either 'bow' or 'arrow_type' (the names live on the
    apollo shot table — there's no FK to bows/arrows, so we group on
    the model-name string the user picked at session time).

    Returns ``{equipment_name: list[shot_dict]}`` where each shot_dict
    has the raw coordinates and the target half-size in mm so the
    caller can normalize distance across targets of different sizes.
    Equipment values that are NULL or empty are excluded — those are
    pre-equipment-tracking sessions and don't belong in a comparison.
    """
    # ``column`` is interpolated directly into the SQL below (DB-API can't
    # bind identifiers). Today every caller passes a hardcoded string, but
    # guard at the door so a future caller can't turn this into injection.
    if column not in ('bow', 'arrow_type'):
        raise ValueError(f"Invalid equipment column: {column!r}")
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            f"SELECT a.{column} AS eq, a.x_coord, a.y_coord, a.is_precise, "
            f"       a.target_id, t.physical_size_mm "
            f"FROM apollo a "
            f"LEFT JOIN targets t ON t.id = a.target_id "
            f"WHERE a.user_id = %s AND a.{column} IS NOT NULL "
            f"      AND a.{column} <> '' ",
            (user_id,)
        ).fetchall()

    groups = {}
    for r in rows:
        name = (r['eq'] or '').strip()
        if not name:
            continue
        # Need a target size to normalize distances; drop shots whose
        # target was deleted (NULL join) so cross-target metrics stay
        # honest.
        try:
            half_mm = float(r['physical_size_mm']) / 2.0
        except (TypeError, ValueError):
            continue
        if half_mm <= 0:
            continue
        groups.setdefault(name, []).append({
            'x_raw': str(r['x_coord']).strip() if r['x_coord'] is not None else '',
            'y_raw': str(r['y_coord']).strip() if r['y_coord'] is not None else '',
            'is_precise': int(r['is_precise'] or 0),
            'target_id': int(r['target_id']) if r['target_id'] is not None else None,
            'half_mm': half_mm,
        })
    return groups


def _style_setting_shot_samples(user_id, setting_key):
    """Like _equipment_shot_samples, but group shots by a bowtype-specific
    setting value parsed from each shot's ``bow_style_settings`` JSON.

    ``setting_key`` is one of the keys in BOWSTYLE_SETTINGS (e.g. 'release_aid').
    Each group is one value the user has actually shot with (e.g. 'index' vs
    'thumb-trigger'), so the head-to-head can ask whether that choice changed
    their grouping. Shots whose snapshot lacks the key are skipped.
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT a.bow_style_settings AS blob, a.x_coord, a.y_coord, "
            "       a.is_precise, a.target_id, t.physical_size_mm "
            "FROM apollo a "
            "LEFT JOIN targets t ON t.id = a.target_id "
            "WHERE a.user_id = %s AND a.bow_style_settings IS NOT NULL "
            "      AND a.bow_style_settings <> '' ",
            (user_id,)
        ).fetchall()

    groups = {}
    for r in rows:
        try:
            settings = json.loads(r['blob']) if r['blob'] else {}
        except (ValueError, TypeError):
            continue
        val = settings.get(setting_key)
        if val in (None, ''):
            continue
        name = str(val).strip()
        if not name:
            continue
        try:
            half_mm = float(r['physical_size_mm']) / 2.0
        except (TypeError, ValueError):
            continue
        if half_mm <= 0:
            continue
        groups.setdefault(name, []).append({
            'x_raw': str(r['x_coord']).strip() if r['x_coord'] is not None else '',
            'y_raw': str(r['y_coord']).strip() if r['y_coord'] is not None else '',
            'is_precise': int(r['is_precise'] or 0),
            'target_id': int(r['target_id']) if r['target_id'] is not None else None,
            'half_mm': half_mm,
        })
    return groups


def _is_auto_tag(tag):
    """True for tags Apollo writes automatically (tournament:<key>, practice).

    The head-to-head picker hides these by default so a user's tag list
    isn't dominated by tournament round identifiers they never typed.
    Comparison is case-insensitive — the storage layer is too."""
    if not tag:
        return False
    t = tag.strip().lower()
    return t == 'practice' or t.startswith('tournament:')


def _tag_inventory(user_id):
    """Per-tag shot counts for the head-to-head tag picker.

    Returns ``[{name, shots, is_auto}, ...]`` sorted by shot count (most
    first, name tiebreaker). A shot tagged "indoor, morning" contributes
    +1 to each of "indoor" and "morning". Case-insensitive dedup within
    a shot; first-seen casing wins the canonical display name."""
    if user_id is None:
        return []
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT session_tags FROM apollo "
                "WHERE user_id = %s AND session_tags IS NOT NULL "
                "      AND session_tags <> ''",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError:
        return []
    counts = {}
    canonical = {}
    for row in rows:
        raw = row[0] or ''
        seen = set()
        for part in raw.split(','):
            t = part.strip()
            if not t:
                continue
            key = t.lower()
            if key in seen:
                continue
            seen.add(key)
            canonical.setdefault(key, t)
            counts[key] = counts.get(key, 0) + 1
    inventory = [
        {'name': canonical[k], 'shots': counts[k], 'is_auto': _is_auto_tag(k)}
        for k in counts
    ]
    inventory.sort(key=lambda x: (-x['shots'], x['name'].lower()))
    return inventory


def _tag_shot_samples(user_id):
    """Pull every shot for ``user_id`` grouped by each tag in session_tags.

    Unlike _equipment_shot_samples this is many-to-many — a shot tagged
    "indoor, morning" lands in both the "indoor" and "morning" groups.
    Tags are deduped case-insensitively within a shot, and the first-seen
    casing becomes the canonical display name for each group.
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT a.session_tags AS tags, a.x_coord, a.y_coord, a.is_precise, "
            "       a.target_id, t.physical_size_mm "
            "FROM apollo a "
            "LEFT JOIN targets t ON t.id = a.target_id "
            "WHERE a.user_id = %s AND a.session_tags IS NOT NULL "
            "      AND a.session_tags <> '' ",
            (user_id,)
        ).fetchall()

    groups = {}
    # lowercase → display name, so "Indoor" and "indoor" collapse to one
    # group with stable casing instead of two silent dupes.
    canonical = {}
    for r in rows:
        try:
            half_mm = float(r['physical_size_mm']) / 2.0
        except (TypeError, ValueError):
            continue
        if half_mm <= 0:
            continue
        shot = {
            'x_raw': str(r['x_coord']).strip() if r['x_coord'] is not None else '',
            'y_raw': str(r['y_coord']).strip() if r['y_coord'] is not None else '',
            'is_precise': int(r['is_precise'] or 0),
            'target_id': int(r['target_id']) if r['target_id'] is not None else None,
            'half_mm': half_mm,
        }
        seen_in_shot = set()
        for part in (r['tags'] or '').split(','):
            t = part.strip()
            if not t:
                continue
            key = t.lower()
            if key in seen_in_shot:
                continue
            seen_in_shot.add(key)
            display = canonical.setdefault(key, t)
            groups.setdefault(display, []).append(shot)
    return groups


def _summarize_equipment(shots):
    """Per-equipment summary stats used by the head-to-head report.

    Distances are tracked in two forms:
      * raw_dist  — millimetres from target center (only comparable
        within a single target size)
      * norm_dist — raw_dist / target_half_size, so 1.0 = the calibrated
        edge of the face. Lets us pool shots across targets of mixed
        sizes (e.g. 40cm vs 60cm) without the bigger target dominating
        the variance.

    Returned ``norm_dists`` feeds Mann-Whitney U on the "total error"
    family. ``dists_from_centroid`` (the same shots referenced to *this*
    group's own centroid, not the bullseye) feeds Brown-Forsythe on the
    "precision" family — without that change, Brown-Forsythe was testing
    equal spread of a quantity that already carried each group's bias.
    """
    n_total = len(shots)
    n_miss = 0
    raw_dists = []
    norm_dists = []
    xs_norm = []
    ys_norm = []
    targets = set()
    for s in shots:
        if s['target_id'] is not None:
            targets.add(s['target_id'])
        if s['x_raw'] == MISS_SENTINEL and s['y_raw'] == MISS_SENTINEL:
            n_miss += 1
            continue
        try:
            x = float(s['x_raw'])
            y = float(s['y_raw'])
        except ValueError:
            continue
        d = math.sqrt(x * x + y * y)
        raw_dists.append(d)
        norm_dists.append(d / s['half_mm'])
        xs_norm.append(x / s['half_mm'])
        ys_norm.append(y / s['half_mm'])

    n_hit = len(raw_dists)
    hit_rate = (n_hit / n_total * 100) if n_total else 0.0
    mean_raw = (sum(raw_dists) / n_hit) if n_hit else 0.0
    mean_norm = (sum(norm_dists) / n_hit) if n_hit else 0.0

    # Hand the 2D point cloud (in normalized units) to the shared archery
    # stats helper. Everything precision-related — R95, MR, σ_x/σ_y,
    # distances-from-centroid — flows from this one fit so every report
    # speaks consistent numbers.
    arch = _archery_stats(xs_norm, ys_norm) if n_hit > 0 else None
    if arch is not None:
        cx, cy = arch['centroid']
        r95_norm = arch['r95']
        mr_norm = arch['mr']
        sigma_x_norm = arch['sigma_x']
        sigma_y_norm = arch['sigma_y']
        dists_from_centroid = arch['dists_from_centroid']
        is_empirical = arch['is_empirical']
    else:
        cx = cy = 0.0
        r95_norm = mr_norm = sigma_x_norm = sigma_y_norm = 0.0
        dists_from_centroid = []
        is_empirical = True

    return {
        'n_total': n_total,
        'n_hit': n_hit,
        'n_miss': n_miss,
        'hit_rate': hit_rate,
        'mean_raw_mm': mean_raw,
        'mean_norm': mean_norm,
        'centroid_norm': (cx, cy),
        'mpi_norm': math.sqrt(cx * cx + cy * cy),
        'r95_norm': r95_norm,
        'mr_norm': mr_norm,
        'sigma_x_norm': sigma_x_norm,
        'sigma_y_norm': sigma_y_norm,
        # Total-error sample (distance from *target center*) — feeds the
        # Mann-Whitney "Total error" family.
        'norm_dists': norm_dists,
        # Pure-precision sample (distance from *each group's own* centroid)
        # — feeds Brown-Forsythe in the "Precision" family.
        'dists_from_centroid': dists_from_centroid,
        # Per-shot normalized coords — Hotelling's T² needs the 2D points
        # to test for centroid-level (accuracy) bias.
        'xs_norm': xs_norm,
        'ys_norm': ys_norm,
        'n_targets': len(targets),
        'precision_is_empirical': is_empirical,
    }


def _report_equipment_head_to_head(user_id, categories=None, tag_filter=None):
    """Pairwise head-to-head comparison of bows, arrows, and session tags.

    Each pair is judged on three independent axes so the user can tell
    *what* differs, not just *that* something does:

      * Accuracy — does the centroid sit in a different place?
        Hotelling's T² on the 2D (x, y) shot vectors.
      * Precision — does the group cluster more tightly about its own
        centroid? Brown-Forsythe on distance-from-each-group's-centroid
        (not from the bullseye; that one mixes accuracy and precision).
      * Total error — does one piece simply land closer to the bull on
        average? Mann-Whitney U on distance-from-center. Useful as a
        practical tiebreaker but cannot attribute the cause.

    All distances are normalized by target half-size so shots from
    mixed face sizes (40cm vs 60cm etc.) pool fairly. Skips pairs where
    either side has <2 hits — the tests are uninformative.
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch

    # One result section per kind (Bow / Arrow / Tag) so the UI can title
    # each card by what's actually being compared instead of lumping them
    # all under a single "Equipment head-to-head" heading.
    sections = []
    # Track per-column inventory so we can explain *why* a column produced no
    # pairs when the whole report ends up empty (e.g. only one bow logged).
    inventory = {}

    # ``categories`` scopes which kinds to compare. ``None`` means "all" so
    # existing callers (and the report-spec default) keep their old behavior.
    # Bowtype-specific gear settings are addressable as 'style:<key>' columns
    # so the user can ask whether, e.g., release-aid choice changed grouping.
    all_kinds = [('bow', 'Bow'), ('arrow_type', 'Arrow'), ('tag', 'Tag')]
    for sk in _HEAD_TO_HEAD_STYLE_KEYS:
        all_kinds.append((f'style:{sk}', _HEAD_TO_HEAD_STYLE_LABELS[sk]))
    if categories is None:
        wanted = {k for k, _ in all_kinds}
    else:
        wanted = set(categories)

    COLOR_A = '#4d6da6'
    COLOR_B = '#fcba03'
    EDGE = '#1a3a5c'

    for column, label_singular in all_kinds:
        if column not in wanted:
            continue
        if column == 'tag':
            groups = _tag_shot_samples(user_id)
            if tag_filter is not None:
                # Whitelist applied case-insensitively against the canonical
                # group names. Empty whitelist = "no tags compared" (e.g.
                # user explicitly cleared the picker).
                allowed = {t.strip().lower() for t in tag_filter}
                groups = {name: shots for name, shots in groups.items()
                          if name.lower() in allowed}
        elif column.startswith('style:'):
            groups = _style_setting_shot_samples(user_id, column.split(':', 1)[1])
        else:
            groups = _equipment_shot_samples(user_id, column)
        eligible = {
            name: shots for name, shots in groups.items()
            if len(shots) >= _HEAD_TO_HEAD_MIN_SHOTS
        }
        inventory[column] = {
            'label': label_singular,
            'distinct': len(groups),
            'eligible': len(eligible),
        }
        if len(eligible) < 2:
            continue
        # Collect this kind's panels into its own list so we can wrap them
        # in a section titled "{Kind} head-to-head" at the end of the loop.
        panels = []
        summaries = {name: _summarize_equipment(shots)
                     for name, shots in eligible.items()}
        # Most-shot first → deterministic ordering and the most-used
        # equipment shows up at the top of the report.
        names_sorted = sorted(summaries.keys(),
                              key=lambda n: -summaries[n]['n_total'])
        # ── Pass 1: compute per-pair statistics ──────────────────────
        # Holm-Bonferroni needs the full set of p-values within a family
        # before it can adjust any of them, so we defer panel rendering
        # until the whole pair grid is computed.
        pair_stats = []
        for i in range(len(names_sorted)):
            for j in range(i + 1, len(names_sorted)):
                a_name = names_sorted[i]
                b_name = names_sorted[j]
                a_sum = summaries[a_name]
                b_sum = summaries[b_name]
                if a_sum['n_hit'] < 2 or b_sum['n_hit'] < 2:
                    continue
                # Total error (radial distance from target center) — for
                # the practical tiebreaker test (Mann-Whitney + Cliff's δ).
                a_d_total = a_sum['norm_dists']
                b_d_total = b_sum['norm_dists']
                u_stat, p_mw = _mann_whitney_u(a_d_total, b_d_total)
                delta = _cliffs_delta(a_d_total, b_d_total)
                # Precision input: distance from each group's *own* centroid.
                # Feeding norm_dists here (as the prior implementation did)
                # let any centroid bias leak into the spread test.
                a_d_prec = a_sum['dists_from_centroid']
                b_d_prec = b_sum['dists_from_centroid']
                f_bf, df1_bf, df2_bf, p_bf = _brown_forsythe(a_d_prec, b_d_prec)
                # Accuracy input: full 2D points for Hotelling.
                a_xy = list(zip(a_sum['xs_norm'], a_sum['ys_norm']))
                b_xy = list(zip(b_sum['xs_norm'], b_sum['ys_norm']))
                t2, f_ht, df1_ht, df2_ht, p_ht = _hotelling_t2(a_xy, b_xy)
                pair_stats.append({
                    'a_name': a_name, 'b_name': b_name,
                    'a_sum': a_sum, 'b_sum': b_sum,
                    'u': u_stat, 'p_mw_raw': p_mw, 'delta': delta,
                    'f_bf': f_bf, 'df1_bf': df1_bf, 'df2_bf': df2_bf,
                    'p_bf_raw': p_bf,
                    't2': t2, 'f_ht': f_ht, 'df1_ht': df1_ht,
                    'df2_ht': df2_ht, 'p_ht_raw': p_ht,
                })

        if not pair_stats:
            continue

        # ── Holm correction per test family ──────────────────────────
        # Each test type is its own family of pairwise comparisons within
        # this section. Correcting per family rather than across all three
        # types keeps the question "did mean accuracy differ?" honest
        # without dragging in the bias and spread tests as bystanders.
        p_mw_adj = _holm_bonferroni([ps['p_mw_raw'] for ps in pair_stats])
        p_bf_adj = _holm_bonferroni([ps['p_bf_raw'] for ps in pair_stats])
        p_ht_adj = _holm_bonferroni([ps['p_ht_raw'] for ps in pair_stats])
        for ps, pmw, pbf, pht in zip(pair_stats, p_mw_adj, p_bf_adj, p_ht_adj):
            ps['p_mw'] = pmw
            ps['p_bf'] = pbf
            ps['p_ht'] = pht

        # ── Pass 2: render each pair's panel ─────────────────────────
        # The bar chart shows the three headline numbers archers care
        # about — MPI (accuracy), R95 (precision), hit rate (practical
        # outcome). Mean-distance-from-center is intentionally dropped:
        # it duplicates the muddled signal the overhaul targets.
        for ps in pair_stats:
            a_name, b_name = ps['a_name'], ps['b_name']
            a_sum, b_sum = ps['a_sum'], ps['b_sum']

            fig, axes = plt.subplots(1, 3, figsize=(11, 4))
            metric_titles = [
                'Accuracy — MPI\n(normalized: 1.0 = target edge)',
                'Precision — R95\n(normalized, about each group’s centroid)',
                'Hit rate (%)',
            ]
            a_vals = [a_sum['mpi_norm'], a_sum['r95_norm'], a_sum['hit_rate']]
            b_vals = [b_sum['mpi_norm'], b_sum['r95_norm'], b_sum['hit_rate']]
            fmt_per_axis = ['{:.3f}', '{:.3f}', '{:.1f}']
            for ax, title, av, bv, fmt in zip(
                axes, metric_titles, a_vals, b_vals, fmt_per_axis
            ):
                bars = ax.bar([0, 1], [av, bv],
                              color=[COLOR_A, COLOR_B], edgecolor=EDGE)
                ax.set_title(title, fontsize=10)
                # Full equipment names live in the figure-level color key
                # above instead of the x-tick labels so long names don't
                # overlap.
                ax.set_xticks([])
                ax.grid(True, axis='y', linestyle='--', alpha=0.4)
                for bar, v in zip(bars, (av, bv)):
                    ax.text(bar.get_x() + bar.get_width() / 2,
                            bar.get_height(),
                            fmt.format(v), ha='center', va='bottom', fontsize=9)
                ymax = max(av, bv)
                if ymax > 0:
                    ax.set_ylim(0, ymax * 1.18)
            legend_handles = [
                Patch(facecolor=COLOR_A, edgecolor=EDGE, label=a_name),
                Patch(facecolor=COLOR_B, edgecolor=EDGE, label=b_name),
            ]
            fig.legend(handles=legend_handles, loc='upper center',
                       bbox_to_anchor=(0.5, 0.90),
                       ncol=2, frameon=False, fontsize=10)
            fig.suptitle(f'{label_singular} head-to-head',
                         fontsize=12, fontweight='bold', y=0.99)
            fig.tight_layout(rect=(0, 0, 1, 0.86))
            svg_b64 = _render_matplotlib_svg(fig)

            def _fmt(v, digits=3):
                return '—' if v is None else f'{v:.{digits}f}'

            def _fmt_p(p):
                if p is None:
                    return '—'
                if p < 0.001:
                    return '< 0.001'
                return f'{p:.4f}'

            def _verdict(p, label):
                if p is None:
                    return f'{label}: insufficient data'
                if p < 0.05:
                    return f'{label}: significant (p = {_fmt_p(p)})'
                return f'{label}: inconclusive (p = {_fmt_p(p)})'

            verdict = ' · '.join([
                _verdict(ps['p_ht'], 'Accuracy'),
                _verdict(ps['p_bf'], 'Precision'),
                _verdict(ps['p_mw'], 'Total error'),
            ])

            ht_tip = (
                "Hotelling's T² on the 2D shot vectors: tests whether the "
                "two groups' centroids sit in different places — i.e. an "
                "*accuracy* (bias) difference, independent of how tight "
                "either group is."
            )
            bf_tip = (
                "Brown-Forsythe on distance-from-each-group's-own-centroid: "
                "tests whether one group clusters more tightly than the "
                "other — pure *precision*, independent of bias. (Earlier "
                "versions fed this distance-from-bullseye, which mixed "
                "accuracy and precision.)"
            )
            mw_tip = (
                "Mann-Whitney U on distance-from-target-center: practical "
                "tiebreaker — does one piece simply land closer to the "
                "bull on average? Cannot attribute *why* (bias vs spread); "
                "use the Accuracy / Precision rows above for that."
            )
            delta_tip = (
                "Cliff's δ: effect size for Mann-Whitney. Range −1…+1. "
                "Positive means this group's distances tend to exceed the "
                "other's (i.e., the other group has lower total error); "
                "0 means no stochastic ordering."
            )
            holm_tip = (
                "Holm-Bonferroni step-down adjustment, applied separately "
                "within each test family (Accuracy / Precision / Total). "
                "Multiplies each raw p by the family size minus its rank. "
                "Keeps the family-wise false-positive rate at α even when "
                "many pairs are tested."
            )
            verdict_tip = (
                "Per-axis verdict at α = 0.05 using Holm-adjusted p-values. "
                "'Inconclusive' rather than 'not significant' because small "
                "samples often can't reject H₀ even when a real effect "
                "exists. Shots within a session are correlated, which "
                "these tests don't model — treat as exploratory."
            )
            mpi_tip = (
                "MPI (Mean Point of Impact): magnitude of the group's "
                "centroid offset from the bullseye. Pure *accuracy* — "
                "sight or anchor adjustment, not equipment tuning."
            )
            r95_tip = (
                "R95: radius about each group's own centroid that contains "
                "95% of its shots. The headline *precision* number — "
                "independent of where the centroid sits."
            )
            mr_tip = (
                "Mean Radius: average distance of each shot from the "
                "group's own centroid. Secondary precision number, "
                "canonical in shooting-sport testing."
            )
            sigma_xy_tip = (
                "σ_x / σ_y: per-axis standard deviation about the group "
                "centroid. A large σ_y with small σ_x is vertical "
                "stringing (release / anchor variation); the reverse "
                "is horizontal stringing (bow torque)."
            )
            bias_tip = (
                "Signed bias of the centroid: (Δx, Δy) in normalized "
                "units. Sign tells the user *which direction* to move "
                "the sight, not just by how much."
            )

            section_a = Markup('<em>— Accuracy (bias) —</em>')
            section_p = Markup('<em>— Precision (spread about centroid) —</em>')
            section_t = Markup('<em>— Total error & basics —</em>')
            section_tests = Markup('<em>— Pairwise tests —</em>')

            columns_out = ['Metric', a_name, b_name]
            rows_out = [
                [section_t, '', ''],
                ['Total shots', a_sum['n_total'], b_sum['n_total']],
                ['Hits', a_sum['n_hit'], b_sum['n_hit']],
                ['Misses', a_sum['n_miss'], b_sum['n_miss']],
                ['Hit rate (%)',
                 f"{a_sum['hit_rate']:.1f}",
                 f"{b_sum['hit_rate']:.1f}"],
                ['Mean distance from center',
                 Markup(f"{_mm_val(a_sum['mean_raw_mm'])} {_mm_unit()}"),
                 Markup(f"{_mm_val(b_sum['mean_raw_mm'])} {_mm_unit()}")],
                ['Distinct targets used',
                 a_sum['n_targets'], b_sum['n_targets']],
                [section_a, '', ''],
                [_tip('MPI (normalized)', mpi_tip),
                 f"{a_sum['mpi_norm']:.3f}",
                 f"{b_sum['mpi_norm']:.3f}"],
                [_tip('Bias Δx, Δy (normalized, signed)', bias_tip),
                 f"({a_sum['centroid_norm'][0]:.3f}, "
                 f"{a_sum['centroid_norm'][1]:.3f})",
                 f"({b_sum['centroid_norm'][0]:.3f}, "
                 f"{b_sum['centroid_norm'][1]:.3f})"],
                [section_p, '', ''],
                [_tip('R95 (normalized)', r95_tip),
                 f"{a_sum['r95_norm']:.3f}",
                 f"{b_sum['r95_norm']:.3f}"],
                [_tip('Mean Radius (normalized)', mr_tip),
                 f"{a_sum['mr_norm']:.3f}",
                 f"{b_sum['mr_norm']:.3f}"],
                [_tip('σ_x, σ_y (normalized)', sigma_xy_tip),
                 f"({a_sum['sigma_x_norm']:.3f}, "
                 f"{a_sum['sigma_y_norm']:.3f})",
                 f"({b_sum['sigma_x_norm']:.3f}, "
                 f"{b_sum['sigma_y_norm']:.3f})"],
                [section_tests, '', ''],
                [_tip("Accuracy — Hotelling's T²", ht_tip),
                 (f"{ps['t2']:.3f} (F={ps['f_ht']:.3f}, "
                  f"df={ps['df1_ht']},{ps['df2_ht']})"
                  if ps['t2'] is not None else '—'),
                 ''],
                [_tip('Accuracy — p, Holm-adjusted', holm_tip),
                 _fmt_p(ps['p_ht']), ''],
                [_tip('Precision — Brown-Forsythe F', bf_tip),
                 (f"{ps['f_bf']:.3f} (df={ps['df1_bf']},{ps['df2_bf']})"
                  if ps['f_bf'] is not None else '—'),
                 ''],
                [_tip('Precision — p, Holm-adjusted', holm_tip),
                 _fmt_p(ps['p_bf']), ''],
                [_tip('Total error — Mann-Whitney U', mw_tip),
                 _fmt(ps['u'], 1), ''],
                [_tip('Total error — p, Holm-adjusted', holm_tip),
                 _fmt_p(ps['p_mw']), ''],
                [_tip("Total error — Cliff's δ", delta_tip),
                 _fmt(ps['delta']), ''],
                [_tip('Verdict (α = 0.05)', verdict_tip), verdict, ''],
            ]

            panels.append({
                'key': f'head_to_head__{column}__{a_name}__vs__{b_name}',
                'title': f'{label_singular}: {a_name} vs {b_name}',
                'svg_b64': svg_b64,
                'columns': columns_out,
                'rows': rows_out,
            })

        if panels:
            # ``key`` stays as the catalog key so the template's per-report
            # lookups (date_ranges, categories, tag_selections, download
            # URL) all resolve. ``title`` is what changes per section.
            sections.append({
                'key': 'equipment_head_to_head',
                'title': f'{label_singular} head-to-head',
                'panels': panels,
            })

    if not sections:
        # Build a precise explanation so users know what to add (a second bow
        # or arrow type, more shots on one they already have, etc.) instead
        # of the generic "shoot a session first" fallback.
        bits = []
        for col in ('bow', 'arrow_type', 'tag'):
            info = inventory.get(col) or {}
            label = info.get('label', col)
            distinct = info.get('distinct', 0)
            eligible = info.get('eligible', 0)
            noun = label.lower()
            if distinct == 0:
                bits.append(f'no {noun}s logged on any shot')
            elif distinct == 1:
                bits.append(
                    f'only 1 {noun} logged — need at least 2 to compare'
                )
            elif eligible < 2:
                bits.append(
                    f'{distinct} {noun}s logged but fewer than 2 '
                    f'have ≥{_HEAD_TO_HEAD_MIN_SHOTS} shots'
                )
            else:
                # Eligible pairs existed but every pair had <2 hits to compare.
                bits.append(
                    f'{noun} pairs had too few hits for a t-test'
                )
        reason = (
            'Nothing to compare yet — ' + '; '.join(bits) + '. '
            'Log shots with a second bow, arrow type, or session tag to '
            'enable this report.'
        )
        return {
            'key': 'equipment_head_to_head',
            'title': 'Head-to-head comparisons',
            'empty': True,
            'empty_reason': reason,
        }

    # List return — the route flattens this into the per-kind result cards
    # (one ``<section>`` per Bow / Arrow / Tag head-to-head).
    return sections


# ---------------------------------------------------------------------------
# Quiver-indexed reports (within-session drift, cold-bore vs warmed-up)
# ---------------------------------------------------------------------------
# Arrows aren't entered in shot-order — but quivers are, since the user has
# to commit one quiver before starting the next. So aggregate by
# (session_id, quiver_index_in_session) rather than arrow_id.


def _iter_quivers(user_id):
    """Yield ``(session_id, quiver_idx_1based, shots_in_quiver, target_id)``
    for every *completed* quiver across all of ``user_id``'s sessions.

    The slicer mirrors ``get_stats`` — walk shots in id order, group by
    ``quiver_size`` declared on the first shot of each group, close the
    group when ``len(group) == quiver_size``. A trailing partial quiver
    (in-progress or abandoned mid-batch) is intentionally skipped:
    per-quiver metrics aren't comparable when the group is short.
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT a.session_id, a.target_id, a.quiver_size, "
            "       a.x_coord, a.y_coord, a.arrow_shaft_diameter, "
            "       t.physical_size_mm AS half_src "
            "FROM apollo a "
            "LEFT JOIN targets t ON t.id = a.target_id AND t.user_id = a.user_id "
            "WHERE a.user_id = %s "
            "ORDER BY a.session_id ASC, a.id ASC",
            (user_id,)
        ).fetchall()

    current_session = None
    quiver_idx = 0
    buf = []
    buf_size = 0
    buf_target = None
    for r in rows:
        sid = r['session_id']
        if sid != current_session:
            # New session — reset the quiver walker. Any buffered partial
            # quiver from the previous session is dropped (won't appear).
            current_session = sid
            quiver_idx = 0
            buf = []
            buf_size = 0
        try:
            row_qs = int(r['quiver_size']) if r['quiver_size'] else 0
        except (TypeError, ValueError):
            row_qs = 0
        if row_qs <= 0:
            buf = []
            buf_size = 0
            continue
        if not buf:
            buf_size = row_qs
            buf_target = r['target_id']
        buf.append(r)
        if len(buf) >= buf_size:
            quiver_idx += 1
            yield sid, quiver_idx, list(buf), buf_target
            buf = []
            buf_size = 0


def _quiver_xy(shots):
    """Extract normalized (x_norm, y_norm) pairs from a quiver's shots,
    dropping misses and shots on targets with unknown size. Returns
    (xs, ys) parallel lists in normalized units (1.0 = target edge)."""
    xs = []
    ys = []
    for s in shots:
        xraw = str(s['x_coord']).strip() if s['x_coord'] is not None else ''
        yraw = str(s['y_coord']).strip() if s['y_coord'] is not None else ''
        if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
            continue
        try:
            x = float(xraw)
            y = float(yraw)
        except ValueError:
            continue
        try:
            half = float(s['half_src']) / 2.0
        except (TypeError, ValueError):
            continue
        if half <= 0:
            continue
        xs.append(x / half)
        ys.append(y / half)
    return xs, ys


def _report_within_session_drift(user_id):
    """MPI and R95 by quiver-index-in-session, pooled across sessions.

    Reveals whether the user tightens up after a warm-up quiver or
    starts to loosen late in the session (fatigue). Each quiver index
    contributes shots from every session that reached at least that
    many completed quivers — so later indices come from progressively
    fewer sessions. A "n sessions" bar is rendered alongside MPI/R95
    so the user can see where confidence drops off.
    """
    by_idx = {}  # quiver_idx → {sessions: set, xs: [...], ys: [...]}
    for sid, qidx, shots, _tgt in _iter_quivers(user_id):
        xs, ys = _quiver_xy(shots)
        if not xs:
            continue
        bucket = by_idx.setdefault(qidx, {'sessions': set(),
                                          'xs': [], 'ys': []})
        bucket['sessions'].add(sid)
        bucket['xs'].extend(xs)
        bucket['ys'].extend(ys)

    if not by_idx:
        return None

    indices = sorted(by_idx.keys())
    rows_out = []
    line_x = []
    line_mpi = []
    line_r95 = []
    line_sessions = []
    for q in indices:
        b = by_idx[q]
        s = _archery_stats(b['xs'], b['ys'])
        if s is None:
            continue
        line_x.append(q)
        line_mpi.append(s['mpi'])
        line_r95.append(s['r95'])
        line_sessions.append(len(b['sessions']))
        rows_out.append([
            q, len(b['sessions']), s['n'],
            round(s['mpi'], 3), round(s['r95'], 3),
            round(s['mr'], 3),
        ])

    if not line_x:
        return None

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(9, 4.5))
    # Twin-axis "n sessions" bars in the background so they don't visually
    # compete with the metric lines; metric scale stays unambiguous.
    ax_bar = ax.twinx()
    ax_bar.bar(line_x, line_sessions, color='#c9d3e3',
               edgecolor='#95abcf', zorder=1, alpha=0.55,
               label='Sessions reaching this quiver')
    ax_bar.set_ylabel('Sessions reaching this quiver', color='#5a6b8a')
    ax_bar.tick_params(axis='y', labelcolor='#5a6b8a')
    ax.plot(line_x, line_mpi, color='#1a3a5c', marker='o',
            linewidth=1.8, markersize=5, zorder=3,
            label='Accuracy — MPI')
    ax.plot(line_x, line_r95, color='#c79b5a', marker='s',
            linewidth=1.8, markersize=5, zorder=3,
            label='Precision — R95')
    ax.set_xlabel('Quiver index within session (1 = first quiver)')
    ax.set_ylabel('Normalized units (1.0 = target edge)\nlower is better')
    ax.set_title('Within-session drift')
    ax.set_xticks(line_x)
    ax.grid(True, axis='y', linestyle='--', alpha=0.4)
    ax.set_zorder(ax_bar.get_zorder() + 1)
    ax.patch.set_visible(False)
    # Combine legends from both axes.
    h1, l1 = ax.get_legend_handles_labels()
    h2, l2 = ax_bar.get_legend_handles_labels()
    ax.legend(h1 + h2, l1 + l2, loc='upper left', fontsize=9)
    svg_b64 = _render_matplotlib_svg(fig)

    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this answers:</strong> do you warm up into a tighter '
        'group as the session goes on, or do you fatigue and open up? '
        'Each quiver position pools its shots across every session that '
        'reached that many quivers — so the leftmost bucket sees every '
        'session, the rightmost only your longest ones. The grey bars '
        'show how many sessions contributed to each bucket.'
        '</p>'
    )

    return {
        'key': 'within_session_drift',
        'title': 'Within-session drift',
        'intro_html': intro,
        'svg_b64': svg_b64,
        'columns': [
            'Quiver index', 'Sessions', 'Shots',
            _tip('MPI (norm)',
                 'Mean Point of Impact at this quiver position, pooled '
                 'across sessions. Lower = better accuracy.'),
            _tip('R95 (norm)',
                 'R95 about each quiver-position\'s pooled centroid. '
                 'Lower = tighter group.'),
            _tip('MR (norm)',
                 'Mean Radius about the pooled centroid.'),
        ],
        'rows': rows_out,
    }


# The six toggleable traces for the accuracy/precision report. Each tuple is
# (category_key, granularity, metric, legend_label); the registry derives its
# category pills from this list, and the report reads it to decide which
# series to draw.
_TRACE_DEFS = [
    ('session_acc',  'session', 'acc',  'Accuracy — per session (MPI)'),
    ('session_prec', 'session', 'prec', 'Precision — per session (R95)'),
    ('quiver_acc',   'quiver',  'acc',  'Accuracy — per quiver (MPI)'),
    ('quiver_prec',  'quiver',  'prec', 'Precision — per quiver (R95)'),
    ('alltime_acc',  'alltime', 'acc',  'Accuracy — all-time rolling (MPI)'),
    ('alltime_prec', 'alltime', 'prec', 'Precision — all-time rolling (R95)'),
]
_TRACE_CATEGORIES = [
    {'key': 'session_acc',  'label': 'Session · accuracy'},
    {'key': 'session_prec', 'label': 'Session · precision'},
    {'key': 'quiver_acc',   'label': 'Quiver · accuracy'},
    {'key': 'quiver_prec',  'label': 'Quiver · precision'},
    {'key': 'alltime_acc',  'label': 'All-time · accuracy'},
    {'key': 'alltime_prec', 'label': 'All-time · precision'},
]


def _report_accuracy_precision_traces(user_id, date_from=None, date_to=None,
                                      categories=None, bow_filter=None,
                                      arrow_filter=None, tag_filter=None):
    """Combined accuracy + precision traces over time, three granularities.

    Plots up to six lines on one date-axis, each independently toggleable
    via ``categories`` (see ``_TRACE_DEFS``):
      * Accuracy per session  (MPI of every shot in the session)
      * Precision per session (R95 about the session centroid)
      * Accuracy per quiver   (MPI of every shot in each completed quiver)
      * Precision per quiver  (R95 about the quiver centroid)
      * Accuracy all-time     (running MPI through every shot to date)
      * Precision all-time    (running R95 through every shot to date)

    Optional head-to-head: ``bow_filter`` / ``arrow_filter`` / ``tag_filter``
    each name specific subjects to overlay. When any are given, every
    selected bow / arrow / tag becomes its own colored set of traces
    (filtered to that subject's shots) on the shared timeline; with none
    given the report draws a single combined "All shots" set, preserving
    its original behavior.

    All values are normalized by the target half-width (1.0 = target edge)
    so mixed-target histories are comparable, matching the convention in
    the other normalized reports. Misses are excluded.
    """
    range_from = None
    range_to = None
    if date_from:
        try:
            range_from = datetime.strptime(date_from, '%Y-%m-%d')
        except ValueError:
            range_from = None
    if date_to:
        try:
            range_to = datetime.strptime(date_to, '%Y-%m-%d') \
                + timedelta(days=1) - timedelta(microseconds=1)
        except ValueError:
            range_to = None

    # Which trace series to draw. Empty / None ≡ all six (the route hands us
    # the full list when the user ticks nothing, but guard here too so the
    # export path and direct callers keep the same default).
    valid_keys = {d[0] for d in _TRACE_DEFS}
    enabled = {c for c in (categories or []) if c in valid_keys}
    if not enabled:
        enabled = set(valid_keys)

    # Build the subject list for the optional head-to-head overlay. Each
    # selected bow / arrow / tag is one subject; de-duped on (kind, key) so
    # a name picked twice doesn't double-plot. No selection → a single
    # combined subject that includes every shot (the classic behavior).
    subjects = []
    seen_subj = set()
    for kind, names in (('bow', bow_filter), ('arrow', arrow_filter),
                        ('tag', tag_filter)):
        for raw in (names or []):
            key = (raw or '').strip().lower()
            if not key or (kind, key) in seen_subj:
                continue
            seen_subj.add((kind, key))
            subjects.append({'kind': kind, 'name': raw.strip(), 'key': key})
    split = bool(subjects)
    if not subjects:
        subjects = [{'kind': None, 'name': 'All shots', 'key': None}]

    # Pull every shot with its session_id, quiver_size, target width and
    # the parent session's begin time, plus the equipment / tag columns so
    # we can filter per subject in Python. Quiver bookkeeping needs the rows
    # in (session, id) order so the slicer reproduces the same quiver
    # grouping the per-shot machinery used at recording time.
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        sql = (
            "SELECT a.session_id, a.id, a.quiver_size, "
            "       a.x_coord, a.y_coord, "
            "       a.bow, a.arrow_type, a.session_tags, "
            "       t.physical_size_mm AS width_mm, "
            "       st.session_begin_time "
            "FROM apollo a "
            "LEFT JOIN session_times st "
            "  ON st.session_id = a.session_id AND st.user_id = a.user_id "
            "LEFT JOIN targets t "
            "  ON t.id = a.target_id AND t.user_id = a.user_id "
            "WHERE a.user_id = %s "
            "  AND st.session_begin_time IS NOT NULL "
            "  AND t.physical_size_mm IS NOT NULL"
        )
        params = [user_id]
        if range_from is not None:
            sql += " AND st.session_begin_time >= %s"
            params.append(range_from)
        if range_to is not None:
            sql += " AND st.session_begin_time <= %s"
            params.append(range_to)
        sql += " ORDER BY a.session_id ASC, a.id ASC"
        rows = cur.execute(sql, tuple(params)).fetchall()

    # First pass: per-shot normalized coords (plus the attrs needed to test
    # subject membership), in (session, id) order so the quiver slicer can
    # walk them naturally below.
    master = {}  # session_id → list of shot dicts
    for r in rows:
        xraw = str(r['x_coord']).strip() if r['x_coord'] is not None else ''
        yraw = str(r['y_coord']).strip() if r['y_coord'] is not None else ''
        if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
            continue
        try:
            x = float(xraw)
            y = float(yraw)
        except ValueError:
            continue
        try:
            width_mm = float(r['width_mm'])
        except (TypeError, ValueError):
            continue
        if width_mm <= 0:
            continue
        half = width_mm / 2.0
        begin_dt = _utc_to_user(r['session_begin_time'])
        if begin_dt is None:
            continue
        begin_dt = begin_dt.replace(tzinfo=None)
        try:
            qs = int(r['quiver_size']) if r['quiver_size'] else 0
        except (TypeError, ValueError):
            qs = 0
        # Tokenize session_tags the same way _tag_shot_samples does so the
        # tag picker's values match what we filter on here.
        tagset = {t.strip().lower()
                  for t in (r['session_tags'] or '').split(',')
                  if t.strip()}
        master.setdefault(r['session_id'], []).append({
            'dt': begin_dt,
            'nx': x / half,
            'ny': y / half,
            'qs': qs,
            'bow': (r['bow'] or '').strip().lower(),
            'arrow': (r['arrow_type'] or '').strip().lower(),
            'tags': tagset,
        })

    if not master:
        return None

    def _shot_matches(shot, subj):
        kind = subj['kind']
        if kind is None:
            return True
        if kind == 'bow':
            return shot['bow'] == subj['key']
        if kind == 'arrow':
            return shot['arrow'] == subj['key']
        return subj['key'] in shot['tags']  # tag

    def _compute_series(sessions_sorted):
        """Per-session / per-quiver / all-time MPI & R95 series + table rows
        for one subject's sessions (already date-sorted). Mirrors the
        original single-group aggregation."""
        out = {
            'session': ([], [], []),   # dates, mpi, r95
            'quiver':  ([], [], []),
            'alltime': ([], [], []),
            'table': [],
        }
        rolling_xs, rolling_ys = [], []
        for sid, shots in sessions_sorted:
            if not shots:
                continue
            sess_dt = shots[0]['dt']
            xs_sess = [s['nx'] for s in shots]
            ys_sess = [s['ny'] for s in shots]
            s_sess = _archery_stats(xs_sess, ys_sess)
            if s_sess is None:
                continue
            out['session'][0].append(sess_dt)
            out['session'][1].append(s_sess['mpi'])
            out['session'][2].append(s_sess['r95'])

            # Walk quivers using the recorded quiver_size on the first shot
            # of each group — same slicer as _iter_quivers. Trailing partial
            # quiver is intentionally skipped.
            buf_x, buf_y = [], []
            buf_size = 0
            n_quivers_in_session = 0
            for shot in shots:
                qs = shot['qs']
                if qs <= 0:
                    buf_x, buf_y, buf_size = [], [], 0
                    continue
                if not buf_x:
                    buf_size = qs
                buf_x.append(shot['nx'])
                buf_y.append(shot['ny'])
                if len(buf_x) >= buf_size:
                    s_q = _archery_stats(buf_x, buf_y)
                    if s_q is not None:
                        out['quiver'][0].append(sess_dt)
                        out['quiver'][1].append(s_q['mpi'])
                        out['quiver'][2].append(s_q['r95'])
                        n_quivers_in_session += 1
                    buf_x, buf_y, buf_size = [], [], 0

            rolling_xs.extend(xs_sess)
            rolling_ys.extend(ys_sess)
            s_all = _archery_stats(rolling_xs, rolling_ys)
            if s_all is not None:
                out['alltime'][0].append(sess_dt)
                out['alltime'][1].append(s_all['mpi'])
                out['alltime'][2].append(s_all['r95'])

            out['table'].append([
                sess_dt.strftime('%Y-%m-%d'),
                sid,
                s_sess['n'],
                n_quivers_in_session,
                round(s_sess['mpi'], 3),
                round(s_sess['r95'], 3),
                round(s_all['mpi'], 3) if s_all else '',
                round(s_all['r95'], 3) if s_all else '',
            ])
        return out

    # Compute each subject's series, dropping subjects that yield no data.
    subjects_with_data = []
    table_rows = []
    for subj in subjects:
        subset = {}
        for sid, shots in master.items():
            kept = ([s for s in shots if _shot_matches(s, subj)]
                    if split else shots)
            if kept:
                subset[sid] = kept
        if not subset:
            continue
        sessions_sorted = sorted(subset.items(),
                                 key=lambda kv: kv[1][0]['dt'])
        series = _compute_series(sessions_sorted)
        if not series['session'][0]:
            continue
        subj['series'] = series
        subjects_with_data.append(subj)
        for row in series['table']:
            table_rows.append([subj['name']] + row if split else row)

    if not subjects_with_data:
        return None

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 5.4))

    # Visual encoding. Single combined subject keeps the classic look:
    # blue = accuracy, green = precision, line style by granularity. Each
    # granularity gets its own shade so overlapping traces of the same
    # metric stay distinguishable beyond line style alone. Split mode uses
    # color = subject; marker = metric; line style = granularity.
    GRAN_LS = {'session': '-', 'quiver': ':', 'alltime': '--'}
    METRIC_MARKER = {'acc': 'o', 'prec': 's'}
    # Accuracy → shades of blue; precision → shades of #667867 (green),
    # darkest for per-session, lightest for all-time rolling.
    ACC_SHADES = {'session': '#13314f', 'quiver': '#2e6ca3', 'alltime': '#7aaad6'}
    PREC_SHADES = {'session': '#3f4d40', 'quiver': '#667867', 'alltime': '#9cb29d'}
    SUBJECT_COLORS = ['#1a3a5c', '#c0392b', '#27ae60', '#8e44ad',
                      '#d68910', '#16a085', '#2c3e50', '#e74c3c']

    for si, subj in enumerate(subjects_with_data):
        series = subj['series']
        subj_color = SUBJECT_COLORS[si % len(SUBJECT_COLORS)]
        for key, gran, metric, base_label in _TRACE_DEFS:
            if key not in enabled:
                continue
            dates, mpis, r95s = series[gran]
            if not dates:
                continue
            yvals = mpis if metric == 'acc' else r95s
            color = subj_color if split else (
                ACC_SHADES[gran] if metric == 'acc' else PREC_SHADES[gran])
            label = f"{subj['name']} — {base_label}" if split else base_label
            ax.plot(dates, yvals, color=color,
                    marker=METRIC_MARKER[metric], markersize=4,
                    linewidth=1.8 if gran == 'alltime' else 1.3,
                    linestyle=GRAN_LS[gran],
                    alpha=0.7 if gran == 'quiver' else 0.95,
                    label=label)

    ax.set_xlabel('Date')
    ax.set_ylabel('Normalized units (1.0 = target edge)\nlower is better')
    ax.set_title('Accuracy & precision traces')
    ax.grid(True, axis='y', linestyle='--', alpha=0.4)
    fig.autofmt_xdate()
    ax.legend(loc='upper left', fontsize=7.5, ncol=2, framealpha=0.85)
    svg_b64 = _render_matplotlib_svg(fig)

    if split:
        overlay = (
            ' Each selected subject (bow / arrow / tag) is a separate color; '
            'marker shape is the metric (accuracy ○ vs precision ▢) and line '
            'style is the granularity (session solid, quiver dotted, all-time '
            'dashed).')
    else:
        overlay = (
            ' Blue traces are accuracy (MPI); green traces are precision '
            '(R95) — each granularity is a distinct shade. Line style also '
            'tracks the granularity (session solid, quiver dotted, all-time '
            'dashed). Pick bows, arrows, or tags to overlay them '
            'head-to-head.')
    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this answers:</strong> are you getting more '
        'accurate (centroid closer to bullseye) or more precise '
        '(tighter group), and on what timescale?' + overlay +
        ' All values are normalized by target half-width so mixed targets '
        'are comparable. Misses are excluded.'
        '</p>'
    )

    base_columns = [
        'Session date', 'Session id',
        _tip('Shots',
             'Hits in this session that contributed to the stats. '
             'Misses excluded.'),
        _tip('Quivers',
             'Number of completed quivers in this session — a '
             'partial last quiver is intentionally skipped.'),
        _tip('Session MPI (norm)',
             'Accuracy for this session: distance from origin to '
             "this session's shot centroid."),
        _tip('Session R95 (norm)',
             'Precision for this session: 95% radius about the '
             "session's centroid."),
        _tip('All-time MPI (norm)',
             'Rolling accuracy across every shot through this '
             'session, inclusive.'),
        _tip('All-time R95 (norm)',
             'Rolling precision across every shot through this '
             'session, inclusive.'),
    ]
    columns = (['Subject'] + base_columns) if split else base_columns

    return {
        'key': 'accuracy_precision_traces',
        'title': 'Accuracy & precision traces',
        'intro_html': intro,
        'svg_b64': svg_b64,
        'columns': columns,
        'rows': table_rows,
    }


def _report_cold_bore_vs_warmed(user_id):
    """Compare quiver 1 of each session against quivers 2+ of that
    session. Pools across all sessions, then runs the same
    accuracy / precision tests the head-to-head report uses.
    """
    cold_xs, cold_ys = [], []
    warm_xs, warm_ys = [], []
    sessions_with_cold = set()
    sessions_with_warm = set()
    for sid, qidx, shots, _tgt in _iter_quivers(user_id):
        xs, ys = _quiver_xy(shots)
        if not xs:
            continue
        if qidx == 1:
            cold_xs.extend(xs)
            cold_ys.extend(ys)
            sessions_with_cold.add(sid)
        else:
            warm_xs.extend(xs)
            warm_ys.extend(ys)
            sessions_with_warm.add(sid)

    if len(cold_xs) < 2 or len(warm_xs) < 2:
        # Not enough data to compare. Return a typed empty so the user
        # gets a precise reason instead of a silent blank card.
        bits = []
        if len(cold_xs) < 2:
            bits.append('not enough first-quiver shots')
        if len(warm_xs) < 2:
            bits.append('not enough later-quiver shots (sessions need 2+ '
                       'completed quivers)')
        return {
            'key': 'cold_bore_vs_warmed',
            'title': 'Cold bore vs warmed up',
            'empty': True,
            'empty_reason': 'Nothing to compare — ' + '; '.join(bits) + '.',
        }

    cold = _archery_stats(cold_xs, cold_ys)
    warm = _archery_stats(warm_xs, warm_ys)

    # Pure-precision test: distance-from-each-group's-own-centroid.
    f_bf, df1_bf, df2_bf, p_bf = _brown_forsythe(
        cold['dists_from_centroid'], warm['dists_from_centroid']
    )
    # Accuracy test: 2D centroid difference.
    a_xy = list(zip(cold_xs, cold_ys))
    b_xy = list(zip(warm_xs, warm_ys))
    t2, f_ht, df1_ht, df2_ht, p_ht = _hotelling_t2(a_xy, b_xy)

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch

    COLOR_A = '#4d6da6'
    COLOR_B = '#fcba03'
    EDGE = '#1a3a5c'
    fig, axes = plt.subplots(1, 3, figsize=(11, 4))
    titles = [
        'Accuracy — MPI\n(normalized: 1.0 = target edge)',
        'Precision — R95\n(normalized, about each pool\'s centroid)',
        'Shots in pool',
    ]
    a_vals = [cold['mpi'], cold['r95'], cold['n']]
    b_vals = [warm['mpi'], warm['r95'], warm['n']]
    fmts = ['{:.3f}', '{:.3f}', '{:.0f}']
    for ax, ti, av, bv, fmt in zip(axes, titles, a_vals, b_vals, fmts):
        bars = ax.bar([0, 1], [av, bv],
                      color=[COLOR_A, COLOR_B], edgecolor=EDGE)
        ax.set_title(ti, fontsize=10)
        ax.set_xticks([])
        ax.grid(True, axis='y', linestyle='--', alpha=0.4)
        for bar, v in zip(bars, (av, bv)):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    fmt.format(v), ha='center', va='bottom', fontsize=9)
        ymax = max(av, bv)
        if ymax > 0:
            ax.set_ylim(0, ymax * 1.18)
    fig.legend(
        handles=[Patch(facecolor=COLOR_A, edgecolor=EDGE,
                       label=f'Cold bore (quiver 1 of {len(sessions_with_cold)} sessions)'),
                 Patch(facecolor=COLOR_B, edgecolor=EDGE,
                       label=f'Warmed up (quivers 2+ of {len(sessions_with_warm)} sessions)')],
        loc='upper center', bbox_to_anchor=(0.5, 0.96),
        ncol=2, frameon=False, fontsize=9)
    fig.tight_layout(rect=(0, 0, 1, 0.86))
    svg_b64 = _render_matplotlib_svg(fig)

    def _fmt_p(p):
        if p is None:
            return '—'
        if p < 0.001:
            return '< 0.001'
        return f'{p:.4f}'

    def _verdict(p, label):
        if p is None:
            return f'{label}: insufficient data'
        if p < 0.05:
            return f'{label}: significant (p = {_fmt_p(p)})'
        return f'{label}: inconclusive (p = {_fmt_p(p)})'

    verdict = ' · '.join([
        _verdict(p_ht, 'Accuracy'),
        _verdict(p_bf, 'Precision'),
    ])

    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this answers:</strong> is your first quiver of '
        'each session enough to warm you up, or are you losing points '
        'before you settle in? Pool 1 = the first completed quiver of '
        'every session; pool 2 = every quiver after that. Two '
        'independent tests run: <em>Hotelling\'s T²</em> for an '
        'accuracy shift, <em>Brown-Forsythe</em> for a precision shift.'
        '</p>'
    )

    section_basics = Markup('<em>— Pool basics —</em>')
    section_acc = Markup('<em>— Accuracy —</em>')
    section_prec = Markup('<em>— Precision —</em>')
    section_tests = Markup('<em>— Pairwise tests —</em>')

    columns_out = ['Metric', 'Cold bore', 'Warmed up']
    rows_out = [
        [section_basics, '', ''],
        ['Sessions contributing',
         len(sessions_with_cold), len(sessions_with_warm)],
        ['Shots', cold['n'], warm['n']],
        [section_acc, '', ''],
        [_tip('MPI (normalized)',
              'Mean Point of Impact: magnitude of the centroid offset. '
              'Lower = more accurate.'),
         f"{cold['mpi']:.3f}", f"{warm['mpi']:.3f}"],
        [_tip('Bias Δx, Δy (normalized, signed)',
              'Signed centroid offset; direction tells you which way the '
              'warm-up shifts your group.'),
         f"({cold['centroid'][0]:.3f}, {cold['centroid'][1]:.3f})",
         f"({warm['centroid'][0]:.3f}, {warm['centroid'][1]:.3f})"],
        [section_prec, '', ''],
        [_tip('R95 (normalized)',
              '95% containment radius about each pool\'s own centroid. '
              'Lower = tighter group, independent of accuracy.'),
         f"{cold['r95']:.3f}", f"{warm['r95']:.3f}"],
        [_tip('Mean Radius (normalized)',
              'Average distance from each pool\'s own centroid.'),
         f"{cold['mr']:.3f}", f"{warm['mr']:.3f}"],
        [section_tests, '', ''],
        [_tip("Accuracy — Hotelling's T²",
              "Hotelling's T² on the 2D shot vectors: tests whether the "
              "two pools' centroids sit in different places — i.e. an "
              "accuracy shift after warm-up, independent of spread."),
         (f"{t2:.3f} (F={f_ht:.3f}, df={df1_ht},{df2_ht})"
          if t2 is not None else '—'),
         ''],
        [_tip('Accuracy — p-value', 'Two-sided p from Hotelling F-test.'),
         _fmt_p(p_ht), ''],
        [_tip('Precision — Brown-Forsythe F',
              "Brown-Forsythe on distance-from-own-centroid: tests "
              "whether one pool clusters more tightly than the other — "
              "pure precision, independent of bias."),
         (f"{f_bf:.3f} (df={df1_bf},{df2_bf})"
          if f_bf is not None else '—'),
         ''],
        [_tip('Precision — p-value', 'Two-sided p from Brown-Forsythe.'),
         _fmt_p(p_bf), ''],
        [_tip('Verdict (α = 0.05)',
              "Per-axis verdict at α = 0.05. Shots within a quiver are "
              "correlated, so treat as exploratory."),
         verdict, ''],
    ]

    return {
        'key': 'cold_bore_vs_warmed',
        'title': 'Cold bore vs warmed up',
        'intro_html': intro,
        'svg_b64': svg_b64,
        'columns': columns_out,
        'rows': rows_out,
    }


# ---------------------------------------------------------------------------
# Shot density heatmap
# ---------------------------------------------------------------------------


def _report_shot_density_heatmap(user_id):
    """Hexbin density of every hit on each target, overlaid on the face.

    Complements the existing scatter report — once a target has more
    than ~500 shots, dots overlap so heavily that structure is lost.
    The hexbin shows where shots actually cluster.
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        target_rows = cur.execute(
            "SELECT id AS rowid, name, image_filename, "
            "       physical_size_mm, image_size_px "
            "FROM targets WHERE user_id = %s "
            "ORDER BY id ASC",
            (user_id,)
        ).fetchall()
    if not target_rows:
        return None

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    panels = []
    for trow in target_rows:
        target_id = int(trow['rowid'])
        target_name = trow['name']
        target_cfg = target_to_config(trow)
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            shots = cur.execute(
                "SELECT x_coord, y_coord FROM apollo "
                "WHERE user_id = %s AND target_id = %s",
                (user_id, target_id)
            ).fetchall()
        xs, ys = [], []
        for s in shots:
            xraw = str(s['x_coord']).strip() if s['x_coord'] is not None else ''
            yraw = str(s['y_coord']).strip() if s['y_coord'] is not None else ''
            if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
                continue
            try:
                xs.append(float(xraw))
                ys.append(float(yraw))
            except ValueError:
                continue
        if len(xs) < 25:
            # Sparse target — the scatter report already covers this case;
            # a heatmap on <25 shots is just a bunch of singletons.
            continue
        half = float(target_cfg['target_width_mm']) / 2.0

        img_disk_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'static',
            target_cfg['target_image']
        )
        try:
            from PIL import Image as PILImage
            bg = PILImage.open(img_disk_path)
        except (FileNotFoundError, OSError):
            bg = None

        fig, ax = plt.subplots(figsize=(7, 7))
        if bg is not None:
            ax.imshow(bg, extent=[-half, half, -half, half], origin='upper',
                      alpha=0.45)
        else:
            ax.set_facecolor('#1a3a5c')
        ax.set_xlim(-half, half)
        ax.set_ylim(-half, half)
        ax.set_aspect('equal')
        # gridsize tuned so each hex covers ~5–6% of the face — fine
        # enough to show structure, coarse enough to hide single-shot noise.
        hb = ax.hexbin(xs, ys, gridsize=22,
                       cmap='YlOrRd', mincnt=1, edgecolors='none',
                       extent=(-half, half, -half, half))
        cb = fig.colorbar(hb, ax=ax, fraction=0.046, pad=0.04)
        cb.set_label('Shots per hex')
        ax.axhline(0, color=(1, 1, 1, 0.35), linewidth=0.6, linestyle=':')
        ax.axvline(0, color=(1, 1, 1, 0.35), linewidth=0.6, linestyle=':')
        ax.set_title(f'{target_name} — shot density ({len(xs)} hits)')
        ax.set_xlabel('X (mm from center)')
        ax.set_ylabel('Y (mm from center)')
        svg_b64 = _render_matplotlib_svg(fig)

        # Tabulate a quadrant breakdown — actionable summary that the
        # heatmap itself only hints at.
        q_counts = {'UR': 0, 'UL': 0, 'LL': 0, 'LR': 0,
                    'on_axis': 0}
        for x, y in zip(xs, ys):
            if x == 0 and y == 0:
                q_counts['on_axis'] += 1
            elif x > 0 and y > 0:
                q_counts['UR'] += 1
            elif x < 0 and y > 0:
                q_counts['UL'] += 1
            elif x < 0 and y < 0:
                q_counts['LL'] += 1
            else:
                q_counts['LR'] += 1
        n = len(xs)
        rows_out = [
            ['Total hits', n],
            ['Upper-right quadrant',
             f"{q_counts['UR']} ({q_counts['UR'] / n * 100:.1f}%)"],
            ['Upper-left quadrant',
             f"{q_counts['UL']} ({q_counts['UL'] / n * 100:.1f}%)"],
            ['Lower-left quadrant',
             f"{q_counts['LL']} ({q_counts['LL'] / n * 100:.1f}%)"],
            ['Lower-right quadrant',
             f"{q_counts['LR']} ({q_counts['LR'] / n * 100:.1f}%)"],
        ]

        panels.append({
            'key': f'shot_density__{target_id}',
            'title': f'{target_name} — shot density ({n} hits)',
            'svg_b64': svg_b64,
            'columns': ['Metric', 'Value'],
            'rows': rows_out,
        })

    if not panels:
        return None

    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this shows:</strong> where your shots actually '
        'cluster, not just where they have landed. The scatter report '
        'plots every shot as a dot; once you have hundreds of shots on '
        'one face the dots overlap and structure is lost. A hexbin '
        'density map keeps the picture readable at any shot count. '
        'Brighter cells = more shots in that area.'
        '</p>'
    )

    return {
        'key': 'shot_density_heatmap',
        'title': 'Shot density heatmap',
        'intro_html': intro,
        'panels': panels,
    }


# ---------------------------------------------------------------------------
# Expected score from fitted bivariate normal
# ---------------------------------------------------------------------------


def _report_expected_score(user_id):
    """For each scoring target, fit a 2D Gaussian to the user's shots,
    Monte-Carlo sample it, and report the expected score per arrow plus
    expected per-end scores at common end lengths.

    Closes the loop between the practice metrics in the rest of /analyze
    and what the archer actually scores on tournament day.
    """
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        target_rows = cur.execute(
            "SELECT t.id AS rowid, t.name, t.image_filename, "
            "       t.physical_size_mm, t.image_size_px "
            "FROM targets t "
            "WHERE t.user_id = %s "
            "  AND EXISTS (SELECT 1 FROM target_zones z "
            "              WHERE z.target_id = t.id AND z.user_id = %s) "
            "ORDER BY t.id ASC",
            (user_id, user_id)
        ).fetchall()
    if not target_rows:
        return None

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import random as _rng

    N_SAMPLES = 20000  # Monte-Carlo budget per target
    END_LENGTHS = (3, 6, 10)

    panels = []
    for trow in target_rows:
        target_id = int(trow['rowid'])
        target_name = trow['name']
        target_cfg = target_to_config(trow)
        zones = _fetch_target_zones(target_id, user_id)
        if not _zones_define_scoring(zones):
            continue

        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            shots = cur.execute(
                "SELECT x_coord, y_coord, arrow_shaft_diameter "
                "FROM apollo WHERE user_id = %s AND target_id = %s",
                (user_id, target_id)
            ).fetchall()
        xs_mm, ys_mm = [], []
        for s in shots:
            xraw = str(s['x_coord']).strip() if s['x_coord'] is not None else ''
            yraw = str(s['y_coord']).strip() if s['y_coord'] is not None else ''
            if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
                continue
            try:
                xs_mm.append(float(xraw))
                ys_mm.append(float(yraw))
            except ValueError:
                continue
        if len(xs_mm) < 10:
            continue

        stats = _archery_stats(xs_mm, ys_mm)
        if stats is None:
            continue
        cx, cy = stats['centroid']
        sx = stats['sigma_x']
        sy = stats['sigma_y']
        rho = stats['rho']
        if sx <= 0 or sy <= 0:
            continue
        # Cholesky of the 2×2 covariance for correlated sampling:
        #   X = cx + sx · z1
        #   Y = cy + sy · (ρ z1 + √(1 − ρ²) z2)
        rho_clip = max(-0.999, min(0.999, rho))
        rho_perp = math.sqrt(max(0.0, 1.0 - rho_clip * rho_clip))

        # Empirical miss rate (sentinels were dropped from xs_mm but
        # we re-count them here so the score model honors misses).
        n_total = len(shots)
        n_misses = sum(
            1 for s in shots
            if (str(s['x_coord']).strip() if s['x_coord'] else '') == MISS_SENTINEL
            and (str(s['y_coord']).strip() if s['y_coord'] else '') == MISS_SENTINEL
        )
        miss_rate = n_misses / n_total if n_total else 0.0

        ring_hits = [0] * len(zones)
        ring_points = [int(z['point_value'] or 0) for z in zones]
        outside = 0
        shaft_d = None  # No per-shot shaft diameter at sampling time —
        # use a typical 6mm shaft for the line-cutter test (most arrows
        # are between 5 and 7 mm; the small mis-classification rate at
        # ring boundaries is well within Monte-Carlo noise).
        for _ in range(N_SAMPLES):
            z1 = _rng.gauss(0, 1)
            z2 = _rng.gauss(0, 1)
            x = cx + sx * z1
            y = cy + sy * (rho_clip * z1 + rho_perp * z2)
            idx = _classify_shot(f'{x:.4f}', f'{y:.4f}', zones, shaft_d)
            if idx is None:
                outside += 1
            else:
                ring_hits[idx] += 1

        # Blend in the empirical miss rate: shots that scored zero in
        # the user's actual history (sentinel-misses) are extra outside
        # mass that the Gaussian can't model.
        eff_samples = N_SAMPLES + int(N_SAMPLES * miss_rate /
                                       max(1e-9, 1 - miss_rate))
        outside_with_misses = outside + (eff_samples - N_SAMPLES)
        denom = eff_samples

        p_per_ring = [h / denom for h in ring_hits]
        p_outside = outside_with_misses / denom

        expected_score = sum(p * pts for p, pts in zip(p_per_ring, ring_points))
        max_ring_points = max(ring_points) if ring_points else 0

        # Per-end expected scores at common end lengths.
        end_table = [
            (n, expected_score * n, max_ring_points * n)
            for n in END_LENGTHS
        ]

        # Bar chart: expected % of arrows in each ring.
        labels = [z['name'] or f'Zone {i + 1}' for i, z in enumerate(zones)]
        labels.append('Miss / outside')
        pcts = [p * 100 for p in p_per_ring]
        pcts.append(p_outside * 100)
        # Reuse target-face color sampler for visual consistency with
        # hits-by-boundaries.
        colors = _sample_zone_colors(zones, target_cfg)
        colors.append('#e53935')

        fig, ax = plt.subplots(figsize=(7.5, 4.5))
        bars = ax.bar(labels, pcts, color=colors, edgecolor='#1a3a5c')
        ax.set_ylabel('Expected % of arrows')
        ax.set_title(f'Expected hit distribution — {target_name}')
        ax.grid(True, axis='y', linestyle='--', alpha=0.4)
        for bar, p in zip(bars, pcts):
            if p > 0.1:
                ax.text(bar.get_x() + bar.get_width() / 2,
                        bar.get_height(),
                        f'{p:.1f}%', ha='center', va='bottom', fontsize=9)
        fig.autofmt_xdate(rotation=20)
        svg_b64 = _render_matplotlib_svg(fig)

        rows_out = [[Markup('<em>— Expected per arrow —</em>'), '', '']]
        rows_out.append(['Expected score (points / arrow)',
                         f'{expected_score:.3f}',
                         f'(max {max_ring_points})'])
        rows_out.append([Markup('<em>— Expected per end —</em>'), '', ''])
        for n, exp, max_n in end_table:
            pct = (exp / max_n * 100) if max_n else 0.0
            rows_out.append([f'{n}-arrow end (expected)',
                             f'{exp:.1f}',
                             f'(max {max_n}, {pct:.1f}% of max)'])
        rows_out.append([Markup('<em>— Ring breakdown —</em>'), '', ''])
        for lab, p, pts in zip(labels[:-1], p_per_ring, ring_points):
            rows_out.append([lab, f'{p * 100:.2f}%',
                             f'×{pts} = {p * pts:.3f}'])
        rows_out.append(['Miss / outside zones',
                         f'{p_outside * 100:.2f}%', '×0 = 0.000'])

        panels.append({
            'key': f'expected_score__{target_id}',
            'title': f'{target_name} — expected score',
            'svg_b64': svg_b64,
            'columns': ['Metric', 'Value', 'Notes'],
            'rows': rows_out,
        })

    if not panels:
        return None

    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this answers:</strong> if your current group held, '
        'what would you score? Each target with scoring zones gets its '
        'own bivariate-normal fit to your shot history; Monte-Carlo '
        'sampling integrates that fit over the rings and projects an '
        'expected score per arrow and per end. Empirical miss rate is '
        'blended in so total-flyers count against the projection.'
        '</p>'
    )

    return {
        'key': 'expected_score',
        'title': 'Expected score from fit',
        'intro_html': intro,
        'panels': panels,
    }


# ---------------------------------------------------------------------------
# Calendar heatmap
# ---------------------------------------------------------------------------


def _report_calendar_heatmap(user_id):
    """GitHub-style year grid of shots-per-day across the user's history."""
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT st.session_begin_time, COUNT(a.id) AS shots "
            "FROM session_times st "
            "LEFT JOIN apollo a ON a.session_id = st.session_id "
            "                   AND a.user_id   = st.user_id "
            "WHERE st.user_id = %s AND st.session_begin_time IS NOT NULL "
            "GROUP BY st.session_id, st.session_begin_time "
            "ORDER BY st.session_begin_time ASC",
            (user_id,)
        ).fetchall()
    if not rows:
        return None

    # Aggregate to per-day shot counts in the user's timezone.
    from collections import defaultdict
    per_day = defaultdict(int)
    for r in rows:
        dt = _utc_to_user(r['session_begin_time'])
        if dt is None:
            continue
        d = dt.date()
        per_day[d] += int(r['shots'] or 0)
    if not per_day:
        return None

    sorted_dates = sorted(per_day.keys())
    first_d = sorted_dates[0]
    last_d = sorted_dates[-1]

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.colors import LinearSegmentedColormap, Normalize
    from matplotlib.patches import Rectangle

    # Build a year-by-year grid. Each year is a separate calendar so
    # longer histories don't smear into one impossible-to-read strip.
    years = list(range(first_d.year, last_d.year + 1))
    cmap = LinearSegmentedColormap.from_list(
        'apollo_volume',
        ['#eef3fa', '#a8c4e8', '#4d6da6', '#c79b5a'],
    )
    max_shots = max(per_day.values())
    norm = Normalize(vmin=0, vmax=max(1, max_shots))

    fig, axes = plt.subplots(
        len(years), 1,
        figsize=(11, max(2.4, 1.6 * len(years))),
        squeeze=False,
    )
    for ax, yr in zip(axes[:, 0], years):
        jan1 = datetime(yr, 1, 1).date()
        dec31 = datetime(yr, 12, 31).date()
        n_days = (dec31 - jan1).days + 1
        # The grid: 7 rows (weekdays, Mon top), columns = ISO weeks.
        # We map each date to (weekday_idx, week_idx_within_year).
        first_weekday = jan1.weekday()  # Mon=0
        ax.set_xlim(-0.5, 53.5)
        ax.set_ylim(-0.5, 6.5)
        ax.set_aspect('equal')
        ax.invert_yaxis()  # Mon at top
        ax.set_yticks(range(7))
        ax.set_yticklabels(['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
                           fontsize=7)
        ax.tick_params(axis='y', length=0)
        # Month ticks along the bottom.
        from datetime import timedelta as _td
        month_starts = []
        for m in range(1, 13):
            md = datetime(yr, m, 1).date()
            offset = (md - jan1).days
            col = (offset + first_weekday) // 7
            month_starts.append((col, md.strftime('%b')))
        ax.set_xticks([c for c, _ in month_starts])
        ax.set_xticklabels([lab for _, lab in month_starts], fontsize=8)
        ax.tick_params(axis='x', length=0)
        for spine in ax.spines.values():
            spine.set_visible(False)
        # Draw the cells.
        for i in range(n_days):
            d = jan1 + _td(days=i)
            col = (i + first_weekday) // 7
            row = d.weekday()
            count = per_day.get(d, 0)
            if d > last_d:
                # Future day in the same year as last_d — leave blank.
                color = '#ffffff'
            else:
                color = cmap(norm(count)) if count > 0 else '#eef3fa'
            ax.add_patch(Rectangle(
                (col - 0.45, row - 0.45), 0.9, 0.9,
                facecolor=color, edgecolor='#dde4ee', linewidth=0.35,
            ))
        ax.set_title(f'{yr}', loc='left', fontsize=10,
                     fontweight='bold', color='#1a3a5c')

    fig.suptitle('Shot volume calendar', fontsize=12,
                 fontweight='bold', color='#1a3a5c')
    # Colorbar on the side.
    cax = fig.add_axes([0.92, 0.15, 0.012, 0.7])
    import matplotlib.cm as cm
    sm = cm.ScalarMappable(norm=norm, cmap=cmap)
    sm.set_array([])
    cb = fig.colorbar(sm, cax=cax)
    cb.set_label('Shots / day', fontsize=8)
    cb.ax.tick_params(labelsize=7)
    # Manual colorbar axes are incompatible with tight_layout; lay out
    # the main grid via subplots_adjust instead so matplotlib stops
    # warning.
    fig.subplots_adjust(left=0.06, right=0.9, top=0.92, bottom=0.08,
                        hspace=0.55)
    svg_b64 = _render_matplotlib_svg(fig)

    # Summary table.
    total_shots = sum(per_day.values())
    active_days = sum(1 for v in per_day.values() if v > 0)
    span_days = (last_d - first_d).days + 1
    busiest_d, busiest_n = max(per_day.items(), key=lambda kv: kv[1])
    rows_out = [
        ['Span', f'{first_d.isoformat()} → {last_d.isoformat()} '
                 f'({span_days} day{"s" if span_days != 1 else ""})'],
        ['Active days', f'{active_days} of {span_days} '
                       f'({active_days / span_days * 100:.1f}%)'],
        ['Total shots', total_shots],
        ['Busiest day',
         f'{busiest_d.isoformat()} — {busiest_n} shot'
         f'{"s" if busiest_n != 1 else ""}'],
        ['Average on active days',
         f'{total_shots / active_days:.1f} shots/day' if active_days else '—'],
    ]

    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this shows:</strong> a GitHub-style year grid of '
        'shots per calendar day. Streaks and gaps are visible at a '
        'glance — useful for honest answers to "am I really practicing '
        'as much as I think?" Brighter cells = busier days.'
        '</p>'
    )

    return {
        'key': 'calendar_heatmap',
        'title': 'Shot volume calendar',
        'intro_html': intro,
        'svg_b64': svg_b64,
        'columns': ['Metric', 'Value'],
        'rows': rows_out,
    }


# ---------------------------------------------------------------------------
# Handicap trend
# ---------------------------------------------------------------------------
# Apollo computes an Archery GB 2023 handicap for every completed AGB-target
# round (see ``_session_handicap``), but only ever as a per-round footnote.
# This report aggregates those per-round handicaps into a trend over time and
# the three headline figures archers actually care about.


def _handicap_summary(points):
    """Headline handicap figures from a list of completed-round handicaps.

    Pure (no DB / Flask) so it unit-tests like ``handicap.py``. ``points`` is
    a list of ``{'date': datetime, 'handicap': int, ...}`` sorted ascending by
    date (lower handicap = better). Returns a dict with three figures, each an
    int or ``None`` when there isn't enough data:

    * ``latest`` — the most recent round's handicap.
    * ``best_recent`` — the best (lowest) single handicap in the 12 months up
      to the most recent round. A recent personal best.
    * ``agb`` — Archery GB's official figure: the **average of the best three
      handicaps, rounded down** (the 2023 scheme rounds averages down, not up).
      Taken from the current season (the calendar year of the latest round); if
      that season has fewer than three rounds it falls back to the trailing 12
      months, then to all rounds, and ``agb_basis`` records which was used.

    Only competed rounds feed this figure: match play and score sheets from
    real competitions. Practice — both the live on-screen Practice mode and
    paper practice scorecards — is excluded by the caller
    (``_report_handicap_trend``), so casual shooting never moves the number.
    It is still a personal tracking figure, not a club-submitted one.
    """
    if not points:
        return {'latest': None, 'best_recent': None,
                'agb': None, 'agb_basis': None}

    latest = points[-1]['handicap']
    ref = points[-1]['date']
    trailing_12mo = [p['handicap'] for p in points
                     if (ref - p['date']).days <= 365]
    best_recent = min(trailing_12mo) if trailing_12mo else latest

    # AGB official: best three, averaged, rounded down. Prefer the current
    # season (calendar year of the latest round); degrade to wider pools when
    # the season is too thin, flagging the result as provisional.
    season = sorted(p['handicap'] for p in points if p['date'].year == ref.year)
    if len(season) >= 3:
        pool, basis = season, f'season {ref.year}'
    elif len(trailing_12mo) >= 3:
        pool, basis = sorted(trailing_12mo), 'last 12 months (provisional)'
    else:
        pool, basis = sorted(p['handicap'] for p in points), 'all rounds (provisional)'
    if len(pool) >= 3:
        agb = math.floor(sum(pool[:3]) / 3)
    else:
        agb, basis = None, None

    return {'latest': latest, 'best_recent': best_recent,
            'agb': agb, 'agb_basis': basis}


def _report_handicap_trend(user_id):
    """Archery GB handicap for every completed AGB-target round, over time."""
    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT st.session_id, st.session_begin_time, "
            "       (SELECT a.session_tags FROM apollo a "
            "          WHERE a.session_id = st.session_id "
            "            AND a.user_id = st.user_id "
            "          ORDER BY a.id LIMIT 1) AS tags "
            "FROM session_times st "
            "WHERE st.user_id = %s AND st.session_begin_time IS NOT NULL "
            "ORDER BY st.session_begin_time ASC",
            (user_id,)
        ).fetchall()

    points = []
    for r in rows:
        # Only competed rounds feed the running handicap — match play or a
        # score sheet logged from a real competition. Practice never counts,
        # neither the live on-screen Practice mode nor a paper practice
        # scorecard.
        if not _counts_toward_handicap(r['tags']):
            continue
        round_key = _round_key_from_tags(r['tags'])
        # Cheap filter first: skip non-tournament and non-AGB rounds before
        # paying for a per-session progress recompute.
        if not round_key or round_key not in classifications.data.AGB_TARGET_ROUNDS:
            continue
        round_def = _tournament_round_def(round_key)
        if not round_def:
            continue
        prog = _compute_tournament_progress(r['session_id'], user_id, round_def)
        if not prog.get('is_complete'):
            continue
        hc = _session_handicap(round_key, prog.get('total_score'))
        if hc is None:
            continue
        begin_dt = _utc_to_user(r['session_begin_time'])
        if begin_dt is None:
            continue
        begin_dt = begin_dt.replace(tzinfo=None)
        points.append({
            'date': begin_dt,
            'handicap': hc,
            'round_name': round_def.get('name', round_key),
            'score': prog.get('total_score'),
        })

    if not points:
        return {
            'key': 'handicap_trend',
            'title': 'Handicap over time',
            'empty': True,
            'empty_reason': 'No completed Archery GB target rounds yet. '
                            'Shoot a recognised AGB round to completion and '
                            'its handicap will appear here.',
        }

    summary = _handicap_summary(points)

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates

    fig, ax = plt.subplots(figsize=(9, 4.5))
    xs = [p['date'] for p in points]
    ys = [p['handicap'] for p in points]
    ax.plot(xs, ys, '-', color='#9bb0d0', linewidth=1.0, zorder=1)
    ax.scatter(xs, ys, s=36, color='#4d6da6', edgecolor='#1a3a5c',
               zorder=3, label='Completed round')

    # Least-squares trend line (improvement direction). Fit on ordinal days so
    # the slope is handicap-points per day; only meaningful with ≥2 rounds.
    if len(points) >= 2:
        import numpy as np
        t0 = xs[0]
        days = np.array([(x - t0).total_seconds() / 86400.0 for x in xs])
        slope, intercept = np.polyfit(days, np.array(ys, dtype=float), 1)
        ax.plot(xs, intercept + slope * days, '--', color='#c0504d',
                linewidth=1.4, zorder=2,
                label=f'Trend ({slope * 365.0:+.1f} / yr)')

    # AGB official figure as a reference line.
    if summary['agb'] is not None:
        ax.axhline(summary['agb'], color='#3f7d3f', linewidth=1.2,
                   linestyle=':', zorder=2,
                   label=f"AGB handicap {summary['agb']}")

    # Lower is better → invert so improvement reads as "up".
    ax.invert_yaxis()
    ax.set_xlabel('Date')
    ax.set_ylabel('Handicap (lower is better)')
    ax.set_title('Archery GB handicap over time')
    ax.grid(True, axis='y', linestyle='--', alpha=0.4)
    ax.legend(loc='best', fontsize=9)
    locator = mdates.AutoDateLocator()
    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))
    fig.autofmt_xdate()
    svg_b64 = _render_matplotlib_svg(fig)

    def _fig(v):
        return '—' if v is None else str(v)

    agb_cell = _fig(summary['agb'])
    if summary['agb'] is not None and summary['agb_basis']:
        agb_cell = f"{summary['agb']} ({summary['agb_basis']})"

    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this shows:</strong> the Archery GB 2023 handicap for '
        'every recognised AGB target round you have completed, plotted over '
        'time (the y-axis is inverted, so a rising line means you are '
        'improving). Three headline figures sit below.'
        '</p>'
        '<div class="handicap-headlines" style="display:flex;flex-wrap:wrap;'
        'gap:14px;margin:6px 0 4px 0;">'
        + ''.join(
            f'<div style="flex:1 1 150px;min-width:150px;padding:10px 12px;'
            f'border:1px solid #d6deeb;border-radius:8px;background:#f6f9fd;">'
            f'<div style="font-size:0.78em;text-transform:uppercase;'
            f'letter-spacing:0.04em;color:#5a6a82;">{label}</div>'
            f'<div style="font-size:1.6em;font-weight:700;color:#1a3a5c;">'
            f'{value}</div>'
            f'<div style="font-size:0.74em;color:#7a879c;">{sub}</div>'
            f'</div>'
            for label, value, sub in (
                ('Latest', _fig(summary['latest']),
                 'most recent round'),
                ('Best (recent)', _fig(summary['best_recent']),
                 'lowest in last 12 months'),
                ('AGB handicap', _fig(summary['agb']),
                 summary['agb_basis'] or 'needs 3 rounds'),
            )
        )
        + '</div>'
        '<p class="report-intro" style="font-size:0.82em;color:#7a879c;">'
        'The AGB figure is the average of your best three handicaps, rounded '
        'down (per the 2023 scheme). Every completed round counts — Apollo '
        'does not distinguish practice from record shoots — so treat it as a '
        'personal tracking number.'
        '</p>'
    )

    columns = ['Date', 'Round', 'Score', 'Handicap']
    rows_out = [
        [p['date'].strftime('%Y-%m-%d'), p['round_name'], p['score'],
         p['handicap']]
        for p in reversed(points)  # newest first in the table
    ]
    # Summary rows up top so the table mirrors the headline cards.
    rows_out = [
        [Markup('<em>— Headline —</em>'), '', '', ''],
        ['Latest', '', '', _fig(summary['latest'])],
        ['Best (last 12 months)', '', '', _fig(summary['best_recent'])],
        ['AGB handicap', agb_cell, '', _fig(summary['agb'])],
        [Markup('<em>— Per round (newest first) —</em>'), '', '', ''],
    ] + rows_out

    return {
        'key': 'handicap_trend',
        'title': 'Handicap over time',
        'intro_html': intro,
        'svg_b64': svg_b64,
        'columns': columns,
        'rows': rows_out,
    }


def _report_quiver_spread(user_id, date_from=None, date_to=None,
                          tag_filter=None, categories=None,
                          bow_filter=None, arrow_filter=None):
    """Per-quiver chart of the biggest and smallest pairwise arrow distance.

    For each completed quiver matching the active filters, compute the
    max and min Euclidean distance between any two hit arrows in the
    quiver, normalized by the target face width so mixed face sizes
    are comparable on one chart. A spread of 1.0 means the pair sat
    at opposite edges of the target; 0.0 means they landed in the
    same hole. Plot the two series on a shared chronological x-axis —
    the band between them is the within-quiver consistency.

    Quivers with fewer than 2 hits (all misses, or only one arrow on
    the face) are skipped — pairwise distance is undefined there. So
    are quivers whose target has no recorded ``physical_size_mm``;
    we'd have nothing to normalize by.

    Filter semantics for this report are all *additive narrowing* on top
    of the default "every quiver ever":
      * ``date_from`` / ``date_to``: YYYY-MM-DD; either side may be
        omitted for an open bound. A dated filter drops quivers whose
        session has no recorded begin time.
      * ``categories``: subset of session-type keys to include —
        ``regular`` (no auto-tags) or ``tournament_practice`` (any
        ``tournament:*`` or ``practice`` auto-tag, since live-shot
        tournament rounds are all practice in Apollo's model).
        Empty / ``None`` = include both.
      * ``bow_filter`` / ``arrow_filter``: allow-list of bow models or
        arrow names (case-insensitive). Empty / ``None`` = include all
        equipment. Matched against the bow / arrow recorded on the
        first shot of the quiver.
      * ``tag_filter``: allow-list of ``session_tags``. Empty /
        ``None`` = include every tag (matches the user's "default to
        all shots" requirement — note this is *different* from
        head-to-head's semantics, where an empty list means "match
        nothing").
    """
    range_from = None
    range_to = None
    if date_from:
        try:
            range_from = datetime.strptime(date_from, '%Y-%m-%d')
        except ValueError:
            range_from = None
    if date_to:
        try:
            range_to = datetime.strptime(date_to, '%Y-%m-%d') \
                + timedelta(days=1) - timedelta(microseconds=1)
        except ValueError:
            range_to = None

    def _norm_set(values):
        """None or empty list → no filter; otherwise a lowercased set."""
        if not values:
            return None
        norm = {str(v).strip().lower() for v in values if str(v).strip()}
        return norm or None

    wanted_tags = _norm_set(tag_filter)
    wanted_bows = _norm_set(bow_filter)
    wanted_arrows = _norm_set(arrow_filter)

    # All live-target tournament rounds are practice — actual
    # competition data only enters Apollo via the scoresheet flow,
    # which doesn't produce per-shot rows in the apollo table. So the
    # quiver-spread world only has two flavors of session.
    ALL_TYPES = {'regular', 'tournament_practice'}
    if categories:
        wanted_types = {c for c in categories if c in ALL_TYPES}
        if not wanted_types:
            wanted_types = set(ALL_TYPES)
    else:
        wanted_types = set(ALL_TYPES)

    with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
        rows = cur.execute(
            "SELECT a.session_id, a.id, a.quiver_size, "
            "       a.x_coord, a.y_coord, a.session_tags, "
            "       a.bow, a.arrow_type, "
            "       st.session_begin_time, "
            "       t.physical_size_mm AS face_mm "
            "FROM apollo a "
            "LEFT JOIN session_times st "
            "  ON st.session_id = a.session_id AND st.user_id = a.user_id "
            "LEFT JOIN targets t "
            "  ON t.id = a.target_id AND t.user_id = a.user_id "
            "WHERE a.user_id = %s "
            "ORDER BY a.session_id ASC, a.id ASC",
            (user_id,)
        ).fetchall()

    # Slice into completed quivers — same rule as _iter_quivers (group by
    # the first shot's declared quiver_size, close the group when it
    # hits that size). Filters are applied per quiver at close-time.
    quivers = []
    current_session = None
    quiver_idx = 0
    buf = []
    buf_size = 0
    buf_begin = None
    buf_tags = None
    for r in rows:
        sid = r['session_id']
        if sid != current_session:
            current_session = sid
            quiver_idx = 0
            buf = []
            buf_size = 0
        try:
            row_qs = int(r['quiver_size']) if r['quiver_size'] else 0
        except (TypeError, ValueError):
            row_qs = 0
        if row_qs <= 0:
            buf = []
            buf_size = 0
            continue
        if not buf:
            buf_size = row_qs
            raw_begin = r['session_begin_time']
            # SQLite returns the DateTime column as a naïve string under
            # some adapter configs; coerce it once here so every
            # downstream comparison/sort gets a real datetime (or None).
            if isinstance(raw_begin, str):
                buf_begin = None
                for fmt in ('%Y-%m-%d %H:%M:%S.%f',
                            '%Y-%m-%d %H:%M:%S',
                            '%Y-%m-%dT%H:%M:%S.%f',
                            '%Y-%m-%dT%H:%M:%S'):
                    try:
                        buf_begin = datetime.strptime(raw_begin, fmt)
                        break
                    except ValueError:
                        continue
            else:
                buf_begin = raw_begin
            tag_str = r['session_tags'] or ''
            buf_tags = {t.strip().lower()
                        for t in tag_str.split(',') if t.strip()}
            # Equipment is recorded per-shot; use the first shot's bow
            # and arrow_type as the quiver's representative kit. (Most
            # users don't swap mid-quiver; for the rare ones who do
            # the first shot is the right tiebreaker for the filter.)
            buf_bow = (r['bow'] or '').strip().lower()
            buf_arrow = (r['arrow_type'] or '').strip().lower()
            # Target face size in mm — used to normalize the pairwise
            # distance so mixed face sizes are comparable. Quivers
            # whose target has no recorded size are dropped at close
            # time; a normalized distance is meaningless without it.
            try:
                buf_face_mm = float(r['face_mm']) if r['face_mm'] else 0.0
            except (TypeError, ValueError):
                buf_face_mm = 0.0
        buf.append(r)
        if len(buf) >= buf_size:
            quiver_idx += 1
            keep = True
            # Date filter: a quiver without a usable begin_dt can't be
            # ordered chronologically, so drop it when the user has
            # actively scoped to a date range. Without a filter it's
            # still useful — keep it and let it sort to the end.
            if (range_from or range_to) and not buf_begin:
                keep = False
            if keep and range_from and buf_begin and buf_begin < range_from:
                keep = False
            if keep and range_to and buf_begin and buf_begin > range_to:
                keep = False
            # Two session-type buckets — anything carrying a
            # tournament:* or practice auto-tag is treated as
            # tournament practice (the only kind of tournament data
            # that lands here as live shots).
            if keep:
                has_auto = any(
                    t == 'practice' or t.startswith('tournament:')
                    for t in buf_tags
                )
                stype = 'tournament_practice' if has_auto else 'regular'
                if stype not in wanted_types:
                    keep = False
            if keep and wanted_bows is not None and buf_bow not in wanted_bows:
                keep = False
            if (keep and wanted_arrows is not None
                    and buf_arrow not in wanted_arrows):
                keep = False
            if keep and wanted_tags is not None and not (buf_tags & wanted_tags):
                keep = False
            # Without a target face size we can't normalize. Mixing
            # raw mm and normalized values on one chart would be
            # misleading, so drop these quivers entirely.
            if keep and buf_face_mm <= 0:
                keep = False
            if keep:
                hits = []
                for s in buf:
                    xraw = (str(s['x_coord']).strip()
                            if s['x_coord'] is not None else '')
                    yraw = (str(s['y_coord']).strip()
                            if s['y_coord'] is not None else '')
                    # A miss is stored as the sentinel on both axes —
                    # not a real coordinate, so it can't contribute to
                    # a pairwise spread.
                    if xraw == MISS_SENTINEL and yraw == MISS_SENTINEL:
                        continue
                    try:
                        hits.append((float(xraw), float(yraw)))
                    except ValueError:
                        continue
                if len(hits) >= 2:
                    quivers.append({
                        'begin_dt': buf_begin,
                        'sid': sid,
                        'qidx': quiver_idx,
                        'hits': hits,
                        'face_mm': buf_face_mm,
                    })
            buf = []
            buf_size = 0

    if not quivers:
        return {
            'key': 'quiver_spread',
            'title': 'Biggest vs smallest spread per quiver',
            'empty': True,
            'empty_reason': ('No completed quivers with ≥2 hits match '
                             'these filters.'),
        }

    # Chronological x-axis; undated quivers sort to the end so they
    # don't break the time ordering for everything else.
    quivers.sort(key=lambda q: (q['begin_dt'] or datetime.max, q['qidx']))

    # Spread is reported normalized by the full target face width so
    # quivers shot on different face sizes are directly comparable.
    # A value of 1.0 means the pair sat at opposite edges of the
    # target; 0.0 means the pair landed in the same hole.
    biggest = []
    smallest = []
    rows_out = []
    for i, q in enumerate(quivers, start=1):
        hits = q['hits']
        face_mm = q['face_mm']  # guaranteed > 0 by the keep gate above
        dmax = 0.0
        dmin = None
        for a_i in range(len(hits)):
            for b_i in range(a_i + 1, len(hits)):
                dx = hits[a_i][0] - hits[b_i][0]
                dy = hits[a_i][1] - hits[b_i][1]
                d = math.hypot(dx, dy) / face_mm
                if d > dmax:
                    dmax = d
                if dmin is None or d < dmin:
                    dmin = d
        if dmin is None:
            dmin = 0.0
        biggest.append(dmax)
        smallest.append(dmin)
        date_str = (q['begin_dt'].strftime('%Y-%m-%d')
                    if q['begin_dt'] else '—')
        rows_out.append([
            i, date_str, q['qidx'], len(hits),
            round(face_mm, 0),
            round(dmax, 3), round(dmin, 3),
        ])

    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    xs = list(range(1, len(quivers) + 1))
    fig, ax = plt.subplots(figsize=(9, 4.5))
    ax.fill_between(xs, smallest, biggest, color='#b3c6e3', alpha=0.25,
                    label='Within-quiver gap')
    ax.plot(xs, biggest, color='#1a3a5c', marker='o',
            linewidth=1.6, markersize=4,
            label='Biggest spread (max pair)')
    ax.plot(xs, smallest, color='#c79b5a', marker='s',
            linewidth=1.6, markersize=4,
            label='Smallest spread (min pair)')
    ax.set_xlabel('Session date (sequential quivers in date range)')
    ax.set_ylabel('Arrow-pair distance ÷ target face width\n(1.0 = opposite edges of target)')
    ax.set_title('Biggest vs smallest spread per quiver')
    ax.grid(True, axis='y', linestyle='--', alpha=0.4)
    ax.legend(loc='upper left', fontsize=9)
    # X-axis: each quiver is one evenly-spaced point, but the tick
    # *labels* show the session date so the user can read off when
    # each quiver happened. On long histories we thin the ticks so
    # labels don't overrun each other.
    date_labels = [
        q['begin_dt'].strftime('%Y-%m-%d') if q['begin_dt'] else '—'
        for q in quivers
    ]
    if len(xs) <= 14:
        tick_positions = xs
    else:
        # Aim for ~12 evenly-spaced ticks across the series.
        step = max(1, len(xs) // 12)
        tick_positions = list(range(1, len(xs) + 1, step))
        # Always include the last quiver so the right end of the chart
        # has a real anchor label.
        if tick_positions[-1] != len(xs):
            tick_positions.append(len(xs))
    tick_labels = [date_labels[p - 1] for p in tick_positions]
    ax.set_xticks(tick_positions)
    ax.set_xticklabels(tick_labels, rotation=35, ha='right')
    fig.subplots_adjust(bottom=0.22)
    svg_b64 = _render_matplotlib_svg(fig)

    intro = Markup(
        '<p class="report-intro">'
        '<strong>What this answers:</strong> for each completed quiver, '
        'how spread out were the arrows across the <em>whole target</em>? '
        'The blue trace is the worst pair (biggest gap between any two '
        'arrows in the quiver); the amber trace is the tightest pair. '
        'The shaded band between them is the within-quiver spread '
        'gap — a thinner band means a more consistent group. '
        'Distances are normalized by the target face width, so '
        '<strong>1.0</strong> = arrows at opposite edges of the target '
        'and <strong>0.0</strong> = arrows in the same hole. Mixed face '
        'sizes are directly comparable on this scale. Quivers shot at '
        'a target without a recorded physical size are skipped.'
        '</p>'
    )

    return {
        'key': 'quiver_spread',
        'title': 'Biggest vs smallest spread per quiver',
        'intro_html': intro,
        'svg_b64': svg_b64,
        'columns': [
            '#', 'Session date', 'Quiver in session', 'Hits',
            'Face width (mm)',
            'Biggest spread (norm)', 'Smallest spread (norm)',
        ],
        'rows': rows_out,
    }


REPORTS = {
    'arrows_vs_time': {
        'label': 'Arrows shot vs time',
        'description': 'Arrows shot per calendar day (empty days included '
                       'so gaps are visible). Table lists the underlying '
                       'per-session counts.',
        'fn': _report_arrows_vs_time,
    },
    'all_shots_per_target': {
        'label': 'Shots per target (date range)',
        'description': 'Every shot you have taken on each target — overlaid '
                       'on the target face with centroid and group-spread '
                       'circle so trends (left/right bias, cluster size) are '
                       'obvious. Optionally restrict to a date range; leave '
                       'both empty to include all shots ever.',
        'fn': _report_all_shots_per_target,
        # Marker used by the template to render the date-range inputs and
        # by analyze() / analyze_export() to forward dates into the report.
        'accepts_date_range': True,
    },
    'hits_by_boundaries': {
        'label': 'Hits by boundaries',
        'description': 'For each target with scoring zones, count hits per '
                       'ring (plus outside / miss). Includes a replay of '
                       'every shot on that target.',
        'fn': _report_hits_by_boundaries,
    },
    'accuracy_precision_traces': {
        'label': 'Accuracy & precision traces',
        'description': 'Accuracy and precision on a single timeline at '
                       'three granularities: per session, per quiver, and '
                       'as an all-time rolling pool. Tick which traces to '
                       'plot. Optionally pick one or more bows, arrows, or '
                       'tags to overlay them head-to-head — each subject '
                       'gets its own colored set of traces on the shared '
                       'timeline. Values are normalized by target '
                       'half-width so mixed-target histories are '
                       'comparable. Misses excluded.',
        'fn': _report_accuracy_precision_traces,
        'accepts_date_range': True,
        # Trace toggles — which of the six (granularity × metric) series to
        # draw. The route defaults to "all" when none are ticked, so the
        # report keeps its original six-line output by default.
        'category_label': 'Traces to plot:',
        'categories': _TRACE_CATEGORIES,
        # Optional head-to-head: pick specific bows / arrows (equipment
        # picker) and / or tags (tag picker) to overlay as separate
        # subjects. Leave all empty for a single combined "All shots" trace.
        'equipment_picker': True,
        'tag_picker': True,
    },
    'equipment_head_to_head': {
        'label': 'Head-to-head comparisons',
        'description': 'Pairwise comparison of every bow, arrow, and '
                       'session tag you have shot with. Three tests per '
                       'pair: Mann-Whitney U (mean accuracy, with '
                       "Cliff's δ effect size), Brown-Forsythe (group "
                       "spread / consistency), and Hotelling's T² (2D "
                       "centroid bias). p-values are Holm-corrected within "
                       "each test family. Distances are normalized by "
                       "target size so mixed targets are comparable. "
                       "Caveat: shots within a session are correlated, so "
                       "treat p-values as exploratory.",
        'fn': _report_equipment_head_to_head,
        # Sub-options revealed when this report is ticked. Each entry is
        # forwarded into the report fn as ``categories=[...]``. Defaults
        # are applied when the form omits the field (e.g. first GET).
        'categories': [
            {'key': 'bow',        'label': 'Bows'},
            {'key': 'arrow_type', 'label': 'Arrows'},
            {'key': 'tag',        'label': 'Tags'},
        ] + [
            # Bowtype-specific gear dimensions (opt-in). 'style:<key>' is what
            # the report fn matches against in its all_kinds loop.
            {'key': f'style:{sk}', 'label': _HEAD_TO_HEAD_STYLE_LABELS[sk]}
            for sk in _HEAD_TO_HEAD_STYLE_KEYS
        ],
        # Tags can grow into the dozens; render an in-form picker so the
        # user can scope the comparison instead of generating C(n,2) panels
        # for every tag they've ever used. See ``_tag_inventory``.
        'tag_picker': True,
    },
    'quiver_spread': {
        'label': 'Biggest vs smallest spread per quiver',
        'description': 'For each completed quiver, plot the biggest and '
                       'smallest pairwise distance between arrows on the '
                       'target face. Top trace is the worst pair in each '
                       'quiver; bottom trace is the tightest pair. '
                       'Defaults to every quiver ever shot — narrow with '
                       'the session-type pills, equipment picker, date '
                       'range, or tag picker.',
        'fn': _report_quiver_spread,
        'accepts_date_range': True,
        'tag_picker': True,
        'equipment_picker': True,
        # Session-type buckets, derived from auto-tags. Defaults to all
        # checked so the unfiltered run includes everything.
        'category_label': 'Include session types:',
        'categories': [
            {'key': 'regular',             'label': 'Regular sessions'},
            {'key': 'tournament_practice', 'label': 'Tournament practice'},
        ],
    },
    'within_session_drift': {
        'label': 'Within-session drift',
        'description': 'How MPI and R95 change by quiver position across '
                       'all your sessions. Reveals warm-up gains and '
                       'late-session fatigue.',
        'fn': _report_within_session_drift,
    },
    'cold_bore_vs_warmed': {
        'label': 'Cold bore vs warmed up',
        'description': 'Pool the first quiver of every session against '
                       'later quivers, then run independent accuracy and '
                       'precision tests on the two pools.',
        'fn': _report_cold_bore_vs_warmed,
    },
    'shot_density_heatmap': {
        'label': 'Shot density heatmap',
        'description': 'Hexbin density of every hit on each target, '
                       'overlaid on the face. Better than the scatter '
                       'report once a target has hundreds of shots.',
        'fn': _report_shot_density_heatmap,
    },
    'expected_score': {
        'label': 'Expected score from fit',
        'description': 'For each target with scoring zones, fit a '
                       'bivariate normal to your shots, Monte-Carlo '
                       'sample it, and project expected points per arrow '
                       'and per end.',
        'fn': _report_expected_score,
    },
    'calendar_heatmap': {
        'label': 'Shot volume calendar',
        'description': 'GitHub-style year-grid of shots per calendar day. '
                       'Streaks and gaps are visible at a glance.',
        'fn': _report_calendar_heatmap,
    },
    'handicap_trend': {
        'label': 'Handicap over time',
        'description': 'Your Archery GB 2023 handicap for every completed '
                       'AGB target round, plotted over time with a trend '
                       'line. Headline figures: latest, best of the last 12 '
                       'months, and the official AGB number (average of your '
                       'best three, rounded down).',
        'fn': _report_handicap_trend,
    },
}


# ---------------------------------------------------------------------------
# Monte-Carlo performance prediction (/predict)
# ---------------------------------------------------------------------------
# Fit an angular-dispersion distribution to a slice of the user's shot data,
# then hand the parameters + target zones to the browser, which runs the
# simulation in JS (so the histogram can grow live). The server only fits.
#
# The angular-dispersion model: treat each historical hit as a sample from a
# 2D Gaussian in milliradians (linear mm divided by the shot's distance in
# metres ≈ mrad). To extrapolate to a different distance D', multiply the
# sampled mrad offset by D' (in mm) — that's the standard linear-with-range
# spread an angular dispersion produces. Gravity drop and wind aren't
# modelled; the user sees the σ-mrad in the results so they can sanity-check.

# Minimum number of hits the fitter needs before it'll produce a covariance.
# Below this the sample covariance is too noisy to be worth simulating.
_PREDICT_MIN_HITS = 30

# Distance-trend fit tunables. The trend (bias + dispersion growth vs.
# distance) is only fit when the slice spans at least this many distinct
# distances, each carrying at least this many hits — below that the
# per-distance covariance is too noisy to regress against, and the model
# falls back to the flat angular fit.
_PREDICT_TREND_MIN_DISTANCES = 2
_PREDICT_TREND_MIN_PER_DIST = 15
# James-Stein-style shrinkage softness: a departure from the prior with
# |t| = sqrt(c) earns half weight. Larger c → more conservative (the trend
# stays nearer its prior unless the data strongly says otherwise).
_PREDICT_TREND_SHRINK_C = 4.0


def _predict_user_bows(user_id):
    """List of distinct bow_model names the user has registered."""
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT DISTINCT bow_model FROM bows "
                "WHERE user_id = %s ORDER BY bow_model",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError:
        return []
    return [r[0] for r in rows if r[0]]


def _predict_user_arrows(user_id):
    """List of distinct arrow names the user has registered."""
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT DISTINCT arrow FROM arrows "
                "WHERE user_id = %s ORDER BY arrow",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError:
        return []
    return [r[0] for r in rows if r[0]]


def _predict_user_targets(user_id):
    """User's custom targets (with physical size) for the custom endpoint."""
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT id, name, physical_size_mm FROM targets "
                "WHERE user_id = %s ORDER BY name",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError:
        return []
    out = []
    for r in rows:
        zones = _fetch_target_zones(int(r['id']), user_id)
        if not _zones_define_scoring(zones):
            continue
        out.append({
            'id': int(r['id']),
            'name': _display_target_name(r['name']) or f'Target {r["id"]}',
            'physical_size_mm': float(r['physical_size_mm'] or 0),
        })
    return out


def _predict_row_is_scorecard(tags_raw):
    """True when a shot row was entered from a scorecard rather than plotted.

    Scorecards — competition score sheets *and* paper practice scorecards
    alike — don't store real shot positions; they write *synthetic*
    ring-midpoint coordinates via ``_coords_for_ring_label`` (y≡0, x≡the
    ring's radius). Fitting those would inject a fake sideways bias and
    collapse vertical dispersion, so the predict fit excludes them. Both
    paths carry a ``match:`` tag (the same marker ``_counts_toward_handicap``
    keys on); this is independent of bow/arrow, so it still holds once
    scorecards capture equipment data.
    """
    if not tags_raw:
        return False
    return any(t.strip().startswith('match:')
               for t in str(tags_raw).split(','))


def _predict_session_matches_practice(tags_raw, mode):
    """Tournament/practice filter applied to a shot's session_tags string.

    ``mode`` is one of 'all', 'practice_only', 'tournament_only'.
      - 'practice_only': the row carries an explicit ``practice`` tag.
        Untagged sessions are excluded — use 'all' to include those.
      - 'tournament_only': the row carries a ``tournament:*`` tag AND does
        not carry the ``practice`` tag.
    """
    if mode == 'all':
        return True
    tags = [t.strip().lower() for t in (tags_raw or '').split(',')]
    has_tournament = any(t.startswith('tournament:') for t in tags)
    has_practice = 'practice' in tags
    if mode == 'tournament_only':
        return has_tournament and not has_practice
    if mode == 'practice_only':
        return has_practice
    return True


def _wls_centered(pts):
    """Weighted least squares for ``y = α + β·(x − x_c)``, centered at the
    inverse-variance weighted mean ``x_c`` so that ``α`` (the fitted value at
    ``x_c``) and ``β`` (the slope) are *uncorrelated*. That lets each be shrunk
    toward its own prior independently without the intercept/slope coupling a
    raw ``y = a + b·x`` fit suffers from (an unstable intercept extrapolated
    far off the data drags the slope with it).

    ``pts`` is ``(x, y, var)`` with ``var`` the measurement variance of ``y``.
    Returns ``(x_c, α, β, var_α, var_β)``, or ``None`` when singular.
    """
    sw = swx = 0.0
    valid = []
    for x, y, var in pts:
        if var is None or var <= 0:
            continue
        w = 1.0 / var
        sw += w
        swx += w * x
        valid.append((x, y, w))
    if sw <= 0 or len(valid) < 2:
        return None
    x_c = swx / sw
    swy = sxx = sxy = 0.0
    for x, y, w in valid:
        dx = x - x_c
        swy += w * y
        sxx += w * dx * dx
        sxy += w * dx * y
    if sxx <= 0:
        return None
    return x_c, swy / sw, sxy / sxx, 1.0 / sw, 1.0 / sxx


def _shrink_weight(estimate, prior, var):
    """James-Stein-style weight ``w = t²/(t² + c)`` for shrinking
    ``estimate`` toward ``prior``, with ``t = (estimate − prior)/√var``.

    Returns ``w`` in [0, 1): ``w→0`` when the departure from the prior sits
    within its own noise (use the prior), ``w→1`` when it's well-resolved
    (trust the data).
    """
    if var is None or var <= 0:
        return 0.0
    t2 = (estimate - prior) ** 2 / var
    return t2 / (t2 + _PREDICT_TREND_SHRINK_C)


def _fit_angular_trend(samples):
    """Angular-dispersion fit (flat fit + distance trend) over hit samples.

    ``samples`` is a list of ``(x_mm, y_mm, dist_m)`` for *hits* only
    (``dist_m > 0``, at least ``_PREDICT_MIN_HITS`` of them). Pure math — no
    DB, no Flask — so it unit-tests like ``handicap.py``.

    Returns::

        {'mean_mrad', 'cov_mrad', 'n_hits', 'distances_m',
         'trend': {...}, 'per_distance': [...]}

    The flat ``mean_mrad``/``cov_mrad`` reproduce the legacy pooled fit
    exactly; ``trend`` and ``per_distance`` are additive. ``trend`` is
    ``{'ok': False, 'reason': str}`` when the slice can't support one.

    The trend models bias as per-axis affine in mm, ``μ_mm(d) = a + b·d``,
    and dispersion as ``σ_mrad(d) ∝ e^{k·d}`` (the AGB form in
    ``handicap.sigma_t``). Departures from today's flat model are shrunk
    toward their priors: the bias intercept ``a`` toward 0 (no population
    prior for an aim offset) and the growth ``k`` toward ``handicap.KD``
    (AGB's population-fitted angular growth), so the low-data fallback is
    "grows like a typical archer," not a known-wrong flat line.
    """
    xs_mrad = [x / d for (x, y, d) in samples]
    ys_mrad = [y / d for (x, y, d) in samples]
    n = len(samples)

    mx = sum(xs_mrad) / n
    my = sum(ys_mrad) / n
    vxx = sum((x - mx) ** 2 for x in xs_mrad) / (n - 1)
    vyy = sum((y - my) ** 2 for y in ys_mrad) / (n - 1)
    vxy = sum((x - mx) * (y - my)
              for x, y in zip(xs_mrad, ys_mrad)) / (n - 1)
    s_global = math.sqrt(max(0.0, 0.5 * (vxx + vyy)))

    # Group angular offsets by distance.
    by_dist = {}
    for (x, y, d) in samples:
        by_dist.setdefault(round(d, 2), []).append((x / d, y / d))
    distances = sorted(by_dist)

    # Per-distance estimates (n_d, angular mean, scalar angular size).
    grp = {}
    for d, pts in by_dist.items():
        nd = len(pts)
        mxd = sum(p[0] for p in pts) / nd
        myd = sum(p[1] for p in pts) / nd
        if nd >= 2:
            cxx = sum((p[0] - mxd) ** 2 for p in pts) / (nd - 1)
            cyy = sum((p[1] - myd) ** 2 for p in pts) / (nd - 1)
            sd = math.sqrt(max(0.0, 0.5 * (cxx + cyy)))
        else:
            cxx = cyy = sd = None
        grp[d] = {'n': nd, 'mx': mxd, 'my': myd, 's': sd,
                  'cxx': cxx, 'cyy': cyy}

    # Per-distance diagnostic: each distance vs. the rest of the slice.
    # Bias via Hotelling's T² on the mrad offsets; spread via Brown-Forsythe
    # on the per-axis deviations. This is the "feedback, made visible".
    per_distance = []
    for d in distances:
        g = grp[d]
        this_pts = by_dist[d]
        other_pts = [p for od in distances if od != d for p in by_dist[od]]
        bias_p = spread_p = None
        # Need ≥3 shots at this distance for a meaningful test: fewer makes
        # the within-group deviation collapse to ~0 and the spread test fire
        # spuriously (a single shot is "infinitely tight").
        if other_pts and g['n'] >= 3:
            bias_p = _hotelling_t2(this_pts, other_pts)[4]
            this_dev = ([p[0] - g['mx'] for p in this_pts]
                        + [p[1] - g['my'] for p in this_pts])
            omx = sum(p[0] for p in other_pts) / len(other_pts)
            omy = sum(p[1] for p in other_pts) / len(other_pts)
            other_dev = ([p[0] - omx for p in other_pts]
                         + [p[1] - omy for p in other_pts])
            spread_p = _brown_forsythe(this_dev, other_dev)[3]
        per_distance.append({
            'd': d,
            'n': g['n'],
            's_mrad': g['s'],
            'mean_mrad': [g['mx'], g['my']],
            'bias_p': bias_p,
            'spread_ratio': (g['s'] / s_global
                             if g['s'] and s_global > 0 else None),
            'spread_p': spread_p,
        })

    out = {
        'mean_mrad':  [mx, my],
        'cov_mrad':   [[vxx, vxy], [vxy, vyy]],
        'n_hits':     n,
        'distances_m': distances,
        'per_distance': per_distance,
    }

    # --- Distance trend --------------------------------------------------
    qual = [d for d in distances
            if grp[d]['n'] >= _PREDICT_TREND_MIN_PER_DIST
            and grp[d]['s'] is not None and grp[d]['s'] > 0]
    if len(qual) < _PREDICT_TREND_MIN_DISTANCES:
        out['trend'] = {
            'ok': False,
            'reason': (f'Need ≥{_PREDICT_TREND_MIN_DISTANCES} distances '
                       f'with ≥{_PREDICT_TREND_MIN_PER_DIST} hits each to '
                       f'fit a distance trend; this slice qualifies at '
                       f'{len(qual)}.'),
        }
        return out

    # Bias: per axis, regress the per-distance mean offset (mm) on distance,
    # centered so the value-at-centroid (α) and slope (β) are uncorrelated and
    # shrink cleanly toward the flat model. The flat bias is μ_mm = mean_mrad·d,
    # so α shrinks toward mean_mrad·d_c and β toward mean_mrad; with no signal
    # the line collapses *exactly* to flat. Each point's measurement variance
    # is Var(mean) = σ²_ang · d² / n_d.
    def _bias_axis(key_m, key_v, mean_ang):
        pts = []
        for d in qual:
            g = grp[d]
            var_mm = g[key_v] * d * d / g['n'] if g[key_v] > 0 else None
            pts.append((d, g[key_m] * d, var_mm))
        fit = _wls_centered(pts)
        if fit is None:
            return None
        d_c, alpha, beta, va, vb = fit
        w_a = _shrink_weight(alpha, mean_ang * d_c, va)
        alpha = w_a * alpha + (1.0 - w_a) * (mean_ang * d_c)
        w_b = _shrink_weight(beta, mean_ang, vb)
        beta = w_b * beta + (1.0 - w_b) * mean_ang
        # Back to intercept/slope form (a + b·d) for the payload.
        return alpha - beta * d_c, beta

    bx = _bias_axis('mx', 'cxx', mx)
    by = _bias_axis('my', 'cyy', my)

    # Growth: regress ln σ_mrad on distance, centered. Var(ln σ) ≈ 1/(2ν) with
    # ν ≈ 2·(n_d − 1) (the two axes pooled). The slope k shrinks toward the
    # handicap population prior; the centroid (d_ref, lns_ref) the line pivots
    # through pins σ at d_ref.
    g_pts = [(d, math.log(grp[d]['s']), 1.0 / (4.0 * max(1, grp[d]['n'] - 1)))
             for d in qual]
    growth = _wls_centered(g_pts)

    if bx is None or by is None or growth is None:
        out['trend'] = {'ok': False,
                        'reason': 'Distance-trend regression was singular.'}
        return out

    a_x, b_x = bx
    a_y, b_y = by
    d_ref, lns_ref, k_fit, _va, vk = growth
    w_k = _shrink_weight(k_fit, handicap.KD, vk)
    k_used = w_k * k_fit + (1.0 - w_k) * handicap.KD

    # cov_ref keeps the pooled global shape, rescaled to σ_mrad at d_ref.
    f = math.exp(lns_ref) / s_global if s_global > 0 else 1.0
    f2 = f * f
    out['trend'] = {
        'ok':           True,
        'mean_mm':      {'ax': a_x, 'bx': b_x, 'ay': a_y, 'by': b_y},
        'growth_k':     k_used,
        'growth_k_raw': k_fit,
        'growth_prior': handicap.KD,
        'd_ref':        d_ref,
        'cov_ref_mrad': [[vxx * f2, vxy * f2], [vxy * f2, vyy * f2]],
        'n_distances':  len(qual),
    }
    return out


def _fit_shot_distribution(user_id, bows=None, arrows=None, tags=None,
                            date_from=None, date_to=None,
                            practice_mode='all'):
    """Return a 2D-Gaussian angular dispersion fit over the filtered shots.

    Handles the DB query + filtering, then delegates the math to
    ``_fit_angular_trend``. Output shape::

        {'ok': True,
         'mean_mrad': [mx, my],
         'cov_mrad': [[vxx, vxy], [vxy, vyy]],
         'miss_rate': float,
         'n_hits': int,
         'n_misses': int,
         'distances_m': [d1, d2, ...],
         'shaft_mm': float,
         'trend': {...}, 'per_distance': [...]}   # see _fit_angular_trend

    or ``{'ok': False, 'reason': str}`` when the slice is too small.
    """
    # Single query, then filter in Python — the filter combinations
    # (multi-bow, multi-arrow, tag intersection, practice/tournament mode)
    # are awkward to express as parameterized SQL across both backends, and
    # the row volumes here are well within memory.
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT x_coord, y_coord, distance, bow, arrow_type, "
                "       session_tags, is_precise, timestamp, "
                "       arrow_shaft_diameter "
                "FROM apollo WHERE user_id = %s",
                (user_id,)
            ).fetchall()
    except SQLAlchemyError as e:
        return {'ok': False, 'reason': f'database error: {e}'}

    bow_set = {b for b in (bows or []) if b}
    arr_set = {a for a in (arrows or []) if a}
    tag_set = {t.strip().lower() for t in (tags or []) if t and t.strip()}

    # Date bounds — strings 'YYYY-MM-DD' compare lexicographically against
    # the timestamp string prefix without any datetime parsing.
    df = (date_from or '').strip() or None
    dt = (date_to or '').strip() or None

    samples = []     # (x_mm, y_mm, dist_m) per fitted hit
    shaft_mms = []   # defined (positive) shaft diameters among fitted hits
    n_misses = 0
    for r in rows:
        # Scorecard rows store synthetic ring-midpoint coords (y≡0, x≡the
        # ring radius), not plotted positions — fitting them would inject a
        # fake bias and collapse vertical dispersion. Drop them outright (no
        # hit, no miss); the coordinate system itself is left untouched.
        if _predict_row_is_scorecard(_row_get(r, 'session_tags')):
            continue
        # Equipment / tag / date filters
        if bow_set and (_row_get(r, 'bow') or '') not in bow_set:
            continue
        if arr_set and (_row_get(r, 'arrow_type') or '') not in arr_set:
            continue
        if tag_set:
            row_tags = {t.strip().lower()
                        for t in (_row_get(r, 'session_tags') or '').split(',')
                        if t.strip()}
            if not (tag_set & row_tags):
                continue
        if not _predict_session_matches_practice(
                _row_get(r, 'session_tags'), practice_mode):
            continue
        ts = _row_get(r, 'timestamp')
        if df or dt:
            ts_str = _ts_date_str(ts)
            if df and ts_str < df:
                continue
            if dt and ts_str > dt:
                continue

        xraw = _row_get(r, 'x_coord')
        yraw = _row_get(r, 'y_coord')
        xs = str(xraw).strip() if xraw is not None else ''
        ys = str(yraw).strip() if yraw is not None else ''
        if xs == MISS_SENTINEL and ys == MISS_SENTINEL:
            n_misses += 1
            continue
        try:
            x_mm = float(xs)
            y_mm = float(ys)
            dist_m = float(str(_row_get(r, 'distance') or '').strip())
        except (TypeError, ValueError):
            continue
        if dist_m <= 0:
            continue
        # The fit works in milliradians (1 mm at 1 m ≈ 1 mrad); the angular
        # conversion + grouping happens in _fit_angular_trend. The small-angle
        # approximation is exact at the magnitudes we care about (a few mm at
        # tens of metres → tens of mrad, well within the linear regime).
        samples.append((x_mm, y_mm, dist_m))
        # Track the arrow's real shaft diameter when this shot defines one;
        # an undefined / blank / non-positive value contributes nothing, so
        # the client falls back to the default shaft only if *no* fitted shot
        # carries a real diameter.
        try:
            d_mm = float(str(_row_get(r, 'arrow_shaft_diameter') or '').strip())
            if d_mm > 0:
                shaft_mms.append(d_mm)
        except (TypeError, ValueError):
            pass

    n_hits = len(samples)
    if n_hits < _PREDICT_MIN_HITS:
        return {'ok': False,
                'reason': (f'Need at least {_PREDICT_MIN_HITS} hits to fit '
                           f'a distribution; only found {n_hits} with the '
                           f'current filters.')}

    fit = _fit_angular_trend(samples)

    total = n_hits + n_misses
    fit['miss_rate'] = n_misses / total if total > 0 else 0.0
    fit['n_misses'] = n_misses
    # Shaft diameter for the simulator's line-cutter rule: the mean of the
    # real diameters the fitted shots define, or the default shaft only when
    # none of them do.
    fit['shaft_mm'] = (sum(shaft_mms) / len(shaft_mms)
                       if shaft_mms else DEFAULT_SHAFT_DIAMETER_MM)
    fit['ok'] = True
    return fit


def _predict_zones_for_face(face_key):
    """Normalize TOURNAMENT_FACES[face_key]['zones'] into the JSON shape the
    JS simulator expects: ``[{radius_mm, point_value}, ...]`` sorted
    innermost-out. Returns ``(zones_norm, target_physical_mm)`` or
    ``(None, None)`` for an unknown face.
    """
    face = _tournament_face_def(face_key)
    if not face:
        return None, None
    raw = face['zones']
    # Tournament face zones are tuples (point_value, radius_mm, color)
    # already sorted innermost-out by construction.
    zones_norm = [{'radius_mm': float(z[1]), 'point_value': int(z[0])}
                  for z in raw]
    return zones_norm, float(face['physical_size_mm'])


def _build_predict_segments(form, user_id):
    """Construct the list of simulation segments from POST form fields.

    Returns ``(segments, label, max_score)`` or raises ``ValueError`` with
    a user-readable message.

    Each segment is::
        {'distance_m', 'ends', 'arrows_per_end', 'target_physical_mm',
         'zones': [{'radius_mm', 'point_value'}, ...]}
    """
    mode = (form.get('endpoint_mode') or 'round').strip()

    if mode == 'round':
        key = (form.get('round_key') or '').strip()
        rd = _tournament_round_def(key)
        if not rd:
            raise ValueError(f'Unknown tournament round: {key!r}')
        segs_in = rd.get('segments') or [{
            'distance_m':    rd['distance_m'],
            'ends':          rd['ends'],
            'face_key':      rd['face_key'],
        }]
        out = []
        for s in segs_in:
            zones, phys = _predict_zones_for_face(s['face_key'])
            if zones is None:
                raise ValueError(f'Unknown face: {s["face_key"]!r}')
            out.append({
                'distance_m':         float(s['distance_m']),
                'ends':               int(s['ends']),
                'arrows_per_end':     int(rd['arrows_per_end']),
                'target_physical_mm': phys,
                'zones':              zones,
            })
        return out, rd['name'], int(rd.get('max_score') or 0)

    # Custom endpoint
    try:
        distance_m = float(form.get('custom_distance_m') or '')
        ends = int(form.get('custom_ends') or '')
        arrows_per_end = int(form.get('custom_arrows_per_end') or '')
    except ValueError:
        raise ValueError('Custom endpoint: distance, ends, and arrows '
                         'per end must all be positive numbers.')
    if distance_m <= 0 or ends <= 0 or arrows_per_end <= 0:
        raise ValueError('Custom endpoint: distance, ends, and arrows '
                         'per end must all be positive numbers.')

    face_source = (form.get('custom_face_source') or 'tournament').strip()
    if face_source == 'tournament':
        face_key = (form.get('custom_face_key') or '').strip()
        zones, phys = _predict_zones_for_face(face_key)
        if zones is None:
            raise ValueError(f'Unknown face: {face_key!r}')
        face_name = _tournament_face_def(face_key)['name']
    else:
        try:
            target_id = int(form.get('custom_target_id') or 0)
        except ValueError:
            target_id = 0
        if target_id <= 0:
            raise ValueError('Pick one of your targets for the custom face.')
        zone_rows = _fetch_target_zones(target_id, user_id)
        if not _zones_define_scoring(zone_rows):
            raise ValueError('That target has no scoring zones with point '
                             'values — it can\'t be simulated.')
        zones = [{'radius_mm': float(z['radius_mm']),
                  'point_value': int(z['point_value'] or 0)}
                 for z in zone_rows]
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            t = cur.execute(
                "SELECT name, physical_size_mm FROM targets "
                "WHERE id = %s AND user_id = %s",
                (target_id, user_id)
            ).fetchone()
        if not t:
            raise ValueError('That target doesn\'t exist.')
        phys = float(t['physical_size_mm'] or 0)
        face_name = _display_target_name(t['name']) or f'Target {target_id}'

    seg = {
        'distance_m':         distance_m,
        'ends':               ends,
        'arrows_per_end':     arrows_per_end,
        'target_physical_mm': phys,
        'zones':              zones,
    }
    label = (f'Custom: {face_name} @ {distance_m:g} m '
             f'({ends} × {arrows_per_end})')
    max_zone = max((z['point_value'] for z in zones), default=0)
    max_score = max_zone * ends * arrows_per_end
    return [seg], label, max_score


@app.route('/tools', methods=['GET'])
@login_required
def tools():
    """Standalone archery calculators. Pure client-side math — no DB.

    Six tools on one page (wind drift, sight-mark interpolator, spine
    selector, FOC, arrow speed, kinetic energy & momentum). The template
    does all the math in JS so the user gets live updates and we don't
    burn a server round-trip per keystroke."""
    return render_template('tools.html')


# Pre-built, app-styled HTML documentation lives here (see
# documentation/build_docs.py). Served read-only and public: it's the
# open-source project's own docs, linked from both the public splash page
# and the in-app side nav, so it must work for logged-out visitors too.
DOCS_HTML_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), 'documentation', 'html'
)


# Registered WITH a trailing slash on purpose: the index page uses relative
# links (readme.html, formulas.html, …), which only resolve under /docs/ when
# the index URL itself ends in a slash. Flask 308-redirects a bare /docs here.
@app.route('/docs/', methods=['GET'])
def docs_index():
    """Documentation home — the generated index page."""
    return send_from_directory(DOCS_HTML_DIR, 'index.html')


@app.route('/docs/<path:page>', methods=['GET'])
def docs_page(page):
    """Serve a generated documentation page. Only .html files; the path is
    constrained by send_from_directory so it can't escape the docs dir."""
    if not page.endswith('.html'):
        abort(404)
    return send_from_directory(DOCS_HTML_DIR, page)


@app.route('/predict', methods=['GET', 'POST'])
@login_required
def predict():
    """Performance-prediction wizard. Fits a 2D angular Gaussian to the
    user's filtered shot data, then hands the parameters + endpoint config
    to a client-side Monte-Carlo simulator (templates/predict.html +
    static/apollo-predict.js)."""
    user_id = current_user_id()
    rounds = [
        {'key': k,
         'name': v['name'],
         'org':  v.get('org', ''),
         'description': v.get('description', '')}
        for k, v in TOURNAMENT_ROUNDS.items()
    ]
    faces = [
        {'key': k,
         'name': v['name'],
         'physical_size_mm': v['physical_size_mm']}
        for k, v in TOURNAMENT_FACES.items()
    ]
    ctx = {
        'rounds':         rounds,
        'faces':          faces,
        'user_bows':      _predict_user_bows(user_id),
        'user_arrows':    _predict_user_arrows(user_id),
        'user_targets':   _predict_user_targets(user_id),
        'tag_inventory':  _tag_inventory(user_id),
        # Echo back POSTed selections so the form sticks across the
        # render. On GET these are empty / defaulted.
        'form':           {},
        'error':          None,
        'payload':        None,
        'endpoint_label': None,
        'endpoint_max':   None,
    }

    if request.method != 'POST':
        return render_template('predict.html', **ctx)

    form = request.form
    ctx['form'] = {
        'bows':         form.getlist('bows'),
        'arrows':       form.getlist('arrows'),
        'tags':         form.getlist('tags'),
        'date_from':    (form.get('date_from') or '').strip(),
        'date_to':      (form.get('date_to') or '').strip(),
        'practice_mode': (form.get('practice_mode') or 'all').strip(),
        'endpoint_mode': (form.get('endpoint_mode') or 'round').strip(),
        'round_key':    (form.get('round_key') or '').strip(),
        'custom_face_source': (form.get('custom_face_source') or 'tournament').strip(),
        'custom_face_key':  (form.get('custom_face_key') or '').strip(),
        'custom_target_id': (form.get('custom_target_id') or '').strip(),
        'custom_distance_m': (form.get('custom_distance_m') or '').strip(),
        'custom_ends':      (form.get('custom_ends') or '').strip(),
        'custom_arrows_per_end': (form.get('custom_arrows_per_end') or '').strip(),
        'score_target':     (form.get('score_target') or '').strip(),
        'n_runs':           (form.get('n_runs') or '').strip(),
    }

    # Build the endpoint first so a bad endpoint is reported even when
    # the filters would also fail.
    try:
        segments, endpoint_label, endpoint_max = _build_predict_segments(
            form, user_id)
    except ValueError as e:
        ctx['error'] = str(e)
        return render_template('predict.html', **ctx)

    fit = _fit_shot_distribution(
        user_id,
        bows=ctx['form']['bows'],
        arrows=ctx['form']['arrows'],
        tags=ctx['form']['tags'],
        date_from=ctx['form']['date_from'],
        date_to=ctx['form']['date_to'],
        practice_mode=ctx['form']['practice_mode'],
    )
    if not fit['ok']:
        ctx['error'] = fit['reason']
        return render_template('predict.html', **ctx)

    # Sim runs entirely in the browser, so no server-side capacity cap.
    # Floor at 1; trust the user not to type something silly.
    try:
        n_runs = int(ctx['form']['n_runs'])
    except ValueError:
        n_runs = 100
    n_runs = max(1, n_runs)

    try:
        score_target = (int(ctx['form']['score_target'])
                        if ctx['form']['score_target'] else None)
    except ValueError:
        score_target = None

    ctx['payload'] = {
        'dist':          fit,
        'segments':      segments,
        'n_runs':        n_runs,
        'score_target':  score_target,
        'endpoint_label': endpoint_label,
        'endpoint_max':  endpoint_max,
    }
    ctx['endpoint_label'] = endpoint_label
    ctx['endpoint_max'] = endpoint_max
    return render_template('predict.html', **ctx)


# ─── Form analysis (motion capture) ──────────────────────────────────────
# Pose extraction and scoring happen entirely client-side (static/apollo-form.js
# + a vendored markerless pose model). The server only (a) hands the page the
# checkpoint reference spec for the chosen bowstyle, and (b) optionally persists
# the derived numbers. No video ever reaches the server.

def _form_payload(bowstyle):
    """Build the JSON payload the /form page hands to the client analyzer."""
    bs = (bowstyle or '').strip().lower()
    return {
        'bowstyle':    bs,
        'checkpoints': form_checkpoints.checkpoints_for(bs),
    }


@app.route('/form', methods=['GET', 'POST'])
@login_required
def form_analysis():
    """Motion-capture form analysis (analysis mode).

    GET renders the capture/analyze page seeded with the checkpoint spec for
    the requested (or the archer's default) bowstyle. POST persists a finished
    analysis — only the derived angles + scores, never video — so the archer
    can track form over time. The POST is fetch-based JSON from the client and
    rides CSRFProtect via the X-CSRFToken header, same as /api/sync_shots.
    """
    user_id = current_user_id()
    user = current_user()

    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        bs = (data.get('bowstyle') or '').strip().lower()
        metrics = data.get('metrics')
        scores = data.get('scores')
        overall = data.get('overall_score')
        if not isinstance(metrics, dict) or not isinstance(scores, dict):
            return jsonify(ok=False, msg='Expected metrics and scores objects.'), 400
        try:
            overall = float(overall)
        except (TypeError, ValueError):
            overall = None
        # Optional replay payload: a compact pose sequence (anonymous skeleton
        # points only — never video). Serialise it, but drop it rather than
        # reject the save if it's implausibly large (keeps it under a TEXT
        # column and guards against a runaway client).
        frames = data.get('frames')
        frames_json = None
        if isinstance(frames, dict):
            try:
                candidate = json.dumps(frames, separators=(',', ':'))
            except (TypeError, ValueError):
                candidate = None
            if candidate is not None and len(candidate) <= 60000:
                frames_json = candidate
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                cur.execute(
                    "INSERT INTO form_captures "
                    "(user_id, created_at, bowstyle, overall_score, "
                    " metrics_json, scores_json, frames_json) "
                    "VALUES (%s, %s, %s, %s, %s, %s, %s)",
                    (user_id, _app_now(), bs, overall,
                     json.dumps(metrics), json.dumps(scores), frames_json)
                )
                con.commit()
        except SQLAlchemyError as e:
            print(f"❌ Form-capture save error: {e}")
            return jsonify(ok=False, msg='Could not save analysis.'), 500
        return jsonify(ok=True)

    # GET — pick the bowstyle: explicit ?bowstyle=, else the archer's profile
    # default, else recurve. The picker on the page re-fetches the spec on
    # change without a round-trip (all profiles are embedded in the payload).
    requested = (request.args.get('bowstyle') or '').strip().lower()
    default_bs = requested or (user.get('default_bowstyle') if user else '') or 'recurve'
    specs = {bs: form_checkpoints.checkpoints_for(bs)
             for bs in form_checkpoints.all_bowstyles()}
    payload = {
        'bowstyle':     default_bs if default_bs in specs else 'recurve',
        'specs':        specs,
        'save_url':     url_for('form_analysis'),
    }
    shown_bs = payload['bowstyle']
    return render_template(
        'form.html',
        payload=payload,
        bowstyle_labels=BOWSTYLE_LABELS,
        bowstyles=list(specs.keys()),
        recent=_form_recent_captures(user_id),
        consistency=_form_consistency(user_id, shown_bs, specs[shown_bs]),
        consistency_bowstyle=shown_bs,
    )


def _form_consistency(user_id, bowstyle, checkpoints, limit=10):
    """Shot-to-shot consistency: per-metric mean and (sample) standard deviation
    across the archer's recent saved analyses for a bowstyle.

    The lower the SD, the more repeatable that element of form is shot to shot —
    which, for archery, matters more than any single perfect frame. Reads the
    stored metrics_json (no video), needs ≥2 captures of a metric to report it,
    and joins each to its checkpoint label/unit so the template can render a
    table in checkpoint order. Returns [] when there isn't enough history.
    """
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                "SELECT metrics_json FROM form_captures "
                "WHERE user_id = %s AND bowstyle = %s "
                "ORDER BY created_at DESC LIMIT %s",
                (user_id, (bowstyle or '').strip().lower(), int(limit))
            ).fetchall()
    except SQLAlchemyError:
        return []
    series = {}
    for r in rows or []:
        try:
            metrics = json.loads(r['metrics_json'] or '{}')
        except (ValueError, TypeError):
            continue
        if not isinstance(metrics, dict):
            continue
        for key, val in metrics.items():
            if isinstance(val, (int, float)):
                series.setdefault(key, []).append(float(val))
    out = []
    for cp in checkpoints:
        vals = series.get(cp['id'])
        if not vals or len(vals) < 2:
            continue
        mean = sum(vals) / len(vals)
        sd = (sum((x - mean) ** 2 for x in vals) / (len(vals) - 1)) ** 0.5
        out.append({'label': cp['label'], 'unit': cp.get('unit', ''),
                    'mean': mean, 'sd': sd, 'n': len(vals)})
    return out


def _form_recent_captures(user_id, limit=10):
    """Most recent saved form analyses for the archer (for the history strip).

    Includes the row id (so a click can fetch the full capture for replay) and a
    has_replay flag (whether a pose sequence was stored).
    """
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            rows = cur.execute(
                # Test frames_json for presence in SQL rather than selecting the
                # whole (potentially ~tens-of-KB) blob for every history row.
                "SELECT id, created_at, bowstyle, overall_score, "
                "(frames_json IS NOT NULL) AS has_replay "
                "FROM form_captures WHERE user_id = %s "
                "ORDER BY created_at DESC LIMIT %s",
                (user_id, int(limit))
            ).fetchall()
    except SQLAlchemyError:
        return []
    return [{'id': r['id'], 'created_at': r['created_at'],
             'bowstyle': r['bowstyle'], 'overall_score': r['overall_score'],
             'has_replay': bool(r['has_replay'])} for r in (rows or [])]


@app.route('/form/capture/<int:capture_id>', methods=['GET'])
@login_required
def form_capture(capture_id):
    """Return one saved analysis as JSON for replay — scoped to the owner.

    Hands back the stored derived data (bowstyle, scores, overall) plus the
    compact pose sequence (frames) so the client can re-draw the shot as a
    silhouette with the angle lines over it. 404s on a capture that isn't this
    user's, so ids can't be enumerated across accounts.
    """
    user_id = current_user_id()
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT bowstyle, overall_score, metrics_json, scores_json, "
                "frames_json FROM form_captures WHERE id = %s AND user_id = %s",
                (capture_id, user_id)
            ).fetchone()
    except SQLAlchemyError:
        row = None
    if row is None:
        return jsonify(ok=False, msg='Not found.'), 404

    def _loads(raw, default):
        try:
            return json.loads(raw) if raw else default
        except (ValueError, TypeError):
            return default

    return jsonify(
        ok=True,
        bowstyle=row['bowstyle'],
        overall_score=row['overall_score'],
        metrics=_loads(row['metrics_json'], {}),
        scores=_loads(row['scores_json'], {}),
        frames=_loads(row['frames_json'], None),
    )


@app.route('/form/author', methods=['GET'])
@root_required
def form_author():
    """Learning mode (dev/admin only): capture a known-good archer to author
    the reference checkpoint targets.

    This is the *config-writing* half of the feature, deliberately gated to the
    root user and kept separate from the user-facing analyzer so no user's form
    can ever influence the reference. It reuses the same client pipeline but,
    instead of scoring against the spec, it reports the raw measured angles so
    the operator can transcribe tuned targets into form_checkpoints.py. v1 has
    no write-back UI on purpose — editing the spec is a code change, reviewed
    like any other.
    """
    specs = {bs: form_checkpoints.checkpoints_for(bs)
             for bs in form_checkpoints.all_bowstyles()}
    payload = {
        'bowstyle': 'recurve',
        'specs':    specs,
        'author':   True,
    }
    return render_template(
        'form.html',
        payload=payload,
        bowstyle_labels=BOWSTYLE_LABELS,
        bowstyles=list(specs.keys()),
        recent=[],
        author_mode=True,
    )


@app.route('/analyze', methods=['GET', 'POST'])
@login_required
def analyze():
    """Data analysis page — pick reports, render their charts and tables."""
    selected = []
    results = []
    error = None
    # Per-report date-range inputs. Currently only `all_shots_per_target`
    # accepts them; using a per-key naming convention (`<key>_date_from`)
    # keeps the wiring extensible if other reports adopt the same option.
    date_ranges = {}
    # Per-report category checkboxes (e.g. which kinds to compare in the
    # head-to-head report). Same naming convention: `<key>_categories`.
    categories = {}
    # Per-report tag selections for reports that expose a tag picker.
    # Populated only on POST so the template can distinguish "submitted
    # nothing" (explicit empty) from "first GET" (use defaults).
    tag_selections = {}
    # Per-report bow / arrow selections for reports with equipment_picker.
    # Empty list ≡ no filter applied (include every bow / arrow).
    bow_selections = {}
    arrow_selections = {}
    if request.method == 'POST':
        selected = [k for k in request.form.getlist('reports') if k in REPORTS]
        for k, spec in REPORTS.items():
            if spec.get('accepts_date_range'):
                date_ranges[k] = {
                    'date_from': (request.form.get(f'{k}_date_from') or '').strip(),
                    'date_to':   (request.form.get(f'{k}_date_to')   or '').strip(),
                }
            if spec.get('categories'):
                valid = {c['key'] for c in spec['categories']}
                chosen = [c for c in request.form.getlist(f'{k}_categories')
                          if c in valid]
                # No box ticked → fall back to "all" so the user gets a
                # report instead of a silent empty.
                if not chosen:
                    chosen = list(valid)
                categories[k] = chosen
            if spec.get('tag_picker'):
                tag_selections[k] = request.form.getlist(f'{k}_tags')
            if spec.get('equipment_picker'):
                bow_selections[k] = request.form.getlist(f'{k}_bows')
                arrow_selections[k] = request.form.getlist(f'{k}_arrows')
        if not selected:
            error = 'Pick at least one report to generate.'
        else:
            try:
                user_id = current_user_id()
                for key in selected:
                    spec = REPORTS[key]
                    kwargs = {}
                    if spec.get('accepts_date_range'):
                        dr = date_ranges.get(key) or {}
                        kwargs['date_from'] = dr.get('date_from') or None
                        kwargs['date_to']   = dr.get('date_to') or None
                    if spec.get('categories'):
                        kwargs['categories'] = categories.get(key)
                    if spec.get('tag_picker'):
                        # Submitted list is the source of truth — even an
                        # empty list means "user picked no tags" (no
                        # default fill-in here; the template handles the
                        # first-GET case by pre-checking the top-N).
                        kwargs['tag_filter'] = tag_selections.get(key, [])
                    if spec.get('equipment_picker'):
                        # Empty submitted lists are forwarded as-is — the
                        # report fn's contract is "empty = no filter,
                        # include every bow / arrow," which matches the
                        # "default to all shots ever" behavior the user
                        # asked for on the quiver_spread report.
                        kwargs['bow_filter'] = bow_selections.get(key, [])
                        kwargs['arrow_filter'] = arrow_selections.get(key, [])
                    out = spec['fn'](user_id, **kwargs)
                    if out is None:
                        results.append({
                            'key': key,
                            'title': REPORTS[key]['label'],
                            'svg_b64': None,
                            'columns': [],
                            'rows': [],
                            'empty': True,
                        })
                        continue
                    # Reports may return a single result-dict OR a list of
                    # them (head-to-head splits per-kind sections). Normalize
                    # so each item gets the same post-processing.
                    items = out if isinstance(out, list) else [out]
                    for item in items:
                        if item.get('empty'):
                            # Report ran but found nothing worth rendering
                            # and supplied its own diagnostic message — pass
                            # it through with sensible defaults.
                            item.setdefault('title', REPORTS[key]['label'])
                            item.setdefault('key', key)
                            results.append(item)
                        else:
                            item['empty'] = False
                            results.append(item)
            except ImportError as e:
                print(f"❌ Analyze missing dependency: {e}")
                error = ('matplotlib is not installed. Install it with '
                         '`pip install matplotlib` and retry.')
                results = []
            except SQLAlchemyError as e:
                print(f"❌ Analyze read error: {e}")
                error = 'Database error while building report.'
                results = []
    catalog = [
        {
            'key': k,
            'label': v['label'],
            'description': v['description'],
            'accepts_date_range': v.get('accepts_date_range', False),
            'categories': v.get('categories', []),
            'category_label': v.get('category_label', 'Compare:'),
            'tag_picker': v.get('tag_picker', False),
            'equipment_picker': v.get('equipment_picker', False),
        }
        for k, v in REPORTS.items()
    ]
    # The picker UI needs the user's tag inventory regardless of POST
    # state. Cheap one-query call; no-op if no report uses the picker.
    if any(v.get('tag_picker') for v in REPORTS.values()):
        tag_inventory = _tag_inventory(current_user_id())
    else:
        tag_inventory = []
    # Default selection = top 5 most-shot *user* tags (auto-tags excluded
    # so the picker doesn't open with tournament:* pre-selected).
    default_top_tags = [t['name'] for t in tag_inventory
                        if not t['is_auto']][:5]
    # Equipment inventory: distinct bow models + arrow names the user has
    # logged. Reuses the predict helpers — they already return the
    # user's full set, sorted alphabetically.
    if any(v.get('equipment_picker') for v in REPORTS.values()):
        bow_inventory = _predict_user_bows(current_user_id())
        arrow_inventory = _predict_user_arrows(current_user_id())
    else:
        bow_inventory = []
        arrow_inventory = []
    return render_template('analyze.html',
                           catalog=catalog,
                           selected=selected,
                           date_ranges=date_ranges,
                           categories=categories,
                           tag_selections=tag_selections,
                           tag_inventory=tag_inventory,
                           default_top_tags=default_top_tags,
                           bow_selections=bow_selections,
                           arrow_selections=arrow_selections,
                           bow_inventory=bow_inventory,
                           arrow_inventory=arrow_inventory,
                           numeric_options={},
                           results=results,
                           error=error)


@app.route('/notes', methods=['GET', 'POST'])
@login_required
def notes():
    """Per-user scratchpad — free-form text not tied to any session.

    GET renders the current note; POST upserts the textarea contents.
    Stored as a single row per user in ``user_notes`` (see schema above).
    """
    user_id = current_user_id()
    saved = False
    error = None

    if request.method == 'POST':
        content = request.form.get('content') or ''
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                existing = cur.execute(
                    "SELECT id FROM user_notes WHERE user_id = %s",
                    (user_id,)
                ).fetchone()
                now = _app_now()
                if existing is None:
                    cur.execute(
                        "INSERT INTO user_notes (user_id, content, updated_at) "
                        "VALUES (%s, %s, %s)",
                        (user_id, content, now)
                    )
                else:
                    cur.execute(
                        "UPDATE user_notes SET content = %s, updated_at = %s "
                        "WHERE user_id = %s",
                        (content, now, user_id)
                    )
                con.commit()
            saved = True
        except SQLAlchemyError as e:
            print(f"❌ Notes save error: {e}")
            error = "Could not save notes — please try again."

    content = ''
    updated_at = None
    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT content, updated_at FROM user_notes WHERE user_id = %s",
                (user_id,)
            ).fetchone()
            if row is not None:
                content = row['content'] or ''
                updated_at = row['updated_at']
    except SQLAlchemyError as e:
        print(f"❌ Notes load error: {e}")
        if error is None:
            error = "Could not load notes."

    return render_template('notes.html',
                           content=content,
                           updated_at=updated_at,
                           saved=saved,
                           error=error)


@app.route('/notes/api', methods=['GET', 'POST'])
@login_required
def notes_api():
    """JSON load/save for the side-nav Notes popup.

    GET returns {ok, content, updated_at}. POST accepts JSON or form
    body with 'content' and upserts it.
    """
    user_id = current_user_id()

    if request.method == 'POST':
        if request.is_json:
            content = (request.get_json(silent=True) or {}).get('content', '')
        else:
            content = request.form.get('content', '')
        try:
            with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
                existing = cur.execute(
                    "SELECT id FROM user_notes WHERE user_id = %s",
                    (user_id,)
                ).fetchone()
                now = _app_now()
                if existing is None:
                    cur.execute(
                        "INSERT INTO user_notes (user_id, content, updated_at) "
                        "VALUES (%s, %s, %s)",
                        (user_id, content, now)
                    )
                else:
                    cur.execute(
                        "UPDATE user_notes SET content = %s, updated_at = %s "
                        "WHERE user_id = %s",
                        (content, now, user_id)
                    )
                con.commit()
            return jsonify(ok=True, content=content, updated_at=str(now))
        except SQLAlchemyError as e:
            print(f"❌ Notes API save error: {e}")
            return jsonify(ok=False, error="Could not save notes."), 500

    try:
        with closing(get_db_connection()) as con, closing(con.cursor()) as cur:
            row = cur.execute(
                "SELECT content, updated_at FROM user_notes WHERE user_id = %s",
                (user_id,)
            ).fetchone()
        content = (row['content'] or '') if row is not None else ''
        updated_at = str(row['updated_at']) if row is not None and row['updated_at'] else None
        return jsonify(ok=True, content=content, updated_at=updated_at)
    except SQLAlchemyError as e:
        print(f"❌ Notes API load error: {e}")
        return jsonify(ok=False, error="Could not load notes."), 500


@app.route('/analyze/export', methods=['GET'])
@login_required
def analyze_export():
    """Download one report's tabular data as CSV or Excel."""
    key = (request.args.get('report') or '').strip()
    fmt = (request.args.get('format') or '').strip().lower()
    if key not in REPORTS:
        return "Unknown report", 400
    if fmt not in ('csv', 'xlsx'):
        return "Format must be csv or xlsx", 400
    try:
        spec = REPORTS[key]
        kwargs = {}
        if spec.get('accepts_date_range'):
            kwargs['date_from'] = (request.args.get('date_from') or '').strip() or None
            kwargs['date_to']   = (request.args.get('date_to')   or '').strip() or None
        if spec.get('categories'):
            valid = {c['key'] for c in spec['categories']}
            chosen = [c for c in request.args.getlist('categories') if c in valid]
            # Always forward — even an empty list — so the export matches
            # what the rendered report showed (which also honored an
            # explicit empty selection).
            kwargs['categories'] = chosen if chosen else list(valid)
        if spec.get('tag_picker'):
            # Always forward the submitted tag list so the export matches
            # the rendered report; an empty list explicitly means "no tags".
            kwargs['tag_filter'] = request.args.getlist('tags')
        if spec.get('equipment_picker'):
            # Same shape as tags — empty list ≡ no filter ≡ include all.
            kwargs['bow_filter'] = request.args.getlist('bows')
            kwargs['arrow_filter'] = request.args.getlist('arrows')
        out = spec['fn'](current_user_id(), **kwargs)
    except ImportError:
        return "matplotlib is not installed on the server.", 500
    except SQLAlchemyError as e:
        print(f"❌ Analyze export read error: {e}")
        return "Database error while building report.", 500
    if out is None:
        return "No data to export for this report yet.", 404

    # Reports can return:
    #   * a single dict with `panels` (per-target style, head-to-head pre-split)
    #   * a single dict with columns/rows (single chart+table)
    #   * a list of dicts, each with `panels` (head-to-head per-kind)
    #   * a single dict flagged `empty: True`
    # Flatten everything into one list of (title, columns, rows) sections.
    #
    # Cells may be Markup-wrapped (data-tip spans in the head-to-head
    # tables, italic <em> section-header markers in the shots-per-target
    # table) — the HTML render leaves those alone, but Excel and CSV
    # readers shouldn't see raw tags. Strip them to plain text on the
    # way out.
    _tag_re = re.compile(r'<[^>]+>')

    def _plain(cell):
        if isinstance(cell, Markup):
            from html import unescape
            return unescape(_tag_re.sub('', str(cell)))
        return cell

    def _plain_row(r):
        return [_plain(c) for c in r]

    raw_items = out if isinstance(out, list) else [out]
    sections = []
    multi = False
    for item in raw_items:
        if item.get('empty'):
            continue
        if item.get('panels'):
            multi = True
            for p in item['panels']:
                # Interactive / chart-only panels can omit a table —
                # skip them in the export rather than emit empty sheets.
                if not p.get('columns'):
                    continue
                sections.append((
                    p['title'], _plain_row(p['columns']),
                    [_plain_row(r) for r in p['rows']],
                ))
        else:
            sections.append((
                item['title'], _plain_row(item['columns']),
                [_plain_row(r) for r in item['rows']],
            ))
    if not sections:
        return "No data to export for this report yet.", 404

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename_base = f'apollo_{key}_{timestamp}'

    if fmt == 'csv':
        sbuf = io.StringIO()
        w = csv.writer(sbuf)
        if multi:
            # Prepend a "Section" column so concatenated panels stay
            # readable in a single flat CSV.
            for i, (title, cols, rows) in enumerate(sections):
                if i == 0:
                    w.writerow(['Section'] + list(cols))
                for r in rows:
                    w.writerow([title] + list(r))
        else:
            title, cols, rows = sections[0]
            w.writerow(cols)
            for r in rows:
                w.writerow(r)
        return Response(sbuf.getvalue(), mimetype='text/csv', headers={
            'Content-Disposition': f'attachment; filename={filename_base}.csv'
        })

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()
    for title, cols, rows in sections:
        # Excel sheet names: 31 chars max, no [ ] : * ? / \. Dedup with a
        # numeric suffix when sanitization collides.
        clean = re.sub(r'[\[\]:*?/\\]', '_', title)[:31] or 'sheet'
        name = clean
        n = 2
        while name in used_names:
            suffix = f'_{n}'
            name = (clean[:31 - len(suffix)] + suffix)
            n += 1
        used_names.add(name)
        ws = wb.create_sheet(title=name)
        ws.append(list(cols))
        for r in rows:
            ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return Response(buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={filename_base}.xlsx'
        })


@app.route('/import_data', methods=['POST'])
@login_required
def import_data():
    """Accept a previously-exported file and merge its rows into the
    current user's data. Returns JSON so the modal can show the result
    inline — the templates don't render Flask flash messages."""
    fmt = (request.form.get('format') or '').strip().lower()
    if fmt not in ('sql', 'csv', 'xlsx'):
        return jsonify(ok=False,
                       error="Pick a format (sql, csv, or xlsx)."), 400

    upload = request.files.get('file')
    if upload is None or not upload.filename:
        return jsonify(ok=False, error="No file selected."), 400

    blob = upload.read()
    if not blob:
        return jsonify(ok=False, error="File is empty."), 400

    try:
        if fmt == 'sql':
            data = _parse_sql_import(blob.decode('utf-8', errors='replace'))
        elif fmt == 'csv':
            data = _parse_csv_zip_import(blob)
        else:
            data = _parse_xlsx_import(blob)
    except (zipfile.BadZipFile, ValueError, UnicodeDecodeError) as e:
        print(f"❌ Import parse error: {e}")
        return jsonify(ok=False,
                       error=f"Could not read the file ({e})."), 400
    except Exception as e:  # openpyxl raises its own subclasses
        print(f"❌ Import parse error: {e}")
        return jsonify(ok=False, error="Could not read the file."), 400

    try:
        counts = _apply_import(data, current_user_id())
    except SQLAlchemyError as e:
        print(f"❌ Import write error: {e}")
        return jsonify(ok=False,
                       error="Database error while inserting rows."), 500
    except (ValueError, TypeError, KeyError) as e:
        # _apply_import re-raises non-DB failures (e.g. the unsafe-table
        # guard, a bad coerced value) after rolling back. Surface a clean
        # JSON error instead of letting it escape as an unhandled 500.
        print(f"❌ Import parse/validation error: {e}")
        return jsonify(ok=False,
                       error="Import data was malformed and could not be applied."), 400

    total = sum(counts.values())
    return jsonify(ok=True, total=total, counts=counts)


if __name__ == "__main__":
    # Debug defaults ON for local dev convenience but is force-disabled
    # whenever FLASK_ENV=production, so the Werkzeug debugger (RCE risk)
    # can never ship to a real deployment by accident.
    debug_mode = os.environ.get('FLASK_DEBUG', '1') == '1' \
                 and os.environ.get('FLASK_ENV') != 'production'
    app.run(debug=debug_mode)